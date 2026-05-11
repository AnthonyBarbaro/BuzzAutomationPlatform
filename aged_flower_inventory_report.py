#!/usr/bin/env python3
"""
Build an aged flower inventory report from live Dutchie inventory.
"""

from __future__ import annotations

import argparse
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, datetime
from pathlib import Path
from typing import Any, Sequence
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dutchie_api_reports import (
    STORE_CODES,
    canonical_env_map,
    create_session,
    discover_configured_store_codes,
    request_json,
    resolve_integrator_key,
    resolve_store_keys,
)


BASE_DIR = Path(__file__).resolve().parent
DEFAULT_ENV_FILE = BASE_DIR / ".env"
DEFAULT_OUTPUT_ROOT = BASE_DIR / "reports" / "aged_inventory"
DEFAULT_TIMEZONE = "America/Los_Angeles"
DEFAULT_AGE_DAYS = 90
DEFAULT_WORKERS = 6
DEFAULT_DRIVE_PARENT_FOLDER = "aged_inventory"
CREDENTIALS_FILE = BASE_DIR / "credentials.json"
TOKEN_DRIVE_FILE = BASE_DIR / "token_drive.json"
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]
DATE_COLUMNS = ("packagedDate", "manufacturingDate", "testedDate", "sampleDate", "lastModifiedDateUtc")
FLOWER_CATEGORY_HINTS = (
    "flower",
    "eighth",
    "eighths",
    "quarter",
    "quarters",
    "half",
    "halves",
    "ounce",
    "ounces",
)
SAMPLE_PROMO_PATTERN = re.compile(r"\b(sample|samples|promo|promos|promotional|display|tester)\b", re.IGNORECASE)


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(value).lower())


def to_float(value: Any) -> float:
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def parse_dutchie_date(value: Any) -> pd.Timestamp | pd.NaT:
    if value in (None, ""):
        return pd.NaT
    parsed = pd.to_datetime(value, errors="coerce", utc=True)
    if pd.isna(parsed):
        return pd.NaT
    return parsed


def parse_store_codes(values: list[str] | None) -> list[str]:
    if not values:
        return []
    requested: list[str] = []
    for raw in values:
        for piece in str(raw).replace(",", " ").split():
            code = piece.strip().upper()
            if not code:
                continue
            if code not in STORE_CODES:
                allowed = ", ".join(STORE_CODES)
                raise ValueError(f"Unknown store code '{piece}'. Expected one of: {allowed}")
            if code not in requested:
                requested.append(code)
    return requested


def is_brand_match(row: dict[str, Any], aliases: list[str]) -> bool:
    brand_key = normalize_key(row.get("brandName"))
    alias_keys = {normalize_key(alias) for alias in aliases if normalize_key(alias)}
    if brand_key and brand_key in alias_keys:
        return True

    product_name = clean_text(row.get("productName"))
    product_key = product_name.casefold()
    normalized_product = normalize_key(product_name)
    for alias in aliases:
        alias_text = clean_text(alias)
        alias_key = normalize_key(alias_text)
        if not alias_key:
            continue
        if product_key.startswith(f"{alias_text.casefold()} |"):
            return True
        if normalized_product.startswith(alias_key):
            return True
    return False


def is_flower_row(row: dict[str, Any], include_prerolls: bool = False) -> bool:
    category = clean_text(row.get("category"))
    master_category = clean_text(row.get("masterCategory"))
    product_name = clean_text(row.get("productName"))
    category_key = category.casefold()
    master_key = master_category.casefold()
    product_key = product_name.casefold()

    if "preroll" in category_key or "pre-roll" in product_key or "pre roll" in product_key:
        return bool(include_prerolls)

    if master_key == "flower":
        return True
    if "| flower " in product_key or product_key.startswith("flower "):
        return True
    return any(hint in category_key for hint in FLOWER_CATEGORY_HINTS)


def is_sample_or_promo(row: dict[str, Any]) -> bool:
    text = " ".join(
        clean_text(row.get(column))
        for column in ("productName", "alternateName", "category", "tags", "packageStatus")
    )
    return bool(SAMPLE_PROMO_PATTERN.search(text))


def first_available_date(row: dict[str, Any]) -> tuple[pd.Timestamp | pd.NaT, str]:
    for column in DATE_COLUMNS:
        parsed = parse_dutchie_date(row.get(column))
        if not pd.isna(parsed):
            return parsed, column
    return pd.NaT, ""


def row_to_report_record(
    row: dict[str, Any],
    store_code: str,
    brand_label: str,
    as_of_day: date,
    age_days: int,
) -> dict[str, Any] | None:
    quantity = to_float(row.get("quantityAvailable"))
    if quantity <= 0:
        return None

    age_date, date_source = first_available_date(row)
    if pd.isna(age_date):
        return None

    age_day = age_date.date()
    days_old = (as_of_day - age_day).days
    if days_old < age_days:
        return None

    unit_cost = to_float(row.get("unitCost"))
    unit_price = to_float(row.get("unitPrice") or row.get("recUnitPrice") or row.get("medUnitPrice"))
    return {
        "Store": store_code,
        "Location": STORE_CODES.get(store_code, store_code),
        "Brand": clean_text(row.get("brandName")) or brand_label,
        "Product": clean_text(row.get("productName") or row.get("alternateName")),
        "Category": clean_text(row.get("category")),
        "Master Category": clean_text(row.get("masterCategory")),
        "Available": quantity,
        "Unit Cost": unit_cost,
        "Cost Value": round(quantity * unit_cost, 2),
        "Unit Price": unit_price,
        "Retail Value": round(quantity * unit_price, 2),
        "Packaged/Date Used": age_day.isoformat(),
        "Age Days": days_old,
        "Date Source": date_source,
        "SKU": clean_text(row.get("sku")),
        "Package ID": clean_text(row.get("packageId")),
        "External Package ID": clean_text(row.get("externalPackageId")),
        "Batch": clean_text(row.get("batchName") or row.get("batchId")),
        "Vendor": clean_text(row.get("vendor") or row.get("producer")),
        "Strain": clean_text(row.get("strain")),
        "Strain Type": clean_text(row.get("strainType")),
        "Package Status": clean_text(row.get("packageStatus")),
    }


def fetch_inventory_for_store(store_code: str, location_key: str, integrator_key: str) -> list[dict[str, Any]]:
    session = create_session(location_key, integrator_key)
    try:
        payload = request_json(
            session,
            "/reporting/inventory",
            params={
                "includeLabResults": False,
                "includeRoomQuantities": False,
                "includeAllocated": True,
                "includeLineage": False,
            },
        )
    finally:
        session.close()

    if not isinstance(payload, list):
        return []
    return [row for row in payload if isinstance(row, dict)]


def build_report_frame(
    store_codes: list[str],
    env_file: Path,
    brand_label: str,
    brand_aliases: list[str],
    age_days: int,
    as_of_day: date,
    include_prerolls: bool,
    include_samples: bool,
    workers: int,
) -> pd.DataFrame:
    env_map = canonical_env_map(env_file)
    store_keys = resolve_store_keys(env_map, store_codes)
    missing = [code for code in store_codes if code not in store_keys]
    if missing:
        missing_text = ", ".join(missing)
        raise RuntimeError(f"Missing Dutchie API location key(s) for: {missing_text}")

    integrator_key = resolve_integrator_key(env_map)
    records: list[dict[str, Any]] = []
    worker_count = max(1, min(int(workers or 1), len(store_codes)))

    def process_store(code: str) -> tuple[str, list[dict[str, Any]]]:
        rows = fetch_inventory_for_store(code, store_keys[code], integrator_key)
        store_records: list[dict[str, Any]] = []
        for row in rows:
            if not is_brand_match(row, brand_aliases):
                continue
            if not is_flower_row(row, include_prerolls=include_prerolls):
                continue
            if not include_samples and is_sample_or_promo(row):
                continue
            record = row_to_report_record(row, code, brand_label, as_of_day, age_days)
            if record:
                store_records.append(record)
        return code, store_records

    if worker_count == 1:
        for store_code in store_codes:
            print(f"[FETCH] {store_code} inventory")
            _, store_records = process_store(store_code)
            print(f"[FILTER] {store_code}: {len(store_records)} aged row(s)")
            records.extend(store_records)
    else:
        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            futures = {executor.submit(process_store, code): code for code in store_codes}
            for future in as_completed(futures):
                code = futures[future]
                print(f"[FETCH] {code} inventory")
                _, store_records = future.result()
                print(f"[FILTER] {code}: {len(store_records)} aged row(s)")
                records.extend(store_records)

    frame = pd.DataFrame(records)
    if frame.empty:
        return pd.DataFrame(columns=report_columns())

    frame = frame[report_columns()].sort_values(
        by=["Age Days", "Store", "Product", "Package ID"],
        ascending=[False, True, True, True],
    )
    return frame.reset_index(drop=True)


def report_columns() -> list[str]:
    return [
        "Store",
        "Location",
        "Brand",
        "Product",
        "Category",
        "Master Category",
        "Available",
        "Unit Cost",
        "Cost Value",
        "Unit Price",
        "Retail Value",
        "Packaged/Date Used",
        "Age Days",
        "Date Source",
        "SKU",
        "Package ID",
        "External Package ID",
        "Batch",
        "Vendor",
        "Strain",
        "Strain Type",
        "Package Status",
    ]


def build_summary(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(columns=["Store", "Location", "Rows", "Available", "Cost Value", "Retail Value", "Oldest Age Days"])

    summary = (
        frame.groupby(["Store", "Location"], as_index=False)
        .agg(
            Rows=("Product", "count"),
            Available=("Available", "sum"),
            **{
                "Cost Value": ("Cost Value", "sum"),
                "Retail Value": ("Retail Value", "sum"),
                "Oldest Age Days": ("Age Days", "max"),
            },
        )
        .sort_values(["Cost Value", "Available"], ascending=[False, False])
    )
    return summary


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="12302A")
    header_font = Font(color="FFFFFF", bold=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.auto_filter.ref = ws.dimensions

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_len = max(len(str(cell.value or "")) for cell in column_cells[:200])
            ws.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 42)

        for header in ws[1]:
            if header.value in {"Unit Cost", "Cost Value", "Unit Price", "Retail Value"}:
                for cell in ws[get_column_letter(header.column)][1:]:
                    cell.number_format = "$#,##0.00"
            elif header.value in {"Available", "Age Days", "Rows", "Oldest Age Days"}:
                for cell in ws[get_column_letter(header.column)][1:]:
                    cell.number_format = "#,##0.##"

    wb.save(path)


def write_report(frame: pd.DataFrame, output_root: Path, as_of_day: date, brand_label: str, age_days: int) -> tuple[Path, Path]:
    report_dir = output_root / as_of_day.isoformat()
    report_dir.mkdir(parents=True, exist_ok=True)
    safe_brand = re.sub(r"[^A-Za-z0-9]+", "_", brand_label).strip("_").lower()
    stem = f"{safe_brand}_aged_flower_inventory_{age_days}d_{as_of_day.isoformat()}"
    xlsx_path = report_dir / f"{stem}.xlsx"
    csv_path = report_dir / f"{stem}.csv"

    summary = build_summary(frame)
    frame.to_csv(csv_path, index=False)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        frame.to_excel(writer, sheet_name="Aged Inventory", index=False)

    style_workbook(xlsx_path)
    return xlsx_path, csv_path


def authenticate_drive_api():
    import google.auth.transport.requests
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    creds = None
    if TOKEN_DRIVE_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_DRIVE_FILE), DRIVE_SCOPES)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            if not CREDENTIALS_FILE.exists():
                raise RuntimeError(f"{CREDENTIALS_FILE.name} not found. Cannot upload to Google Drive.")
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        TOKEN_DRIVE_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)


def escape_drive_query_text(value: str) -> str:
    return str(value).replace("\\", "\\\\").replace("'", "\\'")


def find_or_create_drive_folder(service, folder_name: str, parent_id: str | None = None) -> str:
    escaped_name = escape_drive_query_text(folder_name)
    query = (
        "mimeType='application/vnd.google-apps.folder' "
        f"and name='{escaped_name}' and trashed=false"
    )
    if parent_id:
        query += f" and '{parent_id}' in parents"

    response = service.files().list(q=query, spaces="drive", fields="files(id,name)", pageSize=10).execute()
    folders = response.get("files", [])
    if folders:
        return folders[0]["id"]

    metadata = {"name": folder_name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        metadata["parents"] = [parent_id]
    created = service.files().create(body=metadata, fields="id").execute()
    return created["id"]


def make_drive_item_public(service, file_id: str) -> str | None:
    try:
        service.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"},
            fields="id",
        ).execute()
    except Exception as exc:
        print(f"[WARN] Could not make Drive item public: {exc}")

    try:
        info = service.files().get(fileId=file_id, fields="webViewLink").execute()
        return info.get("webViewLink")
    except Exception as exc:
        print(f"[WARN] Could not read Drive link: {exc}")
        return None


def upload_or_update_drive_file(service, file_path: Path, folder_id: str) -> str:
    from googleapiclient.http import MediaFileUpload

    escaped_name = escape_drive_query_text(file_path.name)
    query = f"name='{escaped_name}' and '{folder_id}' in parents and trashed=false"
    response = service.files().list(q=query, spaces="drive", fields="files(id,name)", pageSize=10).execute()
    matches = response.get("files", [])
    media = MediaFileUpload(str(file_path), resumable=True)

    if matches:
        file_id = matches[0]["id"]
        service.files().update(fileId=file_id, media_body=media, fields="id").execute()
        return file_id

    metadata = {"name": file_path.name, "parents": [folder_id]}
    uploaded = service.files().create(body=metadata, media_body=media, fields="id").execute()
    return uploaded["id"]


def upload_reports_to_drive(
    report_paths: Sequence[Path],
    drive_parent_folder: str,
    as_of_day: date,
    brand_label: str,
    make_public: bool,
) -> dict[str, str]:
    service = authenticate_drive_api()
    parent_id = find_or_create_drive_folder(service, drive_parent_folder)
    date_id = find_or_create_drive_folder(service, as_of_day.isoformat(), parent_id=parent_id)
    brand_id = find_or_create_drive_folder(service, brand_label, parent_id=date_id)

    links: dict[str, str] = {
        "folder": f"https://drive.google.com/drive/folders/{brand_id}",
    }
    if make_public:
        folder_link = make_drive_item_public(service, brand_id)
        if folder_link:
            links["folder"] = folder_link

    for path in report_paths:
        file_id = upload_or_update_drive_file(service, path, brand_id)
        link = make_drive_item_public(service, file_id) if make_public else None
        links[path.name] = link or f"https://drive.google.com/file/d/{file_id}/view"
        print(f"[DRIVE] Uploaded {path.name}: {links[path.name]}")

    return links


def write_drive_links(report_dir: Path, links: dict[str, str]) -> Path:
    links_path = report_dir / "drive_links.txt"
    with links_path.open("w", encoding="utf-8") as handle:
        for name, link in links.items():
            handle.write(f"{name}: {link}\n")
    return links_path


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build aged flower inventory report from Dutchie API inventory.")
    parser.add_argument("--env-file", default=str(DEFAULT_ENV_FILE), help="Path to .env with Dutchie API keys.")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT), help="Root folder for report output.")
    parser.add_argument("--stores", nargs="*", help="Store codes to include. Defaults to configured stores.")
    parser.add_argument("--brand", required=True, help="Brand to match, e.g. Hashish, 710 Labs, Jeeter.")
    parser.add_argument(
        "--brand-alias",
        action="append",
        dest="brand_aliases",
        help="Extra brand alias to match. Repeatable. Example: --brand '710 Labs' --brand-alias 710",
    )
    parser.add_argument("--age-days", type=int, default=DEFAULT_AGE_DAYS, help="Minimum package age in days. Default: 90.")
    parser.add_argument("--as-of-date", help="Age cutoff date in YYYY-MM-DD. Defaults to today in local timezone.")
    parser.add_argument("--timezone", default=DEFAULT_TIMEZONE, help=f"Local timezone. Default: {DEFAULT_TIMEZONE}.")
    parser.add_argument("--include-prerolls", action="store_true", help="Include prerolls whose master category is Flower.")
    parser.add_argument("--include-samples", action="store_true", help="Include sample/promo/tester rows. Excluded by default.")
    parser.add_argument("--workers", type=int, default=DEFAULT_WORKERS, help="Concurrent store fetch workers.")
    parser.add_argument(
        "--drive-folder",
        default=DEFAULT_DRIVE_PARENT_FOLDER,
        help=f"Google Drive parent folder for uploads. Default: {DEFAULT_DRIVE_PARENT_FOLDER}",
    )
    parser.add_argument("--no-drive-upload", action="store_true", help="Write local files only; skip Google Drive upload.")
    parser.add_argument("--private-drive", action="store_true", help="Do not make uploaded Drive folder/files public.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    env_file = Path(args.env_file)
    if not env_file.is_absolute():
        env_file = BASE_DIR / env_file

    env_map = canonical_env_map(env_file)
    requested_stores = parse_store_codes(args.stores)
    store_codes = requested_stores or discover_configured_store_codes(env_map) or list(STORE_CODES)

    if args.as_of_date:
        as_of_day = datetime.fromisoformat(args.as_of_date).date()
    else:
        as_of_day = datetime.now(ZoneInfo(args.timezone)).date()

    brand_aliases = list(args.brand_aliases or [])
    if args.brand not in brand_aliases:
        brand_aliases.insert(0, args.brand)

    print(
        f"[INFO] Building {args.brand} aged flower inventory report: "
        f"stores={','.join(store_codes)}, age_days>={args.age_days}, as_of={as_of_day}"
    )

    frame = build_report_frame(
        store_codes=store_codes,
        env_file=env_file,
        brand_label=args.brand,
        brand_aliases=brand_aliases,
        age_days=args.age_days,
        as_of_day=as_of_day,
        include_prerolls=args.include_prerolls,
        include_samples=args.include_samples,
        workers=args.workers,
    )
    xlsx_path, csv_path = write_report(
        frame=frame,
        output_root=Path(args.output_root),
        as_of_day=as_of_day,
        brand_label=args.brand,
        age_days=args.age_days,
    )

    print(f"[SAVED] Excel: {xlsx_path}")
    print(f"[SAVED] CSV:   {csv_path}")
    if not args.no_drive_upload:
        links = upload_reports_to_drive(
            report_paths=[xlsx_path, csv_path],
            drive_parent_folder=args.drive_folder,
            as_of_day=as_of_day,
            brand_label=args.brand,
            make_public=not args.private_drive,
        )
        links_path = write_drive_links(xlsx_path.parent, links)
        print(f"[DRIVE] Folder: {links.get('folder')}")
        print(f"[DRIVE] Links file: {links_path}")
    print(f"[DONE] {len(frame)} aged inventory row(s), {frame['Available'].sum() if not frame.empty else 0:g} unit(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
