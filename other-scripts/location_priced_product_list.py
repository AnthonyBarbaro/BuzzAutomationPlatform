#!/usr/bin/env python3
"""
Generate a location-priced product list from Dutchie catalog CSV/XLSX exports.

Default input:
    files/MM-DD-YYYY_<STORE>.csv

Default output:
    reports/generated/location_priced_product_list_<timestamp>.xlsx

Only true overrides are shown: Location Price must be set and must differ from
the regular Price. The workbook includes one clean product-price sheet per
store. Sample/promo rows and penny-priced rows are hidden by default.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

DEFAULT_INPUT_DIR = REPO_ROOT / "files"
DEFAULT_OUTPUT_ROOT = REPO_ROOT / "reports" / "generated"
DEFAULT_SHEET_URL_FILE = REPO_ROOT / "location_priced_product_list_sheet_url.txt"
DEFAULT_SHEET_TITLE = "Buzz Location Price List"
DEFAULT_INPUT_GLOB = "[0-9][0-9]-[0-9][0-9]-[0-9][0-9][0-9][0-9]_*.csv"
DEFAULT_STORE_ORDER = ["MV", "LG", "LM", "WP", "SV", "NC"]
DEFAULT_MIN_LOCATION_PRICE = 1.0
SHEET_URL_ENV = "LOCATION_PRICED_PRODUCT_LIST_SHEET_URL"
DEFAULT_MIN_COST = 1.0
STORE_NAMES = {
    "MV": "Mission Valley",
    "LG": "Lemon Grove",
    "LM": "La Mesa",
    "WP": "Wildomar Palomar",
    "SV": "Sorrento Valley",
    "NC": "National City",
}
STORE_PRICE_COLUMNS = [
    "Location",
    "Product",
    "Location Price",
    "Regular Price",
    "Difference",
    "Brand",
    "Category",
    "Variant Count",
    "Available",
    "Cost",
]
GOOGLE_COLUMN_PIXEL_WIDTHS = {
    "Location": 150,
    "Product": 380,
    "Location Price": 135,
    "Regular Price": 135,
    "Difference": 120,
    "Brand": 155,
    "Category": 155,
    "Variant Count": 135,
    "Available": 120,
    "Cost": 160,
}

DATE_STORE_RE = re.compile(r"^(?P<date>\d{2}-\d{2}-\d{4})_(?P<store>[A-Za-z]{2,3})$")
BAD_TEXT_VALUES = {"", "nan", "none", "null", "<na>"}
SAMPLE_PROMO_RE = re.compile(r"\b(?:sample|promo)\b", re.IGNORECASE)

BRAND_COLS = ["Brand", "brandName", "Brand Name"]
PRODUCT_COLS = ["Product", "Product Name", "productName", "Item", "Item Name"]
CATEGORY_COLS = ["Category", "Product Category", "Product Category Name", "masterCategory"]
COST_COLS = ["Cost", "Unit Cost", "unitCost", "Last Wholesale Cost", "Inventory Cost"]
LOCATION_PRICE_COLS = ["Location price", "Location Price", "location price", "location_price"]
PRICE_COLS = ["Price", "Retail Price", "MSRP", "unitPrice", "recUnitPrice", "medUnitPrice"]
AVAILABLE_COLS = ["Available", "Quantity on Hand", "quantityAvailable", "On Hand", "Qty", "Quantity"]
SKU_COLS = ["SKU", "sku", "UPC/GTIN", "UPC", "Barcode"]
STORE_COLS = ["Store", "Store Code", "Location Code"]
LOCATION_COLS = ["Location", "Location Name"]


def _clean_text(value: object) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value).strip()
    if text.casefold() in BAD_TEXT_VALUES:
        return ""
    return text


def _norm_key(value: object) -> str:
    return re.sub(r"\s+", " ", _clean_text(value)).casefold()


def _first_present_column(df: pd.DataFrame, candidates: Iterable[str]) -> str | None:
    exact = {str(col).strip(): col for col in df.columns}
    folded = {str(col).strip().casefold(): col for col in df.columns}

    for candidate in candidates:
        if candidate in exact:
            return exact[candidate]
        match = folded.get(candidate.casefold())
        if match is not None:
            return match
    return None


def _to_number(series: pd.Series | None, index: pd.Index) -> pd.Series:
    if series is None:
        return pd.Series(float("nan"), index=index, dtype="float64")
    if series.dtype == object:
        cleaned = series.astype(str).str.replace(r"[$,]", "", regex=True)
    else:
        cleaned = series
    return pd.to_numeric(cleaned, errors="coerce")


def _store_from_filename(path: Path) -> str:
    match = DATE_STORE_RE.match(path.stem)
    if match:
        return match.group("store").upper()

    parts = [part.strip().upper() for part in re.split(r"[_\-\s]+", path.stem) if part.strip()]
    for part in reversed(parts):
        if part in DEFAULT_STORE_ORDER:
            return part
    return path.stem


def _location_label(store: object) -> str:
    store_text = _clean_text(store)
    return STORE_NAMES.get(store_text.upper(), store_text)


def _date_from_catalog_filename(path: Path) -> datetime | None:
    match = DATE_STORE_RE.match(path.stem)
    if not match:
        return None
    try:
        return datetime.strptime(match.group("date"), "%m-%d-%Y")
    except ValueError:
        return None


def _select_latest_catalog_files(paths: list[Path]) -> list[Path]:
    dated = [(path, _date_from_catalog_filename(path)) for path in paths]
    dated = [(path, dt) for path, dt in dated if dt is not None]
    if not dated:
        return paths

    latest = max(dt for _, dt in dated)
    latest_paths = {path for path, dt in dated if dt == latest}
    undated_paths = {path for path in paths if _date_from_catalog_filename(path) is None}
    return sorted(latest_paths | undated_paths)


def _read_file(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, low_memory=False)
    return pd.read_excel(path)


def _sheet_safe_value(value: object) -> object:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, (pd.Timestamp, datetime)):
        return value.isoformat(sep=" ", timespec="seconds")
    return value


def _quote_sheet_name(title: str) -> str:
    return "'" + str(title).replace("'", "''") + "'"


def _sheet_all_range(title: str) -> str:
    return _quote_sheet_name(title)


def _sheet_start_range(title: str) -> str:
    return f"{_quote_sheet_name(title)}!A1"


def _read_sheet_url_file(path: Path) -> str:
    if not path.exists():
        return ""
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if line and not line.startswith("#"):
            return line
    return ""


def resolve_sheet_target(sheet_url: str | None, sheet_url_file: Path) -> str:
    explicit = str(sheet_url or "").strip()
    if explicit:
        return explicit

    env_value = os.environ.get(SHEET_URL_ENV, "").strip()
    if env_value:
        return env_value

    return _read_sheet_url_file(sheet_url_file)


def _looks_like_catalog_file(path: Path, df: pd.DataFrame) -> bool:
    if DATE_STORE_RE.match(path.stem):
        return True
    return _first_present_column(df, LOCATION_PRICE_COLS) is not None


def _extract_store_values(df: pd.DataFrame, path: Path) -> pd.Series:
    store_col = _first_present_column(df, STORE_COLS)
    if store_col:
        stores = df[store_col].map(_clean_text)
        return stores.mask(stores == "", _store_from_filename(path)).astype(str).str.upper()

    location_col = _first_present_column(df, LOCATION_COLS)
    if location_col:
        locations = df[location_col].map(_clean_text)
        fallback = _store_from_filename(path)
        return locations.mask(locations == "", fallback)

    return pd.Series(_store_from_filename(path), index=df.index)


def extract_rows_from_file(
    path: Path,
    *,
    include_zero_cost: bool,
    include_zero_price: bool,
    require_location_price_file: bool,
) -> pd.DataFrame:
    df = _read_file(path)
    if df is None or df.empty:
        return pd.DataFrame()

    df = df.loc[:, ~df.columns.astype(str).str.contains(r"^Unnamed", case=False, regex=True)].copy()
    if require_location_price_file and not _looks_like_catalog_file(path, df):
        return pd.DataFrame()

    brand_col = _first_present_column(df, BRAND_COLS)
    product_col = _first_present_column(df, PRODUCT_COLS)
    category_col = _first_present_column(df, CATEGORY_COLS)
    cost_col = _first_present_column(df, COST_COLS)
    loc_price_col = _first_present_column(df, LOCATION_PRICE_COLS)
    price_col = _first_present_column(df, PRICE_COLS)
    available_col = _first_present_column(df, AVAILABLE_COLS)
    sku_col = _first_present_column(df, SKU_COLS)

    required = {
        "brand": brand_col,
        "product": product_col,
        "category": category_col,
        "cost": cost_col,
    }
    missing = [name for name, col in required.items() if col is None]
    if missing or (loc_price_col is None and price_col is None):
        return pd.DataFrame()

    index = df.index
    base_price = _to_number(df[price_col] if price_col else None, index)
    location_price = _to_number(df[loc_price_col] if loc_price_col else None, index)
    use_location = location_price.notna() & (location_price > 0)
    is_location_override = use_location & base_price.notna() & (location_price.round(2) != base_price.round(2))
    price_used = base_price.where(~use_location, location_price)
    if price_col is None:
        price_used = location_price

    out = pd.DataFrame(
        {
            "Store": _extract_store_values(df, path).map(_clean_text),
            "Location": _extract_store_values(df, path).map(_location_label),
            "Brand": df[brand_col].map(_clean_text),
            "Product": df[product_col].map(_clean_text),
            "Product Family": df[product_col].map(_product_family_name),
            "Category": df[category_col].map(_clean_text),
            "Cost": _to_number(df[cost_col], index).round(2),
            "Location Price": location_price.round(2),
            "Base Price": base_price.round(2),
            "Price Difference": (location_price - base_price).round(2),
            "Price Used": price_used.round(2),
            "Available": _to_number(df[available_col] if available_col else None, index),
            "SKU": df[sku_col].map(_clean_text) if sku_col else "",
            "Source File": path.name,
            "Price Source": use_location.map({True: "Location Price", False: "Price"}),
            "Has Location Price": use_location,
            "Has Location Override": is_location_override,
        }
    )

    out = out[
        (out["Store"] != "")
        & (out["Brand"] != "")
        & (out["Product"] != "")
        & (out["Category"] != "")
        & out["Cost"].notna()
        & out["Price Used"].notna()
    ].copy()

    if not include_zero_cost:
        out = out[out["Cost"] >= DEFAULT_MIN_COST].copy()
    if not include_zero_price:
        out = out[out["Price Used"] > 0].copy()

    return out.reset_index(drop=True)


def _mode_text(values: Iterable[object]) -> str:
    cleaned = [_clean_text(value) for value in values]
    cleaned = [value for value in cleaned if value]
    if not cleaned:
        return ""
    counts = Counter(cleaned)
    return counts.most_common(1)[0][0]


def _sort_key_for_store(store: str, store_order: list[str]) -> tuple[int, str]:
    store_upper = store.upper()
    if store_upper in store_order:
        return (store_order.index(store_upper), store_upper)
    return (len(store_order), store_upper)


def _unique_sorted_text(values: Iterable[object]) -> list[str]:
    seen = {_clean_text(value) for value in values}
    return sorted(value for value in seen if value)


def _join_text(values: Iterable[object], sep: str = ", ") -> str:
    return sep.join(_unique_sorted_text(values))


def _money(value: float) -> str:
    return f"${value:,.2f}"


def _money_list(values: Iterable[object]) -> str:
    nums = pd.to_numeric(pd.Series(list(values)), errors="coerce").dropna().round(2)
    unique = sorted(set(float(value) for value in nums))
    return ", ".join(_money(value) for value in unique)


def _money_range(values: Iterable[object]) -> str:
    nums = pd.to_numeric(pd.Series(list(values)), errors="coerce").dropna().round(2)
    unique = sorted(set(float(value) for value in nums))
    if not unique:
        return ""
    if len(unique) == 1:
        return _money(unique[0])
    return f"{_money(unique[0])} - {_money(unique[-1])}"


def _display_product_name(product_names: list[str]) -> str:
    if not product_names:
        return ""
    if len(product_names) == 1:
        return product_names[0]
    return f"{product_names[0]} (+{len(product_names) - 1} more)"


def _product_family_name(product_name: object) -> str:
    """
    Collapse Dutchie product names to the reusable price-list family.
    Example:
        CAM | Flower 3.5G | H | 92 OG -> CAM | Flower 3.5G
    """
    text = _clean_text(product_name)
    if not text:
        return ""

    parts = [part.strip() for part in text.split("|") if part.strip()]
    if len(parts) >= 2:
        return " | ".join(parts[:2])
    return text


def _remove_sample_promo_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "Product" not in df.columns:
        return df
    return df[~df["Product"].astype(str).str.contains(SAMPLE_PROMO_RE, na=False)].copy()


def build_store_order(raw: pd.DataFrame) -> list[str]:
    discovered = _unique_sorted_text(raw.get("Store", pd.Series(dtype=object)))
    ordered = list(DEFAULT_STORE_ORDER)
    known = {store.casefold() for store in ordered}
    for store in sorted(discovered, key=lambda value: _sort_key_for_store(value, DEFAULT_STORE_ORDER)):
        if store.casefold() not in known:
            ordered.append(store)
            known.add(store.casefold())
    return ordered


def build_store_price_sheet(raw: pd.DataFrame, store: str) -> pd.DataFrame:
    if raw is None or raw.empty:
        return pd.DataFrame(columns=STORE_PRICE_COLUMNS)

    store_rows = raw[raw["Store"].astype(str).str.casefold() == store.casefold()].copy()
    if "Has Location Override" in store_rows.columns:
        store_rows = store_rows[store_rows["Has Location Override"].fillna(False)].copy()
    if store_rows.empty:
        return pd.DataFrame(columns=STORE_PRICE_COLUMNS)

    store_rows["_brand_key"] = store_rows["Brand"].map(_norm_key)
    store_rows["_category_key"] = store_rows["Category"].map(_norm_key)
    store_rows["_product_family_key"] = store_rows["Product Family"].map(_norm_key)
    store_rows["_price_key"] = pd.to_numeric(store_rows["Location Price"], errors="coerce").round(2)
    store_rows["_base_price_key"] = pd.to_numeric(store_rows["Base Price"], errors="coerce").round(2)

    rows: list[dict[str, object]] = []
    group_cols = ["_product_family_key", "_price_key", "_base_price_key", "_brand_key", "_category_key"]
    for (_, price, base_price, _, _), grp in store_rows.groupby(group_cols, dropna=False, sort=False):
        available = pd.to_numeric(grp["Available"], errors="coerce").fillna(0).sum()
        product_names = _unique_sorted_text(grp["Product"])
        rows.append(
            {
                "Location": _location_label(store),
                "Product": _mode_text(grp["Product Family"]),
                "Location Price": float(price) if pd.notna(price) else None,
                "Regular Price": float(base_price) if pd.notna(base_price) else None,
                "Difference": float(price - base_price) if pd.notna(price) and pd.notna(base_price) else None,
                "Brand": _mode_text(grp["Brand"]),
                "Category": _mode_text(grp["Category"]),
                "Variant Count": len(product_names),
                "Available": float(available),
                "Cost": _money_list(grp["Cost"]),
            }
        )

    out = pd.DataFrame(rows)
    if out.empty:
        return pd.DataFrame(columns=STORE_PRICE_COLUMNS)

    out.sort_values(["Brand", "Category", "Product", "Location Price"], inplace=True, na_position="last")
    return out.reindex(columns=STORE_PRICE_COLUMNS).reset_index(drop=True)


def build_all_store_price_rows(raw: pd.DataFrame, store_order: list[str]) -> pd.DataFrame:
    frames = [build_store_price_sheet(raw, store) for store in store_order]
    frames = [frame for frame in frames if not frame.empty]
    if not frames:
        return pd.DataFrame(columns=STORE_PRICE_COLUMNS)
    return pd.concat(frames, ignore_index=True)


def build_summary(raw: pd.DataFrame, output_rows: pd.DataFrame, input_files: list[Path]) -> pd.DataFrame:
    stores = _join_text(raw["Store"]) if "Store" in raw.columns else ""
    brands = int(raw["Brand"].nunique()) if "Brand" in raw.columns else 0
    categories = int(raw["Category"].nunique()) if "Category" in raw.columns else 0
    return pd.DataFrame(
        [
            {"Metric": "Generated At", "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
            {"Metric": "Input Files", "Value": len(input_files)},
            {"Metric": "Raw Rows", "Value": len(raw)},
            {"Metric": "Output Rows", "Value": len(output_rows)},
            {"Metric": "Brands", "Value": brands},
            {"Metric": "Categories", "Value": categories},
            {"Metric": "Stores", "Value": stores},
            {"Metric": "Input File Names", "Value": ", ".join(path.name for path in input_files)},
        ]
    )


def dataframe_to_sheet_values(df: pd.DataFrame, generated_at: str) -> list[list[object]]:
    frame = df.copy()
    headers = list(frame.columns)
    rows: list[list[object]] = [
        ["Updated At", generated_at],
        [],
        headers,
    ]
    for values in frame.itertuples(index=False, name=None):
        rows.append([_sheet_safe_value(value) for value in values])
    return rows


def build_google_readme_values(
    generated_at: str,
    input_files: list[Path],
    output_rows: pd.DataFrame,
    store_order: list[str],
) -> list[list[object]]:
    return [
        ["Buzz Location Price List", ""],
        ["Updated At", generated_at],
        ["Input Files", len(input_files)],
        ["Store Tab Rows", len(output_rows)],
        ["Stores", ", ".join(store_order)],
        [
            "What Is Included",
            "Only products where Location Price is set and differs from regular Price.",
        ],
        ["Default Exclusions", "Sample/promo rows, penny-priced rows, and cost below $1.00."],
        ["How To Use", "Use each store tab as the price guide for products that need a Location Price."],
        ["Input File Names", ", ".join(path.name for path in input_files)],
    ]


def get_sheet_info_by_title(service: object, spreadsheet_id: str, title: str) -> dict[str, object] | None:
    metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    for sheet in metadata.get("sheets", []):
        properties = sheet.get("properties", {})
        if properties.get("title") == title:
            return properties
    return None


def ensure_google_sheet_tab(service: object, spreadsheet_id: str, title: str) -> dict[str, object]:
    existing = get_sheet_info_by_title(service, spreadsheet_id, title)
    if existing:
        return existing

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "addSheet": {
                        "properties": {
                            "title": title,
                            "gridProperties": {"rowCount": 200, "columnCount": 30},
                        }
                    }
                }
            ]
        },
    ).execute()

    created = get_sheet_info_by_title(service, spreadsheet_id, title)
    if not created:
        raise RuntimeError(f"Unable to create Google Sheets tab: {title}")
    return created


def delete_google_sheet_tab(service: object, spreadsheet_id: str, title: str) -> None:
    metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
    sheets = metadata.get("sheets", [])
    target = None
    for sheet in sheets:
        properties = sheet.get("properties", {})
        if properties.get("title") == title:
            target = properties
            break
    if not target or len(sheets) <= 1:
        return

    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"deleteSheet": {"sheetId": int(target["sheetId"])}}]},
    ).execute()


def create_location_price_spreadsheet(service: object, title: str) -> tuple[str, str]:
    result = (
        service.spreadsheets()
        .create(
            body={"properties": {"title": title}},
            fields="spreadsheetId,spreadsheetUrl",
        )
        .execute()
    )
    spreadsheet_id = str(result.get("spreadsheetId", "")).strip()
    spreadsheet_url = str(result.get("spreadsheetUrl", "")).strip()
    if not spreadsheet_id:
        raise RuntimeError("Google Sheets did not return a spreadsheet ID.")
    if not spreadsheet_url:
        spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
    return spreadsheet_id, spreadsheet_url


def write_sheet_url_file(path: Path, spreadsheet_url: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(
            [
                "# Google Sheet used by other-scripts/location_priced_product_list.py",
                "# This file is reused by cron so each run updates the same spreadsheet.",
                spreadsheet_url,
                "",
            ]
        ),
        encoding="utf-8",
    )


def google_column_width_requests(sheet_id: int, headers: list[object]) -> list[dict[str, object]]:
    requests: list[dict[str, object]] = []
    for index, header in enumerate(headers):
        pixel_size = GOOGLE_COLUMN_PIXEL_WIDTHS.get(_clean_text(header))
        if not pixel_size:
            continue
        requests.append(
            {
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": sheet_id,
                        "dimension": "COLUMNS",
                        "startIndex": index,
                        "endIndex": index + 1,
                    },
                    "properties": {"pixelSize": pixel_size},
                    "fields": "pixelSize",
                }
            }
        )
    return requests


def format_google_sheet_tab(
    service: object,
    spreadsheet_id: str,
    sheet_id: int,
    rows: list[list[object]],
    *,
    data_tab: bool,
) -> None:
    total_rows = max(len(rows), 1)
    total_columns = max(max((len(row) for row in rows), default=1), 1)
    header_row_index = 2 if data_tab else 0

    requests: list[dict[str, object]] = [
        {
            "updateSheetProperties": {
                "properties": {
                    "sheetId": sheet_id,
                    "gridProperties": {
                        "frozenRowCount": 3 if data_tab else 1,
                        "rowCount": max(total_rows + 25, 100),
                        "columnCount": max(total_columns + 5, 20),
                    },
                },
                "fields": "gridProperties.frozenRowCount,gridProperties.rowCount,gridProperties.columnCount",
            }
        },
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": header_row_index,
                    "endRowIndex": header_row_index + 1,
                    "startColumnIndex": 0,
                    "endColumnIndex": total_columns,
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": {"red": 0.12, "green": 0.16, "blue": 0.22},
                        "textFormat": {"foregroundColor": {"red": 1, "green": 1, "blue": 1}, "bold": True},
                        "horizontalAlignment": "CENTER",
                        "verticalAlignment": "MIDDLE",
                    }
                },
                "fields": "userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)",
            }
        },
        {
            "autoResizeDimensions": {
                "dimensions": {
                    "sheetId": sheet_id,
                    "dimension": "COLUMNS",
                    "startIndex": 0,
                    "endIndex": total_columns,
                }
            }
        },
    ]
    if data_tab and len(rows) > 2:
        requests.extend(google_column_width_requests(sheet_id, rows[2]))

    if data_tab and total_rows > 3:
        requests.append(
            {
                "setBasicFilter": {
                    "filter": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": 2,
                            "endRowIndex": total_rows,
                            "startColumnIndex": 0,
                            "endColumnIndex": total_columns,
                        }
                    }
                }
            }
        )

    service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body={"requests": requests}).execute()


def update_google_sheet_tab(
    service: object,
    spreadsheet_id: str,
    title: str,
    rows: list[list[object]],
    *,
    data_tab: bool,
) -> None:
    sheet_info = ensure_google_sheet_tab(service, spreadsheet_id, title)
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id,
        range=_sheet_all_range(title),
        body={},
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=_sheet_start_range(title),
        valueInputOption="USER_ENTERED",
        body={"values": rows},
    ).execute()
    format_google_sheet_tab(
        service=service,
        spreadsheet_id=spreadsheet_id,
        sheet_id=int(sheet_info["sheetId"]),
        rows=rows,
        data_tab=data_tab,
    )


def sync_outputs_to_google_sheets(
    *,
    raw: pd.DataFrame,
    output_rows: pd.DataFrame,
    input_files: list[Path],
    store_order: list[str],
    generated_at: str,
    sheet_url: str,
    sheet_url_file: Path,
    sheet_title: str,
) -> tuple[str, str]:
    from deals_brand_config_sync import _parse_sheet_target, authenticate_sheets

    service = authenticate_sheets()
    target = str(sheet_url or "").strip()
    if target:
        spreadsheet_id, _gid = _parse_sheet_target(target) if "/spreadsheets/d/" in target else (target, None)
        spreadsheet_url = (
            target
            if "/spreadsheets/d/" in target
            else f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
        )
    else:
        spreadsheet_id, spreadsheet_url = create_location_price_spreadsheet(service, sheet_title)
        write_sheet_url_file(sheet_url_file, spreadsheet_url)

    delete_google_sheet_tab(service, spreadsheet_id, "Reference")
    delete_google_sheet_tab(service, spreadsheet_id, "Merged Products")

    update_google_sheet_tab(
        service,
        spreadsheet_id,
        "README",
        build_google_readme_values(generated_at, input_files, output_rows, store_order),
        data_tab=False,
    )

    for store in store_order:
        store_sheet = build_store_price_sheet(raw, store)
        update_google_sheet_tab(
            service,
            spreadsheet_id,
            store[:31],
            dataframe_to_sheet_values(store_sheet, generated_at),
            data_tab=True,
        )

    return spreadsheet_id, spreadsheet_url


def _format_workbook(path: Path, store_order: list[str]) -> None:
    wb = load_workbook(path)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    money_headers = {
        "Cost",
        "Location Price",
        "Regular Price",
        "Base Price",
        "Price Used",
        "Price Difference",
        "Difference",
    }
    number_headers = {"Total Available", "Available", "Store Count", "Variant Count"}
    number_headers.update({f"{store} Available" for store in store_order})
    text_wrap_headers = {"Location", "Product", "Stores", "Source Files", "Input File Names"}
    text_wrap_headers.update({f"{store} Price" for store in store_order})
    tab_colors = {
        "Raw Rows": "059669",
        "Summary": "7C3AED",
    }
    for idx, store in enumerate(store_order):
        tab_colors[store[:31]] = ["0F766E", "0369A1", "B45309", "BE123C", "4D7C0F", "6D28D9"][idx % 6]

    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.sheet_properties.tabColor = tab_colors.get(ws.title, "6B7280")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        headers = {cell.value: cell.column for cell in ws[1] if cell.value}

        # Keep the large Raw Rows sheet light. Styling every raw cell can make
        # openpyxl runs crawl on full-store exports, while filters/widths still
        # give the sheet the important usability wins.
        if ws.title != "Raw Rows":
            for row_idx in range(2, ws.max_row + 1):
                row_fill = stripe_fill if row_idx % 2 == 0 else None
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if row_fill is not None:
                        cell.fill = row_fill
                    header = ws.cell(row=1, column=col_idx).value
                    if header in text_wrap_headers:
                        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    elif header in money_headers:
                        cell.number_format = '"$"#,##0.00'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif header in number_headers:
                        cell.number_format = '#,##0.##'
                        cell.alignment = Alignment(horizontal="right", vertical="center")

        preferred_widths = {
            "Brand": 22,
            "Category": 20,
            "Location": 18,
            "Product": 48,
            "Variant Count": 16,
            "Stores": 28,
            "Total Available": 14,
            "Available": 14,
            "Location Price": 17,
            "Regular Price": 17,
            "Difference": 14,
            "Price Range": 18,
            "All Prices": 28,
            "Location Prices": 20,
            "Store Count": 12,
            "Cost": 16,
            "Source Files": 38,
            "Metric": 20,
            "Value": 90,
        }
        for store in store_order:
            preferred_widths[f"{store} Price"] = 18
            preferred_widths[f"{store} Available"] = 14
            preferred_widths[f"{store} Location Price"] = 18

        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            header = ws.cell(row=1, column=col_idx).value
            if header in preferred_widths:
                ws.column_dimensions[letter].width = preferred_widths[header]
                continue

            max_len = 0
            for row_idx in range(1, min(ws.max_row, 250) + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 38)

    wb.save(path)


def write_outputs(
    raw: pd.DataFrame,
    output_rows: pd.DataFrame,
    input_files: list[Path],
    output_root: Path,
    *,
    output_name: str | None,
    include_raw: bool,
    write_csv: bool,
    store_order: list[str],
) -> list[Path]:
    output_root.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = output_name or f"location_priced_product_list_{timestamp}"
    stem = Path(stem).stem

    xlsx_path = output_root / f"{stem}.xlsx"
    summary = build_summary(raw, output_rows, input_files)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        for store in store_order:
            store_sheet = build_store_price_sheet(raw, store)
            store_sheet.to_excel(writer, sheet_name=store[:31], index=False)
        if include_raw:
            raw.to_excel(writer, sheet_name="Raw Rows", index=False)
    _format_workbook(xlsx_path, store_order)

    written = [xlsx_path]
    if write_csv:
        csv_path = output_root / f"{stem}.csv"
        output_rows.to_csv(csv_path, index=False)
        written.append(csv_path)

    return written


def find_input_files(input_dir: Path, input_glob: str, all_matching: bool) -> list[Path]:
    files = sorted(
        path
        for path in input_dir.glob(input_glob)
        if path.is_file() and path.suffix.lower() in {".csv", ".xlsx", ".xls"}
    )
    if not all_matching:
        files = _select_latest_catalog_files(files)
    return files


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build store tabs of products whose Location Price differs from "
            "the regular Price."
        )
    )
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument(
        "--input-glob",
        default=DEFAULT_INPUT_GLOB,
        help=f"Input file glob relative to --input-dir. Default: {DEFAULT_INPUT_GLOB}",
    )
    parser.add_argument("--output-root", type=Path, default=DEFAULT_OUTPUT_ROOT)
    parser.add_argument("--output-name", help="Optional output filename stem.")
    parser.add_argument(
        "--sync-sheets",
        action="store_true",
        help=(
            "Update Google Sheets. If no sheet URL is configured, a new spreadsheet is created "
            "and saved to --sheet-url-file."
        ),
    )
    parser.add_argument(
        "--no-sync-sheets",
        action="store_true",
        help="Skip Google Sheets even if a sheet URL file or environment variable is configured.",
    )
    parser.add_argument(
        "--sheet-url",
        help=f"Google Sheets URL or spreadsheet ID. Overrides {SHEET_URL_ENV} and --sheet-url-file.",
    )
    parser.add_argument(
        "--sheet-url-file",
        type=Path,
        default=DEFAULT_SHEET_URL_FILE,
        help=f"File containing the stable Google Sheets URL. Default: {DEFAULT_SHEET_URL_FILE}",
    )
    parser.add_argument(
        "--sheet-title",
        default=DEFAULT_SHEET_TITLE,
        help=f"Title to use when creating a new Google Spreadsheet. Default: {DEFAULT_SHEET_TITLE}",
    )
    parser.add_argument(
        "--all-matching",
        action="store_true",
        help="Use every matching file instead of only the latest dated catalog set.",
    )
    parser.add_argument(
        "--include-non-catalog",
        action="store_true",
        help="Allow files without a Location price column or MM-DD-YYYY_STORE filename.",
    )
    parser.add_argument(
        "--include-zero-cost",
        action="store_true",
        help=f"Include penny/zero-cost products. Hidden by default with min cost {DEFAULT_MIN_COST:.2f}.",
    )
    parser.add_argument("--include-zero-price", action="store_true")
    parser.add_argument(
        "--min-location-price",
        type=float,
        default=DEFAULT_MIN_LOCATION_PRICE,
        help=(
            "Minimum Location Price to include after override filtering. "
            f"Default: {DEFAULT_MIN_LOCATION_PRICE:.2f}"
        ),
    )
    parser.add_argument(
        "--include-samples",
        action="store_true",
        help="Include sample/promo products. Hidden by default.",
    )
    parser.add_argument("--no-raw", action="store_true", help="Do not include the Raw Rows sheet.")
    parser.add_argument("--csv", action="store_true", help="Also write the combined store tabs as CSV.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_dir = args.input_dir.expanduser().resolve()
    output_root = args.output_root.expanduser().resolve()

    if not input_dir.exists():
        raise SystemExit(f"Input directory not found: {input_dir}")

    input_files = find_input_files(input_dir, args.input_glob, args.all_matching)
    if not input_files:
        raise SystemExit(f"No input files found in {input_dir} matching {args.input_glob!r}")

    frames: list[pd.DataFrame] = []
    skipped: list[str] = []
    for path in input_files:
        try:
            frame = extract_rows_from_file(
                path,
                include_zero_cost=bool(args.include_zero_cost),
                include_zero_price=bool(args.include_zero_price),
                require_location_price_file=not bool(args.include_non_catalog),
            )
        except Exception as exc:
            skipped.append(f"{path.name}: {exc}")
            continue

        if frame.empty:
            skipped.append(f"{path.name}: no usable product rows")
            continue
        frames.append(frame)

    if not frames:
        detail = "\n".join(f"- {item}" for item in skipped)
        raise SystemExit(f"No usable rows found.\n{detail}")

    raw = pd.concat(frames, ignore_index=True)
    if not bool(args.include_samples):
        raw = _remove_sample_promo_rows(raw)
    if "Has Location Override" in raw.columns:
        raw = raw[raw["Has Location Override"].fillna(False)].copy()
    raw = raw[pd.to_numeric(raw["Location Price"], errors="coerce").fillna(0) >= float(args.min_location_price)].copy()
    if raw.empty:
        raise SystemExit("No products were found where Location Price differs from regular Price.")

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    store_order = build_store_order(raw)
    output_rows = build_all_store_price_rows(raw, store_order)
    written = write_outputs(
        raw,
        output_rows,
        input_files,
        output_root,
        output_name=args.output_name,
        include_raw=not bool(args.no_raw),
        write_csv=bool(args.csv),
        store_order=store_order,
    )

    sheet_url_file = Path(args.sheet_url_file).expanduser().resolve()
    sheet_target = resolve_sheet_target(args.sheet_url, sheet_url_file)
    should_sync_sheets = not bool(args.no_sync_sheets) and (bool(args.sync_sheets) or bool(sheet_target))
    if should_sync_sheets:
        spreadsheet_id, spreadsheet_url = sync_outputs_to_google_sheets(
            raw=raw,
            output_rows=output_rows,
            input_files=input_files,
            store_order=store_order,
            generated_at=generated_at,
            sheet_url=sheet_target,
            sheet_url_file=sheet_url_file,
            sheet_title=str(args.sheet_title or DEFAULT_SHEET_TITLE),
        )
        print(f"[SHEETS] Updated: {spreadsheet_url}")
        if not sheet_target:
            print(f"[SHEETS] Saved sheet URL to: {sheet_url_file}")

    print(f"Processed {len(input_files)} file(s), {len(raw):,} raw row(s), {len(output_rows):,} output row(s).")
    if skipped:
        print("Skipped:")
        for item in skipped:
            print(f"- {item}")
    print("Wrote:")
    for path in written:
        print(f"- {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
