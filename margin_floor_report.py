#!/usr/bin/env python3
"""
Brand Margin Floor Report

Reads product files (CSV/XLSX), separates output by brand, and reports
which products are below a target margin floor.

Scenarios included:
1) 30% discount (no back)
2) 30% discount + 10% back

For each product, this script calculates:
- Margin for each scenario
- Required shelf price to hit target margin
- How much to raise the current shelf price

Outputs are saved under done/ by default.
"""

from __future__ import annotations

import argparse
import re
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


LOCATION_PRICE_ALIASES = [
    "Location price",
    "Location Price",
    "location price",
    "location_price",
]

BRAND_COLS = ["Brand"]
PRODUCT_COLS = ["Product", "Product Name", "Item", "Item Name"]
CATEGORY_COLS = ["Category", "Major Category", "Product Category", "Product Category Name"]
PRICE_COLS = ["Price", "Retail Price", "MSRP"]
COST_COLS = ["Cost", "Inventory Cost", "COGS", "Cost of Goods Sold", "Unit Cost"]
AVAILABLE_COLS = ["Available", "On Hand", "Inventory", "Qty", "Quantity"]
MIN_PRICE_USED = 1.00


def _first_present_column(df: pd.DataFrame, candidates: Iterable[str]) -> Optional[str]:
    for col in candidates:
        if col in df.columns:
            return col
    return None


def _to_num(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce")


def _safe_filename(text: str) -> str:
    return re.sub(r'[\\/*?:"<>|]+', "_", str(text or "")).strip() or "Unknown"


def _store_from_filename(path: Path) -> str:
    base = path.stem
    parts = base.split("_")
    return parts[-1].strip() if len(parts) > 1 else base


def _read_file(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, low_memory=False)
    return pd.read_excel(path)


def _merge_label_cells(values: Iterable[object]) -> str:
    """
    Merge comma-separated labels from multiple rows into one unique, sorted list.
    Example: "MV, LM" + "LM, SV" -> "LM, MV, SV"
    """
    items = set()
    for val in values:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            continue
        for part in str(val).split(","):
            p = part.strip()
            if p:
                items.add(p)
    return ", ".join(sorted(items))


def _extract_rows_from_file(path: Path, min_available: float, exclude_accessories: bool) -> pd.DataFrame:
    df = _read_file(path)
    if df is None or df.empty:
        return pd.DataFrame()

    brand_col = _first_present_column(df, BRAND_COLS)
    product_col = _first_present_column(df, PRODUCT_COLS)
    category_col = _first_present_column(df, CATEGORY_COLS)
    price_col = _first_present_column(df, PRICE_COLS)
    cost_col = _first_present_column(df, COST_COLS)
    available_col = _first_present_column(df, AVAILABLE_COLS)
    loc_price_col = _first_present_column(df, LOCATION_PRICE_ALIASES)

    required = [brand_col, product_col, cost_col]
    if any(col is None for col in required):
        return pd.DataFrame()

    price = _to_num(df[price_col]) if price_col else pd.Series(index=df.index, dtype="float64")
    loc_price = _to_num(df[loc_price_col]) if loc_price_col else pd.Series(index=df.index, dtype="float64")
    use_loc = loc_price.notna() & (loc_price > 0)
    price_used = pd.Series(float("nan"), index=df.index, dtype="float64")
    if price_col:
        price_used = price.where(~use_loc, loc_price)
    else:
        price_used = loc_price

    cost = _to_num(df[cost_col])
    available = _to_num(df[available_col]) if available_col else pd.Series(pd.NA, index=df.index)
    category = df[category_col] if category_col else pd.Series("", index=df.index)

    out = pd.DataFrame({
        "Source File": path.name,
        "Store": _store_from_filename(path),
        "Brand": df[brand_col].astype(str).str.strip(),
        "Product": df[product_col].astype(str).str.strip(),
        "Category": category.astype(str).str.strip(),
        "Available": available,
        "Price": price,
        "Location Price": loc_price,
        "Price Used": price_used,
        "Cost": cost,
        "Price Source": use_loc.map({True: loc_price_col or "Location price", False: price_col or "Price"}),
    })

    # Basic validity filters
    out = out[
        out["Brand"].notna()
        & (out["Brand"] != "")
        & out["Product"].notna()
        & (out["Product"] != "")
        & out["Price Used"].notna()
        & (out["Price Used"] > MIN_PRICE_USED)
        & out["Cost"].notna()
        & (out["Cost"] > 0)
    ].copy()

    # Always exclude zero/negative inventory rows when availability exists.
    if available_col:
        out = out[out["Available"].fillna(0) > 0].copy()
        if min_available > 0:
            out = out[out["Available"] >= float(min_available)].copy()

    if exclude_accessories:
        out = out[~out["Category"].str.contains(r"accessor", case=False, na=False)].copy()

    return out


def _add_margin_scenarios(
    df: pd.DataFrame,
    target_margin: float,
    discount_rate: float,
    back_rate: float,
) -> pd.DataFrame:
    d = df.copy()

    rate_discount_only = 1.0 - discount_rate
    rate_discount_plus_back = rate_discount_only * (1.0 - back_rate)

    shelf = _to_num(d["Price Used"])
    cost = _to_num(d["Cost"])
    rev_30 = shelf * rate_discount_only
    rev_30_back = shelf * rate_discount_plus_back

    d["Shelf Price"] = shelf
    d["Price (30% Off)"] = rev_30
    d["Price (30% Off + 10% Back)"] = rev_30_back
    d["Margin 30% Off"] = (rev_30 - cost) / rev_30
    d["Margin 30% Off + 10% Back"] = (rev_30_back - cost) / rev_30_back

    target_price_30 = np.ceil(cost / ((1.0 - target_margin) * rate_discount_only))
    target_price_30_back = cost / ((1.0 - target_margin) * rate_discount_plus_back)

    d["Target Shelf Price @ 40% (30% Off)"] = target_price_30
    d["Target Shelf Price @ 40% (30% Off + 10% Back)"] = target_price_30_back
    d["Raise Needed (30% Off)"] = (target_price_30 - shelf).clip(lower=0)
    d["Raise Needed (30% Off + 10% Back)"] = (target_price_30_back - shelf).clip(lower=0)
    d["Raise Needed (Worst Case)"] = d[
        ["Raise Needed (30% Off)", "Raise Needed (30% Off + 10% Back)"]
    ].max(axis=1)

    d["Below 40% (30% Off)"] = d["Margin 30% Off"] < target_margin
    d["Below 40% (30% Off + 10% Back)"] = d["Margin 30% Off + 10% Back"] < target_margin
    # "Needs Raise" should be based on the 30%-off margin check only.
    d["Needs Raise"] = d["Below 40% (30% Off)"]

    return d


def _consolidate_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Two-step consolidation for easier reading:
    1) Exact pass: collapse duplicate rows across stores for identical
       brand + product + category + price + cost.
    2) Similar-product pass (modeled after margin_report.py):
       within each brand/category/price/cost group, merge product variants into
       one row with "Product (+N more)" and full "Product List".
    """
    if df is None or df.empty:
        return df

    exact_keys = ["Brand", "Product", "Category", "Price Used", "Cost", "Price Source"]
    exact = (
        df.groupby(exact_keys, dropna=False, as_index=False)
        .agg(
            **{
                "Stores": ("Store", _merge_label_cells),
                "Source Files": ("Source File", _merge_label_cells),
                "Available": ("Available", "sum"),
                "Price": ("Price", "first"),
                "Location Price": ("Location Price", "first"),
            }
        )
    )

    similar_keys = ["Brand", "Category", "Price Used", "Cost", "Price Source"]
    merged_rows = []
    for _, grp in exact.groupby(similar_keys, dropna=False):
        row = grp.iloc[0].copy()

        names = sorted({str(x).strip() for x in grp["Product"].dropna() if str(x).strip()})
        n = len(names)
        if n == 0:
            display_name = ""
        elif n == 1:
            display_name = names[0]
        else:
            display_name = f"{names[0]} (+{n - 1} more)"

        row["Product"] = display_name
        row["Product List"] = "; ".join(names)
        row["Merged Count"] = int(n)
        row["Stores"] = _merge_label_cells(grp["Stores"])
        row["Source Files"] = _merge_label_cells(grp["Source Files"])
        row["Available"] = pd.to_numeric(grp["Available"], errors="coerce").fillna(0).sum()

        merged_rows.append(row)

    out = pd.DataFrame(merged_rows)
    if not out.empty:
        out = out[out["Available"].fillna(0) > 0].copy()
        out.sort_values(
            by=["Brand", "Category", "Price Used", "Cost", "Product"],
            inplace=True,
            na_position="last",
        )
        out.reset_index(drop=True, inplace=True)
    return out


def _ordered_columns(df: pd.DataFrame) -> pd.DataFrame:
    preferred = [
        "Brand",
        "Product",
        "Merged Count",
        "Product List",
        "Category",
        "Stores",
        "Cost",
        "Shelf Price",
        "Price (30% Off)",
        "Price (30% Off + 10% Back)",
        "Available",
        "Margin 30% Off",
        "Margin 30% Off + 10% Back",
        "Below 40% (30% Off)",
        "Below 40% (30% Off + 10% Back)",
        "Needs Raise",
        "Target Shelf Price @ 40% (30% Off)",
        "Target Shelf Price @ 40% (30% Off + 10% Back)",
        "Raise Needed (30% Off)",
        "Raise Needed (30% Off + 10% Back)",
        "Raise Needed (Worst Case)",
    ]
    cols = [c for c in preferred if c in df.columns]
    return df[cols]


def _is_truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if value is None:
        return False
    return str(value).strip().lower() in {"true", "yes", "y", "1"}


def _format_excel(path: Path) -> None:
    wb = load_workbook(path)

    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    danger_row_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            continue

        # Tab colors for quick scan
        if ws.title == "Overview":
            ws.sheet_properties.tabColor = "1D4ED8"
        elif ws.title == "Needs Raise":
            ws.sheet_properties.tabColor = "DC2626"
        else:
            ws.sheet_properties.tabColor = "059669"

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

        # Header styling
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # Base row styling + zebra
        for row_idx in range(2, max_row + 1):
            row_fill = stripe_fill if row_idx % 2 == 0 else None
            for col_idx in range(1, max_col + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.border = thin_border
                if row_fill is not None:
                    c.fill = row_fill

        ws.row_dimensions[1].height = 22

        header_map = {cell.value: cell.column_letter for cell in ws[1] if cell.value}
        text_cols = [
            "Brand",
            "Product",
            "Product List",
            "Category",
            "Stores",
        ]
        pct_cols = [
            "Margin 30% Off",
            "Margin 30% Off + 10% Back",
        ]
        money_cols = [
            "Cost",
            "Shelf Price",
            "Price (30% Off)",
            "Price (30% Off + 10% Back)",
            "Target Shelf Price @ 40% (30% Off)",
            "Target Shelf Price @ 40% (30% Off + 10% Back)",
            "Raise Needed (30% Off)",
            "Raise Needed (30% Off + 10% Back)",
            "Raise Needed (Worst Case)",
        ]
        int_cols = ["Available", "Merged Count", "Products", "Below 40% (30% Off)", "Below 40% (30% Off + 10% Back)"]

        # Text wrapping/alignment for long fields
        for name in text_cols:
            letter = header_map.get(name)
            if not letter:
                continue
            for cell in ws[letter][1:]:
                if cell.value is not None:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Integer columns
        for name in int_cols:
            letter = header_map.get(name)
            if not letter:
                continue
            for cell in ws[letter][1:]:
                if cell.value is not None:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # Percent columns
        for name in pct_cols:
            letter = header_map.get(name)
            if not letter:
                continue
            for cell in ws[f"{letter}"][1:]:
                if cell.value is not None:
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")

            # Heatmap for margin columns: red -> yellow -> green
            if max_row >= 2:
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="FCA5A5",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FDE68A",
                        end_type="max",
                        end_color="86EFAC",
                    ),
                )

        # Currency columns
        for name in money_cols:
            letter = header_map.get(name)
            if not letter:
                continue
            for cell in ws[f"{letter}"][1:]:
                if cell.value is not None:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        # Heatmap for raise columns
        for name in ("Raise Needed (30% Off)", "Raise Needed (30% Off + 10% Back)", "Raise Needed (Worst Case)"):
            letter = header_map.get(name)
            if not letter:
                continue
            if max_row >= 2:
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color="DCFCE7",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FDE68A",
                        end_type="max",
                        end_color="FCA5A5",
                    ),
                )

        # Highlight full row when Needs Raise = True
        needs_col = header_map.get("Needs Raise")
        if needs_col:
            for row_idx in range(2, max_row + 1):
                needs_value = ws[f"{needs_col}{row_idx}"].value
                if _is_truthy(needs_value):
                    for col_idx in range(1, max_col + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = danger_row_fill

        # Column widths with sensible caps
        preferred_widths = {
            "Brand": 20,
            "Product": 42,
            "Product List": 60,
            "Category": 22,
            "Stores": 28,
            "Shelf Price": 12,
            "Price (30% Off)": 14,
            "Price (30% Off + 10% Back)": 20,
            "Cost": 12,
            "Available": 12,
            "Merged Count": 12,
            "Needs Raise": 12,
            "Overview": 18,
        }

        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            header = ws.cell(row=1, column=col_idx).value
            if isinstance(header, str) and header in preferred_widths:
                ws.column_dimensions[letter].width = preferred_widths[header]
                continue

            max_len = 0
            for row_idx in range(1, max_row + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                if value is not None:
                    max_len = max(max_len, len(str(value)))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 42)

    wb.save(path)


def _write_reports(df: pd.DataFrame, run_dir: Path) -> None:
    all_products = df.sort_values(
        by=["Brand", "Needs Raise", "Raise Needed (30% Off)"],
        ascending=[True, False, False],
    ).reset_index(drop=True)
    all_products = _ordered_columns(all_products)
    low_margin = all_products[all_products["Needs Raise"]].copy()

    overview = (
        all_products.groupby("Brand", dropna=False)
        .agg(
            products=("Product", "count"),
            low_30=("Below 40% (30% Off)", "sum"),
            low_30_back=("Below 40% (30% Off + 10% Back)", "sum"),
            avg_margin_30=("Margin 30% Off", "mean"),
            avg_margin_30_back=("Margin 30% Off + 10% Back", "mean"),
            avg_raise_needed=("Raise Needed (30% Off)", "mean"),
        )
        .reset_index()
        .rename(
            columns={
                "Brand": "Brand",
                "products": "Products",
                "low_30": "Below 40% (30% Off)",
                "low_30_back": "Below 40% (30% Off + 10% Back)",
                "avg_margin_30": "Avg Margin (30% Off)",
                "avg_margin_30_back": "Avg Margin (30% Off + 10% Back)",
                "avg_raise_needed": "Avg Raise Needed (30% Off)",
            }
        )
        .sort_values("Below 40% (30% Off)", ascending=False)
    )

    summary_path = run_dir / "margin_floor_summary.xlsx"
    with pd.ExcelWriter(summary_path, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="Overview", index=False)
        low_margin.to_excel(writer, sheet_name="Needs Raise", index=False)
        all_products.to_excel(writer, sheet_name="All Products", index=False)
    _format_excel(summary_path)

    brand_dir = run_dir / "brand_reports"
    brand_dir.mkdir(parents=True, exist_ok=True)
    for brand, bdf in all_products.groupby("Brand", dropna=False):
        safe_brand = _safe_filename(brand if pd.notna(brand) else "Unknown")
        out_path = brand_dir / f"{safe_brand}_margin_floor.xlsx"
        needs_raise = bdf[bdf["Needs Raise"]].copy()
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            bdf.to_excel(writer, sheet_name="All Products", index=False)
            needs_raise.to_excel(writer, sheet_name="Needs Raise", index=False)
        _format_excel(out_path)


def _parse_args() -> argparse.Namespace:
    script_dir = Path(__file__).resolve().parent
    parser = argparse.ArgumentParser(description="Build brand-separated margin floor reports.")
    parser.add_argument("--input-dir", type=Path, default=script_dir / "files")
    parser.add_argument("--output-root", type=Path, default=script_dir / "done")
    parser.add_argument("--target-margin", type=float, default=0.40, help="Margin floor (0.40 = 40%%).")
    parser.add_argument("--discount-rate", type=float, default=0.30, help="Base discount rate.")
    parser.add_argument("--back-rate", type=float, default=0.10, help="Back rate after discount.")
    parser.add_argument("--min-available", type=float, default=0.0, help="Optional minimum available qty.")
    parser.add_argument(
        "--exclude-accessories",
        action="store_true",
        help="Exclude products where category contains 'accessor'.",
    )
    return parser.parse_args()


def main() -> None:
    args = _parse_args()
    input_dir: Path = args.input_dir
    output_root: Path = args.output_root

    if not input_dir.exists():
        raise SystemExit(f"Input directory not found: {input_dir}")

    files = sorted(
        [p for p in input_dir.iterdir() if p.suffix.lower() in {".csv", ".xlsx", ".xls"} and p.is_file()]
    )
    if not files:
        raise SystemExit(f"No CSV/XLSX files found in: {input_dir}")

    frames = []
    for path in files:
        try:
            out = _extract_rows_from_file(
                path,
                min_available=args.min_available,
                exclude_accessories=args.exclude_accessories,
            )
        except Exception as exc:
            print(f"[WARN] Skipped {path.name}: {exc}")
            continue
        if out is not None and not out.empty:
            frames.append(out)

    if not frames:
        raise SystemExit("No usable rows after parsing/filtering.")

    combined = pd.concat(frames, ignore_index=True)
    combined = _consolidate_rows(combined)
    combined = _add_margin_scenarios(
        combined,
        target_margin=float(args.target_margin),
        discount_rate=float(args.discount_rate),
        back_rate=float(args.back_rate),
    )

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = output_root / f"margin_floor_reports_{ts}"
    run_dir.mkdir(parents=True, exist_ok=True)

    _write_reports(combined, run_dir)

    print(f"Done. Reports saved to: {run_dir}")
    print("Includes:")
    print("- margin_floor_summary.xlsx")
    print("- brand_reports/<BRAND>_margin_floor.xlsx")
    print("")
    print("Key checks:")
    print("- Below 40% (30% Off)")
    print("- Needs Raise (based on 30% Off only)")
    print("- Raise Needed (30% Off)")


if __name__ == "__main__":
    main()
