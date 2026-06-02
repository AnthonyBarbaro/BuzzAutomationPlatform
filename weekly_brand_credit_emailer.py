#!/usr/bin/env python3
"""
Send weekly brand credit emails for the brands that need both deal report links
and inventory folder links in the same message.

Workflow modes:
1. Automatic mode pulls last week's sales/inventory exports, generates only the
   Hashish and Treesap reports, uploads them to Drive by brand/week, then emails.
2. Existing-link mode emails the latest already-generated deal and inventory links.
"""

import argparse
import base64
import datetime
import json
import os
import re
import sys
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from email.utils import formatdate
from pathlib import Path
from xml.sax.saxutils import escape

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


BASE_DIR = Path(__file__).resolve().parent
CREDENTIALS_FILE = "credentials.json"
TOKEN_GMAIL_FILE = "token_gmail.json"
TOKEN_DRIVE_FILE = "token_drive.json"
DEFAULT_REPORTS_DIR = "brand_reports"
DEFAULT_LINKS_FILE = "links.txt"
DEFAULT_INVENTORY_LINKS_FILE = os.path.join("inventory_links", "latest.json")
DEFAULT_AUTO_OUTPUT_ROOT = os.path.join("reports", "weekly_brand_credit")
DEFAULT_WEEKLY_DRIVE_PARENT = "Weekly Brand Credit Reports"
DEFAULT_INVENTORY_INPUT_DIR = "files"
DEFAULT_SALES_SOURCE = "api"
DEFAULT_API_ENV_FILE = str(BASE_DIR / ".env")
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]
BUZZ_CC = ["joseph@buzzcannabis.com", "donna@buzzcannabis.com"]
TEST_MODE_EMAIL = "anthony@buzzcannabis.com"
STORE_LABELS = {
    "MV": "Mission Valley",
    "LG": "Lemon Grove",
    "LM": "La Mesa",
    "WP": "Wildomar Palomar",
    "SV": "Sorrento Valley",
    "NC": "National City",
}
HASHISH_SECTION_CATEGORY_ORDER = {
    "Disposables": 80,
    "Pre-Rolls": 90,
    "Wellness": 100,
}

WEEKLY_BRAND_EMAILS = [
    {
        "brand": "Hashish",
        "report_aliases": ["Hashish"],
        "inventory_folder": "Hashish",
        "inventory_aliases": ["Hashish"],
        "drive_folder": "Hashish",
        "to": ["ryanbtcventures@gmail.com"],
    },
    {
        "brand": "TreeSap",
        "report_aliases": ["TreeSap", "Treesap"],
        "inventory_folder": "Treesap",
        "inventory_aliases": ["Treesap", "TreeSap"],
        "drive_folder": "Treesap",
        "to": ["sales@treesapsyrup.com"],
    },
]


def normalize_key(value):
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def safe_filename_part(value):
    safe = re.sub(r"[^A-Za-z0-9._ -]+", "_", str(value or "").strip())
    safe = re.sub(r"\s+", " ", safe).strip(" ._-")
    return safe or "file"


def emit_status(message, status_callback=None):
    if status_callback:
        status_callback(message)
    else:
        print(message)


def resolve_repo_path(path_value):
    path = Path(path_value)
    if path.is_absolute():
        return path
    return BASE_DIR / path


def coerce_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        return datetime.date.fromisoformat(value)
    raise TypeError(f"Unsupported date value: {value!r}")


def get_previous_monday_sunday(reference_date=None):
    today = coerce_date(reference_date) if reference_date else datetime.date.today()
    monday_this_week = today - datetime.timedelta(days=today.weekday())
    last_monday = monday_this_week - datetime.timedelta(days=7)
    last_sunday = last_monday + datetime.timedelta(days=6)
    return last_monday, last_sunday


def week_key_for_range(start_date, end_date):
    start_day = coerce_date(start_date)
    end_day = coerce_date(end_date)
    return f"{start_day.isoformat()}_to_{end_day.isoformat()}"


def selected_brand_configs(selected_brands=None):
    return [
        brand_cfg
        for brand_cfg in WEEKLY_BRAND_EMAILS
        if should_include_brand(brand_cfg, selected_brands)
    ]


def label_for_brand_config(brand_cfg):
    return brand_cfg.get("drive_folder") or brand_cfg.get("inventory_folder") or brand_cfg["brand"]


def capture_printed_status(func, status_callback=None):
    """
    Run a noisy legacy helper while forwarding its print output through the GUI/CLI log.
    """
    import contextlib
    import io

    buffer = io.StringIO()
    result = None
    error = None
    with contextlib.redirect_stdout(buffer):
        try:
            result = func()
        except Exception as exc:
            error = exc

    for raw_line in buffer.getvalue().splitlines():
        line = raw_line.strip()
        if line:
            emit_status(line, status_callback)

    if error:
        raise error
    return result


def call_in_base_dir(func):
    current_dir = os.getcwd()
    os.chdir(BASE_DIR)
    try:
        return func()
    finally:
        os.chdir(current_dir)


def gmail_authenticate():
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    scopes = ["https://www.googleapis.com/auth/gmail.send"]
    creds = None
    token_path = resolve_repo_path(TOKEN_GMAIL_FILE)
    credentials_path = resolve_repo_path(CREDENTIALS_FILE)
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), scopes)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), scopes)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def send_email_with_gmail_html(
    subject,
    html_body,
    recipients,
    cc_recipients=None,
    attachments=None,
    dry_run=False,
    status_callback=None,
):
    if isinstance(recipients, str):
        recipients = [recipients]
    cc_recipients = list(cc_recipients or [])
    attachments = list(attachments or [])

    if dry_run:
        emit_status(f"[DRY RUN] Subject: {subject}", status_callback)
        emit_status(f"[DRY RUN] To: {', '.join(recipients)}", status_callback)
        if cc_recipients:
            emit_status(f"[DRY RUN] Cc: {', '.join(cc_recipients)}", status_callback)
        if attachments:
            emit_status(f"[DRY RUN] Attachments: {attachments}", status_callback)
        return

    service = gmail_authenticate()

    msg = MIMEMultipart("alternative")
    msg["From"] = "me"
    msg["To"] = ", ".join(recipients)
    if cc_recipients:
        msg["Cc"] = ", ".join(cc_recipients)
    msg["Date"] = formatdate(localtime=True)
    msg["Subject"] = subject
    msg.attach(MIMEText(html_body, "html"))

    for file_path in attachments:
        if not os.path.isfile(file_path):
            continue
        with open(file_path, "rb") as fp:
            file_data = fp.read()
        part = MIMEBase("application", "octet-stream")
        part.set_payload(file_data)
        encoders.encode_base64(part)
        part.add_header(
            "Content-Disposition",
            f'attachment; filename="{os.path.basename(file_path)}"',
        )
        msg.attach(part)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    body = {"raw": raw_message}
    all_recipients = recipients + cc_recipients

    sent = service.users().messages().send(userId="me", body=body).execute()
    emit_status(
        f"[GMAIL] Email sent to {all_recipients} | ID: {sent['id']} | Subject: {subject}",
        status_callback,
    )


def drive_authenticate():
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    creds = None
    token_path = resolve_repo_path(TOKEN_DRIVE_FILE)
    credentials_path = resolve_repo_path(CREDENTIALS_FILE)
    if token_path.exists():
        creds = Credentials.from_authorized_user_file(str(token_path), DRIVE_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(str(credentials_path), DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(token_path, "w", encoding="utf-8") as f:
            f.write(creds.to_json())

    return build("drive", "v3", credentials=creds)


def drive_link_for_folder(folder_id):
    return f"https://drive.google.com/drive/folders/{folder_id}"


def escape_drive_query_value(value):
    return str(value or "").replace("\\", "\\\\").replace("'", "\\'")


def find_drive_item(service, name, parent_id=None, mime_type=None):
    name_escaped = escape_drive_query_value(name)
    query_parts = [f"name='{name_escaped}'", "trashed=false"]
    if mime_type:
        query_parts.append(f"mimeType='{mime_type}'")
    if parent_id:
        query_parts.append(f"'{parent_id}' in parents")

    response = service.files().list(
        q=" and ".join(query_parts),
        spaces="drive",
        fields="files(id,name,mimeType,webViewLink)",
        pageSize=10,
    ).execute()
    items = response.get("files", [])
    return items[0] if items else None


def make_drive_item_public(service, item_id, status_callback=None):
    try:
        permissions = service.permissions().list(
            fileId=item_id,
            fields="permissions(id,type,role)",
        ).execute().get("permissions", [])
        for permission in permissions:
            if permission.get("type") == "anyone" and permission.get("role") == "reader":
                return
    except Exception:
        pass

    permission = {"type": "anyone", "role": "reader"}
    try:
        service.permissions().create(fileId=item_id, body=permission).execute()
    except Exception as exc:
        emit_status(f"[WARN] Could not make Drive item public: {item_id} ({exc})", status_callback)


def find_or_create_drive_folder(service, folder_name, parent_id=None, make_public=False, status_callback=None):
    folder_mime = "application/vnd.google-apps.folder"
    existing = find_drive_item(service, folder_name, parent_id=parent_id, mime_type=folder_mime)
    if existing:
        folder_id = existing["id"]
        if make_public:
            make_drive_item_public(service, folder_id, status_callback=status_callback)
        return folder_id

    metadata = {
        "name": folder_name,
        "mimeType": folder_mime,
    }
    if parent_id:
        metadata["parents"] = [parent_id]

    created = service.files().create(body=metadata, fields="id").execute()
    folder_id = created["id"]
    emit_status(f"[DRIVE] Created folder: {folder_name}", status_callback)
    if make_public:
        make_drive_item_public(service, folder_id, status_callback=status_callback)
    return folder_id


def upload_or_update_file_to_drive(service, file_path, parent_id, make_public=True, status_callback=None):
    from googleapiclient.http import MediaFileUpload

    path = Path(file_path)
    existing = find_drive_item(service, path.name, parent_id=parent_id)
    media = MediaFileUpload(str(path), resumable=True)

    if existing:
        uploaded = service.files().update(
            fileId=existing["id"],
            media_body=media,
            fields="id,webViewLink",
        ).execute()
        action = "Updated"
    else:
        metadata = {"name": path.name, "parents": [parent_id]}
        uploaded = service.files().create(
            body=metadata,
            media_body=media,
            fields="id,webViewLink",
        ).execute()
        action = "Uploaded"

    file_id = uploaded["id"]
    if make_public:
        make_drive_item_public(service, file_id, status_callback=status_callback)
        uploaded = service.files().get(fileId=file_id, fields="id,webViewLink").execute()

    link = uploaded.get("webViewLink")
    emit_status(f"[DRIVE] {action} {path.name}", status_callback)
    return {
        "id": file_id,
        "name": path.name,
        "path": str(path),
        "link": link,
    }


def parse_kickback_summary(report_path):
    results = []
    wb = openpyxl.load_workbook(report_path, data_only=True)
    try:
        if "Summary" not in wb.sheetnames:
            return results

        sheet = wb["Summary"]
        for row_idx in range(2, sheet.max_row + 1):
            store_val = sheet.cell(row=row_idx, column=1).value
            owed_val = sheet.cell(row=row_idx, column=2).value
            if store_val is None or owed_val is None:
                continue

            store_str = str(store_val).strip().lower()
            owed_str = str(owed_val).strip().lower()
            if store_str in {"", "store"} or owed_str in {"", "kickback owed"}:
                continue

            results.append((store_val, owed_val))
    finally:
        wb.close()

    return results


def build_kickback_table(rows):
    if not rows:
        return "<p>(No kickback summary data found.)</p>"

    html = [
        "<table border='1' cellpadding='6' cellspacing='0'>",
        "<thead><tr><th>Store</th><th>Kickback Owed</th></tr></thead>",
        "<tbody>",
    ]
    for store, owed in rows:
        try:
            owed_text = f"${float(owed):,.2f}"
        except (TypeError, ValueError):
            owed_text = str(owed)
        html.append(f"<tr><td>{store}</td><td>{owed_text}</td></tr>")
    html.append("</tbody></table>")
    return "\n".join(html)


def parse_report_link_line(line):
    if ":" not in line:
        return None

    filename, url = line.split(":", 1)
    filename = filename.strip()
    url = url.strip()
    if not url.startswith("http"):
        return None

    match = re.match(
        r"^(?P<brand>.+?)_report_(?P<start>\d{4}-\d{2}-\d{2})_to_(?P<end>\d{4}-\d{2}-\d{2})\.xlsx$",
        filename,
        re.IGNORECASE,
    )
    if match:
        return {
            "brand": match.group("brand").strip(),
            "filename": filename,
            "url": url,
            "start_date": match.group("start"),
            "end_date": match.group("end"),
        }

    return None


def load_report_links(links_file):
    entries = []
    links_path = resolve_repo_path(links_file)
    if not links_path.is_file():
        return entries

    with open(links_path, "r", encoding="utf-8") as f:
        for raw_line in f:
            line = raw_line.strip()
            if not line:
                continue
            parsed = parse_report_link_line(line)
            if parsed:
                entries.append(parsed)

    return entries


def build_report_link_list(entries):
    if not entries:
        return "<p>(No deal report links found.)</p>"

    html = ["<ul>"]
    for entry in entries:
        html.append(
            "<li><strong>{filename}</strong>: <a href='{url}'>{url}</a></li>".format(
                filename=entry["filename"],
                url=entry["url"],
            )
        )
    html.append("</ul>")
    return "\n".join(html)


def load_inventory_links(manifest_path):
    payload = load_inventory_manifest(manifest_path)
    folders = payload.get("folders", {})
    out = {}
    for folder_name, info in folders.items():
        if isinstance(info, str):
            info = {"link": info, "emails": []}
        out[normalize_key(folder_name)] = {
            "folder_name": folder_name,
            "link": info.get("link", ""),
            "emails": info.get("emails", []),
        }
    return out


def default_inventory_manifest():
    now = datetime.datetime.now()
    return {
        "date": now.strftime("%Y-%m-%d"),
        "day": now.strftime("%A"),
        "generated_at": now.isoformat(timespec="seconds"),
        "folders": {},
    }


def load_inventory_manifest(manifest_path):
    manifest_path = resolve_repo_path(manifest_path)
    if not manifest_path.is_file():
        return default_inventory_manifest()

    with open(manifest_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    if not isinstance(payload, dict):
        return default_inventory_manifest()

    payload.setdefault("date", datetime.datetime.now().strftime("%Y-%m-%d"))
    payload.setdefault("day", datetime.datetime.now().strftime("%A"))
    payload.setdefault("generated_at", datetime.datetime.now().isoformat(timespec="seconds"))
    payload.setdefault("folders", {})
    return payload


def save_inventory_manifest(manifest_path, payload, status_callback=None):
    manifest_path = resolve_repo_path(manifest_path)
    manifest_path.parent.mkdir(parents=True, exist_ok=True)

    payload["generated_at"] = datetime.datetime.now().isoformat(timespec="seconds")
    payload.setdefault("date", datetime.datetime.now().strftime("%Y-%m-%d"))
    payload.setdefault("day", datetime.datetime.now().strftime("%A"))
    payload.setdefault("folders", {})

    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, sort_keys=True)

    emit_status(f"[INFO] Saved inventory links manifest: {manifest_path}", status_callback)


def set_inventory_link(payload, folder_name, link, emails=None):
    folders = payload.setdefault("folders", {})
    existing = folders.get(folder_name, {})
    existing_emails = []
    if isinstance(existing, dict):
        existing_emails = existing.get("emails", []) or []

    merged_emails = list(dict.fromkeys(list(existing_emails) + list(emails or [])))
    folders[folder_name] = {
        "link": link,
        "emails": merged_emails,
    }

    return {
        "folder_name": folder_name,
        "link": link,
        "emails": merged_emails,
    }


def find_latest_report(reports_dir, aliases):
    pattern = re.compile(
        r"^(?P<brand>.+?)_report_(?P<start>\d{4}-\d{2}-\d{2})_to_(?P<end>\d{4}-\d{2}-\d{2})\.xlsx$",
        re.IGNORECASE,
    )
    alias_keys = {normalize_key(alias) for alias in aliases}
    candidates = []

    reports_path = resolve_repo_path(reports_dir)
    if not reports_path.is_dir():
        return None

    for filename in os.listdir(reports_path):
        match = pattern.match(filename)
        if not match:
            continue
        if normalize_key(match.group("brand")) not in alias_keys:
            continue

        candidates.append(
            {
                "brand": match.group("brand").strip(),
                "filename": filename,
                "path": str(reports_path / filename),
                "start_date": match.group("start"),
                "end_date": match.group("end"),
            }
        )

    if not candidates:
        return None

    candidates.sort(key=lambda item: (item["end_date"], item["start_date"], item["filename"]))
    return candidates[-1]


def select_report_links(report_info, link_entries, aliases):
    alias_keys = {normalize_key(alias) for alias in aliases}
    matching = [
        entry for entry in link_entries
        if entry["filename"] == report_info["filename"]
    ]
    if matching:
        return matching

    return [
        entry for entry in link_entries
        if normalize_key(entry["brand"]) in alias_keys
    ]


def clear_sales_exports(input_dir, status_callback=None):
    input_path = resolve_repo_path(input_dir)
    if not input_path.is_dir():
        return []

    deleted = []
    pattern = re.compile(r"^sales[A-Za-z]+\.xlsx$", re.IGNORECASE)
    for child in input_path.iterdir():
        if child.is_file() and pattern.match(child.name):
            child.unlink()
            deleted.append(str(child))
            emit_status(f"[CLEANUP] Deleted old sales export: {child}", status_callback)
    return deleted


def run_weekly_sales_pull(start_date, end_date, sales_source=DEFAULT_SALES_SOURCE, env_file=DEFAULT_API_ENV_FILE, status_callback=None):
    from autoJob import run_sales_report_api, run_sales_report_browser

    start_day = coerce_date(start_date)
    end_day = coerce_date(end_date)
    if sales_source == "api":
        emit_status(f"[AUTO] Pulling sales from Dutchie API for {start_day} to {end_day}", status_callback)
        return capture_printed_status(
            lambda: call_in_base_dir(lambda: run_sales_report_api(start_day, end_day, env_file=env_file)),
            status_callback=status_callback,
        )

    emit_status(f"[AUTO] Pulling sales with browser flow for {start_day} to {end_day}", status_callback)
    return capture_printed_status(
        lambda: call_in_base_dir(lambda: run_sales_report_browser(start_day, end_day)),
        status_callback=status_callback,
    )


def generate_deal_reports_for_week(brand_cfgs, reports_dir, status_callback=None):
    import deals

    selected = [brand_cfg["brand"] for brand_cfg in brand_cfgs]
    emit_status(f"[AUTO] Generating deal reports for: {', '.join(selected)}", status_callback)

    return capture_printed_status(
        lambda: call_in_base_dir(
            lambda: deals.run_deals_reports(
                selected_brands=selected,
                output_dir=str(resolve_repo_path(reports_dir)),
                old_dir=str(resolve_repo_path("old")),
                archive_existing=True,
                sync_reference=False,
                sync_sheet=False,
            )
        ),
        status_callback=status_callback,
    )


def refresh_inventory_sources(input_dir, include_order_reports=True, status_callback=None):
    from brand_inventory_report_job import refresh_sources

    input_path = resolve_repo_path(input_dir)
    emit_status(f"[AUTO] Refreshing inventory source exports in {input_path}", status_callback)
    return capture_printed_status(
        lambda: refresh_sources(input_path, include_order_reports=include_order_reports),
        status_callback=status_callback,
    )


def clear_inventory_order_report_cache(status_callback=None):
    try:
        from inventory_order_reports import clear_order_report_cache
    except ImportError:
        return

    clear_order_report_cache()
    emit_status("[AUTO] Cleared cached order-report tables before rebuilding inventory.", status_callback)


def generate_inventory_reports_for_week(brand_cfgs, input_dir, output_dir, include_cost=True, status_callback=None):
    from brand_inventory_report_job import build_brand_inventory_reports

    selected = []
    for brand_cfg in brand_cfgs:
        selected.extend(brand_cfg.get("inventory_aliases") or [brand_cfg["inventory_folder"]])
    selected = list(dict.fromkeys(selected))
    output_path = resolve_repo_path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    emit_status(f"[AUTO] Generating inventory reports for: {', '.join(selected)}", status_callback)
    clear_inventory_order_report_cache(status_callback=status_callback)
    return capture_printed_status(
        lambda: build_brand_inventory_reports(
            input_dir=resolve_repo_path(input_dir),
            output_dir=output_path,
            selected_brands=selected,
            include_cost=include_cost,
        ),
        status_callback=status_callback,
    )


def inventory_files_for_brand(brand_cfg, inventory_brand_map):
    aliases = brand_cfg.get("inventory_aliases") or [brand_cfg["inventory_folder"]]
    alias_keys = {str(alias).strip().lower() for alias in aliases}
    files = []
    for brand_key, brand_files in (inventory_brand_map or {}).items():
        if str(brand_key).strip().lower() in alias_keys:
            files.extend(brand_files or [])
    return sorted(dict.fromkeys(files))


def _inventory_store_code_from_path(file_path):
    stem = Path(file_path).stem.upper()
    for token in re.split(r"[^A-Z0-9]+", stem):
        if token in STORE_LABELS:
            return token
    return ""


def _inventory_store_label(file_path):
    code = _inventory_store_code_from_path(file_path)
    if code:
        return f"{code} - {STORE_LABELS.get(code, code)}"
    return Path(file_path).stem


def _format_available_quantity(value):
    if value is None:
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value).strip()
    if number.is_integer():
        return str(int(number))
    return f"{number:g}"


def _find_header_index(header_map, candidates):
    for candidate in candidates:
        idx = header_map.get(normalize_key(candidate))
        if idx is not None:
            return idx
    return None


def read_two_week_sell_thru_by_product(workbook):
    if "Order" not in workbook.sheetnames:
        return {}

    ws = workbook["Order"]
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header:
        return {}

    header_map = {
        normalize_key(value): idx
        for idx, value in enumerate(header)
        if value not in (None, "")
    }
    product_idx = _find_header_index(header_map, ("Product", "Product Name"))
    sell_thru_idx = _find_header_index(header_map, ("Units Sold 14d", "Quantity Sold 14d", "Sold 14d"))
    if product_idx is None or sell_thru_idx is None:
        return {}

    sold_by_product = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        product = row[product_idx] if product_idx < len(row) else None
        product_key = normalize_key(product)
        if not product_key:
            continue
        sold_value = row[sell_thru_idx] if sell_thru_idx < len(row) else None
        sold_by_product[product_key] = _format_available_quantity(sold_value)

    return sold_by_product


def _category_label(value):
    raw = re.sub(r"\s+", " ", str(value or "").strip())
    key = normalize_key(raw)
    if not key:
        return ""
    if "disposable" in key:
        return "Disposables"
    if "preroll" in key or ("pre" in key and "roll" in key):
        return "Pre-Rolls"
    if "wellness" in key:
        return "Wellness"
    if "concentrate" in key:
        return "Concentrate"
    return raw


def read_available_product_rows(inventory_workbook):
    rows = []
    workbook_path = Path(inventory_workbook)
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        two_week_sell_thru = read_two_week_sell_thru_by_product(wb)
        if "Available" not in wb.sheetnames:
            return rows

        ws = wb["Available"]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            return rows

        header_map = {
            normalize_key(value): idx
            for idx, value in enumerate(header)
            if value not in (None, "")
        }
        available_idx = header_map.get("available")
        product_idx = header_map.get("product") if "product" in header_map else header_map.get("productname")
        category_idx = header_map.get("category")
        if available_idx is None or product_idx is None:
            return rows

        store = _inventory_store_label(workbook_path)
        for row in ws.iter_rows(min_row=2, values_only=True):
            product = row[product_idx] if product_idx < len(row) else None
            product_text = re.sub(r"\s+", " ", str(product or "").strip())
            if not product_text or normalize_key(product_text) == "product":
                continue

            available = row[available_idx] if available_idx < len(row) else None
            rows.append(
                {
                    "store": store,
                    "available": _format_available_quantity(available),
                    "sell_thru_14d": two_week_sell_thru.get(normalize_key(product_text), ""),
                    "category": _category_label(row[category_idx] if category_idx is not None and category_idx < len(row) else ""),
                    "product": product_text,
                }
            )
    finally:
        wb.close()

    return rows


def _store_sort_key(store_label):
    code = str(store_label or "").split(" - ", 1)[0].strip().upper()
    store_order = {store_code: idx for idx, store_code in enumerate(STORE_LABELS)}
    return (store_order.get(code, 999), str(store_label or ""))


def _hashish_variant(product_name):
    match = re.search(r"\(([BW])\)\s*$", str(product_name or "").strip(), re.IGNORECASE)
    if not match:
        return ""
    return f"({match.group(1).upper()})"


def _hashish_product_line(product_name):
    parts = [part.strip() for part in str(product_name or "").split("|") if part.strip()]
    if len(parts) >= 2:
        return parts[1]
    return parts[0] if parts else ""


def _hashish_pack_size(product_name):
    match = re.search(r"((?:\d+(?:\.\d+)?)|(?:\.\d+))\s*g\b", str(product_name or ""), re.IGNORECASE)
    if not match:
        return ""
    value = match.group(1)
    if value.startswith("."):
        return f"{value}g"
    try:
        number = float(value)
    except ValueError:
        return f"{value}g"
    if number.is_integer():
        return f"{int(number)}g"
    return f"{number:g}g"


def hashish_product_section(row):
    product = str(row.get("product") or "")
    category = _category_label(row.get("category"))
    if category and category != "Concentrate":
        return category

    line = _hashish_product_line(product)
    line_key = normalize_key(line)
    variant = _hashish_variant(product)
    size = _hashish_pack_size(line or product)
    size_prefix = f"{size} " if size else ""
    variant_prefix = f"{variant} " if variant else ""

    if "liverosin" in line_key or "rosin" in line_key:
        return f"{variant_prefix}{size_prefix}Live Rosin".strip()
    if "templeball" in line_key or ("temple" in line_key and "ball" in line_key):
        return f"{size_prefix}Temple Ball".strip()
    if "topper" in line_key:
        return f"{size_prefix}Topper".strip()
    if category:
        return category
    return line or "Other"


def _hashish_section_sort_key(section_label):
    label = str(section_label or "")
    key = normalize_key(label)
    category_order = HASHISH_SECTION_CATEGORY_ORDER.get(label)
    if category_order is not None:
        return (category_order, 0, 0, label)

    size_match = re.search(r"((?:\d+(?:\.\d+)?)|(?:\.\d+))g", label, re.IGNORECASE)
    size = 999.0
    if size_match:
        try:
            size = float(size_match.group(1))
        except ValueError:
            size = 999.0

    variant = _hashish_variant(label)
    variant_order = {"(B)": 0, "(W)": 1}.get(variant, 2)
    if "liverosin" in key:
        return (10, size, variant_order, label)
    if "templeball" in key:
        return (30, size, variant_order, label)
    if "topper" in key:
        return (40, size, variant_order, label)
    if label == "Concentrate":
        return (70, size, variant_order, label)
    return (75, size, variant_order, label)


def group_hashish_available_rows(rows):
    grouped = {}
    for row in rows:
        grouped.setdefault(hashish_product_section(row), []).append(row)

    ordered = []
    for label in sorted(grouped, key=_hashish_section_sort_key):
        ordered.append(
            (
                label,
                sorted(
                    grouped[label],
                    key=lambda item: (
                        normalize_key(item.get("category")),
                        normalize_key(item.get("product")),
                    ),
                ),
            )
        )
    return ordered


def _available_product_pdf_table(sections, doc, cell_style, header_cell_style, section_style):
    data = [
        [
            Paragraph("Available", header_cell_style),
            Paragraph("2 Wk Sell Thru", header_cell_style),
            Paragraph("Product", header_cell_style),
            Paragraph("Category", header_cell_style),
        ]
    ]
    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2F7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("ALIGN", (0, 0), (1, -1), "CENTER"),
        ("ALIGN", (3, 0), (3, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.4),
        ("LINEABOVE", (0, 0), (-1, 0), 0.45, colors.HexColor("#9CA3AF")),
        ("LINEBELOW", (0, 0), (-1, 0), 0.65, colors.HexColor("#6B7280")),
        ("LINEBELOW", (0, 1), (-1, -1), 0.25, colors.HexColor("#D1D5DB")),
        ("LINEBEFORE", (1, 0), (-1, -1), 0.2, colors.HexColor("#E5E7EB")),
        ("LEFTPADDING", (0, 0), (-1, -1), 2.6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 2.6),
        ("TOPPADDING", (0, 0), (-1, -1), 1.4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.4),
    ]
    for section_label, rows in sections:
        section_row = len(data)
        data.append([Paragraph(escape(str(section_label)), section_style), "", "", ""])
        style_commands.extend(
            [
                ("SPAN", (0, section_row), (-1, section_row)),
                ("BACKGROUND", (0, section_row), (-1, section_row), colors.HexColor("#F8FAFC")),
                ("LINEABOVE", (0, section_row), (-1, section_row), 0.5, colors.HexColor("#9CA3AF")),
                ("LINEBELOW", (0, section_row), (-1, section_row), 0.3, colors.HexColor("#CBD5E1")),
                ("TOPPADDING", (0, section_row), (-1, section_row), 2.0),
                ("BOTTOMPADDING", (0, section_row), (-1, section_row), 2.0),
            ]
        )
        for row in rows:
            data.append(
                [
                    Paragraph(escape(str(row.get("available", ""))), cell_style),
                    Paragraph(escape(str(row.get("sell_thru_14d", ""))), cell_style),
                    Paragraph(escape(row.get("product", "")), cell_style),
                    Paragraph(escape(str(row.get("category", ""))), cell_style),
                ]
            )

    table = Table(
        data,
        colWidths=[0.58 * inch, 0.82 * inch, 5.15 * inch, 0.95 * inch],
        repeatRows=1,
        hAlign="CENTER",
    )
    table.setStyle(TableStyle(style_commands))
    return table


def build_printable_available_inventory_pdf(
    brand_label,
    inventory_files,
    output_dir,
    week_start=None,
    week_end=None,
):
    output_path = resolve_repo_path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    start_text = coerce_date(week_start).isoformat() if week_start else ""
    end_text = coerce_date(week_end).isoformat() if week_end else ""
    week_suffix = f"_{start_text}_to_{end_text}" if start_text and end_text else ""
    pdf_path = output_path / (
        f"{safe_filename_part(brand_label)}_available_products_all_stores{week_suffix}.pdf"
    )

    all_rows = []
    for file_path in sorted(inventory_files, key=lambda path: _store_sort_key(_inventory_store_label(path))):
        if str(file_path).lower().endswith((".xlsx", ".xlsm")):
            all_rows.extend(read_available_product_rows(file_path))

    grouped_rows = {}
    for row in all_rows:
        grouped_rows.setdefault(row["store"], []).append(row)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "PrintableInventoryTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=16,
        alignment=0,
        spaceAfter=5,
    )
    meta_style = ParagraphStyle(
        "PrintableInventoryMeta",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=9,
        textColor=colors.HexColor("#4B5563"),
        spaceAfter=7,
    )
    store_style = ParagraphStyle(
        "PrintableInventoryStore",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=12,
        leading=13,
        spaceBefore=1,
        spaceAfter=5,
    )
    cell_style = ParagraphStyle(
        "PrintableInventoryCell",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7.4,
        leading=8.2,
    )
    header_cell_style = ParagraphStyle(
        "PrintableInventoryHeader",
        parent=cell_style,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#111827"),
    )
    section_style = ParagraphStyle(
        "PrintableInventorySection",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=8.3,
        leading=9.0,
        textColor=colors.HexColor("#111827"),
    )

    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=0.28 * inch,
        rightMargin=0.28 * inch,
        topMargin=0.28 * inch,
        bottomMargin=0.28 * inch,
        title=f"{brand_label} Available Products",
    )

    elements = [
        Paragraph(f"{escape(str(brand_label))} Available Products", title_style),
        Paragraph(
            escape(f"All stores{f' | {start_text} to {end_text}' if start_text and end_text else ''}"),
            meta_style,
        ),
    ]

    if not grouped_rows:
        elements.append(Paragraph("No available inventory rows found.", cell_style))
    else:
        stores = sorted(grouped_rows, key=_store_sort_key)
        for store_index, store in enumerate(stores):
            if store_index:
                elements.append(PageBreak())
            elements.append(Paragraph(escape(store), store_style))
            store_rows = grouped_rows[store]
            if normalize_key(brand_label) == "hashish":
                sections = group_hashish_available_rows(store_rows)
            else:
                sections = [("Available Products", store_rows)]
            elements.append(
                _available_product_pdf_table(
                    sections,
                    doc,
                    cell_style,
                    header_cell_style,
                    section_style,
                )
            )
            elements.append(Spacer(1, 0.03 * inch))

    doc.build(elements)
    return str(pdf_path)


def generate_printable_available_inventory_pdfs(
    brand_cfgs,
    inventory_brand_map,
    output_dir,
    week_start=None,
    week_end=None,
    status_callback=None,
):
    pdf_map = {}
    pdf_output_dir = resolve_repo_path(output_dir) / "printable_available_pdfs"
    for brand_cfg in brand_cfgs:
        if normalize_key(brand_cfg.get("brand")) != "hashish":
            continue

        inventory_files = [
            path for path in inventory_files_for_brand(brand_cfg, inventory_brand_map)
            if str(path).lower().endswith((".xlsx", ".xlsm"))
        ]
        if not inventory_files:
            continue

        pdf_path = build_printable_available_inventory_pdf(
            brand_label=brand_cfg["brand"],
            inventory_files=inventory_files,
            output_dir=pdf_output_dir,
            week_start=week_start,
            week_end=week_end,
        )
        aliases = brand_cfg.get("inventory_aliases") or [brand_cfg["inventory_folder"]]
        brand_key = str(aliases[0]).strip().lower()
        pdf_map.setdefault(brand_key, []).append(pdf_path)
        emit_status(f"[PDF] Created printable available inventory: {pdf_path}", status_callback)

    return pdf_map


def write_report_links_file(links_file, uploaded_report_links, status_callback=None):
    links_path = resolve_repo_path(links_file)
    links_path.parent.mkdir(parents=True, exist_ok=True)
    with open(links_path, "w", encoding="utf-8") as handle:
        for item in uploaded_report_links:
            link = item.get("link")
            if link:
                handle.write(f"{item['name']}: {link}\n")
    emit_status(f"[INFO] Saved report links: {links_path}", status_callback)


def upload_weekly_brand_outputs(
    brand_cfgs,
    reports_dir,
    inventory_brand_map,
    links_file=DEFAULT_LINKS_FILE,
    inventory_links_file=DEFAULT_INVENTORY_LINKS_FILE,
    drive_parent_folder=DEFAULT_WEEKLY_DRIVE_PARENT,
    week_start=None,
    week_end=None,
    status_callback=None,
):
    start_day = coerce_date(week_start) if week_start else get_previous_monday_sunday()[0]
    end_day = coerce_date(week_end) if week_end else get_previous_monday_sunday()[1]
    week_folder = week_key_for_range(start_day, end_day)

    service = drive_authenticate()
    parent_id = find_or_create_drive_folder(
        service,
        drive_parent_folder,
        parent_id=None,
        make_public=False,
        status_callback=status_callback,
    )

    inventory_manifest = load_inventory_manifest(inventory_links_file)
    uploaded_report_links = []
    folders = {}
    failures = []

    for brand_cfg in brand_cfgs:
        report_info = find_latest_report(reports_dir, brand_cfg["report_aliases"])
        if not report_info:
            failures.append(f"{brand_cfg['brand']}: no generated deal report found in {reports_dir}")
            continue

        brand_folder_name = label_for_brand_config(brand_cfg)
        brand_folder_id = find_or_create_drive_folder(
            service,
            brand_folder_name,
            parent_id=parent_id,
            make_public=False,
            status_callback=status_callback,
        )
        week_folder_id = find_or_create_drive_folder(
            service,
            week_folder,
            parent_id=brand_folder_id,
            make_public=True,
            status_callback=status_callback,
        )
        week_folder_link = drive_link_for_folder(week_folder_id)

        uploaded_deal = upload_or_update_file_to_drive(
            service,
            report_info["path"],
            week_folder_id,
            make_public=True,
            status_callback=status_callback,
        )
        uploaded_report_links.append(uploaded_deal)

        inventory_files = inventory_files_for_brand(brand_cfg, inventory_brand_map)
        if not inventory_files:
            failures.append(f"{brand_cfg['brand']}: no inventory workbooks generated")
        for inventory_file in inventory_files:
            upload_or_update_file_to_drive(
                service,
                inventory_file,
                week_folder_id,
                make_public=True,
                status_callback=status_callback,
            )

        folders[brand_cfg["inventory_folder"]] = {
            "folder_id": week_folder_id,
            "link": week_folder_link,
            "report": uploaded_deal,
            "inventory_files": inventory_files,
        }
        set_inventory_link(
            inventory_manifest,
            folder_name=brand_cfg["inventory_folder"],
            link=week_folder_link,
            emails=brand_cfg["to"] + BUZZ_CC,
        )
        emit_status(f"[DRIVE] Weekly folder for {brand_cfg['brand']}: {week_folder_link}", status_callback)

    write_report_links_file(links_file, uploaded_report_links, status_callback=status_callback)
    save_inventory_manifest(inventory_links_file, inventory_manifest, status_callback=status_callback)

    return {
        "folders": folders,
        "report_links": uploaded_report_links,
        "failures": failures,
        "week_folder": week_folder,
    }


def prepare_weekly_brand_credit_run(
    selected_brands=None,
    reports_dir=DEFAULT_REPORTS_DIR,
    links_file=DEFAULT_LINKS_FILE,
    inventory_links_file=DEFAULT_INVENTORY_LINKS_FILE,
    inventory_input_dir=DEFAULT_INVENTORY_INPUT_DIR,
    auto_output_root=DEFAULT_AUTO_OUTPUT_ROOT,
    sales_source=DEFAULT_SALES_SOURCE,
    env_file=DEFAULT_API_ENV_FILE,
    skip_sales_pull=False,
    skip_inventory_refresh=False,
    include_inventory_order_reports=True,
    include_inventory_cost=True,
    no_drive_upload=False,
    drive_parent_folder=DEFAULT_WEEKLY_DRIVE_PARENT,
    status_callback=None,
):
    brand_cfgs = selected_brand_configs(selected_brands)
    if not brand_cfgs:
        raise ValueError("No matching weekly brand configs selected.")

    week_start, week_end = get_previous_monday_sunday()
    week_key = week_key_for_range(week_start, week_end)
    output_root = resolve_repo_path(auto_output_root) / week_key

    if reports_dir == DEFAULT_REPORTS_DIR:
        reports_dir = str(output_root / "deal_reports")
    if links_file == DEFAULT_LINKS_FILE:
        links_file = str(output_root / "links.txt")

    inventory_output_dir = output_root / "inventory_reports"
    output_root.mkdir(parents=True, exist_ok=True)

    emit_status(f"[AUTO] Weekly range: {week_start.isoformat()} to {week_end.isoformat()}", status_callback)

    if not skip_sales_pull:
        clear_sales_exports(inventory_input_dir, status_callback=status_callback)
        run_weekly_sales_pull(
            week_start,
            week_end,
            sales_source=sales_source,
            env_file=env_file,
            status_callback=status_callback,
        )
    else:
        emit_status("[AUTO] Sales pull skipped; using existing files/sales*.xlsx exports.", status_callback)

    generate_deal_reports_for_week(brand_cfgs, reports_dir, status_callback=status_callback)

    if not skip_inventory_refresh:
        refresh_inventory_sources(
            inventory_input_dir,
            include_order_reports=include_inventory_order_reports,
            status_callback=status_callback,
        )
    else:
        emit_status("[AUTO] Inventory source refresh skipped; using existing catalog/order files.", status_callback)

    inventory_brand_map = generate_inventory_reports_for_week(
        brand_cfgs,
        input_dir=inventory_input_dir,
        output_dir=inventory_output_dir,
        include_cost=include_inventory_cost,
        status_callback=status_callback,
    )
    printable_inventory_pdfs = generate_printable_available_inventory_pdfs(
        brand_cfgs,
        inventory_brand_map,
        output_dir=inventory_output_dir,
        week_start=week_start,
        week_end=week_end,
        status_callback=status_callback,
    )
    for brand_key, pdf_paths in printable_inventory_pdfs.items():
        inventory_brand_map.setdefault(brand_key, []).extend(pdf_paths)

    upload_result = {
        "folders": {},
        "report_links": [],
        "failures": [],
        "week_folder": week_key,
    }
    if no_drive_upload:
        emit_status("[AUTO] Drive upload skipped by option.", status_callback)
    else:
        upload_result = upload_weekly_brand_outputs(
            brand_cfgs,
            reports_dir=reports_dir,
            inventory_brand_map=inventory_brand_map,
            links_file=links_file,
            inventory_links_file=inventory_links_file,
            drive_parent_folder=drive_parent_folder,
            week_start=week_start,
            week_end=week_end,
            status_callback=status_callback,
        )

    return {
        "week_start": week_start.isoformat(),
        "week_end": week_end.isoformat(),
        "week_key": week_key,
        "reports_dir": str(reports_dir),
        "links_file": str(links_file),
        "inventory_links_file": str(inventory_links_file),
        "inventory_output_dir": str(inventory_output_dir),
        "inventory_brand_map": inventory_brand_map,
        "printable_inventory_pdfs": printable_inventory_pdfs,
        "upload": upload_result,
    }


def build_email_body(brand_label, report_info, inventory_info, report_links, kickback_rows):
    inventory_folder_name = inventory_info["folder_name"]
    inventory_link = inventory_info["link"]
    report_link_html = build_report_link_list(report_links)
    kickback_html = build_kickback_table(kickback_rows)

    return f"""
    <html>
      <body>
        <p>Hello,</p>
        <p>Please see below the {brand_label} brand deals for <strong>{report_info['start_date']} to {report_info['end_date']}</strong>, along with the weekly Drive folder containing the deal and inventory reports.</p>
        <h3>Folder: {inventory_folder_name}</h3>
        <p>Link: <a href="{inventory_link}">{inventory_link}</a></p>
        <h3>{brand_label}</h3>
        <p><strong>Links:</strong></p>
        {report_link_html}
        <h3>Kickback Summary:</h3>
        {kickback_html}
        <p><strong>please include/contact joseph@buzzcannabis.com &amp; donna@buzzcannabis.com in all emails regarding these credits.</strong></p>
        <p>Regards,<br>Buzz Cannabis</p>
      </body>
    </html>
    """


def parse_args(argv=None):
    parser = argparse.ArgumentParser(description="Send weekly deal + inventory credit emails.")
    parser.add_argument(
        "--brands",
        nargs="*",
        help="Optional subset to send, e.g. --brands Hashish TreeSap",
    )
    parser.add_argument(
        "--reports-dir",
        default=DEFAULT_REPORTS_DIR,
        help=f"Directory containing brand deal reports (default: {DEFAULT_REPORTS_DIR})",
    )
    parser.add_argument(
        "--links-file",
        default=DEFAULT_LINKS_FILE,
        help=f"links.txt path containing Drive URLs for uploaded deal reports (default: {DEFAULT_LINKS_FILE})",
    )
    parser.add_argument(
        "--inventory-links-file",
        default=DEFAULT_INVENTORY_LINKS_FILE,
        help=f"Inventory link manifest written by BrandINVEmailer.py (default: {DEFAULT_INVENTORY_LINKS_FILE})",
    )
    parser.add_argument(
        "--inventory-link",
        action="append",
        default=[],
        metavar="BRAND=URL",
        help="Provide or override an inventory folder link and save it to the manifest. Repeat as needed.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Preview the sends without actually emailing Gmail recipients.",
    )
    parser.add_argument(
        "--test-mode",
        action="store_true",
        help=f"Send only to {TEST_MODE_EMAIL}; no vendor recipients and no CCs.",
    )
    parser.add_argument(
        "--test-email",
        help="Override all recipients with one test email address and do not CC anyone.",
    )
    parser.add_argument(
        "--no-attachments",
        action="store_true",
        help="Do not attach local deal reports or printable inventory PDFs to the email.",
    )
    parser.add_argument(
        "--auto",
        dest="auto",
        action="store_true",
        default=True,
        help="Pull/generate Hashish and Treesap reports, upload weekly Drive folders, then email (default).",
    )
    parser.add_argument(
        "--existing-links",
        dest="auto",
        action="store_false",
        help="Use already-generated deal and inventory links without refreshing inventory.",
    )
    parser.add_argument(
        "--sales-source",
        choices=("api", "browser"),
        default=DEFAULT_SALES_SOURCE,
        help=f"Where to pull weekly sales data from in --auto mode (default: {DEFAULT_SALES_SOURCE}).",
    )
    parser.add_argument(
        "--env-file",
        default=DEFAULT_API_ENV_FILE,
        help=f"Dutchie API .env file for --sales-source api (default: {DEFAULT_API_ENV_FILE}).",
    )
    parser.add_argument(
        "--auto-output-root",
        default=DEFAULT_AUTO_OUTPUT_ROOT,
        help=f"Local output root for automatic weekly reports (default: {DEFAULT_AUTO_OUTPUT_ROOT}).",
    )
    parser.add_argument(
        "--inventory-input-dir",
        default=DEFAULT_INVENTORY_INPUT_DIR,
        help=f"Input folder for sales/catalog/order exports (default: {DEFAULT_INVENTORY_INPUT_DIR}).",
    )
    parser.add_argument(
        "--skip-sales-pull",
        action="store_true",
        help="In --auto mode, use existing files/sales*.xlsx instead of pulling sales.",
    )
    parser.add_argument(
        "--skip-inventory-refresh",
        action="store_true",
        help="In --auto mode, use existing catalog/order source files instead of refreshing inventory.",
    )
    parser.add_argument(
        "--no-inventory-order-reports",
        action="store_true",
        help="In --auto mode, skip 7d/14d/30d inventory order report refresh and tabs.",
    )
    parser.add_argument(
        "--no-inventory-cost",
        action="store_true",
        help="In --auto mode, hide cost columns in generated inventory workbooks.",
    )
    parser.add_argument(
        "--no-drive-upload",
        action="store_true",
        help="In --auto mode, generate local reports but do not upload to Drive.",
    )
    parser.add_argument(
        "--drive-parent-folder",
        default=DEFAULT_WEEKLY_DRIVE_PARENT,
        help=f"Top-level Drive folder for weekly automatic uploads (default: {DEFAULT_WEEKLY_DRIVE_PARENT}).",
    )
    return parser.parse_args(argv)


def should_include_brand(brand_cfg, requested_brands):
    if not requested_brands:
        return True

    requested = {normalize_key(value) for value in requested_brands}
    candidates = {normalize_key(brand_cfg["brand"]), normalize_key(brand_cfg["inventory_folder"])}
    candidates.update(normalize_key(alias) for alias in brand_cfg["report_aliases"])
    candidates.update(normalize_key(alias) for alias in brand_cfg.get("inventory_aliases", []))
    if brand_cfg.get("drive_folder"):
        candidates.add(normalize_key(brand_cfg["drive_folder"]))
    return not candidates.isdisjoint(requested)


def parse_inventory_link_overrides(raw_values):
    overrides = {}
    invalid = []

    for raw in raw_values:
        if "=" not in raw:
            invalid.append(raw)
            continue
        brand_key, link = raw.split("=", 1)
        brand_key = brand_key.strip()
        link = link.strip()
        if not brand_key or not link:
            invalid.append(raw)
            continue
        overrides[normalize_key(brand_key)] = link

    if invalid:
        raise ValueError(
            "Invalid --inventory-link value(s): "
            + ", ".join(invalid)
            + ". Use BRAND=URL."
        )

    return overrides


def get_inventory_override_for_brand(brand_cfg, overrides):
    candidates = [
        brand_cfg["brand"],
        brand_cfg["inventory_folder"],
        brand_cfg.get("drive_folder", ""),
        *brand_cfg["report_aliases"],
        *brand_cfg.get("inventory_aliases", []),
    ]
    for candidate in candidates:
        link = overrides.get(normalize_key(candidate))
        if link:
            return link
    return None


def prompt_for_inventory_link(brand_cfg):
    if not sys.stdin.isatty():
        return None

    prompt = (
        f"Enter inventory folder link for {brand_cfg['brand']} "
        f"({brand_cfg['inventory_folder']}) and press Enter "
        "or leave blank to skip: "
    )
    try:
        entered = input(prompt).strip()
    except EOFError:
        return None

    return entered or None


def normalize_inventory_overrides(inventory_overrides):
    if not inventory_overrides:
        return {}

    if isinstance(inventory_overrides, dict):
        return {
            normalize_key(key): str(value).strip()
            for key, value in inventory_overrides.items()
            if str(value).strip()
        }

    return parse_inventory_link_overrides(inventory_overrides)


def run_weekly_brand_credit_emailer(
    selected_brands=None,
    reports_dir=DEFAULT_REPORTS_DIR,
    links_file=DEFAULT_LINKS_FILE,
    inventory_links_file=DEFAULT_INVENTORY_LINKS_FILE,
    inventory_overrides=None,
    dry_run=False,
    test_mode=False,
    test_email=None,
    no_attachments=False,
    prompt_for_missing=False,
    auto_generate=False,
    sales_source=DEFAULT_SALES_SOURCE,
    env_file=DEFAULT_API_ENV_FILE,
    auto_output_root=DEFAULT_AUTO_OUTPUT_ROOT,
    inventory_input_dir=DEFAULT_INVENTORY_INPUT_DIR,
    skip_sales_pull=False,
    skip_inventory_refresh=False,
    include_inventory_order_reports=True,
    include_inventory_cost=True,
    no_drive_upload=False,
    drive_parent_folder=DEFAULT_WEEKLY_DRIVE_PARENT,
    status_callback=None,
):
    prepare_result = None
    if auto_generate:
        prepare_result = prepare_weekly_brand_credit_run(
            selected_brands=selected_brands,
            reports_dir=reports_dir,
            links_file=links_file,
            inventory_links_file=inventory_links_file,
            inventory_input_dir=inventory_input_dir,
            auto_output_root=auto_output_root,
            sales_source=sales_source,
            env_file=env_file,
            skip_sales_pull=skip_sales_pull,
            skip_inventory_refresh=skip_inventory_refresh,
            include_inventory_order_reports=include_inventory_order_reports,
            include_inventory_cost=include_inventory_cost,
            no_drive_upload=no_drive_upload,
            drive_parent_folder=drive_parent_folder,
            status_callback=status_callback,
        )
        reports_dir = prepare_result["reports_dir"]
        links_file = prepare_result["links_file"]
        inventory_links_file = prepare_result["inventory_links_file"]

    inventory_overrides = normalize_inventory_overrides(inventory_overrides)
    report_links = load_report_links(links_file)
    inventory_manifest = load_inventory_manifest(inventory_links_file)
    inventory_links = load_inventory_links(inventory_links_file)

    failures = list((prepare_result or {}).get("upload", {}).get("failures", []))
    sends = 0
    test_recipient = None
    if test_mode:
        test_recipient = TEST_MODE_EMAIL
    elif test_email and str(test_email).strip():
        test_recipient = str(test_email).strip()

    for brand_cfg in WEEKLY_BRAND_EMAILS:
        if not should_include_brand(brand_cfg, selected_brands):
            continue

        report_info = find_latest_report(reports_dir, brand_cfg["report_aliases"])
        if not report_info:
            failures.append(f"{brand_cfg['brand']}: no report found in {reports_dir}")
            continue

        selected_links = select_report_links(report_info, report_links, brand_cfg["report_aliases"])
        if not selected_links:
            failures.append(f"{brand_cfg['brand']}: no Drive report link found in {links_file}")
            continue

        inventory_info = inventory_links.get(normalize_key(brand_cfg["inventory_folder"]))
        override_link = get_inventory_override_for_brand(brand_cfg, inventory_overrides)
        if override_link:
            inventory_info = set_inventory_link(
                inventory_manifest,
                folder_name=brand_cfg["inventory_folder"],
                link=override_link,
                emails=brand_cfg["to"] + BUZZ_CC,
            )
            save_inventory_manifest(inventory_links_file, inventory_manifest, status_callback=status_callback)
            inventory_links = load_inventory_links(inventory_links_file)

        if (not inventory_info or not inventory_info.get("link")) and prompt_for_missing:
            prompted_link = prompt_for_inventory_link(brand_cfg)
            if prompted_link:
                inventory_info = set_inventory_link(
                    inventory_manifest,
                    folder_name=brand_cfg["inventory_folder"],
                    link=prompted_link,
                    emails=brand_cfg["to"] + BUZZ_CC,
                )
                save_inventory_manifest(inventory_links_file, inventory_manifest, status_callback=status_callback)
                inventory_links = load_inventory_links(inventory_links_file)

        if not inventory_info or not inventory_info.get("link"):
            failures.append(
                f"{brand_cfg['brand']}: no inventory folder link found for '{brand_cfg['inventory_folder']}' in {inventory_links_file}"
            )
            continue

        kickback_rows = parse_kickback_summary(report_info["path"])
        html_body = build_email_body(
            brand_label=brand_cfg["brand"],
            report_info=report_info,
            inventory_info=inventory_info,
            report_links=selected_links,
            kickback_rows=kickback_rows,
        )
        subject = (
            f"{brand_cfg['brand']} Brand Deals for {report_info['start_date']} to "
            f"{report_info['end_date']} and Inventory"
        )
        if test_recipient:
            subject = f"[TEST] {subject}"

        printable_pdfs = inventory_files_for_brand(
            brand_cfg,
            (prepare_result or {}).get("printable_inventory_pdfs", {}),
        )
        attachments = [] if no_attachments else [report_info["path"], *printable_pdfs]
        recipients = [test_recipient] if test_recipient else brand_cfg["to"]
        cc_recipients = [] if test_recipient else BUZZ_CC

        send_email_with_gmail_html(
            subject=subject,
            html_body=html_body,
            recipients=recipients,
            cc_recipients=cc_recipients,
            attachments=attachments,
            dry_run=dry_run,
            status_callback=status_callback,
        )
        sends += 1

    if failures:
        emit_status("[WARN] Some brand emails were skipped:", status_callback)
        for failure in failures:
            emit_status(f"  - {failure}", status_callback)

    return {
        "sends": sends,
        "failures": failures,
        "prepare": prepare_result,
        "reports_dir": reports_dir,
        "links_file": links_file,
        "inventory_links_file": inventory_links_file,
    }


def main(argv=None):
    args = parse_args(argv)
    try:
        inventory_overrides = parse_inventory_link_overrides(args.inventory_link)
    except ValueError as exc:
        print(f"[ERROR] {exc}")
        sys.exit(2)

    result = run_weekly_brand_credit_emailer(
        selected_brands=args.brands,
        reports_dir=args.reports_dir,
        links_file=args.links_file,
        inventory_links_file=args.inventory_links_file,
        inventory_overrides=inventory_overrides,
        dry_run=args.dry_run,
        test_mode=args.test_mode,
        test_email=args.test_email,
        no_attachments=args.no_attachments,
        prompt_for_missing=True,
        auto_generate=args.auto,
        sales_source=args.sales_source,
        env_file=args.env_file,
        auto_output_root=args.auto_output_root,
        inventory_input_dir=args.inventory_input_dir,
        skip_sales_pull=args.skip_sales_pull,
        skip_inventory_refresh=args.skip_inventory_refresh,
        include_inventory_order_reports=not args.no_inventory_order_reports,
        include_inventory_cost=not args.no_inventory_cost,
        no_drive_upload=args.no_drive_upload,
        drive_parent_folder=args.drive_parent_folder,
    )

    if result["sends"] == 0 and not args.dry_run:
        sys.exit(1)


if __name__ == "__main__":
    main()
