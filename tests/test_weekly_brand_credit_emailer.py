import tempfile
import unittest
from datetime import date
from pathlib import Path
from unittest.mock import patch

import openpyxl
from weekly_brand_credit_emailer import (
    TEST_MODE_EMAIL,
    WEEKLY_BRAND_EMAILS,
    build_printable_available_inventory_pdf,
    find_latest_report,
    generate_printable_available_inventory_pdfs,
    get_previous_monday_sunday,
    group_hashish_available_rows,
    inventory_files_for_brand,
    parse_args,
    read_available_product_rows,
    run_weekly_brand_credit_emailer,
    should_include_brand,
    week_key_for_range,
)


class WeeklyBrandCreditEmailerTests(unittest.TestCase):
    def test_previous_week_range_from_monday(self):
        start_day, end_day = get_previous_monday_sunday(date(2026, 5, 18))

        self.assertEqual(start_day, date(2026, 5, 11))
        self.assertEqual(end_day, date(2026, 5, 17))
        self.assertEqual(week_key_for_range(start_day, end_day), "2026-05-11_to_2026-05-17")

    def test_find_latest_report_uses_brand_aliases(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            (root / "Treesap_report_2026-05-04_to_2026-05-10.xlsx").write_text("", encoding="utf-8")
            (root / "TreeSap_report_2026-05-11_to_2026-05-17.xlsx").write_text("", encoding="utf-8")

            report = find_latest_report(str(root), ["TreeSap", "Treesap"])

        self.assertIsNotNone(report)
        self.assertEqual(report["filename"], "TreeSap_report_2026-05-11_to_2026-05-17.xlsx")
        self.assertEqual(report["start_date"], "2026-05-11")
        self.assertEqual(report["end_date"], "2026-05-17")

    def test_treesap_selection_and_inventory_aliases(self):
        treesap_cfg = next(cfg for cfg in WEEKLY_BRAND_EMAILS if cfg["brand"] == "TreeSap")
        brand_map = {
            "treesap": ["/tmp/treesap_mv.xlsx", "/tmp/treesap_lm.xlsx"],
            "hashish": ["/tmp/hashish_mv.xlsx"],
        }

        self.assertTrue(should_include_brand(treesap_cfg, ["Treesap"]))
        self.assertTrue(should_include_brand(treesap_cfg, ["TreeSap"]))
        self.assertEqual(
            inventory_files_for_brand(treesap_cfg, brand_map),
            ["/tmp/treesap_lm.xlsx", "/tmp/treesap_mv.xlsx"],
        )

    def test_cli_defaults_to_auto_refresh(self):
        self.assertTrue(parse_args([]).auto)
        self.assertFalse(parse_args(["--existing-links"]).auto)

    def test_cli_has_test_mode(self):
        args = parse_args(["--test-mode"])

        self.assertTrue(args.test_mode)

    def test_test_mode_sends_only_to_anthony_without_cc(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            reports_dir = root / "reports"
            reports_dir.mkdir()
            report_path = reports_dir / "Hashish_report_2026-05-11_to_2026-05-17.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Summary"
            ws.append(["Store", "Kickback Owed"])
            ws.append(["MV", 12.5])
            wb.save(report_path)

            links_file = root / "links.txt"
            links_file.write_text(
                f"{report_path.name}: https://drive.example/hashish-report\n",
                encoding="utf-8",
            )
            inventory_links = root / "inventory_links.json"
            inventory_links.write_text(
                """
                {
                  "folders": {
                    "Hashish": {
                      "link": "https://drive.example/hashish-week",
                      "emails": []
                    }
                  }
                }
                """,
                encoding="utf-8",
            )
            sends = []

            def capture_send(**kwargs):
                sends.append(kwargs)

            with patch("weekly_brand_credit_emailer.send_email_with_gmail_html", side_effect=capture_send):
                result = run_weekly_brand_credit_emailer(
                    selected_brands=["Hashish"],
                    reports_dir=str(reports_dir),
                    links_file=str(links_file),
                    inventory_links_file=str(inventory_links),
                    test_mode=True,
                )

        self.assertEqual(result["sends"], 1)
        self.assertEqual(sends[0]["recipients"], [TEST_MODE_EMAIL])
        self.assertEqual(sends[0]["cc_recipients"], [])
        self.assertTrue(sends[0]["subject"].startswith("[TEST] "))

    def test_printable_available_pdf_uses_inventory_columns_and_14d_sell_thru(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            workbook_path = root / "06-01-2026_MV_hashish.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Available"
            ws.append(["Available", "Product", "Category", "Brand"])
            ws.append(["Concentrate", None, None, None])
            ws.append([40, "Hashish | CC Live Rosin 1g | Margaritaville (W)", "Concentrate", "Hashish"])
            ws.append([26, "Hashish | CC Live Rosin 1g | Banana Meltshake (B)", "Concentrate", "Hashish"])
            order_ws = wb.create_sheet("Order")
            order_ws.append(["Brand", "Category", "Product", "Units Sold 14d"])
            order_ws.append(["Hashish", "Concentrate", "Hashish | CC Live Rosin 1g | Margaritaville (W)", 9])
            order_ws.append(["Hashish", "Concentrate", "Hashish | CC Live Rosin 1g | Banana Meltshake (B)", 14])
            wb.save(workbook_path)

            rows = read_available_product_rows(workbook_path)
            pdf_path = Path(
                build_printable_available_inventory_pdf(
                    "Hashish",
                    [workbook_path],
                    root / "pdf",
                    week_start=date(2026, 5, 25),
                    week_end=date(2026, 5, 31),
                )
            )

            self.assertEqual(
                rows,
                [
                    {
                        "store": "MV - Mission Valley",
                        "available": "40",
                        "sell_thru_14d": "9",
                        "category": "Concentrate",
                        "product": "Hashish | CC Live Rosin 1g | Margaritaville (W)",
                    },
                    {
                        "store": "MV - Mission Valley",
                        "available": "26",
                        "sell_thru_14d": "14",
                        "category": "Concentrate",
                        "product": "Hashish | CC Live Rosin 1g | Banana Meltshake (B)",
                    },
                ],
            )
            self.assertTrue(pdf_path.exists())
            self.assertGreater(pdf_path.stat().st_size, 0)

    def test_hashish_print_rows_group_by_category_product_size_and_variant(self):
        grouped = group_hashish_available_rows(
            [
                {"category": "Concentrate", "available": "7", "product": "Hashish | Live Temple Ball 2g | Peach Zugar (W)"},
                {"category": "Concentrate", "available": "26", "product": "Hashish | CC Live Rosin 1g | Banana Meltshake (B)"},
                {"category": "Concentrate", "available": "43", "product": "Hashish | IH Topper 1g | Lemon Dawg"},
                {"category": "Concentrate", "available": "19", "product": "Hashish | CC Live Rosin 2g | GMO (W)"},
                {"category": "Concentrate", "available": "54", "product": "Hashish | CC Live Rosin 1g | Zugar Buns (B)"},
                {"category": "Disposables", "available": "12", "product": "Hashish | SLR AIO .5g | Strawberry Guava"},
                {"category": "Pre-Rolls", "available": "90", "product": "Hashish | HH Pre-Roll 1.6g | I | Oz"},
                {"category": "Wellness", "available": "5", "product": "Hashish | Wellness Balm"},
            ]
        )

        self.assertEqual(
            [label for label, _rows in grouped],
            [
                "(B) 1g Live Rosin",
                "(W) 2g Live Rosin",
                "2g Temple Ball",
                "1g Topper",
                "Disposables",
                "Pre-Rolls",
                "Wellness",
            ],
        )
        self.assertEqual(
            [row["product"] for row in grouped[0][1]],
            [
                "Hashish | CC Live Rosin 1g | Banana Meltshake (B)",
                "Hashish | CC Live Rosin 1g | Zugar Buns (B)",
            ],
        )
        self.assertEqual(
            [row["product"] for row in grouped[1][1]],
            [
                "Hashish | CC Live Rosin 2g | GMO (W)",
            ],
        )

    def test_printable_available_pdf_generation_is_hashish_only(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            root = Path(tmp_dir)
            workbook_path = root / "06-01-2026_MV_hashish.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Available"
            ws.append(["Available", "Product", "Category", "Brand"])
            ws.append([26, "Hashish | CC Live Rosin 1g | Banana Meltshake (B)", "Concentrate", "Hashish"])
            wb.save(workbook_path)

            pdf_map = generate_printable_available_inventory_pdfs(
                WEEKLY_BRAND_EMAILS,
                {
                    "hashish": [str(workbook_path)],
                    "treesap": [str(root / "treesap.xlsx")],
                },
                root / "out",
                week_start=date(2026, 5, 25),
                week_end=date(2026, 5, 31),
            )

            self.assertEqual(list(pdf_map), ["hashish"])
            self.assertTrue(Path(pdf_map["hashish"][0]).exists())


if __name__ == "__main__":
    unittest.main()
