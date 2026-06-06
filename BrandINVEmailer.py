#!/usr/bin/env python3
"""
BrandINVEmailer.py

Combined script that:

1. Reads JSON config (including test_mode, test_email, brand list).
2. On the current day, fetches CSVs (via getCatalog.py) [optional].
3. Processes brand inventory, generating Excel files grouped by brand.
4. Uploads those files to a date-based folder inside a parent "INVENTORY" Google Drive folder
   - Each scheduled brand gets its own public subfolder for vendor emails.
   - An optional OTHER folder holds public child folders for every generated brand,
     while the OTHER folder itself is restricted to the Buzz Cannabis domain.
5. Sends an HTML email to the brand's recipients with the scheduled brand links
   and a bottom list of every generated brand folder link.

Requires:
- credentials.json (for Google OAuth Drive + Gmail)
- brand_config2.json (for daily brand scheduling, plus test-mode toggle)
"""

import os
import sys
import json
import subprocess
import datetime
import traceback
import shutil
import re
import argparse
import html
import pandas as pd
import time
from dotenv import load_dotenv

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from brand_inventory_rows import (
    add_product_metadata,
    inventory_columns_or_missing,
    normalize_inventory_base_frame,
    remove_sample_and_promo_rows,
    sort_inventory_report_frame,
    split_available_unavailable,
)
from drive_folder_refresh import clear_drive_folder_contents
from dutchie_api_reports import STORE_CODES, canonical_env_map, resolve_store_keys
from inventory_order_reports import (
    build_brand_order_sections,
    extract_store_code_from_filename,
    format_order_sheet,
    summarize_order_report_files,
    write_order_sections,
)

# ------------------------------------------------------------------------------
# ------------------------- CONFIG / CONSTANTS ----------------------------------
# ------------------------------------------------------------------------------

load_dotenv()
BASE_DIR = os.getenv(
    "BASE_DIR",
    os.path.dirname(os.path.abspath(__file__)),
)

# Folders for CSV input and XLSX output
INPUT_DIRECTORY = os.path.join(BASE_DIR, "files")       # Where CSVs land
LOCAL_REPORTS_FOLDER = os.path.join(BASE_DIR, "brand_reports_tmp")  # Local subfolder for generated reports
INVENTORY_LINKS_DIR = os.path.join(BASE_DIR, "inventory_links")
BRAND_CONFIG_JSON = os.path.join(BASE_DIR, "brand_config2.json")
CATALOG_API_SCRIPT = "getCatalog.py"
CATALOG_BROWSER_SCRIPT = "getCatalog_browser.py"
ORDER_REPORT_API_SCRIPT = "getInventoryOrderReport_api.py"
ORDER_REPORT_BROWSER_SCRIPT = "getInventoryOrderReport.py"
DEFAULT_API_ENV_FILE = os.path.join(BASE_DIR, ".env")
ORDER_REPORT_FILE_PATTERN = re.compile(
    r"^inventory_order_(7d|14d|30d)_[A-Za-z0-9]+\.(xlsx|xls|csv)$",
    re.IGNORECASE,
)
BRAND_REPORT_FILE_PATTERN = re.compile(r"^(.*?)_(.*?)_(\d{2}-\d{2}-\d{4})\.xlsx$", re.IGNORECASE)

# Google Drive parent folder name (where we create subfolders by date)
DRIVE_PARENT_FOLDER_NAME = "INVENTORY"
DRIVE_OTHER_PARENT_FOLDER_NAME = "INVENTORY_OTHER"
DEFAULT_OTHER_FOLDER_NAME = "OTHER"
DEFAULT_OTHER_FOLDER_DOMAIN = "buzzcannabis.com"

# OAuth credential files
CREDENTIALS_FILE = os.path.join(BASE_DIR, "credentials.json")
TOKEN_DRIVE_FILE = os.path.join(BASE_DIR, "token_drive.json")   # Stores Drive API tokens
TOKEN_GMAIL_FILE = os.path.join(BASE_DIR, "token_gmail.json")   # Stores Gmail API tokens

# Google Drive API Scopes
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# Gmail API Scopes
GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]

# ------------------------------------------------------------------------------
# --------------------- GMAIL API SEND HTML HELPER -----------------------------
# ------------------------------------------------------------------------------

def safe_move(src, dst, retries=3, delay=1):
    for _ in range(retries):
        try:
            shutil.move(src, dst)
            return True
        except PermissionError:
            print(f"[WARN] File in use: {src}. Retrying in {delay}s...")
            time.sleep(delay)
    print(f"[ERROR] Could not move {src} to {dst} after {retries} attempts.")
    return False
def gmail_authenticate():
    """
    Authenticate with Gmail API using OAUTH and return a service object.
    """
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(TOKEN_GMAIL_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_GMAIL_FILE, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_GMAIL_FILE, "w") as f:
            f.write(creds.to_json())

    return build("gmail", "v1", credentials=creds)


def send_email_with_gmail_html(subject, html_body, recipients):
    """
    Sends an HTML email via the Gmail API. 
    `recipients` can be a list or a single string.
    """
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if isinstance(recipients, str):
        recipients = [recipients]

    service = gmail_authenticate()

    msg = MIMEMultipart("alternative")
    msg["From"] = "me"  # 'me' means authenticated user
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject

    part_html = MIMEText(html_body, "html")
    msg.attach(part_html)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    body = {"raw": raw_message}

    try:
        sent = service.users().messages().send(userId="me", body=body).execute()
        print(f"[GMAIL] Email sent! ID: {sent['id']} | Subject: {subject}")
    except Exception as e:
        print(f"[ERROR] Could not send HTML email via Gmail API: {e}")


# ------------------------------------------------------------------------------
# ------------------------- GOOGLE DRIVE HELPER ---------------------------------
# ------------------------------------------------------------------------------

def drive_authenticate():
    """
    Authenticate & build the Google Drive service using OAuth.
    """
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    creds = None
    if os.path.exists(TOKEN_DRIVE_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_DRIVE_FILE, DRIVE_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_DRIVE_FILE, "w") as token:
            token.write(creds.to_json())

    return build("drive", "v3", credentials=creds)


def make_folder_public(service, folder_id):
    """
    Make the given Google Drive folder public (viewable by anyone with the link).
    """
    permissions = list_drive_permissions(service, folder_id)
    if permissions is not None:
        for permission in permissions:
            if permission.get("type") == "anyone" and permission.get("role") == "reader":
                print(f"[INFO] Folder ID {folder_id} is already public.")
                return

    try:
        permission = {
            "type": "anyone",
            "role": "reader"
        }
        service.permissions().create(fileId=folder_id, body=permission).execute()
        print(f"[INFO] Folder ID {folder_id} is now public.")
    except Exception as e:
        print(f"[ERROR] Could not make folder public: {e}")


from googleapiclient.errors import HttpError

def list_drive_permissions(service, folder_id):
    try:
        response = service.permissions().list(
            fileId=folder_id,
            fields="permissions(id,type,role,domain)"
        ).execute()
        return response.get("permissions", [])
    except Exception as e:
        print(f"[WARN] Could not inspect permissions for {folder_id}: {e}")
        return None


def remove_anyone_permissions(service, folder_id):
    """
    Remove direct public link permissions from a Drive item.

    This keeps container folders private/internal while child brand folders can
    still be made public with their own direct links.
    """
    permissions = list_drive_permissions(service, folder_id)
    if permissions is None:
        return

    for permission in permissions:
        if permission.get("type") != "anyone":
            continue
        permission_id = permission.get("id")
        if not permission_id:
            continue
        try:
            service.permissions().delete(fileId=folder_id, permissionId=permission_id).execute()
            print(f"[INFO] Removed public access from folder ID {folder_id}.")
        except Exception as e:
            print(f"[WARN] Could not remove public access from {folder_id}: {e}")


def make_folder_domain_viewable(service, folder_id, domain):
    """
    Make the folder viewable only by users in the given Google Workspace domain.
    """
    if not domain:
        return

    remove_anyone_permissions(service, folder_id)
    permissions = list_drive_permissions(service, folder_id)
    if permissions is not None:
        for permission in permissions:
            if (
                permission.get("type") == "domain"
                and permission.get("role") == "reader"
                and permission.get("domain") == domain
            ):
                print(f"[INFO] Folder ID {folder_id} is already viewable by {domain}.")
                return

    try:
        permission = {
            "type": "domain",
            "role": "reader",
            "domain": domain,
            "allowFileDiscovery": False,
        }
        service.permissions().create(fileId=folder_id, body=permission).execute()
        print(f"[INFO] Folder ID {folder_id} is now viewable by {domain} users with the link.")
    except Exception as e:
        print(f"[ERROR] Could not make folder domain-viewable for {domain}: {e}")


def apply_folder_sharing(service, folder_id, make_public=False, domain_view=None, remove_public=False):
    if remove_public:
        remove_anyone_permissions(service, folder_id)
    if domain_view:
        make_folder_domain_viewable(service, folder_id, domain_view)
    if make_public:
        make_folder_public(service, folder_id)


def find_or_create_folder(
    service,
    folder_name,
    parent_id=None,
    retries=5,
    delay=3,
    make_public=False,
    domain_view=None,
    remove_public=False,
):
    folder_name_escaped = folder_name.replace("'", "\\'")
    query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name_escaped}'"
    if parent_id:
        query += f" and '{parent_id}' in parents"

    try:
        response = service.files().list(q=query, spaces="drive", fields="files(id, name)").execute()
        folders = response.get("files", [])
    except HttpError as err:
        print(f"[ERROR] Drive folder lookup failed: {err}")
        return None

    if folders:
        folder_id = folders[0]["id"]
        apply_folder_sharing(
            service,
            folder_id,
            make_public=make_public,
            domain_view=domain_view,
            remove_public=remove_public,
        )
        return folder_id

    # Retry on timeout for folder creation
    for attempt in range(retries):
        try:
            folder_metadata = {
                "name": folder_name,
                "mimeType": "application/vnd.google-apps.folder"
            }
            if parent_id:
                folder_metadata["parents"] = [parent_id]

            new_folder = service.files().create(body=folder_metadata, fields="id").execute()
            folder_id = new_folder.get("id")
            print(f"[INFO] Created new folder '{folder_name}' (ID: {folder_id})")
            apply_folder_sharing(
                service,
                folder_id,
                make_public=make_public,
                domain_view=domain_view,
                remove_public=remove_public,
            )
            return folder_id

        except TimeoutError as e:
            print(f"[WARN] Timeout while creating folder '{folder_name}', attempt {attempt + 1}/{retries}")
            time.sleep(delay)
        except HttpError as e:
            print(f"[ERROR] Google API error while creating folder '{folder_name}': {e}")
            return None

    print(f"[ERROR] Failed to create folder '{folder_name}' after {retries} attempts.")
    return None

def upload_file_to_drive(service, file_path, folder_id):
    """
    Upload local file `file_path` to Google Drive in `folder_id`. Return file ID.
    """
    from googleapiclient.http import MediaFileUpload

    file_name = os.path.basename(file_path)
    file_metadata = {
        "name": file_name,
        "parents": [folder_id]
    }
    media = MediaFileUpload(file_path, resumable=True)
    drive_file = service.files().create(
        body=file_metadata,
        media_body=media,
        fields="id"
    ).execute()
    return drive_file.get("id")


# ------------------------------------------------------------------------------
# ---------------------- INVENTORY PROCESSING FUNCTIONS -------------------------
# ------------------------------------------------------------------------------

INPUT_COLUMNS = ['Available', 'Product', 'Category', 'Brand']

def safe_makedirs(path):
    """Create directory if it doesn't exist."""
    if not os.path.exists(path):
        os.makedirs(path)


def is_order_report_filename(filename):
    return bool(ORDER_REPORT_FILE_PATTERN.match(str(filename or "")))


def list_catalog_csv_files(directory):
    if not os.path.isdir(directory):
        return []
    return sorted(
        filename
        for filename in os.listdir(directory)
        if filename.lower().endswith(".csv") and not is_order_report_filename(filename)
    )


def dutchie_api_readiness(env_file=DEFAULT_API_ENV_FILE):
    expected_codes = list(STORE_CODES.keys())

    try:
        env_map = canonical_env_map(env_file)
        resolved = resolve_store_keys(env_map, expected_codes)
    except Exception as exc:
        return False, [], expected_codes, str(exc)

    available_codes = [code for code in expected_codes if code in resolved]
    missing_codes = [code for code in expected_codes if code not in resolved]
    return not missing_codes, available_codes, missing_codes, ""


def run_refresh_script(script_name, *args):
    script_path = os.path.join(BASE_DIR, script_name)
    if not os.path.exists(script_path):
        raise FileNotFoundError(f"{script_name} not found at {script_path}")

    cmd = [sys.executable, script_path, *[str(arg) for arg in args]]
    print(f"[INFO] Running {' '.join(cmd)}")
    subprocess.check_call(cmd)


def refresh_catalog_exports(output_directory):
    api_ready, available_codes, missing_codes, error_text = dutchie_api_readiness(DEFAULT_API_ENV_FILE)

    if api_ready:
        print(
            "[INFO] Dutchie API catalog refresh is configured for all stores "
            f"({', '.join(available_codes)}). Trying the API exporter first."
        )
        try:
            run_refresh_script(CATALOG_API_SCRIPT, output_directory)
            print("[INFO] Catalog CSV refresh complete via Dutchie API.")
            return "api"
        except subprocess.CalledProcessError as exc:
            print(f"[WARN] Dutchie API catalog refresh failed: {exc}")
            print("[INFO] Falling back to the browser catalog export script.")
    else:
        if error_text:
            print(
                "[WARN] Dutchie API readiness could not be confirmed for catalog refresh. "
                "Falling back to the browser catalog export script."
            )
        else:
            print(
                "[WARN] Dutchie API is missing store credentials for: "
                f"{', '.join(missing_codes)}. Falling back to the browser catalog export script."
            )

    run_refresh_script(CATALOG_BROWSER_SCRIPT, output_directory)
    print("[INFO] Catalog CSV refresh complete via browser export.")
    return "browser"


def refresh_inventory_order_reports(output_directory):
    api_ready, available_codes, missing_codes, error_text = dutchie_api_readiness(DEFAULT_API_ENV_FILE)

    if api_ready:
        print(
            "[INFO] Dutchie API order-report refresh is configured for all stores "
            f"({', '.join(available_codes)}). Trying the API exporter first."
        )
        try:
            run_refresh_script(ORDER_REPORT_API_SCRIPT, output_directory)
            print("[INFO] Inventory order report refresh complete via Dutchie API.")
            return "api"
        except subprocess.CalledProcessError as exc:
            print(f"[WARN] Dutchie API inventory order report refresh failed: {exc}")
            print("[INFO] Falling back to the browser inventory order report script.")
    else:
        if error_text:
            print(
                "[WARN] Dutchie API readiness could not be confirmed for inventory order reports. "
                "Falling back to the browser inventory order report script."
            )
        else:
            print(
                "[WARN] Dutchie API is missing store credentials for: "
                f"{', '.join(missing_codes)}. Falling back to the browser inventory order report script."
            )

    run_refresh_script(ORDER_REPORT_BROWSER_SCRIPT, output_directory)
    print("[INFO] Inventory order report refresh complete via browser export.")
    return "browser"


def clear_old_input_exports(directory):
    """
    Remove prior catalog CSVs and prior inventory-order exports so a fresh run
    does not mix old and new source data.
    """
    if not os.path.exists(directory):
        return

    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if not os.path.isfile(file_path):
            continue
        if (filename.lower().endswith(".csv") and not is_order_report_filename(filename)) or is_order_report_filename(filename):
            try:
                os.remove(file_path)
                print(f"[INFO] Deleted old source export: {file_path}")
            except Exception as e:
                print(f"[ERROR] Could not delete {file_path}: {e}")


def load_other_folder_settings(config):
    """
    Read optional OTHER-folder settings from brand_config2.json.

    Defaults are enabled because the OTHER folder is meant to catch every brand,
    including brands that are not part of the scheduled email list.
    """
    settings = config.get("other_folder", {})
    if settings is False:
        return {
            "enabled": False,
            "parent_folder_name": DRIVE_OTHER_PARENT_FOLDER_NAME,
            "folder_name": DEFAULT_OTHER_FOLDER_NAME,
            "domain": DEFAULT_OTHER_FOLDER_DOMAIN,
        }
    if not isinstance(settings, dict):
        settings = {}

    parent_folder_name = str(settings.get("parent_folder_name", DRIVE_OTHER_PARENT_FOLDER_NAME) or "").strip()
    folder_name = str(settings.get("folder_name", DEFAULT_OTHER_FOLDER_NAME) or "").strip()
    domain = str(settings.get("domain", DEFAULT_OTHER_FOLDER_DOMAIN) or "").strip()
    return {
        "enabled": bool(settings.get("enabled", True)),
        "parent_folder_name": parent_folder_name or DRIVE_OTHER_PARENT_FOLDER_NAME,
        "folder_name": folder_name or DEFAULT_OTHER_FOLDER_NAME,
        "domain": domain or DEFAULT_OTHER_FOLDER_DOMAIN,
    }


def drive_folder_link(folder_id):
    return f"https://drive.google.com/drive/folders/{folder_id}"


def safe_report_filename_part(value):
    """
    Make a brand/store label safe for local workbook filenames.

    Dutchie brand names can contain "/" or trailing whitespace. A slash is fine
    as a Google Drive folder label, but it becomes a path separator locally.
    """
    safe_value = str(value or "").strip()
    safe_value = re.sub(r"[\\/:*?\"<>|]+", " - ", safe_value)
    safe_value = re.sub(r"\s+", " ", safe_value).strip().rstrip(".")
    return safe_value or "Unknown"


def parse_brand_from_report_filename(filename):
    match = BRAND_REPORT_FILE_PATTERN.match(os.path.basename(str(filename or "")))
    if not match:
        return None
    _, brand_name, _ = match.groups()
    return brand_name


def group_generated_files_by_brand(generated_files):
    grouped = {}
    for file_path in generated_files:
        brand_name = parse_brand_from_report_filename(file_path)
        if not brand_name:
            print(f"[WARN] Cannot parse brand from {os.path.basename(file_path)}, skipping OTHER upload.")
            continue
        grouped.setdefault(brand_name, []).append(file_path)
    return grouped


def upload_other_brand_reports_to_drive(service, generated_files, date_folder_id, other_settings):
    """
    Create a domain-only OTHER folder and public child folders for every brand.

    Buzz users can browse the OTHER link. Each child brand folder gets its own
    public link so it can be shared directly with outside recipients.
    """
    if not other_settings.get("enabled"):
        return None, {}

    other_folder_name = other_settings.get("folder_name") or DEFAULT_OTHER_FOLDER_NAME
    other_domain = other_settings.get("domain") or DEFAULT_OTHER_FOLDER_DOMAIN
    other_folder_id = find_or_create_folder(
        service,
        other_folder_name,
        parent_id=date_folder_id,
        domain_view=other_domain,
    )
    if not other_folder_id:
        print(f"[ERROR] Could not create/find {other_folder_name} folder.")
        return None, {}

    other_folder_link = drive_folder_link(other_folder_id)
    other_brand_links = {}
    brand_file_map = group_generated_files_by_brand(generated_files)
    print(
        f"[INFO] Uploading {sum(len(files) for files in brand_file_map.values())} workbook(s) "
        f"across {len(brand_file_map)} brand folder(s) into {other_folder_name}."
    )

    for brand_name, files in sorted(brand_file_map.items()):
        brand_folder_id = find_or_create_folder(
            service,
            brand_name,
            parent_id=other_folder_id,
            make_public=True,
        )
        if not brand_folder_id:
            print(f"[ERROR] Could not create/find OTHER child folder for {brand_name}.")
            continue

        other_brand_links[brand_name] = drive_folder_link(brand_folder_id)
        clear_drive_folder_contents(
            service,
            brand_folder_id,
            folder_label=f"{other_folder_name}/{brand_name}",
        )
        for file_path in files:
            try:
                upload_file_to_drive(service, file_path, brand_folder_id)
                print(f"[OTHER UPLOAD] {os.path.basename(file_path)} uploaded to {other_folder_name}/{brand_name}")
                time.sleep(0.2)
            except Exception as e:
                print(f"[ERROR] Failed to upload {file_path} to {other_folder_name}/{brand_name}: {e}")

    return other_folder_link, other_brand_links


def build_all_brand_links_email_section(brand_links, heading="All generated brand folders"):
    if not brand_links:
        return ""

    rows = []
    for brand_name, link in sorted(brand_links.items(), key=lambda item: item[0].lower()):
        escaped_brand = html.escape(str(brand_name))
        escaped_link = html.escape(str(link), quote=True)
        visible_link = html.escape(str(link))
        rows.append(
            "<tr>"
            f"<td style=\"padding:4px 10px 4px 0; white-space:nowrap;\"><strong>{escaped_brand}</strong></td>"
            f"<td style=\"padding:4px 0;\"><a href=\"{escaped_link}\">{visible_link}</a></td>"
            "</tr>"
        )

    escaped_heading = html.escape(str(heading))
    rows_html = "\n".join(rows)
    return f"""
          <hr>
          <h3>{escaped_heading}</h3>
          <table role="presentation" cellspacing="0" cellpadding="0" border="0">
            {rows_html}
          </table>
        """


def build_inventory_email_subject(today_name):
    return f"Brand Inventory Reports for {today_name}"


def write_inventory_link_manifest(
    date_str,
    today_name,
    brand_folder_links,
    brand_to_emails,
    other_folder_link=None,
    other_brand_folder_links=None,
    other_folder_name=DEFAULT_OTHER_FOLDER_NAME,
    other_domain=DEFAULT_OTHER_FOLDER_DOMAIN,
    other_parent_folder_name=DRIVE_OTHER_PARENT_FOLDER_NAME,
):
    """
    Persist Drive folder links so other scripts can reuse the inventory links
    without scraping prior emails.
    """
    safe_makedirs(INVENTORY_LINKS_DIR)

    manifest = {
        "date": date_str,
        "day": today_name,
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "folders": {
            folder_name: {
                "link": link,
                "emails": brand_to_emails.get(folder_name, []),
            }
            for folder_name, link in sorted(brand_folder_links.items())
        },
    }
    if other_folder_link:
        manifest["other_folder"] = {
            "parent_folder_name": other_parent_folder_name,
            "folder_name": other_folder_name,
            "link": other_folder_link,
            "domain": other_domain,
            "access": f"{other_domain} users with the link can browse this folder.",
            "brand_folders": {
                brand_name: {"link": link}
                for brand_name, link in sorted((other_brand_folder_links or {}).items())
            },
        }

    dated_path = os.path.join(INVENTORY_LINKS_DIR, f"{date_str}.json")
    latest_path = os.path.join(INVENTORY_LINKS_DIR, "latest.json")

    with open(dated_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, sort_keys=True)

    with open(latest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, indent=2, sort_keys=True)

    print(f"[INFO] Wrote inventory link manifest: {dated_path}")
    print(f"[INFO] Updated inventory link manifest: {latest_path}")

    if other_folder_link:
        dated_other_path = os.path.join(INVENTORY_LINKS_DIR, f"{date_str}_other_links.txt")
        latest_other_path = os.path.join(INVENTORY_LINKS_DIR, "latest_other_links.txt")
        lines = [
            f"OTHER ({other_domain} users only): {other_folder_link}",
            "",
            "Public child brand folders:",
        ]
        for brand_name, link in sorted((other_brand_folder_links or {}).items()):
            lines.append(f"{brand_name}: {link}")
        other_links_text = "\n".join(lines) + "\n"

        with open(dated_other_path, "w", encoding="utf-8") as f:
            f.write(other_links_text)
        with open(latest_other_path, "w", encoding="utf-8") as f:
            f.write(other_links_text)

        print(f"[INFO] Wrote OTHER link list: {dated_other_path}")
        print(f"[INFO] Updated OTHER link list: {latest_other_path}")

def format_excel_file(filename: str):
    """
    **ADVANCED** Excel formatting:  
    1) Freeze header row,  
    2) Bold + fill header,  
    3) Auto-fit columns,  
    4) Insert category rows for 'Category' changes,  
    5) Etc.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(filename)

    for ws in wb.worksheets:
        if format_order_sheet(ws):
            continue

        # Freeze the first row
        ws.freeze_panes = "A2"

        # Format the header row
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # Optional: find columns by name
        available_col = None
        category_col = None
        strain_col = None
        product_col = None
        brand_col = None
        for i, cell in enumerate(ws[1], start=1):
            val = (cell.value or '').lower()
            if val == 'category':
                category_col = i
            elif val == 'available':
                available_col = i
            elif val == 'product':
                product_col = i
            elif val == 'brand':
                brand_col = i
            elif val == 'strain_type':
                strain_col = i

        # Widen the "Available" column
        if available_col:
            col_letter = get_column_letter(available_col)
            if ws.column_dimensions[col_letter].width < 20:
                ws.column_dimensions[col_letter].width = 20

        # Insert grouping rows whenever the Category changes
        if category_col:
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            current_type = None
            insert_positions = []
            for idx, row_data in enumerate(rows, start=2):
                cat_val = row_data[category_col - 1]
                if cat_val != current_type:
                    if current_type is not None:
                        insert_positions.append(idx)
                    current_type = cat_val
            if rows:
                # Insert a row at top if there's data
                insert_positions.insert(0, 2)

            category_font = Font(bold=True, size=14)
            fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

            current_type = None
            group_types = []
            row_counter = 2
            for row_data in rows:
                cat_val = row_data[category_col - 1]
                if cat_val != current_type:
                    group_types.append((row_counter, cat_val))
                    current_type = cat_val
                row_counter += 1

            # Insert group headers in reverse order so indexing doesn't shift
            for (pos, cat_value_info) in zip(reversed(insert_positions), reversed(group_types)):
                _, cat_value = cat_value_info
                ws.insert_rows(pos, 1)
                header_cell = ws.cell(row=pos, column=1)
                header_cell.value = f"{cat_value}"
                header_cell.font = category_font
                header_cell.fill = fill
                header_cell.alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)

def process_file(file_path, output_directory, selected_brands):
    """
    Process a single CSV file, filtering to only the selected brands.
    Returns (unavailable_data, processed_file_base_name).
    """
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None, None

    source_columns = list(df.columns)
    df, missing_columns = inventory_columns_or_missing(df, INPUT_COLUMNS, ["Cost"])
    if missing_columns:
        print(
            f"[WARN] {file_path} is missing required columns {INPUT_COLUMNS}. "
            f"Found columns: {source_columns}. Skipped."
        )
        return None, None

    df = remove_sample_and_promo_rows(df)
    df = normalize_inventory_base_frame(df)
    available_data, unavailable_data = split_available_unavailable(df)

    # If we only want certain brands:
    if 'Brand' in available_data.columns and selected_brands:
        available_data = available_data[available_data['Brand'].isin(selected_brands)].copy()

    available_data = add_product_metadata(
        available_data,
        include_details=True,
        filter_empty_products=True,
    )

    available_data = sort_inventory_report_frame(
        available_data,
        include_cost_as_tiebreaker='Cost' in available_data.columns,
    )

    unavailable_data = sort_inventory_report_frame(
        unavailable_data,
        include_cost_as_tiebreaker='Cost' in unavailable_data.columns,
    )

    # Drop Cost column after sorting
    if 'Cost' in available_data.columns:
        available_data = available_data.drop(columns=['Cost'])
    if 'Cost' in unavailable_data.columns:
        unavailable_data = unavailable_data.drop(columns=['Cost'])


    base_name = os.path.splitext(os.path.basename(file_path))[0]
    parts = base_name.split('_')
    store_name = parts[-1] if len(parts) > 1 else "Unknown"
    store_code = extract_store_code_from_filename(base_name)

    # Create subfolder for this CSV
    sub_out = os.path.join(output_directory, base_name)
    safe_makedirs(sub_out)

    today_str = datetime.datetime.now().strftime("%m-%d-%Y")

    if 'Brand' in available_data.columns:
        # Group by brand
        if available_data.empty:
            # If all data was filtered out
            out_xlsx = os.path.join(sub_out, f"{store_name}_{base_name}_{today_str}.xlsx")
            with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
                available_data.to_excel(writer, index=False, sheet_name="Available")
                if not unavailable_data.empty:
                    unavailable_data.to_excel(writer, index=False, sheet_name="Unavailable")
            format_excel_file(out_xlsx)
            print(f"[INFO] Created {out_xlsx} (no brand data after filtering).")
        else:
            for brand_name, brand_data in available_data.groupby('Brand'):
                safe_brand_name = safe_report_filename_part(brand_name)
                out_xlsx = os.path.join(sub_out, f"{store_name}_{safe_brand_name}_{today_str}.xlsx")
                order_sections = build_brand_order_sections(
                    INPUT_DIRECTORY,
                    brand_aliases=[brand_name],
                    store_code=store_code,
                )
                with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
                    brand_data.to_excel(writer, index=False, sheet_name="Available")
                    if not unavailable_data.empty and 'Brand' in unavailable_data.columns:
                        brand_unavail = unavailable_data[unavailable_data['Brand'] == brand_name]
                        if not brand_unavail.empty:
                            brand_unavail.to_excel(writer, index=False, sheet_name="Unavailable")
                    write_order_sections(writer, order_sections)

                format_excel_file(out_xlsx)
                print(f"[INFO] Created {out_xlsx}")
    else:
        # No Brand column
        out_xlsx = os.path.join(sub_out, f"{store_name}_{base_name}_{today_str}.xlsx")
        with pd.ExcelWriter(out_xlsx, engine='openpyxl') as writer:
            available_data.to_excel(writer, index=False, sheet_name="Available")
            if not unavailable_data.empty:
                unavailable_data.to_excel(writer, index=False, sheet_name="Unavailable")
        format_excel_file(out_xlsx)
        print(f"[INFO] Created {out_xlsx}")

    return unavailable_data, base_name

def organize_by_brand(output_directory):
    """
    Moves XLSX files into subfolders named after the brand if their 
    filename is "<Store>_<Brand>_<MM-DD-YYYY>.xlsx".
    """
    pattern = re.compile(r"^(.*?)_(.*?)_(\d{2}-\d{2}-\d{4})\.xlsx$")

    for root, dirs, files in os.walk(output_directory):
        for f in files:
            if f.lower().endswith(".xlsx"):
                match = pattern.match(f)
                if match:
                    _, brand_name, _ = match.groups()
                    if os.path.basename(root) == brand_name:
                        continue
                    brand_folder = os.path.join(output_directory, brand_name)
                    safe_makedirs(brand_folder)

                    old_path = os.path.join(root, f)
                    new_path = os.path.join(brand_folder, f)
                    print(f"Moving {old_path} → {new_path}")
                    safe_move(old_path, new_path)

def process_files(input_directory, output_directory, selected_brands):
    """
    Iterate all CSV files in `input_directory`, process them (filter by `selected_brands`),
    place XLSXs into `output_directory`. Then re-organize by brand subfolders.
    Returns a list of all final XLSX file paths.
    """
    safe_makedirs(output_directory)

    # Process each CSV
    for fn in list_catalog_csv_files(input_directory):
        csv_path = os.path.join(input_directory, fn)
        try:
            process_file(csv_path, output_directory, selected_brands)
        except Exception as e:
            print(f"[ERROR] While processing {fn}: {e}")

    # Re-organize by brand
    organize_by_brand(output_directory)

    # Collect final XLSX
    final_files = []
    for root, dirs, files in os.walk(output_directory):
        for f in files:
            if f.lower().endswith(".xlsx"):
                final_files.append(os.path.join(root, f))

    return final_files


# ------------------------------------------------------------------------------
# --------------------------------- MAIN ---------------------------------------
# ------------------------------------------------------------------------------

def build_arg_parser():
    parser = argparse.ArgumentParser(description="Build, upload, and optionally email brand inventory reports.")
    parser.add_argument(
        "--other-only",
        action="store_true",
        help="Only build/upload the OTHER folder. Skips scheduled brand folders and Gmail emails.",
    )
    parser.add_argument(
        "--no-refresh",
        action="store_true",
        help="Use the existing files/ CSV and order-report exports instead of refreshing Dutchie first.",
    )
    return parser


def main(argv=None):
    args = build_arg_parser().parse_args(argv)

    # 1) Clear out prior catalog/order source exports from the input directory.
    if args.no_refresh:
        print("[INFO] NO REFRESH mode ON => using existing files/ exports.")
    else:
        clear_old_input_exports(INPUT_DIRECTORY)

    # 2) Determine today's day name
    today_name = datetime.datetime.now().strftime("%A")  # e.g. "Monday", "Tuesday"

    # 3) Load brand_config.json
    if not os.path.exists(BRAND_CONFIG_JSON):
        print(f"[ERROR] Cannot find {BRAND_CONFIG_JSON}. Exiting.")
        sys.exit(1)

    with open(BRAND_CONFIG_JSON, "r", encoding="utf-8") as f:
        config = json.load(f)

    # read top-level test_mode, test_email
    test_mode = config.get("test_mode", True)
    test_email = config.get("test_email", "anthony@barbaro.tech")
    other_settings = load_other_folder_settings(config)
    other_only = bool(args.other_only or config.get("other_only", False))
    if other_only:
        other_settings["enabled"] = True

    # read brand array
    brand_cfgs = config.get("brands", [])
    if not brand_cfgs and not other_settings.get("enabled"):
        print("[INFO] No brand definitions found in brand_config.json -> 'brands' array.")
        return
    if not brand_cfgs:
        print("[INFO] No brand definitions found; continuing with OTHER folder only.")

    # ---------------------------------------------------------------
    # 4) Build a dictionary of brand synonyms -> (folder_name, emails)
    #    Also build brand_to_emails keyed by folder_name for emailing
    # ---------------------------------------------------------------
    synonym_to_folder = {}
    brand_to_emails = {}   # key = folder_name, value = final_emails

    for item in brand_cfgs:
        # brand_synonyms is a list of exact brand names from the CSV 'Brand' column
        synonyms = item.get("brand_synonyms", [])
        if isinstance(synonyms, str):
            synonyms = [synonyms]

        # fallback to old "brand" field if brand_synonyms is empty
        if not synonyms and "brand" in item:
            brand_str = item["brand"]
            synonyms = [b.strip() for b in brand_str.split('/')]

        folder_name = item.get("folder_name")
        if not folder_name:
            # if user didn't provide folder_name, fallback to first synonym
            folder_name = synonyms[0] if synonyms else "Unknown"

        real_emails = item.get("emails", [])
        days = item.get("days", [])
        location_str = item.get("location", "")  # optional reference

        # skip if not scheduled today
        if today_name not in days:
            continue

        # if test_mode => override emails
        final_emails = [test_email] if test_mode else real_emails

        # For each synonym brand name, map to (folder_name, final_emails)
        for syn in synonyms:
            synonym_to_folder[syn] = (folder_name, final_emails)
            safe_syn = safe_report_filename_part(syn)
            synonym_to_folder[safe_syn] = (folder_name, final_emails)

        # We'll store folder_name -> final_emails in brand_to_emails
        brand_to_emails[folder_name] = final_emails

    # If no folder_name is active and OTHER is disabled, exit.
    if not brand_to_emails and not other_settings.get("enabled"):
        print(f"[INFO] No brands scheduled for {today_name}.")
        return

    # active_brands is the set of all "folder_name" keys from brand_to_emails
    active_brands = set() if other_only else set(brand_to_emails.keys())

    print(f"[INFO] Today is {today_name}, active brand folders: {active_brands}")
    if test_mode:
        print(f"[INFO] TEST MODE ON => all emails go to {test_email}")
    if other_only:
        print("[INFO] OTHER ONLY mode ON => scheduled brand folders and Gmail emails will be skipped.")
    if other_settings.get("enabled"):
        print(
            f"[INFO] OTHER folder enabled => all brands will be uploaded under "
            f"{other_settings['folder_name']} for {other_settings['domain']} users."
        )

    # 5) Refresh source exports with Dutchie API preference and browser fallback.
    catalog_mode_used = None
    order_mode_used = None

    if args.no_refresh:
        print("[INFO] Source refresh skipped by --no-refresh.")
    else:
        try:
            catalog_mode_used = refresh_catalog_exports(INPUT_DIRECTORY)
        except FileNotFoundError as exc:
            print(f"[WARN] {exc}. Skipping catalog refresh step.")
        except subprocess.CalledProcessError as exc:
            print(f"[ERROR] Catalog refresh failed: {exc}")
        except Exception as exc:
            print(f"[ERROR] Unexpected catalog refresh failure: {exc}")

        try:
            order_mode_used = refresh_inventory_order_reports(INPUT_DIRECTORY)
        except FileNotFoundError as exc:
            print(f"[WARN] {exc}. Skipping inventory order report refresh step.")
        except subprocess.CalledProcessError as exc:
            print(f"[ERROR] Inventory order report refresh failed: {exc}")
        except Exception as exc:
            print(f"[ERROR] Unexpected inventory order report refresh failure: {exc}")

    if catalog_mode_used or order_mode_used:
        print(
            "[INFO] Source refresh summary: "
            f"catalog={catalog_mode_used or 'skipped'}, order_reports={order_mode_used or 'skipped'}"
        )

    # ----------------------------------------------------------------
    # 6) Process CSVs.
    #    OTHER mode needs every brand. Otherwise only scheduled config synonyms.
    # ----------------------------------------------------------------
    synonyms_for_today = [] if other_settings.get("enabled") else list(synonym_to_folder.keys())
    safe_makedirs(LOCAL_REPORTS_FOLDER)
    generated_files = process_files(INPUT_DIRECTORY, LOCAL_REPORTS_FOLDER, synonyms_for_today)

    if not generated_files:
        print("[INFO] No XLSX files were generated. Possibly no data matched.")
        return
    
    # 7) Upload to Google Drive
    drive_service = drive_authenticate()
    date_str = datetime.datetime.now().strftime("%Y-%m-%d")
    brand_folder_links = {}

    if active_brands:
        parent_folder_id = find_or_create_folder(
            drive_service,
            DRIVE_PARENT_FOLDER_NAME,
            parent_id=None,
            remove_public=True,
        )
        if not parent_folder_id:
            print(f"[ERROR] Could not find/create Drive parent folder {DRIVE_PARENT_FOLDER_NAME}.")
            return

        date_folder_id = find_or_create_folder(
            drive_service,
            date_str,
            parent_id=parent_folder_id,
            remove_public=True,
        )
        if not date_folder_id:
            print(f"[ERROR] Could not find/create Drive date folder {date_str}.")
            return

        # For each folder_name in active_brands, create on Drive
        for folder_name in active_brands:
            brand_folder_id = find_or_create_folder(
                drive_service,
                folder_name,
                parent_id=date_folder_id,
                make_public=True,
            )
            if not brand_folder_id:
                print(f"[ERROR] Could not create/find Drive folder for {folder_name}.")
                continue
            clear_drive_folder_contents(
                drive_service,
                brand_folder_id,
                folder_label=f"{DRIVE_PARENT_FOLDER_NAME}/{date_str}/{folder_name}",
            )
            link = drive_folder_link(brand_folder_id)
            brand_folder_links[folder_name] = link

        # Now, parse brand from each generated XLSX => find folder_name => upload
        for file_path in generated_files:
            filename = os.path.basename(file_path)
            brand_syn = parse_brand_from_report_filename(filename)
            if not brand_syn:
                print(f"[WARN] Cannot parse brand from {filename}, skipping.")
                continue

            if brand_syn not in synonym_to_folder:
                print(f"[WARN] brand '{brand_syn}' not recognized. Skipping.")
                continue

            folder_name, _ = synonym_to_folder[brand_syn]
            if folder_name not in active_brands:
                continue

            # Reuse folder ID from earlier lookup.
            if folder_name in brand_folder_links:
                brand_folder_id = brand_folder_links[folder_name].split("/")[-1]
            else:
                print(f"[ERROR] Missing folder ID for {folder_name}, skipping upload.")
                continue

            try:
                upload_file_to_drive(drive_service, file_path, brand_folder_id)
                print(f"[UPLOAD] {filename} uploaded to {folder_name}")
                time.sleep(0.2)  # Google API throttle protection
            except Exception as e:
                print(f"[ERROR] Failed to upload {filename} to {folder_name}: {e}")

    other_folder_link = None
    other_brand_folder_links = {}
    if other_settings.get("enabled"):
        other_parent_folder_name = other_settings.get("parent_folder_name", DRIVE_OTHER_PARENT_FOLDER_NAME)
        other_parent_folder_id = find_or_create_folder(
            drive_service,
            other_parent_folder_name,
            parent_id=None,
            remove_public=True,
        )
        if not other_parent_folder_id:
            print(f"[ERROR] Could not find/create Drive parent folder {other_parent_folder_name}.")
            return

        other_date_folder_id = find_or_create_folder(
            drive_service,
            date_str,
            parent_id=other_parent_folder_id,
            remove_public=True,
        )
        if not other_date_folder_id:
            print(f"[ERROR] Could not find/create Drive date folder {date_str} under {other_parent_folder_name}.")
            return

        other_folder_link, other_brand_folder_links = upload_other_brand_reports_to_drive(
            drive_service,
            generated_files,
            other_date_folder_id,
            other_settings,
        )
        if other_folder_link:
            print(
                f"[INFO] OTHER folder link ({other_settings['domain']} only): "
                f"{other_folder_link}"
            )

    write_inventory_link_manifest(
        date_str=date_str,
        today_name=today_name,
        brand_folder_links=brand_folder_links,
        brand_to_emails=brand_to_emails,
        other_folder_link=other_folder_link,
        other_brand_folder_links=other_brand_folder_links,
        other_folder_name=other_settings.get("folder_name", DEFAULT_OTHER_FOLDER_NAME),
        other_domain=other_settings.get("domain", DEFAULT_OTHER_FOLDER_DOMAIN),
        other_parent_folder_name=other_settings.get("parent_folder_name", DRIVE_OTHER_PARENT_FOLDER_NAME),
    )

    # 8) Email out the folder link
    # Group by unique sets of emails
    email_groups = {}
    email_source = {} if other_only else brand_to_emails
    for folder_name, email_list in email_source.items():
        email_key = frozenset(email_list)
        if email_key not in email_groups:
            email_groups[email_key] = []
        email_groups[email_key].append(folder_name)

    for email_key, folder_list in email_groups.items():
        order_summary = summarize_order_report_files(INPUT_DIRECTORY)
        order_note = ""
        if order_summary:
            order_note = (
                "<p>Matching Dutchie order-report rows were added to the "
                "<strong>Order_7d</strong>, <strong>Order_14d</strong>, and "
                f"<strong>Order_30d</strong> tabs when available. Source windows found: {order_summary}.</p>"
            )

        brand_lines = []
        for f_name in folder_list:
            link = brand_folder_links.get(f_name)
            if link:
                brand_lines.append(f"<h3>Folder: {f_name}</h3>")
                brand_lines.append(f"<p>Link: <a href='{link}'>{link}</a></p>")
            else:
                brand_lines.append(f"<p>No link found for {f_name}</p>")

        brand_html = "\n".join(brand_lines)
        all_brand_links_html = build_all_brand_links_email_section(
            other_brand_folder_links or brand_folder_links
        )
        subject = build_inventory_email_subject(today_name)
        html_body = f"""
        <html>
        <body>
          <p>Hello,</p>
          <p>Below are your brand inventory reports for <strong>{today_name}</strong>.</p>
          {order_note}
          {brand_html}
          <p>All files in the listed brand folders are viewable by anyone with the link.</p>
          {all_brand_links_html}
          <p>Regards,<br>Buzz Cannabis</p>
        </body>
        </html>
        """

        recipients = list(email_key)
        print(f"[INFO] Sending Gmail API email to {recipients} for folders {folder_list} ...")
        send_email_with_gmail_html(subject, html_body, recipients)

    print("[INFO] All done!")

    # 9) Clean up
    if os.path.exists(LOCAL_REPORTS_FOLDER):
        try:
            shutil.rmtree(LOCAL_REPORTS_FOLDER)
            print(f"[INFO] Deleted temporary folder: {LOCAL_REPORTS_FOLDER}")
        except Exception as e:
            print(f"[ERROR] Could not delete {LOCAL_REPORTS_FOLDER}: {e}")



# ------------------------------------------------------------------------------
if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("[FATAL] Unhandled exception in BrandINVEmailer.py:")
        traceback.print_exc()
        sys.exit(1)
