#!/usr/bin/env python3

"""
BrandInventoryGUIAdvanced.py

GUI that:
1) Lets you pick a folder of CSVs and an output folder (loaded from/saved to config.txt).
2) Loads brand names found in the CSV 'Brand' column (lowercased and trimmed).
3) Filters data to the selected brand(s) & splits them into "Available" (>2) and "Unavailable" (<=2),
   generating one XLSX per brand with advanced Excel formatting.
4) Uploads each brand’s XLSX to a date-based folder in Google Drive: 
     INVENTORY -> <YYYY-MM-DD> -> <brandName>  (folder is made public).
5) Sends an HTML email with each brand's public Drive folder link to the specified recipients.

Packages needed:
 - pandas, openpyxl
 - google-auth, google-auth-oauthlib, google-api-python-client
 - credentials.json for Google OAuth (Drive + Gmail)
 - token_drive.json, token_gmail.json are created automatically after first login.
 - config.txt (optional; stores your input/output folder paths).
"""

import os
import re
import json
import shutil
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
import pandas as pd
import traceback
from datetime import datetime
import subprocess
import sys
import time
import queue
import threading
from pathlib import Path
# For Excel formatting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from dutchie_api_reports import STORE_CODES, canonical_env_map, resolve_store_keys
from inventory_order_reports import (
    build_brand_order_sections,
    extract_store_code_from_filename,
    format_order_sheet,
    summarize_order_report_files,
    write_order_sections,
)

# Google API imports
import google.auth.transport.requests
from google_auth_oauthlib.flow import InstalledAppFlow
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

try:
    import ttkbootstrap as tb
except Exception:
    tb = None

# ----------------- CONFIG -----------------

# File where we store the chosen input & output dirs.
CONFIG_FILE = "config.txt"

# Google Drive parent folder name
DRIVE_PARENT_FOLDER_NAME = "INVENTORY"

# OAuth credential files
CREDENTIALS_FILE = "credentials.json"
TOKEN_DRIVE_FILE = "token_drive.json"
TOKEN_GMAIL_FILE = "token_gmail.json"

# Google Drive API Scopes
DRIVE_SCOPES = ["https://www.googleapis.com/auth/drive.file"]

# Gmail API Scopes
GMAIL_SCOPES = ["https://www.googleapis.com/auth/gmail.send"]
ORDER_REPORT_API_SCRIPT = "getInventoryOrderReport_api.py"
ORDER_REPORT_BROWSER_SCRIPT = "getInventoryOrderReport.py"
CATALOG_API_SCRIPT = "getCatalog.py"
CATALOG_BROWSER_SCRIPT = "getCatalog_browser.py"
DEFAULT_API_ENV_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
DUTCHIE_API_WORKERS = 15
DEFAULT_WINDOW_WIDTH = 1180
DEFAULT_WINDOW_HEIGHT = 760
WINDOW_EDGE_PADDING = 24

# Required CSV columns + optional
REQUIRED_COLUMNS = ["Available", "Product", "Brand"]
OPTIONAL_COLUMNS = ["Category", "Cost"]

# We'll consider Available <= 2 => "Unavailable"
MAX_AVAIL_FOR_UNAVAILABLE = 2

DEFAULT_GUI_CONFIG = {
    "input_dir": "",
    "output_dir": "",
    "fetch_order_reports": True,
    "emails": "",
    "include_cost": True,
    "prefer_catalog_api": True,
    "auto_update_on_launch": True,
    "auto_load_brands_after_update": True,
    "auto_load_brands_on_launch": True,
    "show_startup_loading": True,
    "open_output_after_complete": False,
    "theme": "flatly",
    "compact_mode": True,
    "task_eta_seconds": {},
}

THEME_CHOICES = ("flatly", "minty", "litera", "cosmo", "darkly")

THEME_PALETTES = {
    "flatly": {
        "bg": "#F6F8F7",
        "surface": "#FFFFFF",
        "border": "#D8E3DF",
        "text": "#12302A",
        "muted": "#6B7C78",
        "accent": "#0F766E",
        "accent_hover": "#0D5F59",
        "accent_soft": "#E7F4F2",
        "success": "#15803D",
        "warning": "#B45309",
        "danger": "#B91C1C",
        "log_bg": "#FAFBFB",
    },
    "minty": {
        "bg": "#F3FAF7",
        "surface": "#FFFFFF",
        "border": "#CFE6DE",
        "text": "#12302A",
        "muted": "#5F746E",
        "accent": "#12816F",
        "accent_hover": "#0D6659",
        "accent_soft": "#DFF5EF",
        "success": "#15803D",
        "warning": "#A16207",
        "danger": "#B91C1C",
        "log_bg": "#F8FCFA",
    },
    "litera": {
        "bg": "#F7F7F5",
        "surface": "#FFFFFF",
        "border": "#DDDCD5",
        "text": "#26312E",
        "muted": "#747B77",
        "accent": "#0F766E",
        "accent_hover": "#0B5F59",
        "accent_soft": "#E7F0EE",
        "success": "#15803D",
        "warning": "#B45309",
        "danger": "#B91C1C",
        "log_bg": "#FBFBFA",
    },
    "cosmo": {
        "bg": "#F4F7FA",
        "surface": "#FFFFFF",
        "border": "#D5E0E8",
        "text": "#122B39",
        "muted": "#667987",
        "accent": "#0E7490",
        "accent_hover": "#155E75",
        "accent_soft": "#E2F3F7",
        "success": "#15803D",
        "warning": "#B45309",
        "danger": "#B91C1C",
        "log_bg": "#F8FBFD",
    },
    "darkly": {
        "bg": "#111827",
        "surface": "#1F2937",
        "border": "#374151",
        "text": "#F9FAFB",
        "muted": "#CBD5E1",
        "accent": "#2DD4BF",
        "accent_hover": "#14B8A6",
        "accent_soft": "#134E4A",
        "success": "#22C55E",
        "warning": "#F59E0B",
        "danger": "#F87171",
        "log_bg": "#0F172A",
    },
}


def normalize_theme_name(theme_name):
    theme_name = str(theme_name or DEFAULT_GUI_CONFIG["theme"]).strip().lower()
    if theme_name not in THEME_CHOICES:
        return DEFAULT_GUI_CONFIG["theme"]
    return theme_name


def theme_palette(theme_name):
    return dict(THEME_PALETTES[normalize_theme_name(theme_name)])

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
ORDER_REPORT_FILE_PATTERN = re.compile(
    r"^inventory_order_(7d|14d|30d)_[A-Za-z0-9]+\.(xlsx|xls|csv)$",
    re.IGNORECASE,
)


def is_order_report_filename(filename):
    return bool(ORDER_REPORT_FILE_PATTERN.match(str(filename or "")))


def list_catalog_csv_files(directory):
    if not directory or not os.path.isdir(directory):
        return []
    return sorted(
        filename
        for filename in os.listdir(directory)
        if filename.lower().endswith(".csv") and not is_order_report_filename(filename)
    )


def list_order_report_files(directory):
    if not directory or not os.path.isdir(directory):
        return []
    return sorted(
        filename
        for filename in os.listdir(directory)
        if is_order_report_filename(filename)
    )

# ----------------------------------------------------------------------
#                  CONFIG.TXT load/save
# ----------------------------------------------------------------------
def safe_filename(name: str) -> str:
    """
    Make a string safe for filesystem paths.
    - trims whitespace
    - removes illegal characters
    """
    name = name.strip()
    name = re.sub(r"[^\w\-]+", "_", name)
    return name


def parse_bool_value(value, default=True):
    if value is None:
        return default
    return str(value).strip().lower() not in ("0", "false", "no", "off", "")


def parse_task_eta_seconds(value):
    if not value:
        return {}
    try:
        raw_data = json.loads(value) if isinstance(value, str) else value
    except (TypeError, ValueError, json.JSONDecodeError):
        return {}
    if not isinstance(raw_data, dict):
        return {}

    cleaned = {}
    for key, seconds in raw_data.items():
        try:
            parsed_seconds = float(seconds)
        except (TypeError, ValueError):
            continue
        if parsed_seconds > 0:
            cleaned[str(key)] = round(parsed_seconds, 1)
    return cleaned


def load_config():
    """
    Reads config.txt in a backwards-compatible format.
    Legacy shape:
        1) input_dir
        2) output_dir
        3) optional fetch_order_reports flag
    Newer shape:
        key=value lines after the first two path lines
    """
    cfg = dict(DEFAULT_GUI_CONFIG)
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                lines = [line.rstrip("\n") for line in f.readlines() if line.strip()]

            if len(lines) >= 1 and "=" not in lines[0]:
                cfg["input_dir"] = lines[0].strip()
            if len(lines) >= 2 and "=" not in lines[1]:
                cfg["output_dir"] = lines[1].strip()

            extra_lines = lines[2:] if len(lines) >= 2 else lines
            for raw_line in extra_lines:
                if "=" not in raw_line:
                    cfg["fetch_order_reports"] = parse_bool_value(raw_line, True)
                    continue

                key, value = raw_line.split("=", 1)
                key = key.strip().lower()
                value = value.strip()
                if key == "fetch_order_reports":
                    cfg["fetch_order_reports"] = parse_bool_value(value, True)
                elif key == "emails":
                    cfg["emails"] = value
                elif key == "include_cost":
                    cfg["include_cost"] = parse_bool_value(value, True)
                elif key == "prefer_catalog_api":
                    cfg["prefer_catalog_api"] = parse_bool_value(value, True)
                elif key == "auto_update_on_launch":
                    cfg["auto_update_on_launch"] = parse_bool_value(value, True)
                elif key == "auto_load_brands_after_update":
                    cfg["auto_load_brands_after_update"] = parse_bool_value(value, True)
                elif key == "auto_load_brands_on_launch":
                    parsed = parse_bool_value(value, True)
                    cfg["auto_load_brands_on_launch"] = parsed
                    cfg["auto_load_brands_after_update"] = parsed
                elif key == "show_startup_loading":
                    cfg["show_startup_loading"] = parse_bool_value(value, True)
                elif key == "open_output_after_complete":
                    cfg["open_output_after_complete"] = parse_bool_value(value, False)
                elif key == "compact_mode":
                    cfg["compact_mode"] = parse_bool_value(value, True)
                elif key == "theme":
                    cfg["theme"] = normalize_theme_name(value)
                elif key == "task_eta_seconds":
                    cfg["task_eta_seconds"] = parse_task_eta_seconds(value)
        except:
            pass
    return cfg


def clear_old_input_exports(directory, clear_order_reports=True):
    if not directory or not os.path.isdir(directory):
        return []

    deleted_paths = []
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if not os.path.isfile(file_path):
            continue
        is_order_file = is_order_report_filename(filename)
        should_delete = filename.lower().endswith(".csv") and not is_order_file
        if clear_order_reports and is_order_file:
            should_delete = True
        if should_delete:
            os.remove(file_path)
            deleted_paths.append(file_path)
            print(f"[INFO] Deleted old source export: {file_path}")
    return deleted_paths


def fetch_inventory_order_reports(output_directory):
    script_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ORDER_REPORT_BROWSER_SCRIPT)
    if not os.path.exists(script_path):
        print(f"[WARN] {ORDER_REPORT_BROWSER_SCRIPT} not found, skipping order-report fetch.")
        return False

    try:
        print("[INFO] Running browser order-report exporter for 7d/14d/30d source files ...")
        subprocess.check_call([sys.executable, script_path, output_directory])
        print("[INFO] Inventory order report fetch complete.")
        return True
    except subprocess.CalledProcessError as e:
        print(f"[ERROR] {ORDER_REPORT_BROWSER_SCRIPT} failed: {e}")
    except Exception as e:
        print(f"[ERROR] Unexpected order-report fetch failure: {e}")
    return False


def dutchie_api_readiness(env_file=DEFAULT_API_ENV_FILE):
    """
    Returns whether Dutchie API refresh is ready for all configured Buzz stores.
    """
    expected_codes = list(STORE_CODES.keys())

    try:
        env_map = canonical_env_map(env_file)
        resolved = resolve_store_keys(env_map, expected_codes)
    except Exception as exc:
        return False, [], expected_codes, str(exc)

    available_codes = [code for code in expected_codes if code in resolved]
    missing_codes = [code for code in expected_codes if code not in resolved]
    return not missing_codes, available_codes, missing_codes, ""


def catalog_api_readiness(env_file=DEFAULT_API_ENV_FILE):
    return dutchie_api_readiness(env_file)

def save_config(
    input_dir,
    output_dir,
    fetch_order_reports=True,
    emails="",
    include_cost=True,
    prefer_catalog_api=True,
    auto_update_on_launch=True,
    auto_load_brands_after_update=True,
    auto_load_brands_on_launch=True,
    show_startup_loading=True,
    open_output_after_complete=False,
    theme="flatly",
    compact_mode=True,
    task_eta_seconds=None,
):
    """
    Writes GUI settings to config.txt while keeping the first two lines as
    plain paths for backwards compatibility with older runs.
    """
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            f.write(input_dir + "\n")
            f.write(output_dir + "\n")
            f.write(f"fetch_order_reports={'1' if fetch_order_reports else '0'}\n")
            f.write(f"emails={emails.strip()}\n")
            f.write(f"include_cost={'1' if include_cost else '0'}\n")
            f.write(f"prefer_catalog_api={'1' if prefer_catalog_api else '0'}\n")
            f.write(f"auto_update_on_launch={'1' if auto_update_on_launch else '0'}\n")
            f.write(f"auto_load_brands_after_update={'1' if auto_load_brands_after_update else '0'}\n")
            f.write(f"auto_load_brands_on_launch={'1' if auto_load_brands_on_launch else '0'}\n")
            f.write(f"show_startup_loading={'1' if show_startup_loading else '0'}\n")
            f.write(f"open_output_after_complete={'1' if open_output_after_complete else '0'}\n")
            f.write(f"theme={normalize_theme_name(theme)}\n")
            f.write(f"compact_mode={'1' if compact_mode else '0'}\n")
            eta_data = parse_task_eta_seconds(task_eta_seconds)
            f.write(f"task_eta_seconds={json.dumps(eta_data, sort_keys=True, separators=(',', ':'))}\n")
    except Exception as e:
        print(f"[ERROR] Could not write config.txt: {e}")

# ----------------------------------------------------------------------
#                  GOOGLE DRIVE / GMAIL AUTH
# ----------------------------------------------------------------------
def drive_authenticate():
    """Authenticate & build the Google Drive service using OAuth."""
    creds = None
    if os.path.exists(TOKEN_DRIVE_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_DRIVE_FILE, DRIVE_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, DRIVE_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_DRIVE_FILE, "w") as token:
            token.write(creds.to_json())
    return build("drive", "v3", credentials=creds)

def gmail_authenticate():
    """Authenticate with Gmail API (OAuth) and return a service object."""
    creds = None
    if os.path.exists(TOKEN_GMAIL_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_GMAIL_FILE, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_GMAIL_FILE, "w") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)

def make_folder_public(drive_service, folder_id):
    """Make the given folder ID publicly viewable."""
    permission = {"type": "anyone", "role": "reader"}
    drive_service.permissions().create(fileId=folder_id, body=permission).execute()

def find_or_create_folder(drive_service, folder_name, parent_id=None, make_public=False):
    """
    Find or create a folder named folder_name under parent_id.
    If newly created and make_public=True, sets public read permission.
    Returns folder_id or None on error.
    """
    from googleapiclient.errors import HttpError
    folder_name_escaped = folder_name.replace("'", "\\'")
    q = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name_escaped}'"
    if parent_id:
        q += f" and '{parent_id}' in parents"

    try:
        res = drive_service.files().list(q=q, spaces="drive", fields="files(id, name)").execute()
        folders = res.get("files", [])
    except HttpError as e:
        print(f"[ERROR] find_or_create_folder: {e}")
        return None

    if folders:
        return folders[0]["id"]

    meta = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
    }
    if parent_id:
        meta["parents"] = [parent_id]

    new_folder = drive_service.files().create(body=meta, fields="id").execute()
    fid = new_folder.get("id")
    print(f"[INFO] Created new folder '{folder_name}' (ID: {fid})")

    if make_public:
        try:
            make_folder_public(drive_service, fid)
        except Exception as e:
            print(f"[ERROR] Could not make folder public: {e}")

    return fid

def upload_file_to_drive(drive_service, file_path, parent_id):
    """Upload a local file to the given parent folder ID. Return the uploaded file ID."""
    file_name = os.path.basename(file_path)
    meta = {"name": file_name, "parents": [parent_id]}
    media = MediaFileUpload(file_path, resumable=True)
    uploaded = drive_service.files().create(body=meta, media_body=media, fields="id").execute()
    return uploaded.get("id")

def send_email_with_gmail_html(subject, html_body, recipients):
    """
    Sends an HTML email via the Gmail API. 
    recipients can be a list or a single comma-separated string.
    """
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    if isinstance(recipients, str):
        recipients = [r.strip() for r in recipients.split(",") if r.strip()]

    service = gmail_authenticate()

    msg = MIMEMultipart("alternative")
    msg["From"] = "me"
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject

    part_html = MIMEText(html_body, "html")
    msg.attach(part_html)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    body = {"raw": raw_message}

    sent = service.users().messages().send(userId="me", body=body).execute()
    print(f"[GMAIL] Email sent! ID: {sent['id']} | Subject: {subject}")

# ----------------- EXCEL FORMATTING -----------------
def advanced_format_excel(xlsx_path):
    """Freeze top row, bold grey headers, auto-fit columns, group by 'Category'."""
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        if format_order_sheet(ws):
            continue

        # Freeze row 1
        ws.freeze_panes = "A2"

        # Header style
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        header_map = {
            (cell.value or "").strip().lower(): idx
            for idx, cell in enumerate(ws[1], start=1)
            if cell.value
        }

        # Auto-fit columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = cell.value
                if val is not None:
                    length = len(str(val))
                    if length > max_len:
                        max_len = length
            ws.column_dimensions[col_letter].width = max_len + 3

        product_col_idx = header_map.get("product name") or header_map.get("product")
        if product_col_idx:
            product_letter = get_column_letter(product_col_idx)
            ws.column_dimensions[product_letter].width = max(
                42,
                min(ws.column_dimensions[product_letter].width or 42, 52),
            )
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=product_col_idx).alignment = Alignment(
                    wrap_text=True,
                    vertical="top",
                )

        notes_col_idx = header_map.get("reorder notes")
        if notes_col_idx:
            notes_letter = get_column_letter(notes_col_idx)
            ws.column_dimensions[notes_letter].width = max(
                28,
                min(ws.column_dimensions[notes_letter].width or 28, 34),
            )

        suggested_col_idx = next(
            (
                idx for name, idx in header_map.items()
                if name.startswith("suggested order qty")
            ),
            None,
        )
        if suggested_col_idx:
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=suggested_col_idx).font = Font(bold=True)

        priority_col_idx = header_map.get("reorder priority")
        if priority_col_idx:
            color_rules = {
                "Urgent": "FECACA",
                "Reorder Now": "FED7AA",
                "Reorder Soon": "FEF08A",
                "Check PO": "BFDBFE",
                "Watch": "FDE68A",
                "Healthy": "BBF7D0",
                "No Recent Sales": "E5E7EB",
            }
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=priority_col_idx)
                color = color_rules.get(str(cell.value or "").strip())
                if color:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Insert grouping rows for 'Category'
        category_index = header_map.get("category")
        if category_index:
            rows_data = list(ws.iter_rows(min_row=2, values_only=True))
            if rows_data:
                current_cat = None
                insert_positions = []
                row_num = 2
                for row_vals in rows_data:
                    cat_val = row_vals[category_index - 1]
                    if cat_val != current_cat:
                        if current_cat is not None:
                            insert_positions.append(row_num)
                        current_cat = cat_val
                    row_num += 1
                # Insert at the very top
                insert_positions.insert(0, 2)

                cat_font = Font(bold=True, size=14)
                cat_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")

                # We'll also need the cat value
                row_num = 2
                cat_list = []
                cur_cat = None
                for row_vals in rows_data:
                    cat_val = row_vals[category_index - 1]
                    if cat_val != cur_cat:
                        cat_list.append((row_num, cat_val))
                        cur_cat = cat_val
                    row_num += 1

                # Insert from bottom to top
                for (pos, cat_val) in reversed(cat_list):
                    ws.insert_rows(pos, 1)
                    c = ws.cell(row=pos, column=1)
                    c.value = str(cat_val)
                    c.font = cat_font
                    c.fill = cat_fill
                    c.alignment = Alignment(horizontal='center', vertical='center')
    wb.save(xlsx_path)

def extract_strain_type(product_name):
    """Optional: parse 'S', 'H', 'I' from product name, if you want to track strain."""
    if not isinstance(product_name, str):
        return ""
    text = " " + product_name.upper() + " "
    if re.search(r"\bS\b", text):
        return "S"
    if re.search(r"\bH\b", text):
        return "H"
    if re.search(r"\bI\b", text):
        return "I"
    return ""

# ----------------- CSV -> XLSX: Avail + Unavail -----------------
def generate_brand_reports(csv_path, out_dir, selected_brands, include_cost=True, order_reports_dir=None):
    """
    Splits CSV rows into:
      - Available: Available>2
      - Unavailable: Available<=2
    Then for each brand in "available", produce one XLSX with 2 sheets:
      - "Available" (the brand’s rows)
      - "Unavailable" (the brand’s rows from the unavailable set, if any)
    Returns { brand_lower: [list_of_xlsx_paths] } for each brand found.
    """
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print(f"[ERROR] reading {csv_path}: {e}")
        return {}

    if not all(c in df.columns for c in REQUIRED_COLUMNS):
        print(
            f"[WARN] '{csv_path}' is missing required columns {REQUIRED_COLUMNS}. "
            f"Columns found: {list(df.columns)}. Skipping."
        )
        return {}

    # Keep relevant columns only
    keep_cols = [c for c in REQUIRED_COLUMNS + OPTIONAL_COLUMNS if c in df.columns]
    df = df[keep_cols]

    # Remove "sample"/"promo" lines
    if "Product" in df.columns:
        df = df[~df["Product"].str.contains(r"(?i)\bsample\b|\bpromo\b", na=False)]

    # Split into available/unavailable
    unavailable_df = df[df["Available"] <= MAX_AVAIL_FOR_UNAVAILABLE].copy()
    available_df   = df[df["Available"] > MAX_AVAIL_FOR_UNAVAILABLE].copy()
     # Drop Cost column if disabled
    if not include_cost:
        if "Cost" in available_df.columns:
            available_df = available_df.drop(columns=["Cost"])
        if "Cost" in unavailable_df.columns:
            unavailable_df.drop(columns=["Cost"], inplace=True)
    if "Brand" not in available_df.columns or available_df.empty:
        print(f"[INFO] No brand data or empty after filtering in '{csv_path}'")
        return {}

    # Lowercase + strip brand for consistent matching
    available_df["Brand"] = available_df["Brand"].astype(str).str.strip().str.lower()

    # If user selected brand(s), also convert them to lowercase
    if selected_brands:
        # Turn each user brand into a lowercased version
        selected_lower = [b.strip().lower() for b in selected_brands]
        available_df = available_df[available_df["Brand"].isin(selected_lower)]

    # If nothing remains:
    if available_df.empty:
        print(f"[INFO] No matching brand data in '{csv_path}' after brand filter.")
        return {}

    # Example: add "Strain_Type"
    if "Product" in available_df.columns:
        available_df["Strain_Type"] = available_df["Product"].apply(extract_strain_type)

    # Sort
    sort_cols = []
    if "Category" in available_df.columns:
        sort_cols.append("Category")
    if include_cost  and "Cost" in available_df.columns:
        available_df["Cost"] = pd.to_numeric(available_df["Cost"], errors="coerce")
        sort_cols.append("Cost")
    if "Product" in available_df.columns:
        sort_cols.append("Product")
    if sort_cols:
        available_df = available_df.sort_values(by=sort_cols, na_position="last")

    # # Drop "Cost"
    # if "Cost" in available_df.columns:
    #     available_df.drop(columns=["Cost"], inplace=True)
    # if "Cost" in unavailable_df.columns:
    #     unavailable_df = unavailable_df.drop(columns=["Cost"])

    # Also normalize brand in the unavailable set
    if "Brand" in unavailable_df.columns and not unavailable_df.empty:
        unavailable_df.loc[:, "Brand"] = unavailable_df["Brand"].astype(str).str.strip().str.lower()
    if not unavailable_df.empty:
        unavailable_sort_cols = []
        if "Category" in unavailable_df.columns:
            unavailable_sort_cols.append("Category")
        if "Product" in unavailable_df.columns:
            unavailable_sort_cols.append("Product")
        if unavailable_sort_cols:
            unavailable_df = unavailable_df.sort_values(by=unavailable_sort_cols, na_position="last")

    os.makedirs(out_dir, exist_ok=True)
    base_csv_name = os.path.splitext(os.path.basename(csv_path))[0]
    store_code = extract_store_code_from_filename(base_csv_name)

    # Group the *available* portion by brand
    brand_map = {}
    for brand_name_lower, brand_data in available_df.groupby("Brand"):
        # Grab the "Unavailable" rows for that brand
        brand_unavail = pd.DataFrame()
        if not unavailable_df.empty:
            brand_unavail = unavailable_df[unavailable_df["Brand"] == brand_name_lower]

        dt_str = datetime.now().strftime("%m-%d-%Y")
        #out_name = f"{base_csv_name}_{brand_name_lower}_{dt_str}.xlsx"
        safe_brand = safe_filename(brand_name_lower)

        out_name = f"{base_csv_name}_{safe_brand}.xlsx"
        out_path = os.path.join(out_dir, out_name)
        order_sections = build_brand_order_sections(
            order_reports_dir or os.path.dirname(csv_path),
            brand_aliases=[brand_name_lower],
            store_code=store_code,
        )

        # Ensure output directory exists (extra safety)
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            brand_data.to_excel(writer, index=False, sheet_name="Available")
            if not brand_unavail.empty:
                brand_unavail.to_excel(writer, index=False, sheet_name="Unavailable")
            write_order_sections(writer, order_sections)

        advanced_format_excel(out_path)

        if brand_name_lower not in brand_map:
            brand_map[brand_name_lower] = []
        brand_map[brand_name_lower].append(out_path)

        print(f"[INFO] Created {out_path}")

    return brand_map

def upload_brand_reports_to_drive(brand_reports_map):
    """
    brand_reports_map: { brand_name_lower: [list_of_xlsx_paths] }
    1) Create/find top-level "INVENTORY"
    2) Create date subfolder "YYYY-MM-DD"
    3) For each brand, create brand folder (public), upload
    Return: { brand_name_lower: "https://drive.google.com/drive/folders/<id>"}
    """
    drive_svc = drive_authenticate()
    top_id = find_or_create_folder(drive_svc, DRIVE_PARENT_FOLDER_NAME)
    if not top_id:
        print("[ERROR] Could not find/create top-level folder. Aborting.")
        return {}

    date_str = datetime.now().strftime("%Y-%m-%d")
    date_id = find_or_create_folder(drive_svc, date_str, parent_id=top_id)
    if not date_id:
        print("[ERROR] Could not create/find date subfolder. Aborting.")
        return {}

    brand_links = {}
    for brand_lower, xlsx_list in brand_reports_map.items():
        brand_id = find_or_create_folder(drive_svc, brand_lower, parent_id=date_id, make_public=True)
        if not brand_id:
            print(f"[ERROR] Could not create folder for {brand_lower}")
            continue

        for xfile in xlsx_list:
            try:
                upload_file_to_drive(drive_svc, xfile, brand_id)
                print(f"[DRIVE] Uploaded {os.path.basename(xfile)} => {brand_lower}")
            except Exception as e:
                print(f"[ERROR] Uploading {xfile} => {brand_lower}: {e}")

        link = f"https://drive.google.com/drive/folders/{brand_id}"
        brand_links[brand_lower] = link

    return brand_links

# ----------------- LEGACY GUI (kept as a fallback/reference; modern shell below is used) -----------------
class LegacyBrandInventoryGUI:
    def __init__(self, master):
        self.master = master
        self.colors = {
            "bg": "#EEF3EF",
            "card": "#FFFFFF",
            "hero": "#153B34",
            "text": "#15312C",
            "muted": "#64756F",
            "border": "#D7E2DC",
            "accent": "#177A69",
            "accent_dark": "#115E53",
            "accent_soft": "#E3F4EF",
            "log_bg": "#F7FAF8",
            "tab_idle": "#DCE7E2",
            "tab_active": "#F6FBF8",
        }
        self.all_brands = []
        self.filtered_brands = []
        self.selected_brand_names = set()
        self.autosave_job = None
        self.jump_reset_job = None
        self.quick_jump_buffer = ""
        self.quick_jump_last_ts = 0.0

        cfg = load_config()

        self.master.title("Buzz Brand Inventory Studio")
        self.master.geometry(f"{DEFAULT_WINDOW_WIDTH}x{DEFAULT_WINDOW_HEIGHT}")
        self.master.minsize(900, 520)
        self.master.configure(bg=self.colors["bg"])

        self.input_dir_var = tk.StringVar(value=cfg.get("input_dir", ""))
        self.output_dir_var = tk.StringVar(value=cfg.get("output_dir", ""))
        self.emails_var = tk.StringVar(value=cfg.get("emails", ""))
        self.fetch_order_reports_var = tk.BooleanVar(value=cfg.get("fetch_order_reports", True))
        self.include_cost_var = tk.BooleanVar(value=cfg.get("include_cost", True))
        self.prefer_catalog_api_var = tk.BooleanVar(value=cfg.get("prefer_catalog_api", True))
        self.brand_search_var = tk.StringVar()

        self.status_var = tk.StringVar(value="Ready to refresh exports and build reports.")
        self.status_detail_var = tk.StringVar(
            value="Choose your source folders, review recipients, then refresh files or generate the report package."
        )
        self.input_summary_var = tk.StringVar()
        self.output_summary_var = tk.StringVar()
        self.brand_summary_var = tk.StringVar(value="No brands loaded yet.")
        self.email_summary_var = tk.StringVar(value="No recipients saved yet.")
        self.order_reports_caption_var = tk.StringVar()
        self.catalog_refresh_caption_var = tk.StringVar()
        self.settings_state_var = tk.StringVar(
            value="Auto-save is on. Folder choices, recipients, and toggles are stored as you work."
        )
        self.date_var = tk.StringVar(value=datetime.now().strftime("%A, %B %d, %Y"))
        self.source_snapshot_var = tk.StringVar(value="No source folder scanned yet.")
        self.order_window_summary_var = tk.StringVar(value="Dutchie order windows: none detected yet.")
        self.brand_load_status_var = tk.StringVar(value="Brand library not loaded yet.")
        self.brand_hint_var = tk.StringVar(
            value="Ctrl+F focuses search. Type in the list for a quick prefix jump. Space toggles the active brand."
        )
        self.brand_total_var = tk.StringVar(value="0")
        self.brand_visible_var = tk.StringVar(value="0")
        self.brand_selected_var = tk.StringVar(value="0")
        self.catalog_count_var = tk.StringVar(value="0")
        self.order_file_count_var = tk.StringVar(value="0")
        self.recipient_count_var = tk.StringVar(value="0")

        self._configure_styles()
        self._build_layout()
        self._bind_events()
        self._refresh_path_summaries()
        self._refresh_source_snapshot()
        self._update_email_display()
        self._update_catalog_refresh_caption()
        self._update_order_report_caption()
        self._update_brand_summary()
        self.append_log("Workspace ready.")
        self.master.protocol("WM_DELETE_WINDOW", self.on_close)
        self.master.after(0, self._show_full_program_on_launch)
        self.master.after(150, self._autoload_saved_workspace)

    def _configure_styles(self):
        self.style = ttk.Style(self.master)
        if "clam" in self.style.theme_names():
            self.style.theme_use("clam")

        self.style.configure(
            "App.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(14, 10),
            background=self.colors["card"],
            foreground=self.colors["text"],
            borderwidth=0,
        )
        self.style.map(
            "App.TButton",
            background=[("active", "#F1F5F2"), ("pressed", "#E4EBE7")],
        )

        self.style.configure(
            "Primary.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(16, 12),
            background=self.colors["accent"],
            foreground="#FFFFFF",
            borderwidth=0,
        )
        self.style.map(
            "Primary.TButton",
            background=[("active", self.colors["accent_dark"]), ("pressed", self.colors["accent_dark"])],
            foreground=[("disabled", "#FFFFFF")],
        )

        self.style.configure(
            "Quiet.TButton",
            font=("Segoe UI", 9, "bold"),
            padding=(10, 7),
            background=self.colors["accent_soft"],
            foreground=self.colors["accent_dark"],
            borderwidth=0,
        )
        self.style.map(
            "Quiet.TButton",
            background=[("active", "#D6EEE7"), ("pressed", "#C8E8DE")],
        )
        self.style.configure(
            "Loading.Horizontal.TProgressbar",
            troughcolor="#DCE7E2",
            background=self.colors["accent"],
            lightcolor=self.colors["accent"],
            darkcolor=self.colors["accent"],
            bordercolor=self.colors["border"],
        )

        self.style.configure(
            "App.TEntry",
            fieldbackground="#FFFFFF",
            bordercolor=self.colors["border"],
            lightcolor=self.colors["border"],
            darkcolor=self.colors["border"],
            insertcolor=self.colors["text"],
            padding=8,
        )
        self.style.configure(
            "Card.TCheckbutton",
            background=self.colors["card"],
            foreground=self.colors["text"],
            font=("Segoe UI", 10),
        )
        self.style.map(
            "Card.TCheckbutton",
            background=[("active", self.colors["card"])],
        )
        self.style.configure(
            "App.TNotebook",
            background=self.colors["bg"],
            borderwidth=0,
            tabmargins=(0, 0, 0, 0),
        )
        self.style.configure(
            "App.TNotebook.Tab",
            background=self.colors["tab_idle"],
            foreground=self.colors["text"],
            padding=(16, 9),
            font=("Segoe UI", 10, "bold"),
            borderwidth=0,
        )
        self.style.map(
            "App.TNotebook.Tab",
            background=[
                ("selected", "#FFFFFF"),
                ("active", self.colors["tab_active"]),
            ],
            foreground=[
                ("selected", self.colors["accent_dark"]),
                ("active", self.colors["text"]),
            ],
        )

    def _create_card(self, parent, title, subtitle=None):
        card = tk.Frame(
            parent,
            bg=self.colors["card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
            bd=0,
        )
        accent_bar = tk.Frame(card, bg=self.colors["accent"], height=4)
        accent_bar.pack(fill="x")

        header = tk.Frame(card, bg=self.colors["card"])
        header.pack(fill="x", padx=16, pady=(12, 4))

        tk.Label(
            header,
            text=title,
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 12, "bold"),
        ).pack(anchor="w")
        if subtitle:
            tk.Label(
                header,
                text=subtitle,
                bg=self.colors["card"],
                fg=self.colors["muted"],
                font=("Segoe UI", 9),
                wraplength=520,
                justify="left",
            ).pack(anchor="w", pady=(4, 0))

        body = tk.Frame(card, bg=self.colors["card"])
        body.pack(fill="both", expand=True, padx=16, pady=(4, 16))
        return card, body

    def _build_layout(self):
        self.shell = tk.Frame(self.master, bg=self.colors["bg"])
        self.shell.pack(fill="both", expand=True, padx=10, pady=10)
        self.shell.grid_columnconfigure(0, weight=1)
        self.shell.grid_rowconfigure(1, weight=1)

        self._build_header()
        self._build_tabs()

    def _show_full_program_on_launch(self):
        self.master.update_idletasks()

        try:
            self.master.state("zoomed")
            return
        except tk.TclError:
            pass

        try:
            self.master.attributes("-zoomed", True)
            return
        except tk.TclError:
            pass

        screen_width = max(self.master.winfo_screenwidth(), self.master.winfo_reqwidth())
        screen_height = max(self.master.winfo_screenheight(), self.master.winfo_reqheight())
        width = min(DEFAULT_WINDOW_WIDTH, max(900, screen_width - WINDOW_EDGE_PADDING * 2))
        height = min(DEFAULT_WINDOW_HEIGHT, max(520, screen_height - WINDOW_EDGE_PADDING * 2))
        x = max(0, (screen_width - width) // 2)
        y = max(0, (screen_height - height) // 2)
        self.master.geometry(f"{width}x{height}+{x}+{y}")

    def _build_header(self):
        header = tk.Frame(
            self.shell,
            bg=self.colors["card"],
            highlightbackground=self.colors["border"],
            highlightthickness=1,
            padx=12,
            pady=10,
        )
        header.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header.grid_columnconfigure(1, weight=1)

        tk.Label(
            header,
            text="Buzz Brand Inventory Studio",
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI Semibold", 15),
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            header,
            textvariable=self.date_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
        ).grid(row=1, column=0, sticky="w", pady=(2, 0))

        status_box = tk.Frame(header, bg="#F7FAF8", padx=10, pady=7)
        status_box.grid(row=0, column=2, rowspan=2, sticky="e")
        tk.Label(
            status_box,
            textvariable=self.status_var,
            bg="#F7FAF8",
            fg=self.colors["text"],
            font=("Segoe UI", 9, "bold"),
        ).pack(anchor="w")
        tk.Label(
            status_box,
            textvariable=self.status_detail_var,
            bg="#F7FAF8",
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            justify="left",
            wraplength=280,
        ).pack(anchor="w", pady=(2, 0))

    def _build_summary_bar(self):
        summary = tk.Frame(self.shell, bg=self.colors["bg"])
        summary.grid(row=1, column=0, sticky="ew", pady=(14, 14))
        for col in range(4):
            summary.grid_columnconfigure(col, weight=1)

        self._create_metric_tile(summary, "Catalog Exports", self.catalog_count_var).grid(
            row=0,
            column=0,
            sticky="ew",
            padx=(0, 8),
        )
        self._create_metric_tile(summary, "Order Exports", self.order_file_count_var).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=4,
        )
        self._create_metric_tile(summary, "Recipients", self.recipient_count_var).grid(
            row=0,
            column=2,
            sticky="ew",
            padx=4,
        )
        self._create_metric_tile(summary, "Selected Brands", self.brand_selected_var).grid(
            row=0,
            column=3,
            sticky="ew",
            padx=(8, 0),
        )

    def _build_tabs(self):
        self.notebook = ttk.Notebook(self.shell, style="App.TNotebook")
        self.notebook.grid(row=1, column=0, sticky="nsew")

        self.overview_tab = tk.Frame(self.notebook, bg=self.colors["bg"])
        self.brands_tab = tk.Frame(self.notebook, bg=self.colors["bg"])
        self.activity_tab = tk.Frame(self.notebook, bg=self.colors["bg"])
        self.settings_tab = tk.Frame(self.notebook, bg=self.colors["bg"])

        self.notebook.add(self.overview_tab, text="Run")
        self.notebook.add(self.brands_tab, text="Brands")
        self.notebook.add(self.activity_tab, text="Activity")
        self.notebook.add(self.settings_tab, text="Settings")

        self._build_overview_tab()
        self._build_brand_tab()
        self._build_activity_tab()
        self._build_settings_tab()

    def _build_overview_tab(self):
        self.overview_tab.grid_columnconfigure(0, weight=1)
        self.overview_tab.grid_rowconfigure(0, weight=1)

        canvas = tk.Canvas(
            self.overview_tab,
            bg=self.colors["bg"],
            highlightthickness=0,
            bd=0,
        )
        scrollbar = ttk.Scrollbar(self.overview_tab, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")

        content = tk.Frame(canvas, bg=self.colors["bg"])
        content_window = canvas.create_window((0, 0), window=content, anchor="nw")
        content.grid_columnconfigure(0, weight=1)
        content.grid_columnconfigure(1, weight=1)

        content.bind("<Configure>", lambda event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(content_window, width=event.width))

        metrics = tk.Frame(content, bg=self.colors["bg"])
        metrics.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 10))
        for col in range(4):
            metrics.grid_columnconfigure(col, weight=1)
        self._create_metric_tile(metrics, "Catalog", self.catalog_count_var).grid(row=0, column=0, sticky="ew", padx=(0, 6))
        self._create_metric_tile(metrics, "Order", self.order_file_count_var).grid(row=0, column=1, sticky="ew", padx=2)
        self._create_metric_tile(metrics, "Recipients", self.recipient_count_var).grid(row=0, column=2, sticky="ew", padx=2)
        self._create_metric_tile(metrics, "Selected", self.brand_selected_var).grid(row=0, column=3, sticky="ew", padx=(6, 0))

        actions_card, actions_body = self._create_card(
            content,
            "Run Actions",
            "Refresh files, load brands, then generate and email the finished reports.",
        )
        actions_card.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(0, 10))
        self._build_run_actions_card(actions_body)

        delivery_card, delivery_body = self._create_card(
            content,
            "Recipients",
            "Recipient emails save automatically and show up below in an easy-to-check list.",
        )
        delivery_card.grid(row=2, column=0, sticky="nsew", padx=(0, 6), pady=(0, 10))
        self._build_delivery_card(delivery_body)

        status_card, status_body = self._create_card(
            content,
            "Source Snapshot",
            "A quick check of source files, order windows, and the last brand-library refresh.",
        )
        status_card.grid(row=2, column=1, sticky="nsew", padx=(6, 0), pady=(0, 10))
        self._build_snapshot_card(status_body)

    def _build_settings_tab(self):
        self.settings_tab.grid_columnconfigure(0, weight=1)
        self.settings_tab.grid_rowconfigure(0, weight=1)

        self.settings_notebook = ttk.Notebook(self.settings_tab, style="App.TNotebook")
        self.settings_notebook.grid(row=0, column=0, sticky="nsew")

        self.workspace_settings_tab = tk.Frame(self.settings_notebook, bg=self.colors["bg"])
        self.workflow_settings_tab = tk.Frame(self.settings_notebook, bg=self.colors["bg"])

        self.settings_notebook.add(self.workspace_settings_tab, text="Workspace")
        self.settings_notebook.add(self.workflow_settings_tab, text="Workflow")

        self.workspace_settings_tab.grid_columnconfigure(0, weight=1)
        self.workflow_settings_tab.grid_columnconfigure(0, weight=1)

        workspace_card, workspace_body = self._create_card(
            self.workspace_settings_tab,
            "Workspace",
            "Choose folders and save the setup you want to reopen next time.",
        )
        workspace_card.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        self._build_settings_card(workspace_body)

        workflow_card, workflow_body = self._create_card(
            self.workflow_settings_tab,
            "Workflow Settings",
            "Choose how source refreshes and generated brand workbooks should behave before each run.",
        )
        workflow_card.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        self._build_workflow_settings_card(workflow_body)

    def _build_settings_card(self, body):
        body.grid_columnconfigure(1, weight=1)

        tk.Label(
            body,
            text="Input Folder",
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 9, "bold"),
        ).grid(row=0, column=0, sticky="w")
        ttk.Entry(body, textvariable=self.input_dir_var, style="App.TEntry").grid(
            row=1,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(4, 0),
        )
        ttk.Button(body, text="Browse", style="Quiet.TButton", command=self.browse_input).grid(
            row=1,
            column=2,
            sticky="ew",
            padx=(8, 0),
        )
        tk.Label(
            body,
            textvariable=self.input_summary_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            justify="left",
            wraplength=420,
        ).grid(row=2, column=0, columnspan=3, sticky="w", pady=(6, 0))

        tk.Label(
            body,
            text="Output Folder",
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 9, "bold"),
        ).grid(row=3, column=0, sticky="w", pady=(12, 0))
        ttk.Entry(body, textvariable=self.output_dir_var, style="App.TEntry").grid(
            row=4,
            column=0,
            columnspan=2,
            sticky="ew",
            pady=(4, 0),
        )
        ttk.Button(body, text="Browse", style="Quiet.TButton", command=self.browse_output).grid(
            row=4,
            column=2,
            sticky="ew",
            padx=(8, 0),
        )
        tk.Label(
            body,
            textvariable=self.output_summary_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            justify="left",
            wraplength=420,
        ).grid(row=5, column=0, columnspan=3, sticky="w", pady=(6, 0))

        button_row = tk.Frame(body, bg=self.colors["card"])
        button_row.grid(row=6, column=0, columnspan=3, sticky="ew", pady=(12, 0))
        button_row.grid_columnconfigure(0, weight=1)
        ttk.Button(button_row, text="Save Settings", style="App.TButton", command=self.save_settings).grid(
            row=0,
            column=0,
            sticky="ew",
        )
        ttk.Button(
            button_row,
            text="Open Brands Tab",
            style="Quiet.TButton",
            command=lambda: self._select_tab(self.brands_tab),
        ).grid(row=0, column=1, sticky="ew", padx=(8, 0))

        tk.Label(
            body,
            textvariable=self.settings_state_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            justify="left",
            wraplength=420,
        ).grid(row=7, column=0, columnspan=3, sticky="w", pady=(8, 0))

    def _build_workflow_settings_card(self, body):
        body.grid_columnconfigure(0, weight=1)

        ttk.Checkbutton(
            body,
            text="Prefer Dutchie API for source refresh when available",
            variable=self.prefer_catalog_api_var,
            style="Card.TCheckbutton",
        ).grid(row=0, column=0, sticky="w")
        tk.Label(
            body,
            textvariable=self.catalog_refresh_caption_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=520,
            justify="left",
        ).grid(row=1, column=0, sticky="w", pady=(4, 0))

        ttk.Checkbutton(
            body,
            text="Refresh 7d / 14d / 30d Dutchie order reports with Update Files",
            variable=self.fetch_order_reports_var,
            style="Card.TCheckbutton",
        ).grid(row=2, column=0, sticky="w", pady=(12, 0))
        tk.Label(
            body,
            textvariable=self.order_reports_caption_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=520,
            justify="left",
        ).grid(row=3, column=0, sticky="w", pady=(4, 0))

        ttk.Checkbutton(
            body,
            text="Include Cost column in generated brand workbooks",
            variable=self.include_cost_var,
            style="Card.TCheckbutton",
        ).grid(row=4, column=0, sticky="w", pady=(12, 0))

        tk.Label(
            body,
            text="These preferences auto-save as you work, and Save Settings on the Workspace tab stores them immediately.",
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=520,
            justify="left",
        ).grid(row=5, column=0, sticky="w", pady=(10, 0))

    def _build_run_actions_card(self, body):
        body.grid_columnconfigure(0, weight=1)
        body.grid_columnconfigure(1, weight=1)

        ttk.Button(body, text="Update Files", style="App.TButton", command=self.get_files).grid(
            row=0,
            column=0,
            sticky="ew",
            pady=(4, 0),
        )
        ttk.Button(body, text="Load Brands", style="App.TButton", command=self.load_brands).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=(8, 0),
            pady=(4, 0),
        )
        ttk.Button(
            body,
            text="Generate, Upload & Email",
            style="Primary.TButton",
            command=self.run_process,
        ).grid(row=1, column=0, columnspan=2, sticky="ew", pady=(10, 0))

        nav_row = tk.Frame(body, bg=self.colors["card"])
        nav_row.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        nav_row.grid_columnconfigure(0, weight=1)
        nav_row.grid_columnconfigure(1, weight=1)
        ttk.Button(
            nav_row,
            text="Open Settings Tab",
            style="Quiet.TButton",
            command=lambda: self._select_tab(self.settings_tab),
        ).grid(row=0, column=0, sticky="ew")
        ttk.Button(
            nav_row,
            text="Open Brands Tab",
            style="Quiet.TButton",
            command=lambda: self._select_tab(self.brands_tab),
        ).grid(row=0, column=1, sticky="ew", padx=(8, 0))

        tk.Label(
            body,
            text="Folder paths now live in Settings > Workspace. Report behavior toggles live in Settings > Workflow.",
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=520,
            justify="left",
        ).grid(row=3, column=0, columnspan=2, sticky="w", pady=(10, 0))

        tk.Label(
            body,
            text="Shortcuts: Ctrl+F search brands • Ctrl+U update files • Ctrl+L load brands • Ctrl+Enter send reports",
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=520,
            justify="left",
        ).grid(row=4, column=0, columnspan=2, sticky="w", pady=(10, 0))

    def _build_delivery_card(self, body):
        body.grid_columnconfigure(0, weight=1)

        tk.Label(
            body,
            text="Recipient Emails",
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 9, "bold"),
        ).grid(row=0, column=0, sticky="w")
        ttk.Entry(body, textvariable=self.emails_var, style="App.TEntry").grid(
            row=1,
            column=0,
            sticky="ew",
            pady=(4, 0),
        )
        tk.Label(
            body,
            text="Separate multiple addresses with commas. Invalid entries are highlighted before send.",
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=420,
            justify="left",
        ).grid(row=2, column=0, sticky="w", pady=(6, 0))

        tk.Label(
            body,
            text="Saved Recipients",
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 9, "bold"),
        ).grid(row=3, column=0, sticky="w", pady=(12, 0))
        self.email_chip_frame = tk.Frame(body, bg=self.colors["card"])
        self.email_chip_frame.grid(row=4, column=0, sticky="ew", pady=(6, 0))

        tk.Label(
            body,
            textvariable=self.email_summary_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=420,
            justify="left",
        ).grid(row=5, column=0, sticky="w", pady=(8, 0))

    def _build_snapshot_card(self, body):
        body.grid_columnconfigure(0, weight=1)

        snapshot_box = tk.Frame(
            body,
            bg="#F7FAF8",
            highlightbackground=self.colors["border"],
            highlightthickness=1,
            padx=12,
            pady=12,
        )
        snapshot_box.grid(row=0, column=0, sticky="ew")
        tk.Label(
            snapshot_box,
            text="Source Files",
            bg="#F7FAF8",
            fg=self.colors["muted"],
            font=("Segoe UI", 8, "bold"),
        ).pack(anchor="w")
        tk.Label(
            snapshot_box,
            textvariable=self.source_snapshot_var,
            bg="#F7FAF8",
            fg=self.colors["text"],
            font=("Segoe UI", 9),
            wraplength=430,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        order_box = tk.Frame(
            body,
            bg="#F7FAF8",
            highlightbackground=self.colors["border"],
            highlightthickness=1,
            padx=12,
            pady=12,
        )
        order_box.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        tk.Label(
            order_box,
            text="Order Report Coverage",
            bg="#F7FAF8",
            fg=self.colors["muted"],
            font=("Segoe UI", 8, "bold"),
        ).pack(anchor="w")
        tk.Label(
            order_box,
            textvariable=self.order_window_summary_var,
            bg="#F7FAF8",
            fg=self.colors["text"],
            font=("Segoe UI", 9),
            wraplength=430,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        brand_box = tk.Frame(
            body,
            bg="#F7FAF8",
            highlightbackground=self.colors["border"],
            highlightthickness=1,
            padx=12,
            pady=12,
        )
        brand_box.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        tk.Label(
            brand_box,
            text="Brand Library",
            bg="#F7FAF8",
            fg=self.colors["muted"],
            font=("Segoe UI", 8, "bold"),
        ).pack(anchor="w")
        tk.Label(
            brand_box,
            textvariable=self.brand_load_status_var,
            bg="#F7FAF8",
            fg=self.colors["text"],
            font=("Segoe UI", 9),
            wraplength=430,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

    def _build_brand_tab(self):
        self.brands_tab.grid_columnconfigure(0, weight=1)
        self.brands_tab.grid_rowconfigure(1, weight=1)

        toolbar_card, toolbar_body = self._create_card(
            self.brands_tab,
            "Brand Finder",
            "Keyboard-first search and selection. Hidden selections stay saved even when you filter the list.",
        )
        toolbar_card.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        toolbar_body.grid_columnconfigure(1, weight=1)

        tk.Label(
            toolbar_body,
            text="Search",
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 9, "bold"),
        ).grid(row=0, column=0, sticky="w")
        self.brand_search_entry = ttk.Entry(
            toolbar_body,
            textvariable=self.brand_search_var,
            style="App.TEntry",
        )
        self.brand_search_entry.grid(row=0, column=1, sticky="ew", padx=(8, 8))
        ttk.Button(
            toolbar_body,
            text="Focus Search",
            style="Quiet.TButton",
            command=self.focus_brand_search,
        ).grid(row=0, column=2, sticky="ew")
        ttk.Button(
            toolbar_body,
            text="Select Visible",
            style="Quiet.TButton",
            command=self.select_all_brands,
        ).grid(row=0, column=3, sticky="ew", padx=(8, 0))
        ttk.Button(
            toolbar_body,
            text="Clear All",
            style="Quiet.TButton",
            command=self.clear_selected_brands,
        ).grid(row=0, column=4, sticky="ew", padx=(8, 0))

        tk.Label(
            toolbar_body,
            textvariable=self.brand_hint_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=760,
            justify="left",
        ).grid(row=1, column=0, columnspan=5, sticky="w", pady=(8, 0))
        tk.Label(
            toolbar_body,
            textvariable=self.brand_load_status_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=760,
            justify="left",
        ).grid(row=2, column=0, columnspan=5, sticky="w", pady=(4, 0))

        browser_card, browser_body = self._create_card(
            self.brands_tab,
            "Brand Library",
            "Use the A-Z rail, type directly in the list, or press Enter from search to jump to the first visible match.",
        )
        browser_card.grid(row=1, column=0, sticky="nsew")
        browser_body.grid_rowconfigure(1, weight=1)
        browser_body.grid_columnconfigure(0, weight=1)

        stats_row = tk.Frame(browser_body, bg=self.colors["card"])
        stats_row.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        for col in range(3):
            stats_row.grid_columnconfigure(col, weight=1)
        self._create_metric_tile(stats_row, "Loaded", self.brand_total_var).grid(
            row=0,
            column=0,
            sticky="ew",
            padx=(0, 8),
        )
        self._create_metric_tile(stats_row, "Visible", self.brand_visible_var).grid(
            row=0,
            column=1,
            sticky="ew",
            padx=4,
        )
        self._create_metric_tile(stats_row, "Selected", self.brand_selected_var).grid(
            row=0,
            column=2,
            sticky="ew",
            padx=(8, 0),
        )

        browser = tk.Frame(browser_body, bg=self.colors["card"])
        browser.grid(row=1, column=0, sticky="nsew")
        browser.grid_rowconfigure(0, weight=1)
        browser.grid_columnconfigure(0, weight=1)

        list_container = tk.Frame(
            browser,
            bg="#FFFFFF",
            highlightbackground=self.colors["border"],
            highlightthickness=1,
        )
        list_container.grid(row=0, column=0, sticky="nsew")
        list_container.grid_rowconfigure(0, weight=1)
        list_container.grid_columnconfigure(0, weight=1)

        self.brand_listbox = tk.Listbox(
            list_container,
            selectmode=tk.MULTIPLE,
            activestyle="none",
            bg="#FFFFFF",
            fg=self.colors["text"],
            selectbackground=self.colors["accent"],
            selectforeground="#FFFFFF",
            highlightthickness=0,
            borderwidth=0,
            font=("Segoe UI", 10),
            exportselection=False,
        )
        self.brand_listbox.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(list_container, orient="vertical", command=self.brand_listbox.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.brand_listbox.config(yscrollcommand=scroll.set)
        self.brand_listbox.bind("<<ListboxSelect>>", self._on_brand_listbox_select)
        self.brand_listbox.bind("<Key>", self.on_listbox_keypress)
        self.brand_listbox.bind("<space>", self.toggle_active_brand_selection)
        self.brand_listbox.bind("<Return>", self.toggle_active_brand_selection)

        alpha_panel = tk.Frame(browser, bg=self.colors["card"])
        alpha_panel.grid(row=0, column=1, sticky="ns", padx=(10, 0))
        for idx, label in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ"):
            ttk.Button(
                alpha_panel,
                text=label,
                style="Quiet.TButton",
                width=3,
                command=lambda letter=label: self.scroll_to_letter(letter),
            ).grid(row=idx % 9, column=idx // 9, padx=2, pady=2, sticky="ew")

        footer = tk.Frame(browser_body, bg=self.colors["card"])
        footer.grid(row=2, column=0, sticky="ew", pady=(10, 0))
        tk.Label(
            footer,
            textvariable=self.brand_summary_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 9),
        ).pack(side="left")
        ttk.Button(
            footer,
            text="Load Brands",
            style="Quiet.TButton",
            command=self.load_brands,
        ).pack(side="right")

    def _build_activity_tab(self):
        self.activity_tab.grid_columnconfigure(0, weight=3)
        self.activity_tab.grid_columnconfigure(1, weight=2)
        self.activity_tab.grid_rowconfigure(0, weight=1)

        log_card, log_body = self._create_card(
            self.activity_tab,
            "Activity Log",
            "Recent actions in this session. Use this to confirm refreshes, loads, saves, and delivery steps.",
        )
        log_card.grid(row=0, column=0, sticky="nsew", padx=(0, 9))
        self._build_log_card(log_body)

        right_stack = tk.Frame(self.activity_tab, bg=self.colors["bg"])
        right_stack.grid(row=0, column=1, sticky="nsew", padx=(9, 0))
        right_stack.grid_rowconfigure(1, weight=1)
        right_stack.grid_columnconfigure(0, weight=1)

        session_card, session_body = self._create_card(
            right_stack,
            "Session Snapshot",
            "A quick view of the saved workspace and what files are available right now.",
        )
        session_card.grid(row=0, column=0, sticky="ew")
        tk.Label(
            session_body,
            textvariable=self.source_snapshot_var,
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 9),
            wraplength=360,
            justify="left",
        ).pack(anchor="w")
        tk.Label(
            session_body,
            textvariable=self.order_window_summary_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=360,
            justify="left",
        ).pack(anchor="w", pady=(8, 0))
        tk.Label(
            session_body,
            textvariable=self.brand_load_status_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=360,
            justify="left",
        ).pack(anchor="w", pady=(8, 0))
        tk.Label(
            session_body,
            textvariable=self.settings_state_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
            wraplength=360,
            justify="left",
        ).pack(anchor="w", pady=(8, 0))

        help_card, help_body = self._create_card(
            right_stack,
            "Keyboard Shortcuts",
            "Fast paths for common tasks when you are moving through the app all day.",
        )
        help_card.grid(row=1, column=0, sticky="nsew", pady=(12, 0))
        tk.Label(
            help_body,
            text=(
                "Ctrl+F  Focus brand search\n"
                "Ctrl+U  Update files\n"
                "Ctrl+L  Load brands\n"
                "Ctrl+Enter  Generate, upload, and email\n"
                "Alt+1 / Alt+2 / Alt+3 / Alt+4  Switch tabs\n"
                "Enter on search  Jump to first visible brand\n"
                "Type in list  Quick prefix jump\n"
                "Space on list  Toggle active brand"
            ),
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Consolas", 9),
            justify="left",
            anchor="nw",
        ).pack(fill="both", expand=True, anchor="nw")

    def _build_log_card(self, body):
        self.log_text = ScrolledText(
            body,
            height=14,
            wrap="word",
            bg=self.colors["log_bg"],
            fg=self.colors["text"],
            insertbackground=self.colors["text"],
            relief="flat",
            borderwidth=0,
            font=("Consolas", 10),
            padx=10,
            pady=10,
        )
        self.log_text.pack(fill="both", expand=True)
        self.log_text.configure(state="disabled")

    def _bind_events(self):
        self.input_dir_var.trace_add("write", lambda *_: self._refresh_path_summaries())
        self.output_dir_var.trace_add("write", lambda *_: self._refresh_path_summaries())
        self.emails_var.trace_add("write", lambda *_: self._update_email_display())
        self.brand_search_var.trace_add("write", lambda *_: self.filter_brand_list())
        self.prefer_catalog_api_var.trace_add("write", lambda *_: self._update_catalog_refresh_caption())
        self.fetch_order_reports_var.trace_add("write", lambda *_: self._update_order_report_caption())

        for var in (
            self.input_dir_var,
            self.output_dir_var,
            self.emails_var,
            self.prefer_catalog_api_var,
            self.fetch_order_reports_var,
            self.include_cost_var,
        ):
            var.trace_add("write", lambda *_: self._schedule_autosave())

        self.brand_search_entry.bind("<Return>", self._select_first_visible_brand)
        self.brand_search_entry.bind("<Down>", self._move_focus_to_brand_list)
        self.brand_search_entry.bind("<Escape>", self._clear_brand_search)

        self.master.bind_all("<Control-f>", self.focus_brand_search)
        self.master.bind_all("<Control-u>", self._shortcut_update_files)
        self.master.bind_all("<Control-l>", self._shortcut_load_brands)
        self.master.bind_all("<Control-Return>", self._shortcut_run_process)
        self.master.bind_all("<Alt-1>", lambda event: self._select_tab(self.overview_tab))
        self.master.bind_all("<Alt-2>", lambda event: self._select_tab(self.brands_tab))
        self.master.bind_all("<Alt-3>", lambda event: self._select_tab(self.activity_tab))
        self.master.bind_all("<Alt-4>", lambda event: self._select_tab(self.settings_tab))
        self.master.bind_all("<Escape>", self._global_escape)

    def _set_status(self, headline, detail=None):
        self.status_var.set(headline)
        if detail is not None:
            self.status_detail_var.set(detail)

    def append_log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")
        self.master.update_idletasks()

    def _refresh_path_summaries(self):
        input_dir = self.input_dir_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        self.input_summary_var.set(
            input_dir if input_dir else "Choose the folder where catalog CSVs and optional order-report source files live."
        )
        self.output_summary_var.set(
            output_dir if output_dir else "Choose the folder where the generated brand workbooks should be written."
        )
        self._refresh_source_snapshot()

    def _refresh_source_snapshot(self):
        input_dir = self.input_dir_var.get().strip()
        if not input_dir or not os.path.isdir(input_dir):
            self.catalog_count_var.set("0")
            self.order_file_count_var.set("0")
            self.source_snapshot_var.set("Choose a valid input folder to scan catalog CSVs and Dutchie order reports.")
            self.order_window_summary_var.set("Dutchie order windows: none detected yet.")
            return

        try:
            filenames = os.listdir(input_dir)
        except Exception:
            self.catalog_count_var.set("0")
            self.order_file_count_var.set("0")
            self.source_snapshot_var.set("The input folder could not be read.")
            self.order_window_summary_var.set("Dutchie order windows: unavailable.")
            return

        csv_files = sorted(
            fn for fn in filenames if fn.lower().endswith(".csv") and not is_order_report_filename(fn)
        )
        order_files = sorted(fn for fn in filenames if is_order_report_filename(fn))
        self.catalog_count_var.set(str(len(csv_files)))
        self.order_file_count_var.set(str(len(order_files)))

        latest_paths = [
            os.path.join(input_dir, fn)
            for fn in csv_files + order_files
            if os.path.isfile(os.path.join(input_dir, fn))
        ]
        if latest_paths:
            latest_mtime = max(os.path.getmtime(path) for path in latest_paths)
            latest_text = datetime.fromtimestamp(latest_mtime).strftime("%b %d, %Y %I:%M %p").lstrip("0")
        else:
            latest_text = "No source files yet"

        if csv_files or order_files:
            self.source_snapshot_var.set(
                f"{len(csv_files)} catalog CSVs and {len(order_files)} order-report files found. Latest file activity: {latest_text}."
            )
        else:
            self.source_snapshot_var.set("The selected input folder is valid, but it does not contain catalog CSVs or inventory order reports yet.")

        order_windows = summarize_order_report_files(input_dir)
        if order_windows:
            self.order_window_summary_var.set(f"Dutchie order windows available: {order_windows}.")
        else:
            self.order_window_summary_var.set("Dutchie order windows: none detected yet.")

    def _parse_recipients(self):
        return [item.strip() for item in self.emails_var.get().split(",") if item.strip()]

    def _invalid_recipients(self):
        return [email for email in self._parse_recipients() if not EMAIL_REGEX.match(email)]

    def _update_email_display(self):
        for child in self.email_chip_frame.winfo_children():
            child.destroy()

        recipients = self._parse_recipients()
        invalid = set(self._invalid_recipients())
        self.recipient_count_var.set(str(len(recipients)))

        if not recipients:
            tk.Label(
                self.email_chip_frame,
                text="No recipients entered yet.",
                bg=self.colors["card"],
                fg=self.colors["muted"],
                font=("Segoe UI", 9),
            ).grid(row=0, column=0, sticky="w")
            self.email_summary_var.set("No recipients saved yet.")
            return

        max_cols = 2
        for idx, email in enumerate(recipients):
            row = idx // max_cols
            col = idx % max_cols
            is_invalid = email in invalid
            chip = tk.Label(
                self.email_chip_frame,
                text=email,
                bg="#FDE7E7" if is_invalid else self.colors["accent_soft"],
                fg="#9F1239" if is_invalid else self.colors["accent_dark"],
                font=("Segoe UI", 9, "bold"),
                padx=10,
                pady=5,
            )
            chip.grid(row=row, column=col, sticky="w", padx=(0, 8), pady=(0, 8))

        if invalid:
            self.email_summary_var.set(
                f"{len(recipients)} recipient entries saved. {len(invalid)} need attention before sending."
            )
        else:
            self.email_summary_var.set(
                f"{len(recipients)} recipient{'s' if len(recipients) != 1 else ''} ready for the outgoing Drive-link email."
            )

    def _update_catalog_refresh_caption(self):
        api_ready, available_codes, missing_codes, error_text = dutchie_api_readiness(DEFAULT_API_ENV_FILE)

        if self.prefer_catalog_api_var.get():
            if api_ready:
                store_list = ", ".join(available_codes)
                self.catalog_refresh_caption_var.set(
                    "Update Files will try the Dutchie API first for catalog CSVs and inventory order reports "
                    f"for {store_list}, then fall back to the browser exporters if an API refresh fails."
                )
            elif error_text:
                self.catalog_refresh_caption_var.set(
                    "Dutchie API preference is on, but the API configuration could not be read. "
                    "Update Files will fall back to the browser catalog and order-report scripts."
                )
            else:
                missing_list = ", ".join(missing_codes)
                self.catalog_refresh_caption_var.set(
                    "Dutchie API preference is on, but some store keys are missing "
                    f"({missing_list}). Update Files will fall back to the browser catalog and order-report scripts."
                )
        else:
            self.catalog_refresh_caption_var.set(
                "Update Files will use the browser catalog and order-report exporters even if Dutchie API credentials are available."
            )

    def _update_order_report_caption(self):
        if self.fetch_order_reports_var.get():
            self.order_reports_caption_var.set(
                "Update Files will also refresh the Dutchie 7d, 14d, and 30d order-report source files, using the API first when enabled and available."
            )
        else:
            self.order_reports_caption_var.set(
                "Update Files will refresh only catalog CSVs and leave existing Dutchie order-report files untouched."
            )

    def _current_selected_brands(self):
        return set(self.selected_brand_names)

    def _populate_brand_listbox(self, items):
        self.brand_listbox.delete(0, tk.END)

        if not self.all_brands:
            self.filtered_brands = []
            self.brand_listbox.insert(tk.END, "No brands found.")
            return

        if not items:
            self.filtered_brands = []
            self.brand_listbox.insert(tk.END, "No matching brands.")
            return

        self.filtered_brands = list(items)
        for idx, brand in enumerate(self.filtered_brands):
            self.brand_listbox.insert(tk.END, brand)
            if brand in self.selected_brand_names:
                self.brand_listbox.selection_set(idx)

    def _update_brand_summary(self):
        total = len(self.all_brands)
        visible = len(self.filtered_brands)
        selected = len(self.selected_brand_names)
        query = self.brand_search_var.get().strip()

        self.brand_total_var.set(str(total))
        self.brand_visible_var.set(str(visible))
        self.brand_selected_var.set(str(selected))

        if total == 0:
            self.brand_summary_var.set("No brands loaded yet.")
            return

        base = f"{selected} selected • {visible} visible • {total} total"
        if query:
            base += f" • filtered by \"{query}\""
        self.brand_summary_var.set(base)

    def _set_default_brand_hint(self):
        self.brand_hint_var.set(
            "Ctrl+F focuses search. Type in the list for a quick prefix jump. Space toggles the active brand."
        )

    def filter_brand_list(self):
        query = self.brand_search_var.get().strip().lower()
        if not query:
            visible = list(self.all_brands)
        else:
            visible = [brand for brand in self.all_brands if query in brand.lower()]

        self._populate_brand_listbox(visible)
        self._update_brand_summary()
        if not self.quick_jump_buffer:
            self._set_default_brand_hint()

    def _on_brand_listbox_select(self, _event=None):
        if not self.filtered_brands:
            self._update_brand_summary()
            return

        self.selected_brand_names.difference_update(self.filtered_brands)
        for idx in self.brand_listbox.curselection():
            brand = self.brand_listbox.get(idx)
            if brand not in ("No brands found.", "No matching brands."):
                self.selected_brand_names.add(brand)
        self._update_brand_summary()

    def _has_catalog_exports(self):
        input_dir = self.input_dir_var.get().strip()
        return bool(
            input_dir
            and os.path.isdir(input_dir)
            and bool(list_catalog_csv_files(input_dir))
        )

    def _autoload_saved_workspace(self):
        if self._has_catalog_exports():
            self.append_log("Saved catalog exports were found. Auto-loading the brand library.")
            if self.load_brands(silent=True):
                self._set_status(
                    "Saved workspace restored.",
                    "Catalog files were detected in the saved input folder and the brand library was loaded automatically.",
                )

    def show_loading(self, message="Processing...", detail="Working on your request..."):
        if hasattr(self, "loading_window") and self.loading_window.winfo_exists():
            self.hide_loading()

        self.loading_message_var = tk.StringVar(value=message)
        self.loading_detail_var = tk.StringVar(value=detail)
        self.loading_step_var = tk.StringVar(value="Starting...")

        self.loading_window = tk.Toplevel(self.master)
        self.loading_window.title(message)
        self.loading_window.transient(self.master)
        self.loading_window.configure(bg=self.colors["border"])
        self.loading_window.resizable(False, False)
        self.loading_window.protocol("WM_DELETE_WINDOW", lambda: None)

        outer = tk.Frame(self.loading_window, bg=self.colors["card"], padx=0, pady=0)
        outer.pack(fill="both", expand=True, padx=1, pady=1)

        tk.Frame(outer, bg=self.colors["accent"], height=5).pack(fill="x")

        frame = tk.Frame(outer, bg=self.colors["card"], padx=22, pady=20)
        frame.pack(fill="both", expand=True)

        header = tk.Frame(frame, bg=self.colors["card"])
        header.pack(fill="x")

        badge = tk.Frame(header, bg=self.colors["accent_soft"], width=42, height=42)
        badge.pack(side="left", padx=(0, 12))
        badge.pack_propagate(False)
        tk.Label(
            badge,
            text="BZ",
            bg=self.colors["accent_soft"],
            fg=self.colors["accent_dark"],
            font=("Segoe UI", 14, "bold"),
        ).pack(expand=True)

        tk.Label(
            header,
            textvariable=self.loading_message_var,
            bg=self.colors["card"],
            fg=self.colors["text"],
            font=("Segoe UI", 13, "bold"),
        ).pack(anchor="w")
        tk.Label(
            header,
            textvariable=self.loading_detail_var,
            bg=self.colors["card"],
            fg=self.colors["muted"],
            font=("Segoe UI", 9),
            wraplength=390,
            justify="left",
        ).pack(anchor="w", pady=(4, 0))

        step_box = tk.Frame(frame, bg="#F7FAF8", padx=12, pady=10)
        step_box.pack(fill="x", pady=(16, 12))
        tk.Label(
            step_box,
            text="Current step",
            bg="#F7FAF8",
            fg=self.colors["muted"],
            font=("Segoe UI", 8, "bold"),
        ).pack(anchor="w")
        tk.Label(
            step_box,
            textvariable=self.loading_step_var,
            bg="#F7FAF8",
            fg=self.colors["text"],
            font=("Segoe UI", 9),
            wraplength=410,
            justify="left",
        ).pack(anchor="w", pady=(3, 0))

        self.loading_progress = ttk.Progressbar(
            frame,
            mode="indeterminate",
            length=420,
            style="Loading.Horizontal.TProgressbar",
        )
        self.loading_progress.pack(fill="x")
        self.loading_progress.start(10)

        self.loading_window.update_idletasks()
        self._center_loading_window()
        self.loading_window.update()

    def _center_loading_window(self):
        width = 500
        height = 230
        x = self.master.winfo_rootx() + (self.master.winfo_width() // 2) - (width // 2)
        y = self.master.winfo_rooty() + (self.master.winfo_height() // 2) - (height // 2)
        self.loading_window.geometry(f"{width}x{height}+{max(40, x)}+{max(40, y)}")

    def update_loading(self, message=None, detail=None, step=None):
        if not hasattr(self, "loading_window") or not self.loading_window.winfo_exists():
            return
        if message is not None and hasattr(self, "loading_message_var"):
            self.loading_message_var.set(message)
            self.loading_window.title(message)
        if detail is not None and hasattr(self, "loading_detail_var"):
            self.loading_detail_var.set(detail)
        if step is not None and hasattr(self, "loading_step_var"):
            self.loading_step_var.set(step)
        self.loading_window.update_idletasks()
        self.loading_window.update()

    def hide_loading(self):
        if hasattr(self, "loading_progress"):
            try:
                self.loading_progress.stop()
            except Exception:
                pass
        if hasattr(self, "loading_window") and self.loading_window.winfo_exists():
            self.loading_window.destroy()

    def focus_brand_search(self, _event=None):
        self._select_tab(self.brands_tab)
        self.brand_search_entry.focus_set()
        self.brand_search_entry.selection_range(0, tk.END)
        return "break"

    def _move_focus_to_brand_list(self, _event=None):
        if self.filtered_brands:
            self.brand_listbox.focus_set()
            self._focus_listbox_index(0, flash=True)
        return "break"

    def _clear_brand_search(self, _event=None):
        if self.brand_search_var.get():
            self.brand_search_var.set("")
        self._clear_quick_jump()
        return "break"

    def _select_first_visible_brand(self, _event=None):
        if self.filtered_brands:
            self.brand_listbox.focus_set()
            self._focus_listbox_index(0, flash=True)
        return "break"

    def scroll_to_letter(self, letter):
        self._select_tab(self.brands_tab)
        self.brand_listbox.focus_set()
        self._jump_to_brand_prefix(letter.lower())

    def _jump_to_brand_prefix(self, prefix):
        if not prefix:
            return False
        for idx, brand in enumerate(self.filtered_brands):
            if brand.lower().startswith(prefix.lower()):
                self._focus_listbox_index(idx, flash=True)
                return True
        return False

    def on_listbox_keypress(self, event):
        if event.keysym in {"Up", "Down", "Left", "Right", "Home", "End", "Prior", "Next"}:
            return
        if event.keysym in {"space", "Return"}:
            return self.toggle_active_brand_selection()
        if event.keysym == "Escape":
            return self._clear_brand_search()
        if not event.char or not event.char.isalnum():
            return

        now = time.monotonic()
        if now - self.quick_jump_last_ts > 1.1:
            self.quick_jump_buffer = ""
        self.quick_jump_last_ts = now
        self.quick_jump_buffer += event.char.lower()

        if self.jump_reset_job is not None:
            self.master.after_cancel(self.jump_reset_job)
        self.jump_reset_job = self.master.after(1200, self._clear_quick_jump)

        found = self._jump_to_brand_prefix(self.quick_jump_buffer)
        if found:
            self.brand_hint_var.set(f"Quick jump: '{self.quick_jump_buffer}'")
        else:
            self.brand_hint_var.set(f"No visible brands start with '{self.quick_jump_buffer}'.")
        return "break"

    def _clear_quick_jump(self):
        self.quick_jump_buffer = ""
        self.quick_jump_last_ts = 0.0
        if self.jump_reset_job is not None:
            try:
                self.master.after_cancel(self.jump_reset_job)
            except Exception:
                pass
        self.jump_reset_job = None
        self._set_default_brand_hint()

    def _focus_listbox_index(self, index, flash=False):
        if index < 0 or index >= self.brand_listbox.size():
            return
        item = self.brand_listbox.get(index)
        if item in ("No brands found.", "No matching brands."):
            return

        current_selection = self.brand_listbox.curselection()
        self.brand_listbox.activate(index)
        self.brand_listbox.see(index)

        if not flash:
            return

        self.brand_listbox.selection_clear(0, tk.END)
        self.brand_listbox.selection_set(index)

        def restore_selection():
            self.brand_listbox.selection_clear(0, tk.END)
            for idx in current_selection:
                if idx < self.brand_listbox.size():
                    self.brand_listbox.selection_set(idx)
            self.brand_listbox.activate(index)
            self.brand_listbox.see(index)

        self.master.after(650, restore_selection)

    def toggle_active_brand_selection(self, _event=None):
        if not self.filtered_brands:
            return "break"

        active_index = self.brand_listbox.index(tk.ACTIVE)
        if active_index < 0 or active_index >= len(self.filtered_brands):
            return "break"

        brand = self.brand_listbox.get(active_index)
        if brand in ("No brands found.", "No matching brands."):
            return "break"

        if active_index in self.brand_listbox.curselection():
            self.brand_listbox.selection_clear(active_index)
            self.selected_brand_names.discard(brand)
        else:
            self.brand_listbox.selection_set(active_index)
            self.selected_brand_names.add(brand)

        self.brand_listbox.activate(active_index)
        self.brand_listbox.see(active_index)
        self._update_brand_summary()
        return "break"

    def _create_metric_tile(self, parent, title, value_var):
        tile = tk.Frame(
            parent,
            bg="#F7FAF8",
            highlightbackground=self.colors["border"],
            highlightthickness=1,
            padx=10,
            pady=8,
        )
        tk.Label(
            tile,
            text=title,
            bg="#F7FAF8",
            fg=self.colors["muted"],
            font=("Segoe UI", 8, "bold"),
        ).pack(anchor="w")
        tk.Label(
            tile,
            textvariable=value_var,
            bg="#F7FAF8",
            fg=self.colors["text"],
            font=("Segoe UI Semibold", 13),
        ).pack(anchor="w", pady=(2, 0))
        return tile

    def _persist_settings(self, add_log=False, update_status=False):
        save_config(
            self.input_dir_var.get().strip(),
            self.output_dir_var.get().strip(),
            self.fetch_order_reports_var.get(),
            emails=self.emails_var.get().strip(),
            include_cost=self.include_cost_var.get(),
            prefer_catalog_api=self.prefer_catalog_api_var.get(),
            task_eta_seconds=getattr(self, "task_eta_seconds", {}),
        )
        self.settings_state_var.set(
            f"Auto-saved {datetime.now().strftime('%I:%M %p').lstrip('0')}. Use Save Settings if you want a manual checkpoint."
        )
        if add_log:
            self.append_log("Saved workspace settings.")
        if update_status:
            self._set_status(
                "Settings saved.",
                "Your folders, recipients, and report preferences will be restored next time.",
            )

    def _schedule_autosave(self):
        if self.autosave_job is not None:
            self.master.after_cancel(self.autosave_job)
        self.settings_state_var.set("Saving your latest changes...")
        self.autosave_job = self.master.after(900, self._autosave_settings)

    def _autosave_settings(self):
        self.autosave_job = None
        try:
            self._persist_settings(add_log=False, update_status=False)
        except Exception:
            self.settings_state_var.set("Could not auto-save settings. You can still use Save Settings manually.")

    def save_settings(self, quiet=False):
        if self.autosave_job is not None:
            self.master.after_cancel(self.autosave_job)
            self.autosave_job = None
        self._persist_settings(add_log=True, update_status=True)
        if not quiet:
            messagebox.showinfo("Saved", "Workspace settings were saved.")

    def on_close(self):
        try:
            if self.autosave_job is not None:
                self.master.after_cancel(self.autosave_job)
                self.autosave_job = None
            if self.jump_reset_job is not None:
                self.master.after_cancel(self.jump_reset_job)
                self.jump_reset_job = None
            self._persist_settings(add_log=False, update_status=False)
        except Exception:
            pass
        self.master.destroy()

    def browse_input(self):
        folder = filedialog.askdirectory()
        if folder:
            self.input_dir_var.set(folder)
            self.append_log(f"Selected input folder: {folder}")
            if self._has_catalog_exports():
                self.load_brands(silent=True)

    def browse_output(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_dir_var.set(folder)
            self.append_log(f"Selected output folder: {folder}")

    def select_all_brands(self):
        if not self.filtered_brands:
            return
        self.selected_brand_names.update(self.filtered_brands)
        self.brand_listbox.selection_set(0, tk.END)
        self._update_brand_summary()

    def clear_selected_brands(self):
        self.selected_brand_names.clear()
        self.brand_listbox.selection_clear(0, tk.END)
        self._update_brand_summary()

    def _select_tab(self, tab):
        self.notebook.select(tab)
        return "break"

    def _shortcut_update_files(self, _event=None):
        self.get_files()
        return "break"

    def _shortcut_load_brands(self, _event=None):
        self.load_brands()
        return "break"

    def _shortcut_run_process(self, _event=None):
        self.run_process()
        return "break"

    def _global_escape(self, _event=None):
        if self.brand_search_var.get():
            self.brand_search_var.set("")
            self._clear_quick_jump()
            return "break"
        return None

    def _resolve_script_path(self, script_name):
        return Path(__file__).resolve().with_name(script_name)

    def _run_script(self, script_name, *args):
        script_path = self._resolve_script_path(script_name)
        if not script_path.exists():
            raise FileNotFoundError(f"{script_name} was not found at {script_path}.")

        cmd = [sys.executable, str(script_path), *[str(arg) for arg in args]]
        started_at = time.perf_counter()
        self.append_log(f"Running: {' '.join(str(part) for part in cmd)}")
        self.update_loading(step=f"Started {script_name}. Waiting for Dutchie response...")

        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )

        output_lines = []
        if process.stdout is not None:
            for raw_line in process.stdout:
                line = raw_line.rstrip()
                output_lines.append(raw_line)
                if line:
                    print(line, flush=True)
                    self.append_log(line)
                    if line.startswith(("[FETCH]", "[SAVED]", "[INFO]", "[VERIFY]", "[WARN]")):
                        self.update_loading(step=line)

        return_code = process.wait()
        combined_output = "".join(output_lines)
        elapsed = time.perf_counter() - started_at

        if return_code != 0:
            self.append_log(f"{script_name} exited with code {return_code} after {elapsed:.1f}s.")
            raise subprocess.CalledProcessError(
                return_code,
                cmd,
                output=combined_output,
                stderr=combined_output,
            )

        self.append_log(f"{script_name} finished successfully in {elapsed:.1f}s.")
        return subprocess.CompletedProcess(
            cmd,
            return_code,
            stdout=combined_output,
            stderr="",
        )

    def _refresh_catalog_exports(self, input_dir):
        prefer_api = self.prefer_catalog_api_var.get()
        api_ready, available_codes, missing_codes, error_text = dutchie_api_readiness(DEFAULT_API_ENV_FILE)
        available_text = ", ".join(available_codes) if available_codes else "none"
        missing_text = ", ".join(missing_codes) if missing_codes else "none"
        self.append_log(
            f"Catalog refresh readiness: prefer_api={prefer_api}, api_ready={api_ready}, "
            f"available stores={available_text}, missing stores={missing_text}."
        )

        if prefer_api and api_ready:
            self.append_log(
                f"Dutchie API catalog refresh is configured for all stores. Trying the API exporter first with {DUTCHIE_API_WORKERS} workers."
            )
            self.update_loading(
                message="Downloading catalog data...",
                detail=f"Fetching live inventory CSVs from Dutchie with {DUTCHIE_API_WORKERS} worker threads.",
                step=f"Catalog API stores queued: {available_text}",
            )
            try:
                self._run_script(CATALOG_API_SCRIPT, input_dir, "--workers", DUTCHIE_API_WORKERS)
                return "api"
            except subprocess.CalledProcessError:
                self.append_log(
                    "Dutchie API catalog refresh failed. Falling back to the browser catalog export script."
                )
                self.update_loading(
                    message="Switching catalog method...",
                    detail="The API catalog refresh failed, so the browser exporter is taking over.",
                    step="Starting browser catalog fallback.",
                )

        elif prefer_api:
            if error_text:
                self.append_log(
                    f"Dutchie API readiness could not be confirmed: {error_text}. Falling back to the browser catalog export script."
                )
            else:
                self.append_log(
                    f"Dutchie API is missing store credentials for: {missing_text}. "
                    "Falling back to the browser catalog export script."
                )
        else:
            self.append_log("Catalog refresh is set to browser mode. Skipping the Dutchie API path.")

        self.update_loading(
            message="Downloading catalog data...",
            detail="Using the browser catalog exporter.",
            step="Starting browser catalog export.",
        )
        self._run_script(CATALOG_BROWSER_SCRIPT, input_dir)
        return "browser"

    def _refresh_order_report_exports(self, input_dir):
        prefer_api = self.prefer_catalog_api_var.get()
        api_ready, available_codes, missing_codes, error_text = dutchie_api_readiness(DEFAULT_API_ENV_FILE)
        available_text = ", ".join(available_codes) if available_codes else "none"
        missing_text = ", ".join(missing_codes) if missing_codes else "none"
        self.append_log(
            f"Order-report refresh readiness: prefer_api={prefer_api}, api_ready={api_ready}, "
            f"available stores={available_text}, missing stores={missing_text}."
        )

        if prefer_api and api_ready:
            self.append_log(
                f"Dutchie API order-report refresh is configured for all stores. Trying the API exporter first with {DUTCHIE_API_WORKERS} workers."
            )
            self.update_loading(
                message="Downloading order reports...",
                detail="Fetching Dutchie inventory/order source files and building 7d, 14d, and 30d windows.",
                step=f"Order-report API stores queued: {available_text}",
            )
            try:
                self._run_script(ORDER_REPORT_API_SCRIPT, input_dir, "--workers", DUTCHIE_API_WORKERS)
                return "api"
            except subprocess.CalledProcessError:
                self.append_log(
                    "Dutchie API order-report refresh failed. Falling back to the browser order-report export script."
                )
                self.update_loading(
                    message="Switching order-report method...",
                    detail="The API order-report refresh failed, so the browser exporter is taking over.",
                    step="Starting browser order-report fallback.",
                )
        elif prefer_api:
            if error_text:
                self.append_log(
                    f"Dutchie API readiness could not be confirmed for order reports: {error_text}. Falling back to the browser order-report export script."
                )
            else:
                self.append_log(
                    f"Dutchie API is missing store credentials for: {missing_text}. "
                    "Falling back to the browser order-report export script."
                )
        else:
            self.append_log("Order-report refresh is set to browser mode. Skipping the Dutchie API path.")

        self.update_loading(
            message="Downloading order reports...",
            detail="Using the browser order-report exporter.",
            step="Starting browser order-report export.",
        )
        self._run_script(ORDER_REPORT_BROWSER_SCRIPT, input_dir)
        return "browser"

    def get_files(self):
        in_dir = self.input_dir_var.get().strip()
        fetch_order_reports = self.fetch_order_reports_var.get()
        prefer_api = self.prefer_catalog_api_var.get()

        if not in_dir or not os.path.isdir(in_dir):
            messagebox.showerror("Error", "Please choose a valid input folder first.")
            return

        self.append_log("Starting source refresh.")
        self.append_log(f"Input folder: {in_dir}")
        self.append_log(
            f"Refresh options: prefer_api={prefer_api}, refresh_order_reports={fetch_order_reports}, "
            f"api_workers={DUTCHIE_API_WORKERS}."
        )
        self._set_status(
            "Refreshing source files...",
            "Catalog CSVs are being refreshed, and Dutchie order reports will follow if the toggle is enabled.",
        )
        self.show_loading(
            "Updating files...",
            (
                "Refreshing catalog CSV exports and Dutchie order-report source files with API preference and browser fallback."
                if prefer_api
                else "Refreshing catalog CSV exports and Dutchie order-report source files with the browser scripts."
            ),
        )

        try:
            existing_catalog_count = len(list_catalog_csv_files(in_dir))
            existing_order_count = len(list_order_report_files(in_dir))
            self.append_log(
                f"Before refresh: {existing_catalog_count} catalog CSV(s), "
                f"{existing_order_count} order-report file(s) in the input folder."
            )
            self.update_loading(
                message="Preparing source refresh...",
                detail="Clearing old source exports so the next reports use fresh data.",
                step="Removing stale catalog CSVs and selected order-report files.",
            )
            deleted_paths = clear_old_input_exports(in_dir, clear_order_reports=fetch_order_reports)
            deleted_catalog_count = sum(
                1 for path in deleted_paths if str(path).lower().endswith(".csv") and not is_order_report_filename(os.path.basename(path))
            )
            deleted_order_count = sum(1 for path in deleted_paths if is_order_report_filename(os.path.basename(path)))
            self.append_log(
                f"Cleared {deleted_catalog_count} old catalog CSV(s) and {deleted_order_count} old order-report file(s)."
            )

            self.update_loading(
                message="Downloading catalog data...",
                detail="Refreshing store catalog/inventory CSVs now.",
                step="Starting catalog refresh.",
            )
            catalog_mode_used = self._refresh_catalog_exports(in_dir)
            refreshed_catalog_count = len(list_catalog_csv_files(in_dir))
            self.append_log(
                f"Catalog refresh complete via {catalog_mode_used}; "
                f"{refreshed_catalog_count} catalog CSV(s) are now available."
            )

            order_reports_ok = True
            order_report_mode_used = None
            order_report_error = None
            if fetch_order_reports:
                self.update_loading(
                    message="Downloading order reports...",
                    detail="Refreshing Dutchie 7d, 14d, and 30d order-report source files.",
                    step="Starting order-report refresh.",
                )
                try:
                    order_report_mode_used = self._refresh_order_report_exports(in_dir)
                    refreshed_order_count = len(list_order_report_files(in_dir))
                    self.append_log(
                        f"Order-report refresh complete via {order_report_mode_used}; "
                        f"{refreshed_order_count} order-report file(s) are now available."
                    )
                except subprocess.CalledProcessError as exc:
                    order_reports_ok = False
                    order_report_error = exc
                    self.append_log(f"Order-report refresh failed: {exc}")
                    if getattr(exc, "stderr", None):
                        self.append_log(str(exc.stderr).strip())
                except Exception as exc:
                    order_reports_ok = False
                    order_report_error = exc
                    self.append_log(f"Order-report refresh failed: {exc}")
            else:
                self.append_log("Order-report refresh skipped by toggle; existing order-report files were left untouched.")

            self.update_loading(
                message="Finishing source refresh...",
                detail="Updating the source snapshot and reloading the brand library.",
                step="Scanning refreshed files.",
            )
            self.hide_loading()
            self._persist_settings(add_log=False, update_status=False)
            self._refresh_source_snapshot()
            self.load_brands(silent=True)

            catalog_mode_text = "Dutchie API" if catalog_mode_used == "api" else "browser export"
            order_mode_text = "Dutchie API" if order_report_mode_used == "api" else "browser export"

            if not fetch_order_reports:
                self.append_log(
                    f"Catalog CSV refresh finished via {catalog_mode_text}. Order reports were skipped by choice."
                )
                self._set_status(
                    "Catalog refresh complete.",
                    f"Catalog CSVs are current via {catalog_mode_text}. Existing order-report files were left untouched.",
                )
                messagebox.showinfo(
                    "Success",
                    f"Catalog CSVs were refreshed via {catalog_mode_text}. Inventory order report refresh was skipped.",
                )
            elif order_reports_ok:
                self.append_log(
                    f"Catalog CSV refresh via {catalog_mode_text} and inventory order report refresh via {order_mode_text} finished."
                )
                self._set_status(
                    "Source refresh complete.",
                    f"Catalog CSVs ({catalog_mode_text}) and order-report files ({order_mode_text}) are ready, and the brand library has been refreshed.",
                )
                messagebox.showinfo(
                    "Success",
                    f"Catalog CSVs were refreshed via {catalog_mode_text}, and the inventory order report files were refreshed via {order_mode_text}.",
                )
            else:
                self.append_log(
                    f"Catalog CSV refresh via {catalog_mode_text} finished, but the order report refresh failed."
                )
                self._set_status(
                    "Partial refresh complete.",
                    f"Catalog CSVs were updated via {catalog_mode_text}, but the order report refresh did not finish successfully.",
                )
                messagebox.showwarning(
                    "Partial Success",
                    (
                        f"Catalog CSVs were refreshed via {catalog_mode_text}, but the inventory order report refresh failed."
                        if order_report_error is None
                        else (
                            f"Catalog CSVs were refreshed via {catalog_mode_text}, but the inventory order report refresh failed:\n\n"
                            f"{order_report_error}"
                        )
                    ),
                )
        except subprocess.CalledProcessError as e:
            self.hide_loading()
            self.append_log(f"Catalog refresh failed: {e}")
            self._set_status(
                "Refresh failed.",
                "The catalog refresh did not complete. Check the error details and try again.",
            )
            detail = str(e)
            if getattr(e, "stderr", None):
                detail = f"{detail}\n\n{str(e.stderr).strip()}"
            messagebox.showerror("Error", detail)
        except Exception as e:
            self.hide_loading()
            self.append_log(f"Refresh error: {e}")
            self._set_status("Refresh failed.", "An unexpected error interrupted the refresh.")
            messagebox.showerror("Error", str(e))

    def load_brands(self, silent=False):
        in_dir = self.input_dir_var.get().strip()
        if not in_dir or not os.path.isdir(in_dir):
            self.all_brands = []
            self.filtered_brands = []
            self.selected_brand_names.clear()
            self._populate_brand_listbox([])
            self._update_brand_summary()
            self.brand_load_status_var.set("Brand library not loaded yet.")
            if not silent:
                messagebox.showerror("Error", "Invalid input folder.")
            return False

        brand_set = set()
        csv_count = 0
        self.append_log(f"Scanning brands from CSV files in {in_dir}.")
        self._set_status(
            "Loading brands...",
            "Reading the current catalog CSV files to build the selectable brand list.",
        )
        for fn in list_catalog_csv_files(in_dir):
            csv_count += 1
            path = os.path.join(in_dir, fn)
            try:
                df = pd.read_csv(path, nrows=50000)
                if "Brand" in df.columns:
                    new_brands = (
                        df["Brand"]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .str.lower()
                        .unique()
                        .tolist()
                    )
                    brand_set.update(new_brands)
            except Exception:
                pass

        self.all_brands = sorted(brand_set)
        self.selected_brand_names.intersection_update(self.all_brands)
        self.filter_brand_list()

        loaded_text = datetime.now().strftime("%b %d, %Y %I:%M %p").lstrip("0")
        if not self.all_brands:
            self.brand_load_status_var.set(
                f"No brands were found in {csv_count} catalog CSV file{'s' if csv_count != 1 else ''}."
            )
            self.append_log("No brands were found in the current CSV files.")
            self._set_status(
                "No brands found.",
                "Try refreshing files or verify that the input folder contains catalog CSV exports.",
            )
            return False

        self.brand_load_status_var.set(
            f"Loaded {len(self.all_brands)} brands from {csv_count} catalog CSV file{'s' if csv_count != 1 else ''} at {loaded_text}."
        )
        self.append_log(f"Loaded {len(self.all_brands)} brands into the library.")
        self._set_status(
            "Brand library loaded.",
            f"{len(self.all_brands)} brands are available for selection.",
        )
        return True

    def run_process(self):
        in_dir = self.input_dir_var.get().strip()
        out_dir = self.output_dir_var.get().strip()
        emails = self.emails_var.get().strip()
        invalid_recipients = self._invalid_recipients()

        if not (in_dir and out_dir and emails):
            messagebox.showerror("Error", "Need input folder, output folder, and at least one email address.")
            return
        if invalid_recipients:
            messagebox.showerror(
                "Invalid Email",
                "Please fix these email addresses before sending:\n\n" + "\n".join(invalid_recipients),
            )
            return
        if not os.path.isdir(in_dir):
            messagebox.showerror("Error", f"Invalid input folder: {in_dir}")
            return
        if not os.path.isdir(out_dir):
            messagebox.showerror("Error", f"Invalid output folder: {out_dir}")
            return

        selected_brands = sorted(self._current_selected_brands())
        if not selected_brands:
            messagebox.showinfo(
                "No Selection",
                "No brands selected. All brand data found in the CSVs will be processed.",
            )

        all_brand_map = {}
        self.append_log("Starting report generation workflow.")
        self._set_status(
            "Generating reports...",
            "Building the selected brand workbooks, then uploading them to Drive and sending the delivery email.",
        )
        self.show_loading(
            "Generating reports...",
            "This can take a moment while workbooks are built, uploaded, and the email is sent.",
        )

        try:
            for fname in list_catalog_csv_files(in_dir):
                path = os.path.join(in_dir, fname)
                self.update_loading(
                    message="Generating reports...",
                    detail="Building brand workbooks from refreshed catalog CSVs.",
                    step=f"Processing {fname}",
                )
                brand_map = generate_brand_reports(
                    path,
                    out_dir,
                    selected_brands,
                    include_cost=self.include_cost_var.get(),
                    order_reports_dir=in_dir,
                )
                for b_name, xlsx_list in brand_map.items():
                    if b_name not in all_brand_map:
                        all_brand_map[b_name] = []
                    all_brand_map[b_name].extend(xlsx_list)

            if not all_brand_map:
                self.hide_loading()
                self.append_log("No matching workbooks were generated from the current filters.")
                self._set_status(
                    "No reports generated.",
                    "The current brand selection and CSV files did not produce any workbooks.",
                )
                messagebox.showinfo("Done", "No XLSX files generated (possibly no matching data).")
                return

            self.update_loading(
                message="Uploading reports...",
                detail="Creating Drive folders and uploading the finished workbooks.",
                step=f"Uploading {sum(len(v) for v in all_brand_map.values())} workbook(s).",
            )
            brand_links = upload_brand_reports_to_drive(all_brand_map)
            if not brand_links:
                self.hide_loading()
                self.append_log("Drive upload returned no folder links.")
                self._set_status(
                    "Upload failed.",
                    "No Drive folders were created, so the email step was skipped.",
                )
                messagebox.showerror("Error", "No folders created on Drive. Aborting email.")
                return

            lines = []
            for brand_lower, link in brand_links.items():
                lines.append(f"<h3>{brand_lower}</h3>")
                lines.append(f"<p><a href='{link}'>{link}</a></p>")

            joined = "\n".join(lines)
            order_summary = summarize_order_report_files(in_dir)
            order_note = ""
            if order_summary:
                order_note = (
                    "<p>Matching Dutchie order-report rows were added to the "
                    f"<strong>Order</strong> tab when available. Source windows found: {order_summary}.</p>"
                )
            body_html = f"""
            <html>
              <body>
                <p>Hello,</p>
                <p>Here are the public Drive folders (with Available & Unavailable reports) for each brand:</p>
                {order_note}
                {joined}
                <p>Anyone with these links can download the XLSX files.</p>
                <p>Regards,<br>Brand Inventory Bot</p>
              </body>
            </html>
            """
            subject = "Brand Inventory Drive Links"
            self.update_loading(
                message="Sending email...",
                detail="Sending Drive folder links to the saved recipients.",
                step=f"Emailing {len(self._parse_recipients())} recipient(s).",
            )
            send_email_with_gmail_html(subject, body_html, emails)

            self._persist_settings(add_log=False, update_status=False)

            self.hide_loading()
            self.append_log(
                f"Finished report workflow. Uploaded {sum(len(v) for v in all_brand_map.values())} workbook(s) across {len(brand_links)} brand folder(s)."
            )
            self._set_status(
                "Reports delivered.",
                "Drive folders were uploaded and the recipient email was sent successfully.",
            )
            messagebox.showinfo("Success", "All done! Folders uploaded and email sent.")
        except Exception as e:
            self.hide_loading()
            self.append_log(f"Report workflow failed: {e}")
            self._set_status(
                "Workflow failed.",
                "An unexpected error interrupted report generation, upload, or email delivery.",
            )
            traceback.print_exc()
            messagebox.showerror("Error", f"An error occurred:\n{e}")

# ----------------- MODERN GUI SHELL -----------------
class TaskContext:
    def __init__(self, out_queue, task_id):
        self.out_queue = out_queue
        self.task_id = task_id

    def emit(self, kind, **payload):
        payload["kind"] = kind
        payload["task_id"] = self.task_id
        self.out_queue.put(payload)

    def log(self, message, level="info"):
        self.emit("log", message=str(message), level=level)

    def status(self, headline=None, detail=None, step=None):
        self.emit("status", headline=headline, detail=detail, step=step)


class BrandInventoryGUI:
    SCREEN_LABELS = {
        "dashboard": "Dashboard",
        "brands": "Brands",
        "activity": "Activity",
        "settings": "Settings",
    }
    TASK_ETA_ALIASES = {
        "startup-refresh": "update-files",
        "startup-load-brands": "load-brands",
    }

    def __init__(self, master):
        self.master = master
        self.cfg = load_config()
        self.colors = theme_palette(self.cfg.get("theme", DEFAULT_GUI_CONFIG["theme"]))
        self.all_brands = []
        self.filtered_brands = []
        self.selected_brand_names = set()
        self.log_entries = []
        self.recent_events = []
        self.autosave_job = None
        self.task_queue = queue.Queue()
        self.active_task_id = None
        self.active_task_name = ""
        self.active_task_eta_key = ""
        self.active_task_started_at = None
        self.active_task_estimate_seconds = None
        self.loading_eta_job = None
        self.task_eta_seconds = parse_task_eta_seconds(self.cfg.get("task_eta_seconds", {}))
        self.task_running = False
        self.toast_window = None
        self.loading_window = None
        self.loading_details_visible = False
        self.log_auto_scroll = True

        self.master.title("Buzz Brand Inventory Studio")
        self.master.configure(bg=self.colors["bg"])
        self._set_startup_geometry()

        self._init_vars()
        self._configure_styles()
        self._build_layout()
        self._bind_events()
        self._refresh_path_summaries()
        self._refresh_source_snapshot()
        self._update_email_display()
        self._update_brand_summary()
        self._set_status("Ready", "Inventory reporting workflow")
        self.append_log("Workspace ready.")
        self.master.protocol("WM_DELETE_WINDOW", self.on_close)
        self.master.after(200, self._startup_sequence)

    def _init_vars(self):
        cfg = self.cfg
        self.input_dir_var = tk.StringVar(value=cfg.get("input_dir", ""))
        self.output_dir_var = tk.StringVar(value=cfg.get("output_dir", ""))
        self.emails_var = tk.StringVar(value=cfg.get("emails", ""))
        self.fetch_order_reports_var = tk.BooleanVar(value=cfg.get("fetch_order_reports", True))
        self.include_cost_var = tk.BooleanVar(value=cfg.get("include_cost", True))
        self.prefer_catalog_api_var = tk.BooleanVar(value=cfg.get("prefer_catalog_api", True))
        self.auto_update_on_launch_var = tk.BooleanVar(value=cfg.get("auto_update_on_launch", True))
        auto_load_brands = cfg.get(
            "auto_load_brands_after_update",
            cfg.get("auto_load_brands_on_launch", True),
        )
        self.auto_load_brands_after_update_var = tk.BooleanVar(value=auto_load_brands)
        self.auto_load_brands_on_launch_var = self.auto_load_brands_after_update_var
        self.show_startup_loading_var = tk.BooleanVar(value=cfg.get("show_startup_loading", True))
        self.open_output_after_complete_var = tk.BooleanVar(value=cfg.get("open_output_after_complete", False))
        self.compact_mode_var = tk.BooleanVar(value=cfg.get("compact_mode", True))
        self.theme_var = tk.StringVar(value=normalize_theme_name(cfg.get("theme", "flatly")))
        self.brand_search_var = tk.StringVar()
        self.selected_only_var = tk.BooleanVar(value=False)
        self.log_filter_var = tk.StringVar(value="All")

        self.status_var = tk.StringVar(value="Ready")
        self.status_detail_var = tk.StringVar(value="Inventory reporting workflow")
        self.status_pill_var = tk.StringVar(value="Ready")
        self.api_mode_var = tk.StringVar(value="API check pending")
        self.input_summary_var = tk.StringVar()
        self.output_summary_var = tk.StringVar()
        self.email_summary_var = tk.StringVar(value="No recipients")
        self.brand_summary_var = tk.StringVar(value="0 loaded, 0 selected")
        self.brand_total_var = tk.StringVar(value="0")
        self.brand_visible_var = tk.StringVar(value="0")
        self.brand_selected_var = tk.StringVar(value="0")
        self.catalog_count_var = tk.StringVar(value="0")
        self.order_file_count_var = tk.StringVar(value="0")
        self.recipient_count_var = tk.StringVar(value="0")
        self.last_refresh_var = tk.StringVar(value="Never")
        self.last_load_var = tk.StringVar(value="Never")
        self.last_email_var = tk.StringVar(value="Never")
        self.current_state_var = tk.StringVar(value="Idle")
        self.order_windows_var = tk.StringVar(value="None")
        self.source_last_activity_var = tk.StringVar(value="No files")
        self.settings_state_var = tk.StringVar(value="Settings auto-save as you work.")

    def _set_startup_geometry(self):
        self.master.update_idletasks()
        width = min(1280, max(980, self.master.winfo_screenwidth() - 80))
        height = min(820, max(640, self.master.winfo_screenheight() - 90))
        x = max(0, (self.master.winfo_screenwidth() - width) // 2)
        y = max(0, (self.master.winfo_screenheight() - height) // 2)
        self.master.geometry(f"{width}x{height}+{x}+{y}")
        self.master.minsize(1040, 680)

    def _configure_styles(self):
        self.style = tb.Style.get_instance() if tb is not None else ttk.Style(self.master)
        if tb is None and "clam" in self.style.theme_names():
            self.style.theme_use("clam")

        default_font = ("Segoe UI", 9)
        self.master.option_add("*Font", default_font)
        self.master.configure(bg=self.colors["bg"])
        self.style.configure("TFrame", background=self.colors["bg"])
        self.style.configure("Surface.TFrame", background=self.colors["surface"])
        self.style.configure("TLabel", background=self.colors["bg"], foreground=self.colors["text"])
        self.style.configure("Muted.TLabel", background=self.colors["surface"], foreground=self.colors["muted"], font=("Segoe UI", 8))
        self.style.configure("Title.TLabel", background=self.colors["surface"], foreground=self.colors["text"], font=("Segoe UI", 16, "bold"))
        self.style.configure("SectionTitle.TLabel", background=self.colors["surface"], foreground=self.colors["text"], font=("Segoe UI", 11, "bold"))
        self.style.configure("MetricValue.TLabel", background=self.colors["surface"], foreground=self.colors["text"], font=("Segoe UI", 18, "bold"))
        self.style.configure("MetricLabel.TLabel", background=self.colors["surface"], foreground=self.colors["muted"], font=("Segoe UI", 8, "bold"))
        self.style.configure("Primary.TButton", font=("Segoe UI", 9, "bold"), padding=(14, 8))
        self.style.configure("Secondary.TButton", font=("Segoe UI", 9, "bold"), padding=(12, 7))
        self.style.configure("Ghost.TButton", font=("Segoe UI", 9), padding=(10, 6))
        self.style.configure("Danger.TButton", font=("Segoe UI", 9, "bold"), padding=(10, 6))
        self.style.configure("App.TEntry", padding=7)
        self.style.configure("App.TCheckbutton", background=self.colors["surface"], foreground=self.colors["text"], font=("Segoe UI", 9))
        self.style.configure("Treeview", rowheight=28, font=("Segoe UI", 9))
        self.style.configure("Treeview.Heading", font=("Segoe UI", 9, "bold"))
        self.style.configure("Loading.Horizontal.TProgressbar", troughcolor="#E7EFEC", background=self.colors["accent"])

    def _build_layout(self):
        self.root_frame = tk.Frame(self.master, bg=self.colors["bg"])
        self.root_frame.pack(fill="both", expand=True)
        self.root_frame.grid_columnconfigure(0, weight=1)
        self.root_frame.grid_rowconfigure(1, weight=1)

        self._build_header()

        body = tk.Frame(self.root_frame, bg=self.colors["bg"])
        body.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 10))
        body.grid_columnconfigure(1, weight=1)
        body.grid_rowconfigure(0, weight=1)

        self._build_sidebar(body)
        self.content_host = tk.Frame(body, bg=self.colors["bg"])
        self.content_host.grid(row=0, column=1, sticky="nsew")
        self.content_host.grid_rowconfigure(0, weight=1)
        self.content_host.grid_columnconfigure(0, weight=1)

        self.screens = {}
        self._build_dashboard_screen()
        self._build_brands_screen()
        self._build_activity_screen()
        self._build_settings_screen()
        self._build_status_bar()
        self._show_screen("dashboard")

    def _build_header(self):
        header = tk.Frame(self.root_frame, bg=self.colors["surface"], highlightbackground=self.colors["border"], highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", padx=14, pady=14)
        header.grid_columnconfigure(1, weight=1)

        title_block = tk.Frame(header, bg=self.colors["surface"])
        title_block.grid(row=0, column=0, sticky="w", padx=14, pady=10)
        ttk.Label(title_block, text="Buzz Brand Inventory Studio", style="Title.TLabel").pack(anchor="w")
        ttk.Label(title_block, text="Inventory reporting workflow", style="Muted.TLabel").pack(anchor="w", pady=(2, 0))

        right = tk.Frame(header, bg=self.colors["surface"])
        right.grid(row=0, column=2, sticky="e", padx=14, pady=10)
        self.status_pill = tk.Label(right, textvariable=self.status_pill_var, bg=self.colors["accent_soft"], fg=self.colors["accent"], padx=10, pady=4, font=("Segoe UI", 8, "bold"))
        self.status_pill.pack(side="left", padx=(0, 8))
        tk.Label(right, textvariable=self.api_mode_var, bg=self.colors["surface"], fg=self.colors["muted"], font=("Segoe UI", 8)).pack(side="left", padx=(0, 12))
        ttk.Button(right, text="Help", style="Ghost.TButton", command=self.show_help).pack(side="left", padx=(0, 6))
        ttk.Button(right, text="Settings", style="Ghost.TButton", command=lambda: self._show_screen("settings")).pack(side="left")

    def _build_sidebar(self, parent):
        sidebar = tk.Frame(parent, bg=self.colors["surface"], width=150, highlightbackground=self.colors["border"], highlightthickness=1)
        sidebar.grid(row=0, column=0, sticky="ns", padx=(0, 12))
        sidebar.grid_propagate(False)
        self.nav_buttons = {}
        for idx, (key, label) in enumerate(self.SCREEN_LABELS.items()):
            button = tk.Button(
                sidebar,
                text=label,
                anchor="w",
                relief="flat",
                bd=0,
                padx=14,
                pady=10,
                bg=self.colors["surface"],
                fg=self.colors["text"],
                activebackground=self.colors["accent_soft"],
                activeforeground=self.colors["accent"],
                command=lambda name=key: self._show_screen(name),
            )
            button.grid(row=idx, column=0, sticky="ew", padx=8, pady=(8 if idx == 0 else 2, 0))
            self.nav_buttons[key] = button
        sidebar.grid_columnconfigure(0, weight=1)

    def _build_status_bar(self):
        bar = tk.Frame(self.root_frame, bg=self.colors["surface"], highlightbackground=self.colors["border"], highlightthickness=1)
        bar.grid(row=2, column=0, sticky="ew", padx=14, pady=(0, 10))
        bar.grid_columnconfigure(0, weight=1)
        tk.Label(bar, textvariable=self.current_state_var, bg=self.colors["surface"], fg=self.colors["text"], font=("Segoe UI", 8, "bold")).grid(row=0, column=0, sticky="w", padx=12, pady=6)
        tk.Label(bar, textvariable=self.last_refresh_var, bg=self.colors["surface"], fg=self.colors["muted"], font=("Segoe UI", 8)).grid(row=0, column=1, padx=12)
        tk.Label(bar, textvariable=self.brand_selected_var, bg=self.colors["surface"], fg=self.colors["muted"], font=("Segoe UI", 8)).grid(row=0, column=2, padx=12)
        tk.Label(bar, textvariable=self.recipient_count_var, bg=self.colors["surface"], fg=self.colors["muted"], font=("Segoe UI", 8)).grid(row=0, column=3, padx=12)

    def _create_screen(self, name):
        frame = tk.Frame(self.content_host, bg=self.colors["bg"])
        frame.grid(row=0, column=0, sticky="nsew")
        self.screens[name] = frame
        return frame

    def _show_screen(self, name):
        frame = self.screens.get(name)
        if not frame:
            return "break"
        frame.tkraise()
        for key, button in self.nav_buttons.items():
            selected = key == name
            button.configure(
                bg=self.colors["accent_soft"] if selected else self.colors["surface"],
                fg=self.colors["accent"] if selected else self.colors["text"],
                font=("Segoe UI", 9, "bold" if selected else "normal"),
            )
        return "break"

    def _section_card(self, parent, title=None):
        card = tk.Frame(parent, bg=self.colors["surface"], highlightbackground=self.colors["border"], highlightthickness=1)
        if title:
            ttk.Label(card, text=title, style="SectionTitle.TLabel").pack(anchor="w", padx=14, pady=(12, 4))
        body = tk.Frame(card, bg=self.colors["surface"])
        body.pack(fill="both", expand=True, padx=14, pady=(6, 14))
        return card, body

    def _metric_card(self, parent, label, value_var, status_var=None):
        card = tk.Frame(parent, bg=self.colors["surface"], highlightbackground=self.colors["border"], highlightthickness=1)
        ttk.Label(card, text=label.upper(), style="MetricLabel.TLabel").pack(anchor="w", padx=12, pady=(10, 0))
        ttk.Label(card, textvariable=value_var, style="MetricValue.TLabel").pack(anchor="w", padx=12, pady=(2, 0))
        if status_var is not None:
            ttk.Label(card, textvariable=status_var, style="Muted.TLabel").pack(anchor="w", padx=12, pady=(0, 10))
        else:
            tk.Frame(card, height=8, bg=self.colors["surface"]).pack()
        return card

    def _action_card(self, parent, step, title, subtitle, button_text, command, primary=False):
        card = tk.Frame(parent, bg=self.colors["surface"], highlightbackground=self.colors["border"], highlightthickness=1)
        card.grid_columnconfigure(1, weight=1)
        tk.Label(card, text=str(step), bg=self.colors["accent_soft"], fg=self.colors["accent"], width=3, height=1, font=("Segoe UI", 11, "bold")).grid(row=0, column=0, rowspan=2, padx=12, pady=12, sticky="n")
        ttk.Label(card, text=title, style="SectionTitle.TLabel").grid(row=0, column=1, sticky="w", pady=(12, 0))
        ttk.Label(card, text=subtitle, style="Muted.TLabel").grid(row=1, column=1, sticky="w", pady=(2, 12))
        button = ttk.Button(card, text=button_text, style="Primary.TButton" if primary else "Secondary.TButton", command=command)
        button.grid(row=0, column=2, rowspan=2, sticky="e", padx=12, pady=12)
        self.workflow_buttons.append(button)
        return card

    def _build_dashboard_screen(self):
        screen = self._create_screen("dashboard")
        screen.grid_columnconfigure(0, weight=1)
        screen.grid_rowconfigure(2, weight=1)
        self.workflow_buttons = []

        metrics = tk.Frame(screen, bg=self.colors["bg"])
        metrics.grid(row=0, column=0, sticky="ew")
        for col in range(5):
            metrics.grid_columnconfigure(col, weight=1)
        self._metric_card(metrics, "Catalog Files", self.catalog_count_var).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self._metric_card(metrics, "Order Files", self.order_file_count_var).grid(row=0, column=1, sticky="ew", padx=4)
        self._metric_card(metrics, "Brands Loaded", self.brand_total_var).grid(row=0, column=2, sticky="ew", padx=4)
        self._metric_card(metrics, "Selected", self.brand_selected_var).grid(row=0, column=3, sticky="ew", padx=4)
        self._metric_card(metrics, "Recipients", self.recipient_count_var).grid(row=0, column=4, sticky="ew", padx=(8, 0))

        workflow_card, workflow_body = self._section_card(screen, "Workflow")
        workflow_card.grid(row=1, column=0, sticky="ew", pady=12)
        workflow_body.grid_columnconfigure(0, weight=1)
        workflow_body.grid_columnconfigure(1, weight=1)
        workflow_body.grid_columnconfigure(2, weight=1)
        self._action_card(workflow_body, 1, "Update Files", "Catalog + order reports", "Update Files", self.get_files).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self._action_card(workflow_body, 2, "Load Brands", "Refresh brand library", "Load Brands", self.load_brands).grid(row=0, column=1, sticky="ew", padx=4)
        self._action_card(workflow_body, 3, "Generate Reports", "Upload to Drive and email links", "Generate, Upload & Email", self.run_process, primary=True).grid(row=0, column=2, sticky="ew", padx=(8, 0))

        lower = tk.Frame(screen, bg=self.colors["bg"])
        lower.grid(row=2, column=0, sticky="nsew")
        lower.grid_columnconfigure(0, weight=3)
        lower.grid_columnconfigure(1, weight=2)
        lower.grid_rowconfigure(0, weight=1)

        left = tk.Frame(lower, bg=self.colors["bg"])
        left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        left.grid_columnconfigure(0, weight=1)
        recipients_card, recipients_body = self._section_card(left, "Recipients")
        recipients_card.grid(row=0, column=0, sticky="ew")
        recipients_body.grid_columnconfigure(0, weight=1)
        ttk.Entry(recipients_body, textvariable=self.emails_var, style="App.TEntry").grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(recipients_body, text="Save", style="Ghost.TButton", command=self.save_settings).grid(row=0, column=1)
        ttk.Label(recipients_body, textvariable=self.email_summary_var, style="Muted.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))

        source_card, source_body = self._section_card(left, "Source Snapshot")
        source_card.grid(row=1, column=0, sticky="ew", pady=(12, 0))
        self._snapshot_grid(source_body)

        activity_card, activity_body = self._section_card(lower, "Recent Activity")
        activity_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        activity_body.grid_rowconfigure(0, weight=1)
        activity_body.grid_columnconfigure(0, weight=1)
        self.recent_log = ScrolledText(activity_body, height=10, wrap="word", bg=self.colors["log_bg"], fg=self.colors["text"], relief="flat", font=("Consolas", 9), padx=8, pady=8)
        self.recent_log.grid(row=0, column=0, sticky="nsew")
        self.recent_log.configure(state="disabled")

    def _snapshot_grid(self, parent):
        rows = [
            ("Catalog files", self.catalog_count_var),
            ("Order files", self.order_file_count_var),
            ("Order windows", self.order_windows_var),
            ("Brands", self.brand_total_var),
            ("Last refresh", self.last_refresh_var),
            ("Last file activity", self.source_last_activity_var),
        ]
        for idx, (label, var) in enumerate(rows):
            ttk.Label(parent, text=label, style="Muted.TLabel").grid(row=idx, column=0, sticky="w", pady=2)
            tk.Label(parent, textvariable=var, bg=self.colors["surface"], fg=self.colors["text"], font=("Segoe UI", 9, "bold")).grid(row=idx, column=1, sticky="e", padx=(20, 0), pady=2)
        parent.grid_columnconfigure(1, weight=1)

    def _build_brands_screen(self):
        screen = self._create_screen("brands")
        screen.grid_columnconfigure(0, weight=1)
        screen.grid_rowconfigure(2, weight=1)

        toolbar = tk.Frame(screen, bg=self.colors["surface"], highlightbackground=self.colors["border"], highlightthickness=1)
        toolbar.grid(row=0, column=0, sticky="ew")
        toolbar.grid_columnconfigure(1, weight=1)
        ttk.Label(toolbar, text="Search", style="SectionTitle.TLabel").grid(row=0, column=0, padx=(12, 8), pady=12)
        self.brand_search_entry = ttk.Entry(toolbar, textvariable=self.brand_search_var, style="App.TEntry")
        self.brand_search_entry.grid(row=0, column=1, sticky="ew", pady=12)
        ttk.Button(toolbar, text="Clear", style="Ghost.TButton", command=self._clear_brand_search).grid(row=0, column=2, padx=6)
        ttk.Button(toolbar, text="Select Visible", style="Secondary.TButton", command=self.select_all_brands).grid(row=0, column=3, padx=6)
        ttk.Button(toolbar, text="Clear Selected", style="Danger.TButton", command=self.clear_selected_brands).grid(row=0, column=4, padx=(6, 12))

        stats = tk.Frame(screen, bg=self.colors["bg"])
        stats.grid(row=1, column=0, sticky="ew", pady=10)
        stats.grid_columnconfigure(3, weight=1)
        self._metric_card(stats, "Loaded", self.brand_total_var).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self._metric_card(stats, "Visible", self.brand_visible_var).grid(row=0, column=1, sticky="ew", padx=4)
        self._metric_card(stats, "Selected", self.brand_selected_var).grid(row=0, column=2, sticky="ew", padx=4)
        selected_only = ttk.Checkbutton(stats, text="Selected only", variable=self.selected_only_var, style="App.TCheckbutton", command=self.filter_brand_list)
        selected_only.grid(row=0, column=3, sticky="e", padx=(8, 0))

        list_card, list_body = self._section_card(screen)
        list_card.grid(row=2, column=0, sticky="nsew")
        list_body.grid_columnconfigure(0, weight=1)
        list_body.grid_rowconfigure(0, weight=1)
        self.brand_tree = ttk.Treeview(list_body, columns=("selected", "brand"), show="headings", selectmode="browse")
        self.brand_tree.heading("selected", text="")
        self.brand_tree.heading("brand", text="Brand")
        self.brand_tree.column("selected", width=70, minwidth=60, anchor="center", stretch=False)
        self.brand_tree.column("brand", width=500, anchor="w")
        self.brand_tree.grid(row=0, column=0, sticky="nsew")
        scroll = ttk.Scrollbar(list_body, orient="vertical", command=self.brand_tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        self.brand_tree.configure(yscrollcommand=scroll.set)
        self.brand_tree.bind("<ButtonRelease-1>", self._on_brand_tree_click)
        self.brand_tree.bind("<space>", self.toggle_active_brand_selection)
        self.brand_tree.bind("<Return>", self.toggle_active_brand_selection)
        self.brand_tree.bind("<Key>", self.on_listbox_keypress)

        az = tk.Frame(list_body, bg=self.colors["surface"])
        az.grid(row=1, column=0, sticky="ew", pady=(10, 0))
        for idx, letter in enumerate("ABCDEFGHIJKLMNOPQRSTUVWXYZ"):
            ttk.Button(az, text=letter, width=2, style="Ghost.TButton", command=lambda value=letter: self.scroll_to_letter(value)).grid(row=0, column=idx, padx=1)
        ttk.Label(list_body, textvariable=self.brand_summary_var, style="Muted.TLabel").grid(row=2, column=0, sticky="w", pady=(8, 0))

    def _build_activity_screen(self):
        screen = self._create_screen("activity")
        screen.grid_columnconfigure(0, weight=1)
        screen.grid_rowconfigure(1, weight=1)
        cards = tk.Frame(screen, bg=self.colors["bg"])
        cards.grid(row=0, column=0, sticky="ew")
        for col in range(4):
            cards.grid_columnconfigure(col, weight=1)
        self._metric_card(cards, "Last Refresh", self.last_refresh_var).grid(row=0, column=0, sticky="ew", padx=(0, 8))
        self._metric_card(cards, "Last Load", self.last_load_var).grid(row=0, column=1, sticky="ew", padx=4)
        self._metric_card(cards, "Last Email", self.last_email_var).grid(row=0, column=2, sticky="ew", padx=4)
        self._metric_card(cards, "State", self.current_state_var).grid(row=0, column=3, sticky="ew", padx=(8, 0))

        log_card, log_body = self._section_card(screen, "Activity Log")
        log_card.grid(row=1, column=0, sticky="nsew", pady=(12, 0))
        log_body.grid_columnconfigure(0, weight=1)
        log_body.grid_rowconfigure(1, weight=1)
        controls = tk.Frame(log_body, bg=self.colors["surface"])
        controls.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        ttk.Label(controls, text="Filter", style="Muted.TLabel").pack(side="left")
        filter_box = ttk.Combobox(controls, textvariable=self.log_filter_var, values=("All", "Info", "Warning", "Error"), width=10, state="readonly")
        filter_box.pack(side="left", padx=8)
        filter_box.bind("<<ComboboxSelected>>", lambda _event: self._render_activity_log())
        ttk.Button(controls, text="Copy Log", style="Ghost.TButton", command=self.copy_log).pack(side="right", padx=(6, 0))
        ttk.Button(controls, text="Save Log", style="Ghost.TButton", command=self.save_log).pack(side="right", padx=(6, 0))
        ttk.Button(controls, text="Clear Log", style="Danger.TButton", command=self.clear_log).pack(side="right", padx=(6, 0))
        self.log_text = ScrolledText(log_body, wrap="word", bg=self.colors["log_bg"], fg=self.colors["text"], relief="flat", borderwidth=0, font=("Consolas", 9), padx=10, pady=10)
        self.log_text.grid(row=1, column=0, sticky="nsew")
        self.log_text.configure(state="disabled")
        self.log_text.bind("<MouseWheel>", self._update_log_autoscroll)
        self.log_text.bind("<Button-4>", self._update_log_autoscroll)
        self.log_text.bind("<Button-5>", self._update_log_autoscroll)

    def _build_settings_screen(self):
        screen = self._create_screen("settings")
        screen.grid_columnconfigure(0, weight=1)
        canvas = tk.Canvas(screen, bg=self.colors["bg"], highlightthickness=0)
        scroll = ttk.Scrollbar(screen, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        canvas.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")
        screen.grid_rowconfigure(0, weight=1)
        content = tk.Frame(canvas, bg=self.colors["bg"])
        window = canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda _event: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda event: canvas.itemconfigure(window, width=event.width))
        content.grid_columnconfigure(0, weight=1)

        workspace_card, workspace = self._section_card(content, "Workspace")
        workspace_card.grid(row=0, column=0, sticky="ew", pady=(0, 12))
        self._folder_row(workspace, 0, "Input folder", self.input_dir_var, self.browse_input)
        self._folder_row(workspace, 2, "Output folder", self.output_dir_var, self.browse_output)
        ttk.Button(workspace, text="Save Settings", style="Secondary.TButton", command=self.save_settings).grid(row=4, column=0, columnspan=3, sticky="w", pady=(12, 0))

        startup_card, startup = self._section_card(content, "Startup")
        startup_card.grid(row=1, column=0, sticky="ew", pady=(0, 12))
        self._check_row(startup, 0, "Auto-update files on launch", self.auto_update_on_launch_var)
        self._check_row(startup, 1, "Auto-load brands automatically", self.auto_load_brands_after_update_var)
        self._check_row(startup, 2, "Show startup loading screen", self.show_startup_loading_var)

        workflow_card, workflow = self._section_card(content, "Workflow")
        workflow_card.grid(row=2, column=0, sticky="ew", pady=(0, 12))
        self._check_row(workflow, 0, "Prefer Dutchie API", self.prefer_catalog_api_var)
        self._check_row(workflow, 1, "Refresh order reports", self.fetch_order_reports_var)
        self._check_row(workflow, 2, "Include Cost column", self.include_cost_var)
        self._check_row(workflow, 3, "Open output folder after reports complete", self.open_output_after_complete_var)

        email_card, email = self._section_card(content, "Email")
        email_card.grid(row=3, column=0, sticky="ew", pady=(0, 12))
        email.grid_columnconfigure(0, weight=1)
        ttk.Entry(email, textvariable=self.emails_var, style="App.TEntry").grid(row=0, column=0, sticky="ew", padx=(0, 8))
        ttk.Button(email, text="Save", style="Secondary.TButton", command=self.save_settings).grid(row=0, column=1)
        ttk.Label(email, textvariable=self.email_summary_var, style="Muted.TLabel").grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))

        advanced_card, advanced = self._section_card(content, "Advanced")
        advanced_card.grid(row=4, column=0, sticky="ew")
        ttk.Label(advanced, text="Theme", style="Muted.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Combobox(advanced, textvariable=self.theme_var, values=THEME_CHOICES, width=14, state="readonly").grid(row=0, column=1, sticky="w", padx=(10, 0))
        self._check_row(advanced, 1, "Compact mode", self.compact_mode_var)
        ttk.Label(advanced, textvariable=self.settings_state_var, style="Muted.TLabel").grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))

    def _folder_row(self, parent, row, label, var, command):
        parent.grid_columnconfigure(1, weight=1)
        ttk.Label(parent, text=label, style="Muted.TLabel").grid(row=row, column=0, sticky="w", pady=(0, 3))
        ttk.Entry(parent, textvariable=var, style="App.TEntry").grid(row=row + 1, column=0, columnspan=2, sticky="ew", padx=(0, 8))
        ttk.Button(parent, text="Browse", style="Ghost.TButton", command=command).grid(row=row + 1, column=2, sticky="e")

    def _check_row(self, parent, row, label, var):
        ttk.Checkbutton(parent, text=label, variable=var, style="App.TCheckbutton").grid(row=row, column=0, columnspan=2, sticky="w", pady=3)

    def _bind_events(self):
        for var in (
            self.input_dir_var,
            self.output_dir_var,
            self.emails_var,
            self.fetch_order_reports_var,
            self.include_cost_var,
            self.prefer_catalog_api_var,
            self.auto_update_on_launch_var,
            self.auto_load_brands_after_update_var,
            self.show_startup_loading_var,
            self.open_output_after_complete_var,
            self.compact_mode_var,
        ):
            var.trace_add("write", lambda *_: self._schedule_autosave())
        self.theme_var.trace_add("write", self._on_theme_changed)
        self.input_dir_var.trace_add("write", lambda *_: self._refresh_path_summaries())
        self.output_dir_var.trace_add("write", lambda *_: self._refresh_path_summaries())
        self.emails_var.trace_add("write", lambda *_: self._update_email_display())
        self.brand_search_var.trace_add("write", lambda *_: self.filter_brand_list())
        self.master.bind_all("<Control-f>", self.focus_brand_search)
        self.master.bind_all("<Control-u>", self._shortcut_update_files)
        self.master.bind_all("<Control-l>", self._shortcut_load_brands)
        self.master.bind_all("<Control-Return>", self._shortcut_run_process)
        self.master.bind_all("<Alt-1>", lambda event: self._show_screen("dashboard"))
        self.master.bind_all("<Alt-2>", lambda event: self._show_screen("brands"))
        self.master.bind_all("<Alt-3>", lambda event: self._show_screen("activity"))
        self.master.bind_all("<Alt-4>", lambda event: self._show_screen("settings"))
        self.master.bind_all("<Escape>", self._global_escape)

    def _on_theme_changed(self, *_):
        self._apply_theme(self.theme_var.get())
        self._schedule_autosave()
        self.show_toast(f"Theme changed to {normalize_theme_name(self.theme_var.get())}", "info")

    def _apply_theme(self, theme_name):
        theme_name = normalize_theme_name(theme_name)
        if self.theme_var.get() != theme_name:
            self.theme_var.set(theme_name)
            return

        old_colors = dict(self.colors)
        if tb is not None:
            try:
                self.style.theme_use(theme_name)
            except tk.TclError:
                theme_name = DEFAULT_GUI_CONFIG["theme"]
                self.style.theme_use(theme_name)
                self.theme_var.set(theme_name)
                return

        self.colors = theme_palette(theme_name)
        self._configure_styles()
        self._recolor_widget_tree(self.master, old_colors)
        self._set_status(self.status_var.get(), self.status_detail_var.get(), state=self.status_pill_var.get())
        self._render_activity_log()
        self._render_recent_log()

    def _recolor_widget_tree(self, widget, old_colors):
        self._recolor_widget(widget, old_colors)
        for child in widget.winfo_children():
            self._recolor_widget_tree(child, old_colors)

    def _recolor_widget(self, widget, old_colors):
        bg_map = {
            old_colors.get("bg"): self.colors["bg"],
            old_colors.get("surface"): self.colors["surface"],
            old_colors.get("border"): self.colors["border"],
            old_colors.get("accent"): self.colors["accent"],
            old_colors.get("accent_soft"): self.colors["accent_soft"],
            old_colors.get("log_bg"): self.colors["log_bg"],
        }
        fg_map = {
            old_colors.get("text"): self.colors["text"],
            old_colors.get("muted"): self.colors["muted"],
            old_colors.get("accent"): self.colors["accent"],
            old_colors.get("success"): self.colors["success"],
            old_colors.get("warning"): self.colors["warning"],
            old_colors.get("danger"): self.colors["danger"],
        }
        self._replace_widget_color(widget, "background", bg_map)
        self._replace_widget_color(widget, "foreground", fg_map)
        self._replace_widget_color(widget, "activebackground", bg_map)
        self._replace_widget_color(widget, "activeforeground", fg_map)
        self._replace_widget_color(widget, "highlightbackground", bg_map)
        self._replace_widget_color(widget, "highlightcolor", bg_map)
        self._replace_widget_color(widget, "insertbackground", fg_map)

    def _replace_widget_color(self, widget, option, color_map):
        try:
            current = widget.cget(option)
        except tk.TclError:
            return
        replacement = color_map.get(current)
        if replacement is None:
            return
        try:
            widget.configure(**{option: replacement})
        except tk.TclError:
            pass

    def _set_status(self, headline, detail=None, state=None):
        self.status_var.set(headline)
        if detail is not None:
            self.status_detail_var.set(detail)
        self.status_pill_var.set(state or headline)
        self.current_state_var.set(detail or headline)
        color = self.colors["accent"]
        bg = self.colors["accent_soft"]
        if state in ("Error", "Failed") or "fail" in headline.lower() or "error" in headline.lower():
            color = self.colors["danger"]
            bg = "#FEE2E2"
        elif "warning" in headline.lower() or state == "Warning":
            color = self.colors["warning"]
            bg = "#FEF3C7"
        elif "complete" in headline.lower() or "ready" in headline.lower() or state == "Complete":
            color = self.colors["success"]
            bg = "#DCFCE7"
        if hasattr(self, "status_pill"):
            self.status_pill.configure(fg=color, bg=bg)

    def append_log(self, message, level="info"):
        timestamp = datetime.now().strftime("%H:%M:%S")
        level = self._normalize_log_level(message, level)
        self.log_entries.append((timestamp, level, str(message)))
        if len(self.log_entries) > 1000:
            self.log_entries = self.log_entries[-1000:]
        self.recent_events.append((timestamp, level, str(message)))
        self.recent_events = self.recent_events[-5:]
        self._render_activity_log()
        self._render_recent_log()

    def _normalize_log_level(self, message, level):
        text = str(message).lower()
        if level != "info":
            return level
        if "[error]" in text or "failed" in text or "error" in text:
            return "error"
        if "[warn]" in text or "warning" in text or "missing" in text:
            return "warning"
        return "info"

    def _render_activity_log(self):
        if not hasattr(self, "log_text"):
            return
        selected_filter = self.log_filter_var.get().lower()
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        for ts, level, message in self.log_entries:
            if selected_filter != "all" and level != selected_filter:
                continue
            self.log_text.insert("end", f"[{ts}] {level.upper():7} {message}\n")
        if self.log_auto_scroll:
            self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _update_log_autoscroll(self, _event=None):
        self.master.after(50, self._sync_log_autoscroll)

    def _sync_log_autoscroll(self):
        if hasattr(self, "log_text"):
            self.log_auto_scroll = self.log_text.yview()[1] >= 0.98

    def _render_recent_log(self):
        if not hasattr(self, "recent_log"):
            return
        self.recent_log.configure(state="normal")
        self.recent_log.delete("1.0", "end")
        for ts, _level, message in self.recent_events[-5:]:
            self.recent_log.insert("end", f"{ts}  {message}\n")
        self.recent_log.see("end")
        self.recent_log.configure(state="disabled")

    def _refresh_path_summaries(self):
        self.input_summary_var.set(self.input_dir_var.get().strip() or "No input folder")
        self.output_summary_var.set(self.output_dir_var.get().strip() or "No output folder")
        self._refresh_source_snapshot()

    def _refresh_source_snapshot(self):
        input_dir = self.input_dir_var.get().strip()
        if not input_dir or not os.path.isdir(input_dir):
            self.catalog_count_var.set("0")
            self.order_file_count_var.set("0")
            self.order_windows_var.set("None")
            self.source_last_activity_var.set("No files")
            return
        csv_files = list_catalog_csv_files(input_dir)
        order_files = list_order_report_files(input_dir)
        self.catalog_count_var.set(str(len(csv_files)))
        self.order_file_count_var.set(str(len(order_files)))
        order_windows = summarize_order_report_files(input_dir)
        self.order_windows_var.set(order_windows or "None")
        latest_paths = [
            os.path.join(input_dir, fn)
            for fn in csv_files + order_files
            if os.path.isfile(os.path.join(input_dir, fn))
        ]
        if latest_paths:
            latest_mtime = max(os.path.getmtime(path) for path in latest_paths)
            self.source_last_activity_var.set(datetime.fromtimestamp(latest_mtime).strftime("%b %d, %Y %I:%M %p").lstrip("0"))
        else:
            self.source_last_activity_var.set("No files")
        api_ready, available, missing, error = dutchie_api_readiness(DEFAULT_API_ENV_FILE)
        if api_ready:
            self.api_mode_var.set(f"API ready: {', '.join(available)}")
        elif error:
            self.api_mode_var.set("API config error")
        else:
            self.api_mode_var.set(f"API missing: {', '.join(missing)}")

    def _parse_recipients(self):
        return [item.strip() for item in self.emails_var.get().split(",") if item.strip()]

    def _invalid_recipients(self):
        return [email for email in self._parse_recipients() if not EMAIL_REGEX.match(email)]

    def _update_email_display(self):
        recipients = self._parse_recipients()
        invalid = self._invalid_recipients()
        self.recipient_count_var.set(str(len(recipients)))
        if not recipients:
            self.email_summary_var.set("No recipients saved.")
        elif invalid:
            self.email_summary_var.set(f"{len(invalid)} invalid recipient(s): {', '.join(invalid)}")
        else:
            self.email_summary_var.set(", ".join(recipients))

    def _current_selected_brands(self):
        return set(self.selected_brand_names)

    def _apply_brand_result(self, result):
        self.all_brands = result.get("brands", [])
        self.selected_brand_names.intersection_update(self.all_brands)
        self.filter_brand_list()
        self.last_load_var.set(datetime.now().strftime("%b %d, %Y %I:%M %p").lstrip("0"))
        count = len(self.all_brands)
        csv_count = result.get("csv_count", 0)
        self.append_log(f"Loaded {count} brands from {csv_count} catalog CSV file(s).")
        self._set_status("Brands loaded", f"{count} brands available", state="Complete")
        self.show_toast(f"{count} brands loaded", "success")

    def _populate_brand_listbox(self, items):
        if not hasattr(self, "brand_tree"):
            return
        self.brand_tree.delete(*self.brand_tree.get_children())
        self.filtered_brands = list(items)
        for brand in self.filtered_brands:
            marker = "[x]" if brand in self.selected_brand_names else "[ ]"
            self.brand_tree.insert("", "end", iid=brand, values=(marker, brand))
        self._update_brand_summary()

    def filter_brand_list(self):
        query = self.brand_search_var.get().strip().lower()
        items = list(self.all_brands)
        if query:
            items = [brand for brand in items if query in brand.lower()]
        if self.selected_only_var.get():
            items = [brand for brand in items if brand in self.selected_brand_names]
        self._populate_brand_listbox(items)

    def _update_brand_summary(self):
        total = len(self.all_brands)
        visible = len(self.filtered_brands)
        selected = len(self.selected_brand_names)
        self.brand_total_var.set(str(total))
        self.brand_visible_var.set(str(visible))
        self.brand_selected_var.set(str(selected))
        self.brand_summary_var.set(f"{selected} selected | {visible} visible | {total} loaded")

    def _on_brand_tree_click(self, event=None):
        row_id = self.brand_tree.identify_row(event.y) if event else self.brand_tree.focus()
        if not row_id:
            return
        self._toggle_brand(row_id)

    def _toggle_brand(self, brand):
        if brand in self.selected_brand_names:
            self.selected_brand_names.remove(brand)
        else:
            self.selected_brand_names.add(brand)
        if self.brand_tree.exists(brand):
            marker = "[x]" if brand in self.selected_brand_names else "[ ]"
            self.brand_tree.set(brand, "selected", marker)
        self._update_brand_summary()

    def focus_brand_search(self, _event=None):
        self._show_screen("brands")
        self.brand_search_entry.focus_set()
        self.brand_search_entry.selection_range(0, tk.END)
        return "break"

    def _clear_brand_search(self, _event=None):
        self.brand_search_var.set("")
        self.selected_only_var.set(False)
        self.filter_brand_list()
        return "break"

    def scroll_to_letter(self, letter):
        prefix = str(letter).lower()
        for brand in self.filtered_brands:
            if brand.lower().startswith(prefix):
                self.brand_tree.selection_set(brand)
                self.brand_tree.focus(brand)
                self.brand_tree.see(brand)
                break

    def on_listbox_keypress(self, event):
        if not event.char or not event.char.isprintable():
            return
        self.scroll_to_letter(event.char)

    def toggle_active_brand_selection(self, _event=None):
        focused = self.brand_tree.focus()
        if focused:
            self._toggle_brand(focused)
        return "break"

    def select_all_brands(self):
        self.selected_brand_names.update(self.filtered_brands)
        self.filter_brand_list()
        self.show_toast(f"Selected {len(self.filtered_brands)} visible brands", "success")

    def clear_selected_brands(self):
        self.selected_brand_names.clear()
        self.filter_brand_list()
        self.show_toast("Brand selection cleared", "warning")

    def _schedule_autosave(self):
        if self.autosave_job is not None:
            self.master.after_cancel(self.autosave_job)
        self.settings_state_var.set("Saving...")
        self.autosave_job = self.master.after(700, self._autosave_settings)

    def _autosave_settings(self):
        self.autosave_job = None
        try:
            self._persist_settings(add_log=False)
        except Exception as exc:
            self.settings_state_var.set(f"Autosave failed: {exc}")

    def _persist_settings(self, add_log=False):
        save_config(
            self.input_dir_var.get().strip(),
            self.output_dir_var.get().strip(),
            fetch_order_reports=self.fetch_order_reports_var.get(),
            emails=self.emails_var.get().strip(),
            include_cost=self.include_cost_var.get(),
            prefer_catalog_api=self.prefer_catalog_api_var.get(),
            auto_update_on_launch=self.auto_update_on_launch_var.get(),
            auto_load_brands_after_update=self.auto_load_brands_after_update_var.get(),
            auto_load_brands_on_launch=self.auto_load_brands_after_update_var.get(),
            show_startup_loading=self.show_startup_loading_var.get(),
            open_output_after_complete=self.open_output_after_complete_var.get(),
            theme=self.theme_var.get(),
            compact_mode=self.compact_mode_var.get(),
            task_eta_seconds=self.task_eta_seconds,
        )
        self.settings_state_var.set("Settings saved.")
        if add_log:
            self.append_log("Saved workspace settings.")

    def save_settings(self, quiet=False):
        if self.autosave_job is not None:
            self.master.after_cancel(self.autosave_job)
            self.autosave_job = None
        self._persist_settings(add_log=True)
        if not quiet:
            self.show_toast("Settings saved", "success")

    def browse_input(self):
        folder = filedialog.askdirectory()
        if folder:
            self.input_dir_var.set(folder)
            self.append_log(f"Selected input folder: {folder}")
            if self.auto_load_brands_after_update_var.get() and list_catalog_csv_files(folder):
                self.append_log("Catalog CSVs found in selected folder. Auto-loading brands.")
                self.load_brands(silent=True)

    def browse_output(self):
        folder = filedialog.askdirectory()
        if folder:
            self.output_dir_var.set(folder)
            self.append_log(f"Selected output folder: {folder}")

    def _shortcut_update_files(self, _event=None):
        self.get_files()
        return "break"

    def _shortcut_load_brands(self, _event=None):
        self.load_brands()
        return "break"

    def _shortcut_run_process(self, _event=None):
        self.run_process()
        return "break"

    def _global_escape(self, _event=None):
        if self.brand_search_var.get():
            return self._clear_brand_search()
        if self.loading_window and self.loading_window.winfo_exists():
            return "break"
        return None

    def _resolve_script_path(self, script_name):
        return Path(__file__).resolve().with_name(script_name)

    def _run_script_worker(self, ctx, script_name, *args):
        script_path = self._resolve_script_path(script_name)
        if not script_path.exists():
            raise FileNotFoundError(f"{script_name} was not found at {script_path}.")
        cmd = [sys.executable, str(script_path), *[str(arg) for arg in args]]
        started = time.perf_counter()
        ctx.log(f"Running: {' '.join(str(part) for part in cmd)}")
        ctx.status(step=f"Started {script_name}")
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, bufsize=1)
        output_lines = []
        if process.stdout is not None:
            for raw_line in process.stdout:
                line = raw_line.rstrip()
                output_lines.append(raw_line)
                if line:
                    level = "warning" if line.startswith("[WARN]") else "error" if line.startswith("[ERROR]") else "info"
                    ctx.log(line, level=level)
                    if line.startswith(("[FETCH]", "[SAVED]", "[INFO]", "[VERIFY]", "[WARN]")):
                        ctx.status(step=line)
        return_code = process.wait()
        output = "".join(output_lines)
        elapsed = time.perf_counter() - started
        if return_code != 0:
            ctx.log(f"{script_name} exited with code {return_code} after {elapsed:.1f}s.", level="error")
            raise subprocess.CalledProcessError(return_code, cmd, output=output, stderr=output)
        ctx.log(f"{script_name} finished in {elapsed:.1f}s.")
        return output

    def _refresh_catalog_exports_worker(self, ctx, input_dir, prefer_api):
        api_ready, available, missing, error_text = dutchie_api_readiness(DEFAULT_API_ENV_FILE)
        ctx.log(
            f"Catalog readiness: prefer_api={prefer_api}, api_ready={api_ready}, "
            f"available={', '.join(available) if available else 'none'}, missing={', '.join(missing) if missing else 'none'}."
        )
        if prefer_api and api_ready:
            try:
                ctx.status(headline="Refreshing catalog files", step=f"Dutchie API with {DUTCHIE_API_WORKERS} workers")
                self._run_script_worker(ctx, CATALOG_API_SCRIPT, input_dir, "--workers", DUTCHIE_API_WORKERS)
                return "api"
            except subprocess.CalledProcessError:
                ctx.log("Catalog API refresh failed; falling back to browser exporter.", level="warning")
        elif prefer_api:
            reason = error_text or f"missing store credentials: {', '.join(missing)}"
            ctx.log(f"Catalog API unavailable ({reason}); using browser exporter.", level="warning")
        else:
            ctx.log("Catalog refresh is set to browser mode.")
        ctx.status(headline="Refreshing catalog files", step="Browser catalog exporter")
        self._run_script_worker(ctx, CATALOG_BROWSER_SCRIPT, input_dir)
        return "browser"

    def _refresh_order_reports_worker(self, ctx, input_dir, prefer_api):
        api_ready, available, missing, error_text = dutchie_api_readiness(DEFAULT_API_ENV_FILE)
        ctx.log(
            f"Order-report readiness: prefer_api={prefer_api}, api_ready={api_ready}, "
            f"available={', '.join(available) if available else 'none'}, missing={', '.join(missing) if missing else 'none'}."
        )
        if prefer_api and api_ready:
            try:
                ctx.status(headline="Refreshing order reports", step=f"Dutchie API with {DUTCHIE_API_WORKERS} workers")
                self._run_script_worker(ctx, ORDER_REPORT_API_SCRIPT, input_dir, "--workers", DUTCHIE_API_WORKERS)
                return "api"
            except subprocess.CalledProcessError:
                ctx.log("Order-report API refresh failed; falling back to browser exporter.", level="warning")
        elif prefer_api:
            reason = error_text or f"missing store credentials: {', '.join(missing)}"
            ctx.log(f"Order-report API unavailable ({reason}); using browser exporter.", level="warning")
        else:
            ctx.log("Order-report refresh is set to browser mode.")
        ctx.status(headline="Refreshing order reports", step="Browser order-report exporter")
        self._run_script_worker(ctx, ORDER_REPORT_BROWSER_SCRIPT, input_dir)
        return "browser"

    def _scan_brand_library_worker(self, ctx, input_dir):
        if not input_dir or not os.path.isdir(input_dir):
            raise FileNotFoundError("Input folder is missing.")
        brand_set = set()
        csv_count = 0
        errors = []
        ctx.status(headline="Loading brands", step="Reading catalog CSV files")
        for filename in list_catalog_csv_files(input_dir):
            csv_count += 1
            path = os.path.join(input_dir, filename)
            try:
                df = pd.read_csv(path, nrows=50000)
                if "Brand" in df.columns:
                    brand_set.update(
                        df["Brand"].dropna().astype(str).str.strip().str.lower().replace("", pd.NA).dropna().unique().tolist()
                    )
                ctx.status(step=f"Scanned {filename}")
            except Exception as exc:
                errors.append(f"{filename}: {exc}")
                ctx.log(f"Could not scan {filename}: {exc}", level="warning")
        return {"brands": sorted(brand_set), "csv_count": csv_count, "errors": errors}

    def _refresh_sources_worker(self, ctx, input_dir, fetch_order_reports, prefer_api, load_after):
        ctx.status(headline="Preparing source refresh", detail="Clearing old exports", step="Checking source folder")
        before_catalog = len(list_catalog_csv_files(input_dir))
        before_order = len(list_order_report_files(input_dir))
        deleted = clear_old_input_exports(input_dir, clear_order_reports=fetch_order_reports)
        deleted_catalog = sum(1 for path in deleted if str(path).lower().endswith(".csv") and not is_order_report_filename(os.path.basename(path)))
        deleted_order = sum(1 for path in deleted if is_order_report_filename(os.path.basename(path)))
        ctx.log(f"Before refresh: {before_catalog} catalog CSV(s), {before_order} order-report file(s).")
        ctx.log(f"Cleared {deleted_catalog} catalog CSV(s), {deleted_order} order-report file(s).")
        catalog_mode = self._refresh_catalog_exports_worker(ctx, input_dir, prefer_api)
        order_mode = None
        order_error = None
        if fetch_order_reports:
            try:
                order_mode = self._refresh_order_reports_worker(ctx, input_dir, prefer_api)
            except Exception as exc:
                order_error = str(exc)
                ctx.log(f"Order-report refresh failed: {exc}", level="error")
        else:
            ctx.log("Order-report refresh skipped by setting.")
        result = {
            "catalog_mode": catalog_mode,
            "order_mode": order_mode,
            "order_error": order_error,
            "catalog_count": len(list_catalog_csv_files(input_dir)),
            "order_count": len(list_order_report_files(input_dir)),
        }
        if load_after:
            result["brands_result"] = self._scan_brand_library_worker(ctx, input_dir)
        return result

    def _generate_reports_worker(self, ctx, input_dir, output_dir, selected_brands, include_cost, emails):
        all_brand_map = {}
        catalog_files = list_catalog_csv_files(input_dir)
        if not catalog_files:
            raise RuntimeError("No catalog CSV files were found.")
        for filename in catalog_files:
            ctx.status(headline="Generating reports", step=f"Processing {filename}")
            brand_map = generate_brand_reports(
                os.path.join(input_dir, filename),
                output_dir,
                selected_brands,
                include_cost=include_cost,
                order_reports_dir=input_dir,
            )
            for brand, files in brand_map.items():
                all_brand_map.setdefault(brand, []).extend(files)
        if not all_brand_map:
            raise RuntimeError("No matching workbooks were generated.")
        workbook_count = sum(len(files) for files in all_brand_map.values())
        ctx.status(headline="Uploading reports", step=f"Uploading {workbook_count} workbook(s)")
        brand_links = upload_brand_reports_to_drive(all_brand_map)
        if not brand_links:
            raise RuntimeError("No Drive folders were created.")
        lines = []
        for brand_lower, link in brand_links.items():
            lines.append(f"<h3>{brand_lower}</h3>")
            lines.append(f"<p><a href='{link}'>{link}</a></p>")
        joined_links = "\n".join(lines)
        order_summary = summarize_order_report_files(input_dir)
        order_note = ""
        if order_summary:
            order_note = (
                "<p>Matching Dutchie order-report rows were added to the "
                f"<strong>Order</strong> tab when available. Source windows found: {order_summary}.</p>"
            )
        body_html = f"""
        <html>
          <body>
            <p>Hello,</p>
            <p>Here are the public Drive folders for each brand:</p>
            {order_note}
            {joined_links}
            <p>Anyone with these links can download the XLSX files.</p>
            <p>Regards,<br>Brand Inventory Bot</p>
          </body>
        </html>
        """
        ctx.status(headline="Sending email", step=f"Emailing {len([item for item in emails.split(',') if item.strip()])} recipient(s)")
        send_email_with_gmail_html("Brand Inventory Drive Links", body_html, emails)
        recipient_count = len([item for item in emails.split(",") if item.strip()])
        return {"brand_count": len(brand_links), "workbook_count": workbook_count, "recipients": recipient_count, "output_dir": output_dir}

    def _eta_key_for_task(self, task_name):
        return self.TASK_ETA_ALIASES.get(task_name, task_name)

    def _estimate_for_task(self, task_name):
        estimate = self.task_eta_seconds.get(self._eta_key_for_task(task_name))
        try:
            estimate = float(estimate)
        except (TypeError, ValueError):
            return None
        return estimate if estimate > 0 else None

    def _format_duration(self, seconds):
        seconds = max(0, int(round(float(seconds))))
        minutes, seconds = divmod(seconds, 60)
        hours, minutes = divmod(minutes, 60)
        if hours:
            return f"{hours}h {minutes:02d}m"
        if minutes:
            return f"{minutes}m {seconds:02d}s"
        return f"{seconds}s"

    def _loading_eta_text(self):
        if self.active_task_started_at is None:
            return "ETA: starting"
        elapsed = time.perf_counter() - self.active_task_started_at
        if self.active_task_estimate_seconds:
            remaining = max(0, self.active_task_estimate_seconds - elapsed)
            if remaining <= 1:
                return f"ETA: finishing now | elapsed {self._format_duration(elapsed)}"
            return (
                f"ETA: {self._format_duration(remaining)} remaining | "
                f"elapsed {self._format_duration(elapsed)} | based on prior runs"
            )
        return f"Elapsed: {self._format_duration(elapsed)} | ETA appears after one successful run"

    def _refresh_loading_eta(self):
        if hasattr(self, "loading_eta_var"):
            self.loading_eta_var.set(self._loading_eta_text())
        if self.active_task_started_at is None or not self.active_task_estimate_seconds:
            return
        if hasattr(self, "loading_progress"):
            elapsed = time.perf_counter() - self.active_task_started_at
            progress_value = min(98, max(4, (elapsed / self.active_task_estimate_seconds) * 100))
            try:
                self.loading_progress.configure(value=progress_value)
            except Exception:
                pass

    def _start_loading_eta_timer(self):
        self._cancel_loading_eta_timer()
        self._refresh_loading_eta()
        self.loading_eta_job = self.master.after(1000, self._tick_loading_eta)

    def _tick_loading_eta(self):
        self.loading_eta_job = None
        if not self.task_running:
            return
        self._refresh_loading_eta()
        self.loading_eta_job = self.master.after(1000, self._tick_loading_eta)

    def _cancel_loading_eta_timer(self):
        if self.loading_eta_job is not None:
            try:
                self.master.after_cancel(self.loading_eta_job)
            except Exception:
                pass
            self.loading_eta_job = None

    def _record_task_duration(self, task_name, elapsed):
        if elapsed <= 0:
            return
        eta_key = self._eta_key_for_task(task_name)
        previous = self._estimate_for_task(task_name)
        blended = elapsed if previous is None else (previous * 0.65) + (elapsed * 0.35)
        self.task_eta_seconds[eta_key] = round(blended, 1)
        self.append_log(
            f"ETA baseline for {eta_key}: {self._format_duration(blended)} "
            f"(latest run {self._format_duration(elapsed)})."
        )
        try:
            self._persist_settings(add_log=False)
        except Exception as exc:
            self.append_log(f"Could not save ETA history: {exc}", level="warning")

    def run_background_task(self, task_name, worker_fn, on_complete=None, on_error=None, loading_title=None, loading_detail=None):
        if self.task_running:
            self.show_toast("Another task is already running.", "warning")
            return False
        self.task_running = True
        self.active_task_name = task_name
        self.active_task_id = f"{task_name}-{time.time()}"
        self.active_task_eta_key = self._eta_key_for_task(task_name)
        self.active_task_started_at = time.perf_counter()
        self.active_task_estimate_seconds = self._estimate_for_task(task_name)
        self._set_busy(True)
        self.show_loading(loading_title or task_name, loading_detail or "Working...", "Starting")
        self._start_loading_eta_timer()
        ctx = TaskContext(self.task_queue, self.active_task_id)

        def target():
            try:
                result = worker_fn(ctx)
                ctx.emit("success", result=result, on_complete=on_complete)
            except Exception as exc:
                ctx.emit("error", error=str(exc), traceback=traceback.format_exc(), on_error=on_error)

        threading.Thread(target=target, daemon=True).start()
        self.master.after(80, self._poll_task_queue)
        return True

    def _poll_task_queue(self):
        keep_polling = self.task_running
        while True:
            try:
                message = self.task_queue.get_nowait()
            except queue.Empty:
                break
            kind = message.get("kind")
            if message.get("task_id") != self.active_task_id:
                continue
            if kind == "log":
                self.append_log(message.get("message", ""), message.get("level", "info"))
            elif kind == "status":
                headline = message.get("headline")
                detail = message.get("detail")
                step = message.get("step")
                if headline:
                    self._set_status(headline, detail or step or headline, state="Running")
                self.update_loading(message=headline, detail=detail, step=step)
            elif kind == "success":
                self._finish_task()
                callback = message.get("on_complete")
                if callback:
                    callback(message.get("result"))
                keep_polling = False
            elif kind == "error":
                self._finish_task(error=True)
                self.append_log(message.get("traceback", message.get("error", "")), level="error")
                callback = message.get("on_error")
                if callback:
                    callback(message.get("error"), message.get("traceback"))
                else:
                    self._show_error("Task failed", message.get("error", "Unknown error"))
                keep_polling = False
        if keep_polling:
            self.master.after(80, self._poll_task_queue)

    def _finish_task(self, error=False):
        task_name = self.active_task_name
        started_at = self.active_task_started_at
        elapsed = (time.perf_counter() - started_at) if started_at is not None else None
        self.task_running = False
        self._cancel_loading_eta_timer()
        if not error and elapsed is not None:
            if hasattr(self, "loading_progress") and self.active_task_estimate_seconds:
                try:
                    self.loading_progress.configure(value=100)
                except Exception:
                    pass
            self._record_task_duration(task_name, elapsed)
        self.active_task_id = None
        self.active_task_name = ""
        self.active_task_eta_key = ""
        self.active_task_started_at = None
        self.active_task_estimate_seconds = None
        self._set_busy(False)
        self.hide_loading()
        self._refresh_source_snapshot()
        if not error:
            self._set_status("Ready", "Idle", state="Ready")

    def _set_busy(self, busy):
        state = "disabled" if busy else "normal"
        for button in getattr(self, "workflow_buttons", []):
            try:
                button.configure(state=state)
            except Exception:
                pass
        self.current_state_var.set("Running" if busy else "Idle")

    def show_loading(self, title="Working", detail="Please wait...", step="Starting"):
        if not self.show_startup_loading_var.get() and title == "Starting app":
            return
        if self.loading_window and self.loading_window.winfo_exists():
            self.hide_loading()
        self.loading_step_var = tk.StringVar(value=step)
        self.loading_title_var = tk.StringVar(value=title)
        self.loading_detail_var = tk.StringVar(value=detail)
        eta_text = self._loading_eta_text() if self.task_running else "ETA: checking workspace"
        self.loading_eta_var = tk.StringVar(value=eta_text)
        self.loading_window = tk.Toplevel(self.master)
        self.loading_window.title(title)
        self.loading_window.transient(self.master)
        self.loading_window.configure(bg=self.colors["border"])
        self.loading_window.resizable(False, False)
        self.loading_window.protocol("WM_DELETE_WINDOW", lambda: None)
        panel = tk.Frame(self.loading_window, bg=self.colors["surface"], padx=22, pady=18)
        panel.pack(fill="both", expand=True, padx=1, pady=1)
        ttk.Label(panel, text="Buzz Brand Inventory Studio", style="SectionTitle.TLabel").pack(anchor="w")
        ttk.Label(panel, textvariable=self.loading_title_var, style="Muted.TLabel").pack(anchor="w", pady=(3, 12))
        progress_mode = "determinate" if self.active_task_estimate_seconds else "indeterminate"
        self.loading_progress = ttk.Progressbar(
            panel,
            mode=progress_mode,
            maximum=100,
            length=420,
            style="Loading.Horizontal.TProgressbar",
        )
        self.loading_progress.pack(fill="x")
        if self.active_task_estimate_seconds:
            self.loading_progress.configure(value=4)
        else:
            self.loading_progress.start(12)
        tk.Label(panel, textvariable=self.loading_step_var, bg=self.colors["surface"], fg=self.colors["text"], font=("Segoe UI", 9, "bold"), wraplength=420, justify="left").pack(anchor="w", pady=(12, 3))
        ttk.Label(panel, textvariable=self.loading_detail_var, style="Muted.TLabel").pack(anchor="w")
        tk.Label(
            panel,
            textvariable=self.loading_eta_var,
            bg=self.colors["surface"],
            fg=self.colors["muted"],
            font=("Segoe UI", 8),
        ).pack(anchor="w", pady=(6, 0))
        ttk.Button(panel, text="Skip/Close", style="Ghost.TButton", command=self.hide_loading).pack(anchor="e", pady=(14, 0))
        self.loading_window.update_idletasks()
        width, height = 500, 234
        x = self.master.winfo_rootx() + (self.master.winfo_width() // 2) - width // 2
        y = self.master.winfo_rooty() + (self.master.winfo_height() // 2) - height // 2
        self.loading_window.geometry(f"{width}x{height}+{max(30, x)}+{max(30, y)}")

    def update_loading(self, message=None, detail=None, step=None):
        if not self.loading_window or not self.loading_window.winfo_exists():
            return
        if message and hasattr(self, "loading_title_var"):
            self.loading_title_var.set(message)
            self.loading_window.title(message)
        if detail and hasattr(self, "loading_detail_var"):
            self.loading_detail_var.set(detail)
        if step and hasattr(self, "loading_step_var"):
            self.loading_step_var.set(step)
        self.loading_window.update_idletasks()

    def hide_loading(self):
        if hasattr(self, "loading_progress"):
            try:
                self.loading_progress.stop()
            except Exception:
                pass
        if self.loading_window and self.loading_window.winfo_exists():
            self.loading_window.destroy()
        self.loading_window = None

    def show_toast(self, message, level="info"):
        if self.toast_window and self.toast_window.winfo_exists():
            self.toast_window.destroy()
        colors = {
            "success": ("#DCFCE7", self.colors["success"]),
            "warning": ("#FEF3C7", self.colors["warning"]),
            "error": ("#FEE2E2", self.colors["danger"]),
            "info": (self.colors["accent_soft"], self.colors["accent"]),
        }
        bg, fg = colors.get(level, colors["info"])
        self.toast_window = tk.Toplevel(self.master)
        self.toast_window.overrideredirect(True)
        self.toast_window.configure(bg=bg)
        tk.Label(self.toast_window, text=message, bg=bg, fg=fg, padx=14, pady=8, font=("Segoe UI", 9, "bold")).pack()
        self.master.update_idletasks()
        width = 280
        x = self.master.winfo_rootx() + self.master.winfo_width() - width - 28
        y = self.master.winfo_rooty() + self.master.winfo_height() - 92
        self.toast_window.geometry(f"{width}x42+{max(20, x)}+{max(20, y)}")
        self.master.after(3200, lambda: self.toast_window.destroy() if self.toast_window and self.toast_window.winfo_exists() else None)

    def _startup_sequence(self):
        self.show_loading("Starting app", "Reading settings and checking workspace.", "Checking workspace")
        input_dir = self.input_dir_var.get().strip()
        has_input = bool(input_dir and os.path.isdir(input_dir))
        if not has_input:
            self.append_log("Startup auto-update skipped: input folder is missing.", level="warning")
            self.hide_loading()
            self.show_toast("Input folder missing. Startup update skipped.", "warning")
            return
        if self.auto_update_on_launch_var.get():
            self.run_background_task(
                "startup-refresh",
                lambda ctx: self._refresh_sources_worker(
                    ctx,
                    input_dir,
                    self.fetch_order_reports_var.get(),
                    self.prefer_catalog_api_var.get(),
                    self.auto_load_brands_after_update_var.get(),
                ),
                on_complete=self._on_refresh_complete,
                on_error=lambda error, tb_text: self._show_error("Startup refresh failed", error),
                loading_title="Starting app",
                loading_detail="Refreshing files before the dashboard opens.",
            )
        elif self.auto_load_brands_after_update_var.get() and list_catalog_csv_files(input_dir):
            self.run_background_task(
                "startup-load-brands",
                lambda ctx: self._scan_brand_library_worker(ctx, input_dir),
                on_complete=self._on_load_brands_complete,
                on_error=lambda error, tb_text: self._show_error("Startup brand load failed", error),
                loading_title="Starting app",
                loading_detail="Loading saved brand library.",
            )
        else:
            self.hide_loading()
            self._set_status("Ready", "Startup automation is off.", state="Ready")

    def get_files(self):
        input_dir = self.input_dir_var.get().strip()
        if not input_dir or not os.path.isdir(input_dir):
            self.show_toast("Input folder is missing.", "error")
            self._show_error("Input folder is missing", "Choose a valid input folder in Settings.")
            return
        self.append_log("Starting source refresh.")
        self.run_background_task(
            "update-files",
            lambda ctx: self._refresh_sources_worker(
                ctx,
                input_dir,
                self.fetch_order_reports_var.get(),
                self.prefer_catalog_api_var.get(),
                self.auto_load_brands_after_update_var.get(),
            ),
            on_complete=self._on_refresh_complete,
            on_error=lambda error, tb_text: self._show_error("Source refresh failed", error),
            loading_title="Updating files",
            loading_detail="Refreshing Dutchie catalog and order-report files.",
        )

    def _on_refresh_complete(self, result):
        self._refresh_source_snapshot()
        self.last_refresh_var.set(datetime.now().strftime("%b %d, %Y %I:%M %p").lstrip("0"))
        if result.get("brands_result"):
            self._apply_brand_result(result["brands_result"])
        catalog_mode = "Dutchie API" if result.get("catalog_mode") == "api" else "browser"
        order_mode = result.get("order_mode")
        if result.get("order_error"):
            self._set_status("Partial refresh complete", "Catalog updated, order reports failed.", state="Warning")
            self.show_toast("Catalog updated. Order reports failed.", "warning")
        else:
            detail = f"Catalog via {catalog_mode}"
            if order_mode:
                detail += f", orders via {'Dutchie API' if order_mode == 'api' else 'browser'}"
            self._set_status("Files updated", detail, state="Complete")
            self.show_toast("Files updated", "success")
        self.append_log(
            f"Refresh complete: {result.get('catalog_count', 0)} catalog CSV(s), "
            f"{result.get('order_count', 0)} order-report file(s)."
        )

    def load_brands(self, silent=False):
        input_dir = self.input_dir_var.get().strip()
        if not input_dir or not os.path.isdir(input_dir):
            self.show_toast("Input folder is missing.", "error")
            if not silent:
                self._show_error("Input folder is missing", "Choose a valid input folder in Settings.")
            return False
        return self.run_background_task(
            "load-brands",
            lambda ctx: self._scan_brand_library_worker(ctx, input_dir),
            on_complete=self._on_load_brands_complete,
            on_error=lambda error, tb_text: self._show_error("Brand load failed", error),
            loading_title="Loading brands",
            loading_detail="Scanning catalog CSV files.",
        )

    def _on_load_brands_complete(self, result):
        if not result.get("brands"):
            self._set_status("No brands found", "Refresh files or check catalog CSVs.", state="Warning")
            self.show_toast("No brands found", "warning")
            self.append_log("No brands were found in the current catalog CSV files.", level="warning")
            return
        self._apply_brand_result(result)

    def run_process(self):
        input_dir = self.input_dir_var.get().strip()
        output_dir = self.output_dir_var.get().strip()
        emails = self.emails_var.get().strip()
        invalid = self._invalid_recipients()
        if not input_dir or not os.path.isdir(input_dir):
            self._show_error("Input folder is missing", "Choose a valid input folder in Settings.")
            return
        if not output_dir or not os.path.isdir(output_dir):
            self._show_error("Output folder is missing", "Choose a valid output folder in Settings.")
            return
        if not list_catalog_csv_files(input_dir):
            self._show_error("No catalog CSV files found", "Use Update Files before generating reports.")
            return
        if not emails or invalid:
            detail = "No valid recipient emails found." if not emails else "Invalid recipient email(s): " + ", ".join(invalid)
            self._show_error("Recipient email problem", detail)
            return
        selected_brands = sorted(self._current_selected_brands())
        target_text = f"{len(selected_brands)} selected brand(s)" if selected_brands else "all brands"
        recipients = len(self._parse_recipients())
        if not messagebox.askyesno(
            "Generate reports",
            f"Generate reports for {target_text} and email Drive links to {recipients} recipient(s)?",
        ):
            return
        self.append_log("Starting report generation workflow.")
        self.run_background_task(
            "generate-reports",
            lambda ctx: self._generate_reports_worker(
                ctx,
                input_dir,
                output_dir,
                selected_brands,
                self.include_cost_var.get(),
                emails,
            ),
            on_complete=self._on_generate_complete,
            on_error=lambda error, tb_text: self._show_error("Report workflow failed", error),
            loading_title="Generating reports",
            loading_detail="Creating workbooks, uploading to Drive, then sending email.",
        )

    def _on_generate_complete(self, result):
        self.last_email_var.set(datetime.now().strftime("%b %d, %Y %I:%M %p").lstrip("0"))
        self._set_status("Reports delivered", "Drive folders uploaded and email sent.", state="Complete")
        self.append_log(
            f"Finished report workflow: {result.get('workbook_count', 0)} workbook(s), "
            f"{result.get('brand_count', 0)} brand folder(s), {result.get('recipients', 0)} recipient(s)."
        )
        self.show_toast("Reports uploaded and email sent", "success")
        if self.open_output_after_complete_var.get():
            self._open_output_folder(result.get("output_dir"))
        messagebox.showinfo(
            "Reports delivered",
            f"Generated {result.get('workbook_count', 0)} workbook(s) for {result.get('brand_count', 0)} brand folder(s).\n\nEmail sent to {result.get('recipients', 0)} recipient(s).",
        )

    def _show_error(self, title, detail):
        self._set_status(title, detail, state="Error")
        self.show_toast(title, "error")
        messagebox.showerror(title, detail)

    def copy_log(self):
        text = "\n".join(f"[{ts}] {level.upper():7} {msg}" for ts, level, msg in self.log_entries)
        self.master.clipboard_clear()
        self.master.clipboard_append(text)
        self.show_toast("Log copied", "success")

    def save_log(self):
        path = filedialog.asksaveasfilename(
            title="Save Activity Log",
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
        )
        if not path:
            return
        with open(path, "w", encoding="utf-8") as handle:
            for ts, level, msg in self.log_entries:
                handle.write(f"[{ts}] {level.upper():7} {msg}\n")
        self.show_toast("Log saved", "success")

    def clear_log(self):
        self.log_entries.clear()
        self.recent_events.clear()
        self._render_activity_log()
        self._render_recent_log()
        self.show_toast("Log cleared", "warning")

    def show_help(self):
        messagebox.showinfo(
            "Keyboard Shortcuts",
            "Ctrl+F    Search brands\n"
            "Ctrl+U    Update files\n"
            "Ctrl+L    Load brands\n"
            "Ctrl+Enter Generate/upload/email\n"
            "Alt+1-4   Switch screens\n"
            "Esc       Clear search",
        )

    def _open_output_folder(self, path):
        if not path:
            return
        try:
            if sys.platform.startswith("win"):
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])
        except Exception as exc:
            self.append_log(f"Could not open output folder: {exc}", level="warning")

    def on_close(self):
        try:
            if self.autosave_job is not None:
                self.master.after_cancel(self.autosave_job)
            self._persist_settings(add_log=False)
        except Exception:
            pass
        self.master.destroy()


# ----------------- MAIN -----------------
def main():
    cfg = load_config()
    if tb is not None:
        try:
            root = tb.Window(themename=normalize_theme_name(cfg.get("theme", DEFAULT_GUI_CONFIG["theme"])))
        except Exception:
            root = tb.Window(themename=DEFAULT_GUI_CONFIG["theme"])
    else:
        root = tk.Tk()
    app = BrandInventoryGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main()
