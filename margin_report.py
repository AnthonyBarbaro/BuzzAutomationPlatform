import os
import re
import subprocess
import sys
import threading
import tkinter as tk
from contextlib import redirect_stderr, redirect_stdout
from tkinter import filedialog, messagebox, scrolledtext, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import quote_sheetname
import traceback
from datetime import datetime
import shutil
import numpy as np  # used for numeric operations

# =============================================================================
# CONFIG & CONSTANTS
# =============================================================================

CONFIG_FILE = "config.txt"

# Columns we expect from the CSVs
INPUT_COLUMNS = ['Available', 'Product', 'Category', 'Brand', 'Price', 'Cost']

# --- Price selection behavior ---------------------------------------------- #
# Prefer location price if present and > 0, otherwise fall back to base Price.
LOCATION_PRICE_ALIASES = [
    "Location price",
    "Location Price",
    "location price",
    "location_price",
]

# --- Thresholds for filtering products ------------------------------------- #
BASE_MIN_PRICE = 1.01          # Minimum shelf price we care about
MIN_AVAILABLE_QTY = 5          # Minimum inventory units to keep a product
MIN_COST = 1.0                 # Minimum cost to keep a product

# --- Margin & tax/fees configuration --------------------------------------- #
# Everyday scenario: 30% off + 10% back in points ≈ 37% total discount
# Effective revenue is about 63% of shelf price.
EFFECTIVE_REVENUE_RATE = 0.63
OUT_THE_DOOR_MULTIPLIER = 1.33  # multiplier from effective price

# --- Promo definitions ------------------------------------------------------ #
# Scenario 1: 50% discount + 10% back in points (~55% total) + 30% lower cost
PROMO_50 = {
    "label": "50% Off + 10% Back + 30% Cost Relief",
    "total_discount": 0.55,     # overall customer-facing discount on price
    "cost_reduction": 0.30,     # vendor gives ~30% cost support
}

# Scenario 2: 40% discount + 10% back in points (~46% total) + 25% lower cost
PROMO_40 = {
    "label": "40% Off + 10% Back + 25% Cost Relief",
    "total_discount": 0.46,     # 1 - (0.60 * 0.90) = 0.46
    "cost_reduction": 0.25,     # vendor gives 25% cost support
}

# Scenario 3: 50% discount + 10% back + 25% lower cost (softer vendor support)
PROMO_50_SOFT = {
    "label": "50% Off + 10% Back + 25% Cost Relief",
    "total_discount": 0.55,
    "cost_reduction": 0.25,
}

# Scenario 4: 40% discount + 10% back + 20% lower cost
PROMO_40_SOFT = {
    "label": "40% Off + 10% Back + 20% Cost Relief",
    "total_discount": 0.46,
    "cost_reduction": 0.20,
}

# --- Columns to strip from the final export -------------------------------- #
# (We still use some of these internally, but they won’t show in the Excel.)
COLUMNS_TO_STRIP = [
    "Strain",
    "Location price",
    "Vendor",
    "Tags",
    "Strain_Type",
    "Product_Weight",
    "Product_SubType",
    "Available",
    "Source File",
    "SourceFile",
]

EXPORT_ONLY_COLUMNS_TO_REMOVE = {
    "Brand Strain Type",
    "Price_Used_Source",
    "Is_Store_Specific",
}

# Columns to format as currency / percent in Excel (for Products sheet only)
CURRENCY_COLUMNS = {
    "Price",
    "Cost",
    "Price_Used",
    "Effective_Price",
    "Out-The-Door",
    "TargetPrice_45Margin",
    "DiffTo45Margin",
    "Promo50_Effective_Price",
    "Promo50_Cost",
    "Promo40_Effective_Price",
    "Promo40_Cost",
    "Promo50_Cost_25Relief",
    "Promo40_Cost_20Relief",
}

PERCENT_COLUMNS = {
    "Margin",
    "Margin_Promo50",
    "Margin_Promo40",
    "Margin_Promo50_25Relief",
    "Margin_Promo40_20Relief",
    "AvgMargin",
    "MinMargin",
    "MaxMargin",
}

# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def ensure_dir_exists(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def organize_by_brand(output_directory):
    """
    Legacy helper: Goes through all XLSX files in output_directory and its
    subfolders. If the file name matches "<Store>_<Brand>_<MM-DD-YYYY>.xlsx",
    move it into output_directory/Brand/ preserving the file name.
    (Not used with the new naming scheme, but kept around if needed later.)
    """
    pattern = re.compile(r'^(.*?)_(.*?)_(\d{2}-\d{2}-\d{4})\.xlsx$')

    for root, dirs, files in os.walk(output_directory):
        for filename in files:
            if filename.lower().endswith(".xlsx"):
                match = pattern.match(filename)
                if match:
                    store_name, brand_name, date_str = match.groups()
                    if os.path.basename(root) == brand_name:
                        continue

                    brand_folder = os.path.join(output_directory, brand_name)
                    ensure_dir_exists(brand_folder)

                    old_path = os.path.join(root, filename)
                    new_path = os.path.join(brand_folder, filename)

                    print(f"Moving {old_path} → {new_path}")
                    shutil.move(old_path, new_path)


def extract_strain_type(product_name: str):
    """Identify single-letter strain markers like S, H, I in the product name."""
    if not isinstance(product_name, str):
        return ""
    name = " " + product_name.upper() + " "
    if re.search(r'\bS\b', name):
        return 'S'
    if re.search(r'\bH\b', name):
        return 'H'
    if re.search(r'\bI\b', name):
        return 'I'
    return ""


def extract_product_details(product_name: str):
    """
    Parse weight (e.g. '3.5G', '1G', '28G') and an optional subtype (HH / IN)
    from the product name.
    """
    if not isinstance(product_name, str):
        return "", ""
    name_upper = product_name.upper()
    weight_match = re.search(r'(\d+(\.\d+)?)G', name_upper)
    weight = weight_match.group(0) if weight_match else ""

    sub_type = ""
    if " HH " in f" {name_upper} ":
        sub_type = "HH"
    elif " IN " in f" {name_upper} ":
        sub_type = "IN"

    return weight, sub_type


def is_empty_or_numbers(val):
    """
    If the 'Product' cell is empty or only digits, we consider it invalid.
    """
    if not isinstance(val, str):
        return True
    val_str = val.strip()
    return val_str == "" or val_str.isdigit()


def format_excel_file(filename: str) -> None:
    """
    Generic formatting for all sheets EXCEPT the Summary sheet.
    Summary gets its own, more visual formatting.
    """
    wb = load_workbook(filename)

    for ws in wb.worksheets:
        # Summary sheet is handled by a dedicated function
        if ws.title == "Summary":
            continue

        # Freeze header row
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = header_fill

        # Zebra stripes
        stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        max_row = ws.max_row

        # Go column by column so we can set widths + formats + conditional formatting
        for col_cells in ws.columns:
            max_length = 0
            header_cell = col_cells[0]
            header_text = str(header_cell.value) if header_cell.value is not None else ""
            col_letter = header_cell.column_letter

            for cell in col_cells:
                # Zebra striping on data rows only
                if cell.row >= 2 and cell.row % 2 == 0:
                    if cell.fill is None or cell.fill.fill_type in (None, "none"):
                        cell.fill = stripe_fill

                if cell.value is not None:
                    length = _display_width_hint(cell.value)
                    if length > max_length:
                        max_length = length

            # Auto width
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, len(header_text) + 2), 40)

            # Number formats + right alignment
            if header_text in CURRENCY_COLUMNS:
                for cell in col_cells[1:]:
                    cell.number_format = '"$"#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
            elif header_text in PERCENT_COLUMNS:
                for cell in col_cells[1:]:
                    cell.number_format = '0.0%'
                    cell.alignment = Alignment(horizontal='right')
                # Color scale for percentages: red → yellow → green
                data_range = f"{col_letter}2:{col_letter}{max_row}"
                rule = ColorScaleRule(
                    start_type="min", start_color="F8696B",   # red
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",  # yellow
                    end_type="max", end_color="63BE7B"       # green
                )
                ws.conditional_formatting.add(data_range, rule)

    wb.save(filename)


def _detect_header_row(ws) -> int | None:
    """Find the first non-empty row and treat it as the header row."""
    max_scan_rows = min(ws.max_row, 10)
    for row_idx in range(1, max_scan_rows + 1):
        if any(cell.value not in (None, "") for cell in ws[row_idx]):
            return row_idx
    return None


def _display_width_hint(value) -> int:
    """Estimate visual width without letting long formulas explode column sizing."""
    if value is None:
        return 0
    if isinstance(value, str) and value.startswith("="):
        return 14
    return len(str(value))


def export_schema_cleaner(df: pd.DataFrame) -> pd.DataFrame:
    """Drop columns that should never appear in the exported workbook."""
    if df is None or df.empty:
        return df

    columns_to_remove = set(COLUMNS_TO_STRIP) | EXPORT_ONLY_COLUMNS_TO_REMOVE
    existing = [col for col in df.columns if col in columns_to_remove]
    if existing:
        print(f"[INFO] Removing export-only columns: {existing}")
        return df.drop(columns=existing)
    return df


def build_header_column_map(ws) -> dict[str, str]:
    """
    Build a mapping of header text -> Excel column letter using the detected
    header row rather than hardcoded indexes.
    """
    header_row = _detect_header_row(ws)
    if header_row is None:
        print(f"[INFO] No header row found on sheet '{ws.title}'.")
        return {}

    header_map: dict[str, str] = {}
    for cell in ws[header_row]:
        if cell.value in (None, ""):
            continue
        header_map[str(cell.value).strip()] = get_column_letter(cell.column)

    print(f"[INFO] Header row detected at {header_row} on sheet '{ws.title}'.")
    return header_map


def apply_excel_formulas(ws, header_map: dict[str, str]) -> None:
    """Replace exported static values with dynamic Excel formulas."""
    header_row = _detect_header_row(ws)
    if header_row is None:
        print(f"[INFO] Skipping formulas for '{ws.title}' because no header row was found.")
        return

    first_data_row = header_row + 1
    if ws.max_row < first_data_row:
        print(f"[INFO] Skipping formulas for '{ws.title}' because there are no data rows.")
        return

    print(f"[INFO] Applying formulas to sheet '{ws.title}'...")

    formula_specs = [
        (
            "Effective_Price",
            ["Price_Used"],
            lambda refs: f'=IFERROR({refs["Price_Used"]}*{EFFECTIVE_REVENUE_RATE},"")',
        ),
        (
            "Out-The-Door",
            ["Effective_Price"],
            lambda refs: f'=IFERROR({refs["Effective_Price"]}*{OUT_THE_DOOR_MULTIPLIER},"")',
        ),
        (
            "Margin",
            ["Effective_Price", "Cost"],
            lambda refs: (
                f'=IFERROR(({refs["Effective_Price"]}-{refs["Cost"]})/{refs["Effective_Price"]},"")'
            ),
        ),
        (
            "TargetPrice_45Margin",
            ["Cost"],
            lambda refs: f'=IFERROR({refs["Cost"]}/0.385,"")',
        ),
        (
            "DiffTo45Margin",
            ["TargetPrice_45Margin", "Price_Used"],
            lambda refs: f'=IFERROR({refs["TargetPrice_45Margin"]}-{refs["Price_Used"]},"")',
        ),
        (
            "Promo50_Effective_Price",
            ["Price_Used"],
            lambda refs: f'=IFERROR({refs["Price_Used"]}*(1-{PROMO_50["total_discount"]}),"")',
        ),
        (
            "Promo50_Cost",
            ["Cost"],
            lambda refs: f'=IFERROR({refs["Cost"]}*(1-{PROMO_50["cost_reduction"]}),"")',
        ),
        (
            "Margin_Promo50",
            ["Promo50_Effective_Price", "Promo50_Cost"],
            lambda refs: (
                f'=IFERROR(({refs["Promo50_Effective_Price"]}-{refs["Promo50_Cost"]})/'
                f'{refs["Promo50_Effective_Price"]},"")'
            ),
        ),
        (
            "Promo40_Effective_Price",
            ["Price_Used"],
            lambda refs: f'=IFERROR({refs["Price_Used"]}*(1-{PROMO_40["total_discount"]}),"")',
        ),
        (
            "Promo40_Cost",
            ["Cost"],
            lambda refs: f'=IFERROR({refs["Cost"]}*(1-{PROMO_40["cost_reduction"]}),"")',
        ),
        (
            "Margin_Promo40",
            ["Promo40_Effective_Price", "Promo40_Cost"],
            lambda refs: (
                f'=IFERROR(({refs["Promo40_Effective_Price"]}-{refs["Promo40_Cost"]})/'
                f'{refs["Promo40_Effective_Price"]},"")'
            ),
        ),
        (
            "Promo50_Cost_25Relief",
            ["Cost"],
            lambda refs: f'=IFERROR({refs["Cost"]}*(1-{PROMO_50_SOFT["cost_reduction"]}),"")',
        ),
        (
            "Margin_Promo50_25Relief",
            ["Promo50_Effective_Price", "Promo50_Cost_25Relief"],
            lambda refs: (
                f'=IFERROR(({refs["Promo50_Effective_Price"]}-{refs["Promo50_Cost_25Relief"]})/'
                f'{refs["Promo50_Effective_Price"]},"")'
            ),
        ),
        (
            "Promo40_Cost_20Relief",
            ["Cost"],
            lambda refs: f'=IFERROR({refs["Cost"]}*(1-{PROMO_40_SOFT["cost_reduction"]}),"")',
        ),
        (
            "Margin_Promo40_20Relief",
            ["Promo40_Effective_Price", "Promo40_Cost_20Relief"],
            lambda refs: (
                f'=IFERROR(({refs["Promo40_Effective_Price"]}-{refs["Promo40_Cost_20Relief"]})/'
                f'{refs["Promo40_Effective_Price"]},"")'
            ),
        ),
    ]

    for target_col, required_cols, formula_builder in formula_specs:
        target_letter = header_map.get(target_col)
        missing = [col for col in required_cols if col not in header_map]
        if not target_letter:
            print(f"[INFO] Skipping formula column '{target_col}' because it is not in the export.")
            continue
        if missing:
            print(f"[INFO] Skipping formula column '{target_col}' because dependencies are missing: {missing}")
            continue

        for row_idx in range(first_data_row, ws.max_row + 1):
            refs = {col: f"{header_map[col]}{row_idx}" for col in required_cols}
            ws[f"{target_letter}{row_idx}"] = formula_builder(refs)

    print(f"[INFO] Rows processed: {ws.max_row - header_row}")

# =============================================================================
# PRICE SELECTION & PROMO HELPERS
# =============================================================================

def _first_present_column(df: pd.DataFrame, candidates):
    """Return the first column name from candidates that exists in df."""
    for c in candidates:
        if c in df.columns:
            return c
    return None


def _to_num(series):
    """Coerce to numeric; invalid → NaN."""
    return pd.to_numeric(series, errors="coerce")


def inject_sell_price_columns(df: pd.DataFrame):
    """
    Create two diagnostic columns:
      - Price_Used: numeric value actually used for downstream math
      - Price_Used_Source: string label ('Location price' or 'Price')
    """
    loc_col = _first_present_column(df, LOCATION_PRICE_ALIASES)
    price_col_exists = 'Price' in df.columns

    loc = _to_num(df[loc_col]) if loc_col else pd.Series(np.nan, index=df.index)
    base = _to_num(df['Price']) if price_col_exists else pd.Series(np.nan, index=df.index)

    # prefer location price when > 0, else fall back to base Price
    use_loc_mask = loc.notna() & (loc > 0)

    df['Price_Used'] = np.where(use_loc_mask, loc, base)  # numeric
    df['Price_Used_Source'] = np.where(
        use_loc_mask,
        loc_col if loc_col else 'Price',
        'Price' if price_col_exists else (loc_col or '')
    )

    return df, loc_col

# =============================================================================
# CORE DATA PROCESSING
# =============================================================================

def process_single_file(file_path: str, selected_brands):
    """
    Read one CSV, clean/filter it, compute margins & promo simulations,
    and return a DataFrame of valid products for that file.
    """
    try:
        df = pd.read_csv(file_path)
    except Exception as e:
        print(f"Error reading {file_path}: {e}")
        return None

    # Quick sanity check: make sure at least some expected columns exist
    existing_cols = [c for c in INPUT_COLUMNS if c in df.columns]
    if not existing_cols:
        print(f"No required columns found in {file_path}. Skipping.")
        return None

    # Normalize numeric types
    for col in ['Price', 'Cost', 'Available']:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # 1) EXCLUDE PROMO / SAMPLE by name
    if 'Product' in df.columns:
        df = df[~df['Product'].str.contains(r'(?i)\bpromo(?:s)?\b|\bsample\b', na=False)]

    # 2) EXCLUDE Category = "Accessories"
    if 'Category' in df.columns:
        df = df[~df['Category'].str.contains(r'(?i)\baccessories\b', na=False)]

    if df.empty:
        return None

    # Attach store name & source file (derived from filename)
    base_name = os.path.splitext(os.path.basename(file_path))[0]
    parts = base_name.split('_')
    store_name = parts[-1] if len(parts) > 1 else base_name

    df['Store'] = store_name
    df['SourceFile'] = os.path.basename(file_path)

    # 3) Compute Price_Used (location price preferred)
    df, _ = inject_sell_price_columns(df)

    # Flag whether location_price is used (store-specific pricing)
    df['Is_Store_Specific'] = df['Price_Used_Source'] != 'Price'

    # 4) EXCLUDE rows with too-small price using Price_Used
    if 'Price_Used' in df.columns:
        df = df[(df['Price_Used'].isna()) | (df['Price_Used'] >= BASE_MIN_PRICE)]
    elif 'Price' in df.columns:
        df = df[(df['Price'].isna()) | (df['Price'] >= BASE_MIN_PRICE)]

    # 5) EXCLUDE Available < MIN_AVAILABLE_QTY (if Available exists)
    if 'Available' in df.columns:
        df = df[df['Available'] >= MIN_AVAILABLE_QTY]

    # 6) EXCLUDE Cost <= MIN_COST
    if 'Cost' in df.columns:
        df = df[df['Cost'] > MIN_COST]

    if df.empty:
        return None

    # Ensure Brand & Product columns exist
    if 'Brand' not in df.columns:
        df['Brand'] = 'Unknown'
    if 'Product' not in df.columns:
        df['Product'] = ''

    # Keep only selected brands, if specified
    if selected_brands:
        df = df[df['Brand'].isin(selected_brands)]

    if df.empty:
        return None

    # Remove rows where product name is empty or just numbers
    df = df[~df['Product'].apply(is_empty_or_numbers)].copy()
    if df.empty:
        return None

    # Optional product metadata (we drop them later, but fine to keep for now)
    df['Strain_Type'] = df['Product'].apply(extract_strain_type)
    df[['Product_Weight', 'Product_SubType']] = df['Product'].apply(
        lambda x: pd.Series(extract_product_details(x))
    )

    # Margin & price simulations
    if 'Price_Used' in df.columns and 'Cost' in df.columns:
        df['Price_Used'] = pd.to_numeric(df['Price_Used'], errors='coerce')
        df['Cost'] = pd.to_numeric(df['Cost'], errors='coerce')

        price = df['Price_Used']
        cost = df['Cost']

        # Everyday effective price & out-the-door (30% off + 10% back → 63% of price)
        eff = price * EFFECTIVE_REVENUE_RATE
        df['Effective_Price'] = eff
        df['Out-The-Door'] = eff * OUT_THE_DOOR_MULTIPLIER

        # Everyday (current) margin
        df['Margin'] = np.where(
            eff.notna() & (eff != 0),
            (eff - cost) / eff,
            np.nan
        )

        # Target price for 45% margin
        df['TargetPrice_45Margin'] = np.where(
            cost.notna(),
            cost / 0.385,
            np.nan
        )
        df['DiffTo45Margin'] = df['TargetPrice_45Margin'] - df['Price_Used']

        # --- Promo 50% off + 10% back + 30% lower cost --------------------
        promo50_price = price * (1 - PROMO_50["total_discount"])  # Price_Used * 0.45
        promo50_cost = cost * (1 - PROMO_50["cost_reduction"])    # Cost * 0.70

        df['Promo50_Effective_Price'] = promo50_price
        df['Promo50_Cost'] = promo50_cost

        df['Margin_Promo50'] = np.where(
            promo50_price.notna() & (promo50_price > 0),
            (promo50_price - promo50_cost) / promo50_price,
            np.nan
        )

        # --- Promo 40% off + 10% back + 25% cost relief -------------------
        promo40_price = price * (1 - PROMO_40["total_discount"])  # Price_Used * 0.54
        promo40_cost = cost * (1 - PROMO_40["cost_reduction"])    # Cost * 0.75

        df['Promo40_Effective_Price'] = promo40_price
        df['Promo40_Cost'] = promo40_cost

        df['Margin_Promo40'] = np.where(
            promo40_price.notna() & (promo40_price > 0),
            (promo40_price - promo40_cost) / promo40_price,
            np.nan
        )

        # --- Alternate promo scenarios with less cost relief -------------
        # 50% off + 10% back, but only 25% cost relief
        promo50_cost_soft = cost * (1 - PROMO_50_SOFT["cost_reduction"])  # Cost * 0.75
        df['Promo50_Cost_25Relief'] = promo50_cost_soft
        df['Margin_Promo50_25Relief'] = np.where(
            promo50_price.notna() & (promo50_price > 0),
            (promo50_price - promo50_cost_soft) / promo50_price,
            np.nan
        )

        # 40% off + 10% back, but only 20% cost relief
        promo40_cost_soft = cost * (1 - PROMO_40_SOFT["cost_reduction"])  # Cost * 0.80
        df['Promo40_Cost_20Relief'] = promo40_cost_soft
        df['Margin_Promo40_20Relief'] = np.where(
            promo40_price.notna() & (promo40_price > 0),
            (promo40_price - promo40_cost_soft) / promo40_price,
            np.nan
        )

    else:
        # If we can't compute margins, keep the rows but with NaNs
        df['Effective_Price'] = np.nan
        df['Out-The-Door'] = np.nan
        df['Margin'] = np.nan
        df['TargetPrice_45Margin'] = np.nan
        df['DiffTo45Margin'] = np.nan
        df['Promo50_Effective_Price'] = np.nan
        df['Promo50_Cost'] = np.nan
        df['Margin_Promo50'] = np.nan
        df['Promo40_Effective_Price'] = np.nan
        df['Promo40_Cost'] = np.nan
        df['Margin_Promo40'] = np.nan
        df['Promo50_Cost_25Relief'] = np.nan
        df['Margin_Promo50_25Relief'] = np.nan
        df['Promo40_Cost_20Relief'] = np.nan
        df['Margin_Promo40_20Relief'] = np.nan

    return df


def consolidate_across_stores(df: pd.DataFrame) -> pd.DataFrame:
    """
    Consolidate rows across stores so that:
      - Base-price rows (Price_Used_Source == 'Price') are treated as chain-wide
        and labeled "All Stores" in the Store column.
      - Location-price rows (Price_Used_Source != 'Price') are store-specific;
        they get a comma-separated list of stores.

    Grouping key (per SKU):
      Brand, Product, Category, Cost.
    """
    if df is None or df.empty:
        return df

    # Ensure required columns exist
    for col in ['Brand', 'Product', 'Category', 'Cost', 'Price_Used', 'Price_Used_Source', 'Store']:
        if col not in df.columns:
            df[col] = np.nan

    # Ensure the flag exists
    if 'Is_Store_Specific' not in df.columns:
        if 'Price_Used_Source' in df.columns:
            df['Is_Store_Specific'] = df['Price_Used_Source'] != 'Price'
        else:
            df['Is_Store_Specific'] = False

    key_cols = ['Brand', 'Product', 'Category', 'Cost']

    consolidated_rows = []

    grouped = df.groupby(key_cols, dropna=False)
    for _, sku_df in grouped:
        if sku_df.empty:
            continue

        # One row per unique price & source
        for (price, source), combo_df in sku_df.groupby(['Price_Used', 'Price_Used_Source'], dropna=False):
            if combo_df.empty:
                continue

            stores = combo_df['Store'].dropna().astype(str).unique().tolist()

            # Only make it store-specific if location_price is used,
            # otherwise treat it as chain-wide.
            if source == 'Price':
                store_label = "All Stores"
            else:
                store_label = ", ".join(sorted(stores)) if stores else ""

            row = combo_df.iloc[0].copy()
            row['Store'] = store_label
            row['Is_Store_Specific'] = bool(combo_df['Is_Store_Specific'].any())
            consolidated_rows.append(row)

    if consolidated_rows:
        out_df = pd.DataFrame(consolidated_rows).reset_index(drop=True)
        return out_df

    # Fallback
    return df.reset_index(drop=True)


def merge_similar_products(brand_df: pd.DataFrame) -> pd.DataFrame:
    """
    Inside a single brand, merge 'similar' products when they share:

        Category
        Price_Used
        Cost

    We DON'T care about which stores or whether the price came from
    base 'Price' vs 'Location price' as long as the actual Price_Used
    and Cost are identical.
    """
    if brand_df is None or brand_df.empty:
        return brand_df

    df = brand_df.copy()

    # Make sure the columns we group on exist
    for col in ['Category', 'Price_Used', 'Cost', 'Product']:
        if col not in df.columns:
            df[col] = np.nan

    if 'Is_Store_Specific' not in df.columns:
        df['Is_Store_Specific'] = False

    # Group by Category + Price_Used + Cost.
    group_keys = []
    if 'Category' in df.columns:
        group_keys.append('Category')

    for col in ['Price_Used', 'Cost']:
        if col in df.columns:
            group_keys.append(col)

    group_keys = list(dict.fromkeys(group_keys))  # de-dup, just in case

    if not group_keys:
        return df

    merged_rows = []

    for _, grp in df.groupby(group_keys, dropna=False):
        if grp.empty:
            continue

        row = grp.iloc[0].copy()

        # Merge product names
        product_names = sorted({str(x) for x in grp['Product'].dropna()})
        count = len(product_names)

        if count == 0:
            display_name = ""
        elif count == 1:
            display_name = product_names[0]
        else:
            # First name + how many more we collapsed
            display_name = f"{product_names[0]} (+{count - 1} more)"

        row['Product'] = display_name
        row['Product_List'] = "; ".join(product_names)
        row['Merged_Count'] = count

        # Merge Store label if present
        if 'Store' in grp.columns:
            stores_vals = sorted({str(x) for x in grp['Store'].dropna()})
            row['Store'] = ", ".join(stores_vals)

        # propagate store-specific flag
        row['Is_Store_Specific'] = bool(grp['Is_Store_Specific'].any())

        merged_rows.append(row)

    return pd.DataFrame(merged_rows).reset_index(drop=True)

# =============================================================================
# SUMMARY / DASHBOARD HELPERS
# =============================================================================

def build_scenario_summary(df: pd.DataFrame) -> list[dict]:
    """
    Build one row per pricing scenario (current, promos) with
    Avg / Min / Max margin, SKU count and a simple quality band.
    """
    scenario_rows: list[dict] = []

    scenario_label_map = {
        "Margin": "Current Everyday Pricing",
        "Margin_Promo50": PROMO_50["label"],
        "Margin_Promo40": PROMO_40["label"],
        "Margin_Promo50_25Relief": PROMO_50_SOFT["label"],
        "Margin_Promo40_20Relief": PROMO_40_SOFT["label"],
    }

    for col, label in scenario_label_map.items():
        if col not in df.columns:
            continue
        s = pd.to_numeric(df[col], errors="coerce").dropna()
        if s.empty:
            continue

        avg = s.mean()
        mn = s.min()
        mx = s.max()

        # Simple quality band – tweak thresholds here if you like
        if avg >= 0.60:
            quality = "Excellent"
        elif avg >= 0.45:
            quality = "Good"
        elif avg >= 0.35:
            quality = "OK"
        else:
            quality = "Needs Attention"

        scenario_rows.append({
            "Scenario": label,
            "AvgMargin": avg,
            "MinMargin": mn,
            "MaxMargin": mx,
            "SKUsWithMargin": len(s),
            "Quality": quality,
        })

    return scenario_rows


def build_category_summary(df: pd.DataFrame) -> list[dict]:
    """
    Build one row per product Category with:
      - Total SKU count (using Merged_Count if available)
      - Avg margin under each scenario
    """
    rows: list[dict] = []
    if "Category" not in df.columns:
        return rows

    grouped = df.groupby("Category", dropna=True)

    for cat, grp in grouped:
        if cat is None:
            continue
        cat_label = str(cat).strip()
        if not cat_label:
            continue

        row = {"Category": cat_label}

        # ---------- SKU count (FIXED) ----------
        # If we have Merged_Count, use that to count underlying SKUs.
        if "Merged_Count" in grp.columns:
            # each row already represents N merged SKUs
            skus = grp["Merged_Count"].fillna(1).sum()
        elif "Product" in grp.columns:
            # fallback: count distinct product names
            skus = grp["Product"].nunique()
        else:
            # last-resort fallback
            skus = len(grp)

        row["SKUs"] = int(skus)

        # ---------- Average margins by scenario ----------
        for col, key in [
            ("Margin", "AvgMargin_Current"),
            ("Margin_Promo50", "AvgMargin_Promo50"),
            ("Margin_Promo40", "AvgMargin_Promo40"),
        ]:
            if col in grp.columns:
                s = pd.to_numeric(grp[col], errors="coerce").dropna()
                row[key] = s.mean() if not s.empty else np.nan
            else:
                row[key] = np.nan

        rows.append(row)

    # Sort alphabetically by category name
    rows.sort(key=lambda r: r["Category"])
    return rows


def enhance_summary_and_charts(filename: str,
                               brand: str,
                               data_df: pd.DataFrame,
                               use_excel_formulas: bool = False) -> None:
    """
    Create a 'Summary' sheet that acts as a visual dashboard:
      - Scenario Summary table + bar chart
      - Category Margin Breakdown table + bar chart + pie chart
    """
    wb = load_workbook(filename)
    products_ws = wb["Products"] if "Products" in wb.sheetnames else None
    products_header_map = build_header_column_map(products_ws) if products_ws else {}
    products_header_row = _detect_header_row(products_ws) if products_ws else None
    products_first_data_row = (products_header_row + 1) if products_header_row else 2
    products_max_row = products_ws.max_row if products_ws else 0
    products_sheet_ref = quote_sheetname(products_ws.title) if products_ws else quote_sheetname("Products")

    # Start fresh each time
    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])
    ws = wb.create_sheet("Summary")

    scenario_rows = build_scenario_summary(data_df)
    category_rows = build_category_summary(data_df)
    scenario_specs = [
        ("Margin", "Current Everyday Pricing"),
        ("Margin_Promo50", PROMO_50["label"]),
        ("Margin_Promo40", PROMO_40["label"]),
        ("Margin_Promo50_25Relief", PROMO_50_SOFT["label"]),
        ("Margin_Promo40_20Relief", PROMO_40_SOFT["label"]),
    ]
    scenario_lookup = {label: col for col, label in scenario_specs}

    def product_range(header_name: str) -> str | None:
        col_letter = products_header_map.get(header_name)
        if not col_letter or products_max_row < products_first_data_row:
            return None
        return f"{products_sheet_ref}!${col_letter}${products_first_data_row}:${col_letter}${products_max_row}"

    # Big title row
    max_cols_for_title = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols_for_title)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = f"{brand} - Margin Dashboard"
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    current_row = 3

    # ---------------------------------------------------------------------
    # Scenario summary section
    # ---------------------------------------------------------------------
    if scenario_rows:
        sec_title = ws.cell(row=current_row, column=1)
        sec_title.value = "Scenario Summary"
        sec_title.font = Font(size=13, bold=True)
        current_row += 1

        scen_header_row = current_row
        scen_headers = ["Scenario", "Avg Margin", "Min Margin", "Max Margin", "# SKUs", "Quality"]
        scen_avg_col_letter = get_column_letter(2)

        # Header
        for col_idx, header in enumerate(scen_headers, start=1):
            cell = ws.cell(row=scen_header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.border = thin_border

        scen_data_start = scen_header_row + 1
        r = scen_data_start

        # Data rows
        for rowdata in scenario_rows:
            ws.cell(row=r, column=1, value=rowdata["Scenario"])
            scenario_col = scenario_lookup.get(rowdata["Scenario"])
            scenario_range = product_range(scenario_col) if use_excel_formulas and scenario_col else None

            if scenario_range:
                ws.cell(row=r, column=2, value=f'=IFERROR(AVERAGE({scenario_range}),"")')
                ws.cell(row=r, column=3, value=f'=IFERROR(MIN({scenario_range}),"")')
                ws.cell(row=r, column=4, value=f'=IFERROR(MAX({scenario_range}),"")')
                ws.cell(row=r, column=5, value=f'=IFERROR(COUNT({scenario_range}),"")')
                ws.cell(
                    row=r,
                    column=6,
                    value=(
                        f'=IFERROR(IF({scen_avg_col_letter}{r}>=0.60,"Excellent",'
                        f'IF({scen_avg_col_letter}{r}>=0.45,"Good",'
                        f'IF({scen_avg_col_letter}{r}>=0.35,"OK","Needs Attention"))),"")'
                    ),
                )
            else:
                ws.cell(row=r, column=2, value=rowdata["AvgMargin"])
                ws.cell(row=r, column=3, value=rowdata["MinMargin"])
                ws.cell(row=r, column=4, value=rowdata["MaxMargin"])
                ws.cell(row=r, column=5, value=rowdata["SKUsWithMargin"])
                ws.cell(row=r, column=6, value=rowdata["Quality"])
            r += 1

        scen_data_end = r - 1

        stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Format data rows
        for row_idx in range(scen_data_start, scen_data_end + 1):
            for col_idx in range(1, len(scen_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                if col_idx in (2, 3, 4):      # margin %
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 5:            # SKU count
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Zebra stripe effect
                if (row_idx - scen_data_start) % 2 == 0:
                    cell.fill = stripe_fill

        # Color scale on Avg Margin column
        avg_col_letter = get_column_letter(2)
        data_range = f"{avg_col_letter}{scen_data_start}:{avg_col_letter}{scen_data_end}"
        rule = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"
        )
        ws.conditional_formatting.add(data_range, rule)

        # Bar chart: Avg Margin by Scenario
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Average Margin by Scenario"
        chart.y_axis.title = "Margin"
        chart.x_axis.title = "Scenario"

        data = Reference(ws, min_col=2, max_col=2, min_row=scen_header_row, max_row=scen_data_end)
        cats = Reference(ws, min_col=1, max_col=1, min_row=scen_header_row + 1, max_row=scen_data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws.add_chart(chart, "H3")

        current_row = scen_data_end + 3

    # ---------------------------------------------------------------------
    # Category summary section
    # ---------------------------------------------------------------------
    if category_rows:
        sec_title = ws.cell(row=current_row, column=1)
        sec_title.value = "Category Margin Breakdown"
        sec_title.font = Font(size=13, bold=True)
        current_row += 1

        cat_header_row = current_row
        cat_headers = [
            "Category",
            "Avg Margin (Current)",
            "Avg Margin (Promo 50)",
            "Avg Margin (Promo 40)",
            "# SKUs",
        ]

        # Header
        for col_idx, header in enumerate(cat_headers, start=1):
            cell = ws.cell(row=cat_header_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.border = thin_border

        cat_data_start = cat_header_row + 1
        r = cat_data_start

        # Data rows
        category_range = product_range("Category") if use_excel_formulas else None
        merged_count_range = product_range("Merged_Count") if use_excel_formulas else None
        current_margin_range = product_range("Margin") if use_excel_formulas else None
        promo50_margin_range = product_range("Margin_Promo50") if use_excel_formulas else None
        promo40_margin_range = product_range("Margin_Promo40") if use_excel_formulas else None
        category_label_col_letter = get_column_letter(1)

        for rowdata in category_rows:
            ws.cell(row=r, column=1, value=rowdata["Category"])

            if category_range and current_margin_range:
                ws.cell(
                    row=r,
                    column=2,
                    value=f'=IFERROR(AVERAGEIF({category_range},${category_label_col_letter}{r},{current_margin_range}),"")',
                )
            else:
                ws.cell(row=r, column=2, value=rowdata["AvgMargin_Current"])

            if category_range and promo50_margin_range:
                ws.cell(
                    row=r,
                    column=3,
                    value=f'=IFERROR(AVERAGEIF({category_range},${category_label_col_letter}{r},{promo50_margin_range}),"")',
                )
            else:
                ws.cell(row=r, column=3, value=rowdata["AvgMargin_Promo50"])

            if category_range and promo40_margin_range:
                ws.cell(
                    row=r,
                    column=4,
                    value=f'=IFERROR(AVERAGEIF({category_range},${category_label_col_letter}{r},{promo40_margin_range}),"")',
                )
            else:
                ws.cell(row=r, column=4, value=rowdata["AvgMargin_Promo40"])

            if category_range and merged_count_range:
                ws.cell(
                    row=r,
                    column=5,
                    value=f'=IFERROR(SUMIF({category_range},${category_label_col_letter}{r},{merged_count_range}),"")',
                )
            elif category_range:
                ws.cell(
                    row=r,
                    column=5,
                    value=f'=IFERROR(COUNTIF({category_range},${category_label_col_letter}{r}),"")',
                )
            else:
                ws.cell(row=r, column=5, value=rowdata["SKUs"])
            r += 1

        cat_data_end = r - 1

        stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Format data rows
        for row_idx in range(cat_data_start, cat_data_end + 1):
            for col_idx in range(1, len(cat_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

                if col_idx in (2, 3, 4):      # margin %
                    cell.number_format = "0.0%"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx == 5:            # SKU count
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                if (row_idx - cat_data_start) % 2 == 0:
                    cell.fill = stripe_fill

        # Color scales on margin columns
        for col_idx in (2, 3, 4):
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}{cat_data_start}:{col_letter}{cat_data_end}"
            rule = ColorScaleRule(
                start_type="min", start_color="F8696B",
                mid_type="percentile", mid_value=50, mid_color="FFEB84",
                end_type="max", end_color="63BE7B"
            )
            ws.conditional_formatting.add(data_range, rule)

        # Bar chart: Avg Margin (Current) by Category
        bar = BarChart()
        bar.type = "col"
        bar.style = 11
        bar.title = "Avg Margin by Category (Current)"
        bar.y_axis.title = "Margin"
        bar.x_axis.title = "Category"

        data = Reference(ws, min_col=2, max_col=2, min_row=cat_header_row, max_row=cat_data_end)
        cats = Reference(ws, min_col=1, max_col=1, min_row=cat_header_row + 1, max_row=cat_data_end)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)

        ws.add_chart(bar, f"H{cat_header_row}")

        # Pie chart: SKU mix by Category
        pie = PieChart()
        pie.title = "SKU Mix by Category"

        labels = Reference(ws, min_col=1, max_col=1, min_row=cat_data_start, max_row=cat_data_end)
        data = Reference(ws, min_col=5, max_col=5, min_row=cat_header_row, max_row=cat_data_end)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)

        ws.add_chart(pie, f"N{cat_header_row}")

    # Freeze panes under the first header area
    ws.freeze_panes = "A5" if scenario_rows else "A3"

    # Auto-fit column widths
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_length = max(max_length, _display_width_hint(val))
        header_val = ws.cell(row=1, column=col_idx).value
        header_length = _display_width_hint(header_val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_length + 2, header_length + 2), 40)

    wb.save(filename)

# =============================================================================
# EXCEL WRITER & SUMMARY
# =============================================================================

def write_brand_excel(brand: str,
                      df: pd.DataFrame,
                      output_directory: str,
                      suffix: str,
                      today_str: str,
                      use_excel_formulas: bool = True,
                      keep_raw_data_sheet: bool = False) -> str:
    """
    Write ONE Excel file for a brand subset:

      suffix = "ALL_STORES" or "STORE_SPECIFIC"

    - 'Products' sheet with all SKU-level detail
    - 'Summary' sheet with a visual dashboard (scenarios + categories)
    """
    if df is None or df.empty:
        return ""

    export_df = export_schema_cleaner(df.copy())
    if export_df is None or export_df.empty:
        return ""

    # Sort order: store-specific sorted by Store first, then Category/Price/Product
    sort_cols: list[str] = []
    if suffix == "STORE_SPECIFIC" and 'Store' in export_df.columns:
        sort_cols.append('Store')
    if 'Category' in export_df.columns:
        sort_cols.append('Category')
    if 'Price_Used' in export_df.columns:
        sort_cols.append('Price_Used')
    if 'Product' in export_df.columns:
        sort_cols.append('Product')

    if sort_cols:
        export_df.sort_values(by=sort_cols, inplace=True, na_position='last')

    # Brand folder
    safe_brand = re.sub(r'[\\/*?:"<>|]', "_", str(brand))
    brand_folder = os.path.join(output_directory, safe_brand)
    ensure_dir_exists(brand_folder)

    filename = os.path.join(brand_folder, f"{safe_brand}_{suffix}_{today_str}.xlsx")

    # Write the main Products sheet
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        export_df.to_excel(writer, index=False, sheet_name='Products')
        if keep_raw_data_sheet:
            export_df.to_excel(writer, index=False, sheet_name='Raw Data')

    if use_excel_formulas:
        wb = load_workbook(filename)
        ws = wb["Products"]
        header_map = build_header_column_map(ws)
        apply_excel_formulas(ws, header_map)
        if getattr(wb, "calculation", None):
            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        wb.save(filename)

    # Build the Summary dashboard and format all sheets
    enhance_summary_and_charts(filename, brand, export_df, use_excel_formulas=use_excel_formulas)
    format_excel_file(filename)

    print(f"Created {filename}")
    return filename

# =============================================================================
# MAIN PIPELINE
# =============================================================================

def process_files(input_directory,
                  output_directory,
                  selected_brands,
                  excel_options: dict | None = None,
                  progress_callback=None):
    """
    Pipeline:

      1. Read all CSVs from input_directory using process_single_file(...)
      2. Combine into one DataFrame
      3. consolidate_across_stores(...)  → handles All Stores vs location prices
      4. merge_similar_products(...)     → collapses same price+cost into one row
      5. For each Brand:
           - rows with Is_Store_Specific == False → Brand/Brand_ALL_STORES_*.xlsx
           - rows with Is_Store_Specific == True  → Brand/Brand_STORE_SPECIFIC_*.xlsx
      6. Create a simple done.csv summary of which files processed.
    """
    ensure_dir_exists(output_directory)
    excel_options = excel_options or {}

    summary_records = []
    all_data_frames = []
    generated_files: list[str] = []
    csv_files = sorted(
        [filename for filename in os.listdir(input_directory) if filename.lower().endswith('.csv')]
    )
    total_files = len(csv_files)

    # 1) read all csvs
    for idx, filename in enumerate(csv_files, start=1):
        if progress_callback:
            progress_callback(
                status=f"Processing file {idx}/{total_files}: {filename}",
                current=idx - 1,
                total=max(total_files, 1),
            )
        file_path = os.path.join(input_directory, filename)
        try:
            df = process_single_file(file_path, selected_brands)
            if df is not None and not df.empty:
                all_data_frames.append(df)
                summary_records.append({
                    'File': filename,
                    'Status': "Processed successfully",
                    'RowsKept': len(df)
                })
            else:
                summary_records.append({
                    'File': filename,
                    'Status': "No rows after filtering",
                    'RowsKept': 0
                })
        except Exception as e:
            traceback.print_exc()
            summary_records.append({
                'File': filename,
                'Status': f"Error: {str(e)}",
                'RowsKept': 0
            })
        finally:
            if progress_callback:
                progress_callback(
                    status=f"Processed file {idx}/{total_files}: {filename}",
                    current=idx,
                    total=max(total_files, 1),
                )

    # Save simple per-file summary
    summary_df = pd.DataFrame(summary_records)
    summary_file = os.path.join(output_directory, 'done.csv')
    summary_df.to_csv(summary_file, index=False)
    print(f"Summary results saved to {summary_file}")

    if not all_data_frames:
        print("No data found to build brand reports.")
        return generated_files

    # 2) combine all stores
    combined = pd.concat(all_data_frames, ignore_index=True)

    # Ensure Brand column exists
    if 'Brand' in combined.columns:
        combined = combined[combined['Brand'].notna()].copy()
    else:
        combined['Brand'] = 'Unknown'

    # 3) consolidate cross-store pricing (handles All Stores vs location prices)
    combined = consolidate_across_stores(combined)

    # 4) drop columns you don't want in export
    for col in COLUMNS_TO_STRIP:
        if col in combined.columns:
            combined.drop(columns=col, inplace=True)

    # 5) keep only selected brands if any were chosen
    if selected_brands:
        combined = combined[combined['Brand'].isin(selected_brands)].copy()

    if combined.empty:
        print("Nothing left after brand filtering; no reports generated.")
        return generated_files

    today_str = datetime.now().strftime("%m-%d-%Y")
    if progress_callback:
        progress_callback(
            status="Building brand workbooks...",
            current=max(total_files, 1),
            total=max(total_files, 1),
        )

    # 6) one (or two) files per brand
    for brand, brand_data in combined.groupby('Brand'):
        brand_data = brand_data.copy()

        # merge "similar" products inside the brand (same price & cost)
        brand_data = merge_similar_products(brand_data)

        if 'Is_Store_Specific' not in brand_data.columns:
            brand_data['Is_Store_Specific'] = False

        mask_store_specific = brand_data['Is_Store_Specific'].fillna(False)
        mask_all = ~mask_store_specific

        all_stores_df = brand_data[mask_all].copy()
        store_specific_df = brand_data[mask_store_specific].copy()

        # write separate files into brand folder
        if not all_stores_df.empty:
            report_path = write_brand_excel(
                brand=brand,
                df=all_stores_df,
                output_directory=output_directory,
                suffix="ALL_STORES",
                today_str=today_str,
                use_excel_formulas=excel_options.get("use_excel_formulas", True),
                keep_raw_data_sheet=excel_options.get("keep_raw_data_sheet", False),
            )
            if report_path:
                generated_files.append(report_path)

        if not store_specific_df.empty:
            report_path = write_brand_excel(
                brand=brand,
                df=store_specific_df,
                output_directory=output_directory,
                suffix="STORE_SPECIFIC",
                today_str=today_str,
                use_excel_formulas=excel_options.get("use_excel_formulas", True),
                keep_raw_data_sheet=excel_options.get("keep_raw_data_sheet", False),
            )
            if report_path:
                generated_files.append(report_path)

    print("All brand files written.")
    return generated_files

# =============================================================================
# BRAND DISCOVERY & CONFIG HELPERS
# =============================================================================

def get_all_brands(input_directory):
    brands = set()
    brand_found = False
    for filename in os.listdir(input_directory):
        if filename.endswith('.csv'):
            file_path = os.path.join(input_directory, filename)
            try:
                df = pd.read_csv(file_path)
                if 'Brand' in df.columns:
                    brand_found = True
                    new_brands = df['Brand'].dropna().unique().tolist()
                    brands.update(new_brands)
            except:
                pass

    if not brand_found:
        return []
    return sorted(list(brands))


def save_config(input_dir, output_dir):
    with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
        f.write(f"{input_dir}\n")
        f.write(f"{output_dir}\n")


def load_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            lines = f.read().strip().split('\n')
            if len(lines) >= 2:
                input_dir = lines[0].strip()
                output_dir = lines[1].strip()
                if os.path.isdir(input_dir) and os.path.isdir(output_dir):
                    return input_dir, output_dir
    return None, None


def auto_detect_dirs():
    script_dir = os.path.dirname(os.path.abspath(__file__))
    files_dir = os.path.join(script_dir, 'files')
    done_dir = os.path.join(script_dir, 'done')
    if os.path.isdir(files_dir) and os.path.isdir(done_dir):
        return files_dir, done_dir
    return None, None


def open_with_default_app(path: str) -> None:
    """Open a file with the platform's default application."""
    if not path:
        return

    if os.name == "nt":
        os.startfile(path)
        return

    if sys.platform == "darwin":
        subprocess.Popen(["open", path])
        return

    opener = shutil.which("xdg-open")
    if not opener:
        raise RuntimeError("No default file opener was found on this system.")
    subprocess.Popen([opener, path])

# =============================================================================
# TKINTER UI
# =============================================================================

class GuiLogStream:
    """Write redirected stdout/stderr chunks into the GUI log pane."""
    def __init__(self, write_callback):
        self.write_callback = write_callback

    def write(self, text):
        if text:
            self.write_callback(text)

    def flush(self):
        return None


class BrandInventoryApp:
    """Brand Inventory Report application screen."""
    def __init__(self, master, return_to_main=None):
        self.master = master
        self.return_to_main = return_to_main
        self.master.title("Brand Inventory Report")
        self.master.geometry("1360x860")
        self.master.minsize(1040, 720)

        self.input_dir = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.use_excel_formulas = tk.BooleanVar(value=True)
        self.keep_raw_data_sheet = tk.BooleanVar(value=False)
        self.open_file_after_generation = tk.BooleanVar(value=False)
        self.status_text = tk.StringVar(value="Ready.")
        self.progress_value = tk.DoubleVar(value=0.0)
        self.is_running = False

        i_dir, o_dir = load_config()
        if not i_dir or not o_dir:
            i_dir, o_dir = auto_detect_dirs()

        if i_dir and o_dir:
            self.input_dir.set(i_dir)
            self.output_dir.set(o_dir)

        self._configure_styles()
        self.master.configure(bg=self.colors["bg"])
        self.master.protocol("WM_DELETE_WINDOW", self.return_to_main_hub)

        frame = ttk.Frame(master, padding=(12, 10), style="App.TFrame")
        self.frame = frame
        frame.pack(fill='both', expand=True)
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(1, weight=1)

        header_frame = ttk.Frame(frame, style="App.TFrame")
        header_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        header_frame.columnconfigure(0, weight=1)

        ttk.Label(header_frame, text="Brand Margin Reports", style="Header.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        ttk.Label(
            header_frame,
            text="Formula-driven Excel exports, live dashboards, and a cleaner production workflow.",
            style="Subhead.TLabel",
        ).grid(row=1, column=0, sticky="w", pady=(2, 0))

        content_frame = ttk.Frame(frame, style="App.TFrame")
        content_frame.grid(row=1, column=0, sticky="nsew")
        content_frame.columnconfigure(0, weight=0, minsize=420)
        content_frame.columnconfigure(1, weight=1)
        content_frame.rowconfigure(0, weight=1)

        left_panel = ttk.Frame(content_frame, style="App.TFrame")
        left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
        left_panel.columnconfigure(0, weight=1)
        left_panel.rowconfigure(3, weight=1)

        right_panel = ttk.Frame(content_frame, style="App.TFrame")
        right_panel.grid(row=0, column=1, sticky="nsew")
        right_panel.columnconfigure(0, weight=1)
        right_panel.rowconfigure(0, weight=3)
        right_panel.rowconfigure(1, weight=2)

        file_frame = ttk.LabelFrame(left_panel, text="File Inputs", padding=10, style="Card.TLabelframe")
        file_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)

        ttk.Label(file_frame, text="Input Directory", style="Card.TLabel").grid(
            row=0, column=0, sticky='w', padx=(0, 8), pady=3
        )
        self.input_entry = ttk.Entry(file_frame, textvariable=self.input_dir)
        self.input_entry.grid(row=0, column=1, sticky='ew', pady=3)
        self.browse_input_button = ttk.Button(
            file_frame,
            text="Browse",
            command=self.browse_input,
            style="Secondary.TButton",
        )
        self.browse_input_button.grid(row=0, column=2, padx=(8, 0), pady=3)

        ttk.Label(file_frame, text="Output Directory", style="Card.TLabel").grid(
            row=1, column=0, sticky='w', padx=(0, 8), pady=3
        )
        self.output_entry = ttk.Entry(file_frame, textvariable=self.output_dir)
        self.output_entry.grid(row=1, column=1, sticky='ew', pady=3)
        self.browse_output_button = ttk.Button(
            file_frame,
            text="Browse",
            command=self.browse_output,
            style="Secondary.TButton",
        )
        self.browse_output_button.grid(row=1, column=2, padx=(8, 0), pady=3)

        file_button_frame = ttk.Frame(file_frame, style="Card.TFrame")
        file_button_frame.grid(row=2, column=0, columnspan=3, sticky="w", pady=(8, 0))
        self.get_files_button = ttk.Button(
            file_button_frame,
            text="Get Files",
            command=self.get_files,
            style="Secondary.TButton",
        )
        self.get_files_button.pack(side='left', padx=(0, 8))
        self.clear_files_button = ttk.Button(
            file_button_frame,
            text="Clear Files",
            command=self.clear_files,
            style="Secondary.TButton",
        )
        self.clear_files_button.pack(side='left')

        options_frame = ttk.LabelFrame(left_panel, text="Options", padding=10, style="Card.TLabelframe")
        options_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        options_frame.columnconfigure(0, weight=1)
        options_frame.columnconfigure(1, weight=1)
        ttk.Checkbutton(
            options_frame,
            text="Use Excel Formulas (Dynamic Mode)",
            variable=self.use_excel_formulas,
            style="Card.TCheckbutton",
        ).grid(row=0, column=0, sticky='w', pady=2, padx=(0, 8))
        ttk.Checkbutton(
            options_frame,
            text="Keep Raw Data Sheet",
            variable=self.keep_raw_data_sheet,
            style="Card.TCheckbutton",
        ).grid(row=0, column=1, sticky='w', pady=2)
        ttk.Checkbutton(
            options_frame,
            text="Open file after generation",
            variable=self.open_file_after_generation,
            style="Card.TCheckbutton",
        ).grid(row=1, column=0, columnspan=2, sticky='w', pady=2)

        run_frame = ttk.LabelFrame(left_panel, text="Run Controls", padding=10, style="Card.TLabelframe")
        run_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))
        run_frame.columnconfigure(0, weight=1)

        run_button_frame = ttk.Frame(run_frame, style="Card.TFrame")
        run_button_frame.grid(row=0, column=0, sticky="w")
        self.generate_button = ttk.Button(
            run_button_frame,
            text="Generate Reports",
            command=self.run_process,
            style="Accent.TButton",
        )
        self.generate_button.pack(side='left', padx=(0, 8))
        close_label = "Back" if self.return_to_main else "Close App"
        self.return_button = ttk.Button(
            run_button_frame,
            text=close_label,
            command=self.return_to_main_hub,
            style="Quiet.TButton",
        )
        self.return_button.pack(side='left')

        ttk.Label(run_frame, textvariable=self.status_text, style="Status.TLabel").grid(
            row=1, column=0, sticky='w', pady=(8, 5)
        )
        self.progress_bar = ttk.Progressbar(
            run_frame,
            orient="horizontal",
            mode="determinate",
            variable=self.progress_value,
            maximum=100,
            style="Accent.Horizontal.TProgressbar",
        )
        self.progress_bar.grid(row=2, column=0, sticky="ew")

        brand_frame = ttk.LabelFrame(right_panel, text="Brand Selection", padding=10, style="Card.TLabelframe")
        brand_frame.grid(row=0, column=0, sticky="nsew", pady=(0, 10))
        brand_frame.columnconfigure(0, weight=1)
        brand_frame.rowconfigure(1, weight=1)

        ttk.Label(
            brand_frame,
            text="Select one or more brands. Use Ctrl/Cmd-click to select multiple.",
            style="Muted.TLabel",
        ).grid(row=0, column=0, sticky='w', pady=(0, 6))

        brand_button_frame = ttk.Frame(brand_frame, style="Card.TFrame")
        brand_button_frame.grid(row=0, column=1, sticky="e", pady=(0, 6))
        self.load_brands_button = ttk.Button(
            brand_button_frame,
            text="Load Brands",
            command=self.load_brands,
            style="Secondary.TButton",
        )
        self.load_brands_button.pack(side='left', padx=(0, 8))
        self.select_all_button = ttk.Button(
            brand_button_frame,
            text="Select All",
            command=self.select_all_brands,
            style="Secondary.TButton",
        )
        self.select_all_button.pack(side='left')

        list_container = ttk.Frame(brand_frame, style="Card.TFrame")
        list_container.grid(row=1, column=0, sticky="nsew")
        list_container.columnconfigure(0, weight=1)
        list_container.rowconfigure(0, weight=1)

        self.brand_listbox = tk.Listbox(
            list_container,
            selectmode=tk.MULTIPLE,
            exportselection=False,
            height=10,
            width=52,
            activestyle='none',
            relief='flat',
            bd=0,
            highlightthickness=1,
        )
        self.brand_listbox.grid(row=0, column=0, sticky="nsew")
        brand_scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=self.brand_listbox.yview)
        brand_scrollbar.grid(row=0, column=1, sticky="ns")
        self.brand_listbox.configure(yscrollcommand=brand_scrollbar.set)

        log_frame = ttk.LabelFrame(right_panel, text="Logs & Errors", padding=10, style="Card.TLabelframe")
        log_frame.grid(row=1, column=0, sticky="nsew")
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            wrap="word",
            height=10,
            relief='flat',
            bd=0,
            highlightthickness=1,
            font=("TkFixedFont", 9),
        )
        self.log_text.grid(row=0, column=0, sticky="nsew")
        self.log_text.configure(state='disabled')

        self.brand_listbox.configure(
            bg=self.colors["surface_alt"],
            fg=self.colors["text"],
            selectbackground=self.colors["accent"],
            selectforeground="#ffffff",
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["accent"],
        )
        self.log_text.configure(
            bg=self.colors["surface_alt"],
            fg=self.colors["text"],
            insertbackground=self.colors["text"],
            highlightbackground=self.colors["border"],
            highlightcolor=self.colors["accent"],
        )

        self.run_sensitive_widgets = [
            self.browse_input_button,
            self.browse_output_button,
            self.get_files_button,
            self.clear_files_button,
            self.load_brands_button,
            self.select_all_button,
            self.generate_button,
            self.input_entry,
            self.output_entry,
        ]

        self._log("[INFO] Brand Inventory Report is ready.")
        if self.input_dir.get():
            self._log(f"[INFO] Input directory: {self.input_dir.get()}")
        if self.output_dir.get():
            self._log(f"[INFO] Output directory: {self.output_dir.get()}")

    def _configure_styles(self):
        self.colors = {
            "bg": "#f3f7f5",
            "card": "#fbfdfb",
            "surface_alt": "#eef4f0",
            "accent": "#1f6f5f",
            "accent_hover": "#195b4e",
            "accent_soft": "#d7e9e2",
            "accent_soft_hover": "#c7dfd6",
            "text": "#18342d",
            "muted": "#60766e",
            "border": "#b8cbc3",
        }
        style = ttk.Style(self.master)
        available_themes = style.theme_names()
        if "clam" in available_themes:
            style.theme_use("clam")
        style.configure("App.TFrame", background=self.colors["bg"])
        style.configure("Card.TFrame", background=self.colors["card"])
        style.configure(
            "Card.TLabelframe",
            background=self.colors["card"],
            bordercolor=self.colors["border"],
            borderwidth=1,
            relief="solid",
        )
        style.configure(
            "Card.TLabelframe.Label",
            background=self.colors["card"],
            foreground=self.colors["text"],
            font=("TkDefaultFont", 10, "bold"),
        )
        style.configure("Header.TLabel", background=self.colors["bg"], foreground=self.colors["text"], font=("TkDefaultFont", 17, "bold"))
        style.configure("Subhead.TLabel", background=self.colors["bg"], foreground=self.colors["muted"], font=("TkDefaultFont", 10))
        style.configure("Card.TLabel", background=self.colors["card"], foreground=self.colors["text"])
        style.configure("Muted.TLabel", background=self.colors["card"], foreground=self.colors["muted"])
        style.configure("Status.TLabel", background=self.colors["card"], foreground=self.colors["accent"], font=("TkDefaultFont", 10, "bold"))
        style.configure(
            "Accent.TButton",
            background=self.colors["accent"],
            foreground="#ffffff",
            padding=(12, 6),
            borderwidth=0,
            focusthickness=0,
        )
        style.map(
            "Accent.TButton",
            background=[
                ("active", self.colors["accent_hover"]),
                ("pressed", self.colors["accent_hover"]),
                ("disabled", "#9ab8b0"),
            ],
            foreground=[("disabled", "#eef5f2")],
        )
        style.configure(
            "Secondary.TButton",
            background=self.colors["accent_soft"],
            foreground=self.colors["text"],
            padding=(10, 5),
            bordercolor=self.colors["border"],
            focusthickness=0,
        )
        style.map(
            "Secondary.TButton",
            background=[
                ("active", self.colors["accent_soft_hover"]),
                ("pressed", self.colors["accent_soft_hover"]),
            ],
        )
        style.configure(
            "Quiet.TButton",
            background=self.colors["card"],
            foreground=self.colors["text"],
            padding=(10, 5),
            bordercolor=self.colors["border"],
            focusthickness=0,
        )
        style.map(
            "Quiet.TButton",
            background=[("active", self.colors["surface_alt"]), ("pressed", self.colors["surface_alt"])],
        )
        style.configure(
            "TEntry",
            fieldbackground="#ffffff",
            foreground=self.colors["text"],
            bordercolor=self.colors["border"],
            lightcolor=self.colors["border"],
            darkcolor=self.colors["border"],
            padding=5,
        )
        style.map(
            "TEntry",
            bordercolor=[("focus", self.colors["accent"])],
            lightcolor=[("focus", self.colors["accent"])],
            darkcolor=[("focus", self.colors["accent"])],
        )
        style.configure(
            "Card.TCheckbutton",
            background=self.colors["card"],
            foreground=self.colors["text"],
        )
        style.map(
            "Card.TCheckbutton",
            background=[("active", self.colors["card"])],
            foreground=[("disabled", self.colors["muted"])],
        )
        style.configure(
            "Accent.Horizontal.TProgressbar",
            troughcolor=self.colors["accent_soft"],
            background=self.colors["accent"],
            bordercolor=self.colors["border"],
            lightcolor=self.colors["accent"],
            darkcolor=self.colors["accent"],
        )

    def _append_log(self, text):
        self.log_text.configure(state='normal')
        self.log_text.insert(tk.END, text)
        self.log_text.see(tk.END)
        self.log_text.configure(state='disabled')

    def _log(self, message: str):
        text = message if message.endswith("\n") else f"{message}\n"
        self._append_log(text)

    def _schedule_log(self, text: str):
        self.master.after(0, lambda chunk=text: self._append_log(chunk))

    def _update_progress_ui(self, status: str, current: int, total: int):
        total = max(total, 1)
        progress = max(0.0, min((current / total) * 100, 100))
        self.status_text.set(status)
        self.progress_value.set(progress)

    def _schedule_progress_update(self, status: str, current: int, total: int):
        self.master.after(0, lambda: self._update_progress_ui(status, current, total))

    def _set_running_state(self, is_running: bool):
        self.is_running = is_running
        widget_state = 'disabled' if is_running else 'normal'
        for widget in self.run_sensitive_widgets:
            widget.configure(state=widget_state)

    def _show_error(self, title: str, detail: str):
        self.status_text.set(title)
        self._log(f"[ERROR] {title}\n{detail}\n")
        messagebox.showerror("Error", f"{title}\n\n{detail}")

    def _open_generated_report(self, generated_files: list[str]):
        if not generated_files:
            return
        report_to_open = generated_files[0]
        self._log(f"[INFO] Opening report: {report_to_open}")
        open_with_default_app(report_to_open)

    def _run_process_worker(self, input_dir: str, output_dir: str, selected_brands: list[str]):
        stream = GuiLogStream(self._schedule_log)
        try:
            with redirect_stdout(stream), redirect_stderr(stream):
                generated_files = process_files(
                    input_directory=input_dir,
                    output_directory=output_dir,
                    selected_brands=selected_brands,
                    excel_options={
                        "use_excel_formulas": self.use_excel_formulas.get(),
                        "keep_raw_data_sheet": self.keep_raw_data_sheet.get(),
                    },
                    progress_callback=self._schedule_progress_update,
                )
                save_config(input_dir, output_dir)
            self.master.after(0, lambda: self._on_process_success(generated_files))
        except Exception:
            detail = traceback.format_exc()
            self._schedule_log(f"[ERROR] {detail}\n")
            self.master.after(0, lambda: self._on_process_error(detail))

    def _on_process_success(self, generated_files: list[str]):
        self._set_running_state(False)
        self._update_progress_ui("Report generation complete.", 1, 1)

        if generated_files:
            self._log(f"[INFO] Generated {len(generated_files)} report file(s).")
            if self.open_file_after_generation.get():
                try:
                    self._open_generated_report(generated_files)
                except Exception:
                    self._show_error("Failed to open generated report", traceback.format_exc())
                    return
            messagebox.showinfo("Success", f"Brand reports generated successfully.\nFiles created: {len(generated_files)}")
        else:
            self._log("[INFO] No report files were generated.")
            messagebox.showinfo("Completed", "Processing finished, but no reports were generated.")

    def _on_process_error(self, detail: str):
        self._set_running_state(False)
        self._update_progress_ui("Report generation failed.", 0, 1)
        self._show_error("Error generating reports", detail)

    def return_to_main_hub(self):
        if self.is_running:
            messagebox.showinfo("Busy", "A report run is still in progress.")
            return
        if callable(self.return_to_main):
            self.frame.destroy()
            self.return_to_main()
        else:
            self.master.destroy()

    def get_files(self):
        input_dir = self.input_dir.get()
        if not input_dir:
            messagebox.showerror("Error", "Please select an input directory first.")
            return
        try:
            self.status_text.set("Fetching Dutchie API inventory files...")
            self._log(f"[INFO] Fetching Dutchie API files into {input_dir}")
            result = subprocess.run(
                [sys.executable, "getCatalog.py", input_dir],
                check=True,
                capture_output=True,
                text=True,
            )
            if result.stdout:
                self._log(result.stdout)
            if result.stderr:
                self._log(result.stderr)
            self.status_text.set("Dutchie API files fetched successfully.")
            messagebox.showinfo("Success", "Files successfully fetched from the Dutchie API.")
        except subprocess.CalledProcessError:
            self._show_error("Failed to get files", traceback.format_exc())

    def clear_files(self):
        input_dir = self.input_dir.get()
        if not input_dir:
            messagebox.showerror("Error", "Please select an input directory first.")
            return

        answer = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete all CSV files in the input directory?")
        if not answer:
            return

        count = 0
        for filename in os.listdir(input_dir):
            if filename.endswith('.csv'):
                file_path = os.path.join(input_dir, filename)
                try:
                    os.remove(file_path)
                    count += 1
                except Exception as e:
                    print(f"Error deleting {filename}: {e}")
        self.status_text.set(f"Deleted {count} CSV file(s).")
        self._log(f"[INFO] Deleted {count} CSV file(s) from {input_dir}")
        messagebox.showinfo("Files Deleted", f"Deleted {count} CSV files from {input_dir}.")

    def browse_input(self):
        directory = filedialog.askdirectory()
        if directory:
            self.input_dir.set(directory)
            self._log(f"[INFO] Selected input directory: {directory}")

    def browse_output(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_dir.set(directory)
            self._log(f"[INFO] Selected output directory: {directory}")

    def load_brands(self):
        if not self.input_dir.get():
            messagebox.showerror("Error", "Please select an input directory first.")
            return
        try:
            self.status_text.set("Loading brands...")
            brands = get_all_brands(self.input_dir.get())
            self.brand_listbox.delete(0, tk.END)
            self.brand_listbox.configure(state='normal')
            if not brands:
                self.brand_listbox.insert(tk.END, "No brands found. You can still run the report.")
                self._log("[INFO] No brands found in the selected input directory.")
            else:
                for b in brands:
                    self.brand_listbox.insert(tk.END, b)
                self._log(f"[INFO] Loaded {len(brands)} brand(s).")
            self.status_text.set("Brands loaded.")
        except Exception:
            self._show_error("Failed to load brands", traceback.format_exc())

    def select_all_brands(self):
        if self.brand_listbox['state'] == 'normal':
            if self.brand_listbox.size() > 0:
                first_item = self.brand_listbox.get(0)
                if "No brands found" in first_item:
                    messagebox.showinfo("Info", "No actual brands available to select.")
                else:
                    self.brand_listbox.select_set(0, tk.END)
                    self._log("[INFO] Selected all loaded brands.")
            else:
                messagebox.showinfo("Info", "No brands available to select.")
        else:
            messagebox.showinfo("Info", "No brands to select.")

    def run_process(self):
        input_dir = self.input_dir.get()
        output_dir = self.output_dir.get()
        if not input_dir or not output_dir:
            messagebox.showerror("Error", "Please select both input and output directories.")
            return

        # Collect selected brands
        selected_indices = self.brand_listbox.curselection()
        selected_brands = [self.brand_listbox.get(i) for i in selected_indices
                           if "No brands found" not in self.brand_listbox.get(i)]
        self._set_running_state(True)
        self._update_progress_ui("Preparing report generation...", 0, 1)
        self._log(
            "[INFO] Starting report generation "
            f"(dynamic_mode={self.use_excel_formulas.get()}, "
            f"keep_raw_data={self.keep_raw_data_sheet.get()}, "
            f"selected_brands={len(selected_brands) if selected_brands else 'all'})."
        )
        worker = threading.Thread(
            target=self._run_process_worker,
            args=(input_dir, output_dir, selected_brands),
            daemon=True,
        )
        worker.start()


class MainHub:
    """Main hub with multiple app choices."""
    def __init__(self, master):
        self.master = master
        self.master.title("Main Hub - Multiple Apps")
        frame = ttk.Frame(master, padding=20)
        self.frame = frame
        frame.pack(fill='both', expand=True)

        ttk.Label(frame, text="Main Hub", style="HubTitle.TLabel").pack(pady=10)

        style = ttk.Style(master)
        style.configure("HubTitle.TLabel", font=("TkDefaultFont", 16, "bold"))

        ttk.Button(frame, text="Brand Inventory Report", command=self.open_brand_inventory).pack(pady=5)
        ttk.Button(frame, text="Sales Area (Placeholder)", command=self.sales_area).pack(pady=5)
        ttk.Button(frame, text="Another Feature (Placeholder)", command=self.another_feature).pack(pady=5)
        ttk.Button(frame, text="Exit", command=self.exit_app).pack(pady=5)

    def open_brand_inventory(self):
        self.frame.destroy()
        BrandInventoryApp(self.master, self.return_to_main)

    def sales_area(self):
        messagebox.showinfo("Info", "Sales Area is not implemented yet.")

    def another_feature(self):
        messagebox.showinfo("Info", "Another feature is not implemented yet.")

    def exit_app(self):
        self.master.quit()

    def return_to_main(self):
        MainHub(self.master)


if __name__ == "__main__":
    root = tk.Tk()
    BrandInventoryApp(root)
    root.mainloop()
