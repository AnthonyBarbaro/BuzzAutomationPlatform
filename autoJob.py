#!/usr/bin/env python3

import argparse
import os
import re
import subprocess
import sys
import time
import traceback
import datetime
import calendar
from datetime import date, timedelta, datetime as dt
from pathlib import Path
from typing import Any

import pandas as pd

from dutchie_api_reports import (
    DutchieAPIError,
    canonical_env_map,
    create_session,
    local_date_range_to_utc_strings,
    request_json,
    resolve_integrator_key,
    resolve_store_keys,
)

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_API_ENV_FILE = BASE_DIR / ".env"
SALES_API_MAX_WINDOW_DAYS = 30
AUTOJOB_STORES = [
    ("MV", "Buzz Cannabis - Mission Valley"),
    ("LM", "Buzz Cannabis-La Mesa"),
    ("SV", "Buzz Cannabis - SORRENTO VALLEY"),
    ("LG", "Buzz Cannabis - Lemon Grove"),
    ("NC", "Buzz Cannabis (National City)"),
    ("WP", "Buzz Cannabis Wildomar Palomar"),
]
DEALS_EXPORT_COLUMNS = [
    "Order ID",
    "Order Time",
    "Budtender Name",
    "Customer Name",
    "Customer Type",
    "Vendor Name",
    "Product Name",
    "Category",
    "Package ID",
    "Batch ID",
    "External Package ID",
    "Total Inventory Sold",
    "Unit Weight Sold",
    "Total Weight Sold",
    "Gross Sales",
    "Inventory Cost",
    "Discounted Amount",
    "Loyalty as Discount",
    "Net Sales",
    "Return Date",
    "UPC GTIN (Canada)",
    "Provincial SKU (Canada)",
    "Producer",
    "Order Profit",
]
DEALS_EXPORT_NUMERIC_COLUMNS = [
    "Total Inventory Sold",
    "Unit Weight Sold",
    "Total Weight Sold",
    "Gross Sales",
    "Inventory Cost",
    "Discounted Amount",
    "Loyalty as Discount",
    "Net Sales",
    "Order Profit",
]


def _first_nonempty(*values: Any) -> Any:
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            if value.strip():
                return value.strip()
            continue
        return value
    return ""


def _to_float(value: Any) -> float:
    try:
        if value in (None, ""):
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def _coerce_report_day(value: date | dt | str) -> date:
    if isinstance(value, dt):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return dt.fromisoformat(value).date()
    raise TypeError(f"Unsupported date value: {value!r}")


def _product_lookup_by_id(products_payload: Any) -> dict[int, dict[str, Any]]:
    lookup: dict[int, dict[str, Any]] = {}
    if not isinstance(products_payload, list):
        return lookup

    for row in products_payload:
        if not isinstance(row, dict):
            continue
        try:
            product_id = int(row.get("productId"))
        except Exception:
            continue
        lookup[product_id] = row
    return lookup


def _iter_sales_api_chunks(start_day: date, end_day: date, max_days: int = SALES_API_MAX_WINDOW_DAYS) -> list[tuple[date, date]]:
    chunks: list[tuple[date, date]] = []
    if end_day < start_day:
        return chunks

    window_days = max(1, int(max_days))
    chunk_start = start_day
    while chunk_start <= end_day:
        chunk_end = min(chunk_start + timedelta(days=window_days - 1), end_day)
        chunks.append((chunk_start, chunk_end))
        chunk_start = chunk_end + timedelta(days=1)

    return chunks


def _normalize_sales_api_export_rows(transactions_payload: Any, products_payload: Any) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    product_lookup = _product_lookup_by_id(products_payload)

    for tx in transactions_payload or []:
        if not isinstance(tx, dict):
            continue

        tx_id = str(_first_nonempty(tx.get("transactionId"), tx.get("globalId"), tx.get("referenceId"), ""))
        tx_time = pd.to_datetime(
            _first_nonempty(
                tx.get("transactionDateLocalTime"),
                tx.get("transactionDate"),
                tx.get("lastModifiedDateUTC"),
            ),
            errors="coerce",
        )
        budtender = str(_first_nonempty(tx.get("completedByUser"), tx.get("terminalName"), ""))
        customer_name = str(
            _first_nonempty(
                tx.get("customerName"),
                tx.get("customerFullName"),
                tx.get("customer"),
                "",
            )
        )
        customer_type = str(
            _first_nonempty(
                tx.get("customerTypeName"),
                tx.get("customerType"),
                tx.get("customerTypeId"),
                "",
            )
        )
        tx_is_return = bool(tx.get("isReturn"))

        for item in tx.get("items") or []:
            if not isinstance(item, dict):
                continue

            try:
                product_key = int(item.get("productId") or 0)
            except Exception:
                product_key = 0
            product_info = product_lookup.get(product_key, {})

            product_name = str(
                _first_nonempty(
                    product_info.get("productName"),
                    product_info.get("internalName"),
                    product_info.get("alternateName"),
                    f"Unknown Product {item.get('productId')}",
                )
            )
            category = str(_first_nonempty(product_info.get("category"), product_info.get("masterCategory"), "Unknown"))
            quantity = _to_float(item.get("quantity"))
            unit_weight = _to_float(item.get("unitWeight"))
            gross_sales = _to_float(item.get("totalPrice"))
            discount_amount = _to_float(item.get("totalDiscount"))
            loyalty_discount = _to_float(item.get("loyaltyAsDiscount"))
            net_sales = gross_sales - discount_amount
            unit_cost = _to_float(_first_nonempty(item.get("unitCost"), product_info.get("unitCost")))
            inventory_cost = unit_cost * quantity

            is_return = bool(item.get("isReturned")) or tx_is_return
            sign = -1.0 if is_return else 1.0

            quantity = abs(quantity) * sign
            gross_sales = abs(gross_sales) * sign
            discount_amount = abs(discount_amount) * sign
            loyalty_discount = abs(loyalty_discount) * sign
            net_sales = abs(net_sales) * sign
            inventory_cost = abs(inventory_cost) * sign
            total_weight = abs(quantity) * unit_weight * (1.0 if sign >= 0 else -1.0)
            order_profit = net_sales - inventory_cost
            sku = str(_first_nonempty(product_info.get("sku"), ""))
            upc = str(
                _first_nonempty(
                    product_info.get("upc"),
                    product_info.get("gtin"),
                    product_info.get("barcode"),
                    "",
                )
            )

            rows.append(
                {
                    "Order ID": tx_id,
                    "Order Time": tx_time,
                    "Budtender Name": budtender,
                    "Customer Name": customer_name,
                    "Customer Type": customer_type,
                    "Vendor Name": str(
                        _first_nonempty(
                            item.get("vendor"),
                            product_info.get("vendorName"),
                            product_info.get("producerName"),
                            "",
                        )
                    ),
                    "Product Name": product_name,
                    "Category": category,
                    "Package ID": str(_first_nonempty(item.get("packageId"), "")),
                    "Batch ID": str(_first_nonempty(item.get("batchName"), item.get("batchId"), "")),
                    "External Package ID": str(
                        _first_nonempty(item.get("sourcePackageId"), item.get("externalPackageId"), item.get("packageId"), "")
                    ),
                    "Total Inventory Sold": quantity,
                    "Unit Weight Sold": unit_weight,
                    "Total Weight Sold": total_weight,
                    "Gross Sales": gross_sales,
                    "Inventory Cost": inventory_cost,
                    "Discounted Amount": discount_amount,
                    "Loyalty as Discount": loyalty_discount,
                    "Net Sales": net_sales,
                    "Return Date": _first_nonempty(item.get("returnDate"), tx_time if is_return else None),
                    "UPC GTIN (Canada)": upc,
                    "Provincial SKU (Canada)": sku,
                    "Producer": str(_first_nonempty(product_info.get("producerName"), item.get("vendor"), "")),
                    "Order Profit": order_profit,
                }
            )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return pd.DataFrame(columns=DEALS_EXPORT_COLUMNS)

    for column in ("Order Time", "Return Date"):
        frame[column] = pd.to_datetime(frame[column], errors="coerce")
    for column in DEALS_EXPORT_NUMERIC_COLUMNS:
        frame[column] = pd.to_numeric(frame[column], errors="coerce").fillna(0.0)

    frame = frame.reindex(columns=DEALS_EXPORT_COLUMNS)
    return frame.sort_values(by=["Order Time", "Order ID", "Product Name"], na_position="last").reset_index(drop=True)


def _write_deals_compatible_sales_export(frame: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    export_frame = frame.reindex(columns=DEALS_EXPORT_COLUMNS)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        export_frame.to_excel(writer, index=False, startrow=4)


def run_sales_report_api(start_date: date | dt | str, end_date: date | dt | str, env_file: str | os.PathLike[str] = DEFAULT_API_ENV_FILE) -> None:
    start_day = _coerce_report_day(start_date)
    end_day = _coerce_report_day(end_date)
    if end_day < start_day:
        raise ValueError("end_date cannot be earlier than start_date")

    env_map = canonical_env_map(str(env_file))
    store_codes = [code for code, _store_name in AUTOJOB_STORES]
    store_keys = resolve_store_keys(env_map, store_codes)
    integrator_key = resolve_integrator_key(env_map)
    missing_store_codes = [code for code in store_codes if code not in store_keys]
    if missing_store_codes:
        missing = ", ".join(missing_store_codes)
        raise RuntimeError(
            "Missing Dutchie API location key(s) for: "
            f"{missing}. Add them to {env_file} using names like DUTCHIE_API_KEY_MV or mv."
        )

    files_dir = BASE_DIR / "files"
    files_dir.mkdir(parents=True, exist_ok=True)
    chunks = _iter_sales_api_chunks(start_day, end_day)
    failed_stores: list[str] = []

    for store_code, store_name in AUTOJOB_STORES:
        try:
            print(f"[API] Pulling sales for {store_name} ({store_code})")
            session = create_session(store_keys[store_code], integrator_key)
            products_payload = request_json(session, "/reporting/products")
            transactions_payload: list[dict[str, Any]] = []

            for idx, (chunk_start, chunk_end) in enumerate(chunks, start=1):
                from_utc, to_utc = local_date_range_to_utc_strings(
                    chunk_start.isoformat(),
                    chunk_end.isoformat(),
                )
                sales_params = {
                    "FromDateUTC": from_utc,
                    "ToDateUTC": to_utc,
                    "IncludeDetail": True,
                    "IncludeTaxes": True,
                    "IncludeOrderIds": True,
                    "IncludeFeesAndDonations": True,
                }
                print(f"[API] {store_code} chunk {idx}/{len(chunks)}: {chunk_start.isoformat()} -> {chunk_end.isoformat()}")
                payload = request_json(session, "/reporting/transactions", params=sales_params)
                if isinstance(payload, list) and payload:
                    transactions_payload.extend(item for item in payload if isinstance(item, dict))

            export_frame = _normalize_sales_api_export_rows(transactions_payload, products_payload)
            output_path = files_dir / f"sales{store_code}.xlsx"
            _write_deals_compatible_sales_export(export_frame, output_path)
            print(f"[API] Saved {store_code}: {len(export_frame)} row(s) -> {output_path}")
        except (DutchieAPIError, ValueError, OSError) as exc:
            print(f"[WARN] API sales export failed for {store_name} ({store_code}): {exc}")
            failed_stores.append(store_code)

    if failed_stores:
        raise RuntimeError(f"API export failed for store(s): {', '.join(failed_stores)}")


def run_sales_report_browser(start_date: date | dt | str, end_date: date | dt | str) -> None:
    from getSalesReport import run_sales_report as browser_run_sales_report

    browser_run_sales_report(start_date, end_date)

##############################################################################
# 1) LOGIC FOR LAST MONDAY TO SUNDAY
##############################################################################
def get_last_monday_sunday():
    """
    Returns (start_date, end_date) as Python date objects, representing
    last Monday through Sunday.

    Example: if today is Monday 2025-01-20, 
    this returns Monday 2025-01-13 and Sunday 2025-01-19.
    """
    today = date.today()
    # Monday of THIS current week:
    monday_this_week = today - timedelta(days=today.weekday())
    # Last Monday is 7 days before the Monday of this week
    last_monday = monday_this_week - timedelta(days=7)
    # Last Sunday is last_monday + 6 days
    last_sunday = last_monday + timedelta(days=6)
    return last_monday, last_sunday


##############################################################################
# 2) GETCATALOG LOGIC
##############################################################################
def run_get_catalog():
    """
    Calls getCatalog.py with a simple subprocess. 
    Make sure getCatalog.py is in the same directory or specify the full path.
    """
    print("\n===== Running getCatalog.py to download Catalog files... =====\n")
    script_path = BASE_DIR / "getCatalog.py"
    try:
        subprocess.check_call([sys.executable, str(script_path)], cwd=BASE_DIR)
    except subprocess.CalledProcessError as e:
        print(f"[ERROR] getCatalog.py failed: {e}")
    except FileNotFoundError:
        print("[ERROR] getCatalog.py not found. Please check the script name/path.")


##############################################################################
# 3) GETSALESREPORT LOGIC (HEADLESS, NO GUI)
#    We replicate the essential parts of your getSalesReport.py but skip GUI.
##############################################################################

#in salesReport.py


##############################################################################
# 4) RUN deals.py (Brand-Level Deals Report)
##############################################################################

#in deals.py

##############################################################################
# 5) RUN BRAND_INVENTORY.PY FOR 'Hashish' ONLY
##############################################################################
def run_brand_inventory_hashish():
    """
    We'll replicate a minimal version of brand_inventory.py logic,
    forcing brand='Hashish' only.
    We'll assume we want to parse the 'files' directory for new CSVs,
    output to 'done', and only keep lines for brand 'Hashish'.
    Then we apply openpyxl formatting (freeze panes, column widths, row height).
    """
    print("\n===== Running brand_inventory.py logic ONLY for brand='Hashish'... =====\n")

    import pandas as pd
    import re
    from datetime import datetime as dt
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    
    input_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "files")
    output_directory = os.path.join(os.path.dirname(os.path.abspath(__file__)), "done")
    os.makedirs(output_directory, exist_ok=True)

    def is_empty_or_numbers(val):
        if not isinstance(val, str):
            return True
        val_str = val.strip()
        return val_str == "" or val_str.isdigit()

    def extract_strain_type(product_name: str):
        if not isinstance(product_name, str):
            return ""
        name = " " + product_name.upper() + " "
        if re.search(r'\bS\b', name):
            return 'S'
        if re.search(r'\bH\b', name):
            return 'H'
        if re.search(r'\bI\b', name):
            return 'I'
        return ""

    def extract_product_details(product_name: str):
        if not isinstance(product_name, str):
            return "", ""
        name_upper = product_name.upper()
        weight_match = re.search(r'(\d+(\.\d+)?)G', name_upper)
        weight = weight_match.group(0) if weight_match else ""
        sub_type = ""
        if " HH " in f" {name_upper} ":
            sub_type = "HH"
        elif " IN " in f" {name_upper} ":
            sub_type = "IN"
        return weight, sub_type

    INPUT_COLUMNS = ['Available', 'Product', 'Category', 'Brand']

    for filename in os.listdir(input_directory):
        if filename.lower().endswith('.csv'):
            file_path = os.path.join(input_directory, filename)
            try:
                df = pd.read_csv(file_path)
            except Exception as e:
                print(f"[ERROR] reading CSV {filename}: {e}")
                continue

            # Filter to required columns
            use_cols = [c for c in INPUT_COLUMNS if c in df.columns]
            if not use_cols:
                continue
            df = df[use_cols]

            # Only brand=Hashish
            if 'Brand' in df.columns:
                df = df[df['Brand'] == 'Hashish']

            # Separate available vs. unavailable
            if 'Available' not in df.columns:
                continue
            unavailable_data = df[df['Available'] == 0]
            available_data   = df[df['Available'] != 0]

            # Parse product columns for the 'available' subset
            if not available_data.empty and 'Product' in available_data.columns:
                available_data['Strain_Type'] = available_data['Product'].apply(extract_strain_type)
                available_data[['Product_Weight','Product_SubType']] = available_data['Product'].apply(
                    lambda x: pd.Series(extract_product_details(x))
                )
                # Remove rows with empty or numeric product name
                available_data = available_data[~available_data['Product'].apply(is_empty_or_numbers)]

            # Sort by Category, Strain_Type, Product_Weight, Product_SubType, and Product
            sort_cols = []
            if 'Category' in available_data.columns:
                sort_cols.append('Category')
            sort_cols += ['Strain_Type','Product_Weight','Product_SubType']
            if 'Product' in available_data.columns:
                sort_cols.append('Product')
            available_data.sort_values(by=sort_cols, inplace=True, na_position='last')

            # Prepare output path
            base_name = os.path.splitext(filename)[0]  # e.g. "GreenHalo"
            today_str = dt.now().strftime("%m-%d-%Y")
            out_subdir = os.path.join(output_directory, base_name)
            os.makedirs(out_subdir, exist_ok=True)

            # --- Construct final Excel filename with "Hashish" + "_" + <inputfile base name> ---
            # e.g. "Hashish_GreenHalo.xlsx"
            out_file = os.path.join(out_subdir, f"Hashish_{base_name}.xlsx")

            # Write to Excel using pandas
            with pd.ExcelWriter(out_file) as writer:
                available_data.to_excel(writer, index=False, sheet_name='Available')
                if not unavailable_data.empty:
                    unavailable_data.to_excel(writer, index=False, sheet_name='Unavailable')

            # Apply formatting with openpyxl
            workbook = load_workbook(out_file)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Freeze the first row
                sheet.freeze_panes = "A2"

                # Auto-adjust column widths
                for column in sheet.columns:
                    max_length = max(len(str(cell.value)) if cell.value is not None else 0 
                                     for cell in column)
                    sheet.column_dimensions[column[0].column_letter].width = max_length + 2

                # Set a default row height
                for row in sheet.iter_rows():
                    sheet.row_dimensions[row[0].row].height = 17

                # Make the first row bold & center-aligned
                for cell in sheet["1:1"]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

            workbook.save(out_file)
            print(f"Hashish brand inventory saved & formatted -> {out_file}")


##############################################################################
# 6) GOOGLE DRIVE UPLOADER
##############################################################################
def run_drive_upload():
    """
    Upload brand_reports/*.xlsx + any done/**/*Hashish_*.xlsx to 
    Google Drive folder "2025_Kickback -> <week range>", 
    writing all links into links.txt
    """
    print("\n===== Running googleDriveUploader logic... =====\n")
    import google.auth
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    from google.oauth2.credentials import Credentials

    SCOPES = ['https://www.googleapis.com/auth/drive.file']
    LINKS_FILE = "links.txt"
    PARENT_FOLDER_NAME = "2026_Kickback"
    REPORTS_FOLDER = "brand_reports"

    def authenticate_drive_api():
        creds = None
        token_file = "token.json"
        if os.path.exists(token_file):
            creds = Credentials.from_authorized_user_file(token_file, SCOPES)
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(google.auth.transport.requests.Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file("credentials.json", SCOPES)
                creds = flow.run_local_server(port=0)
            with open(token_file, "w") as token:
                token.write(creds.to_json())
        return build("drive","v3", credentials=creds)

    def get_week_range_str():
        lm, ls = get_last_monday_sunday()
        return f"{lm.strftime('%m-%d')} to {ls.strftime('%m-%d')}"


    def find_or_create_folder(service, folder_name, parent_id=None):
        query = f"mimeType='application/vnd.google-apps.folder' and name='{folder_name}'"
        if parent_id:
            query += f" and '{parent_id}' in parents"
        resp = service.files().list(q=query, spaces='drive', fields='files(id,name)').execute()
        items = resp.get('files', [])
        if items:
            return items[0]['id']
        else:
            meta = {
                "name": folder_name,
                "mimeType": "application/vnd.google-apps.folder"
            }
            if parent_id:
                meta["parents"] = [parent_id]
            f = service.files().create(body=meta, fields="id").execute()
            return f["id"]

    def upload_file(service, path, parent_id):
        fname = os.path.basename(path)
        body = {
            "name": fname,
            "parents": [parent_id]
        }
        media = MediaFileUpload(path, resumable=True)
        f = service.files().create(body=body, media_body=media, fields="id").execute()
        return f["id"]

    def make_public(service, file_id):
        try:
            perm = {"type":"anyone","role":"reader"}
            service.permissions().create(fileId=file_id, body=perm).execute()
            info = service.files().get(fileId=file_id, fields="webViewLink").execute()
            return info.get("webViewLink")
        except:
            return None

    service = authenticate_drive_api()
    parent_id = find_or_create_folder(service, PARENT_FOLDER_NAME, None)

    week_range = get_week_range_str()
    week_folder_id = find_or_create_folder(service, week_range, parent_id)

    with open(LINKS_FILE,"w") as lf:
        # 1) Upload brand_reports
        if os.path.isdir(REPORTS_FOLDER):
            for fname in os.listdir(REPORTS_FOLDER):
                if fname.endswith(".xlsx"):
                    full_path = os.path.join(REPORTS_FOLDER,fname)
                    # Upload to the *same* week folder (no sub-subfolders)
                    file_id = upload_file(service, full_path, week_folder_id)
                    link = make_public(service, file_id)
                    if link:
                        lf.write(f"{fname}: {link}\n")
                        print(f"Uploaded {fname} => {link}")

        # 2) Also upload done/**/*Hashish_*.xlsx
        done_dir = "done"
        if os.path.isdir(done_dir):
            for root, dirs, files in os.walk(done_dir):
                for f in files:
                    if f.endswith(".xlsx") and "Hashish_" in f:
                        full_path = os.path.join(root,f)
                        file_id = upload_file(service, full_path, week_folder_id)
                        link = make_public(service, file_id)
                        if link:
                            lf.write(f"{f}: {link}\n")
                            print(f"Uploaded {f} => {link}")

    print("All files uploaded. Links stored in links.txt.")


##############################################################################
# 7) EMAIL THE LINKS.TXT + HASHISH BRAND REPORT (Optional)
##############################################################################
def send_email_with_gmail(subject, body, recipients, attachments=None):
    """
    Sends an email (plain text) via Gmail API with optional attachments.
    """
    print("\n===== Sending Email via Gmail API... =====\n")
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.utils import formatdate
    from email import encoders
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']

    creds = None
    gmail_token = "token_gmail.json"
    if os.path.exists(gmail_token):
        creds = Credentials.from_authorized_user_file(gmail_token, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(gmail_token,"w") as t:
            t.write(creds.to_json())

    service = build('gmail','v1', credentials=creds)

    if isinstance(recipients, str):
        recipients = [recipients]

    msg = MIMEMultipart()
    msg['From'] = "me"
    msg['To'] = ", ".join(recipients)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(body, 'plain'))

    if attachments:
        for path in attachments:
            if not os.path.isfile(path):
                continue
            fn = os.path.basename(path)
            with open(path,"rb") as f:
                part = MIMEBase("application","octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{fn}"')
            msg.attach(part)

    raw_msg = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    send_msg = {'raw': raw_msg}
    try:
        sent = service.users().messages().send(userId='me', body=send_msg).execute()
        print(f"Email sent! ID: {sent['id']}")
    except Exception as e:
        print("[ERROR] Could not send Gmail:", e)
def send_email_with_gmail_html(subject, html_body, recipients, attachments=None):
    """
    Sends an HTML email via Gmail API with optional attachments.
    """
    import base64
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email.mime.text import MIMEText
    from email.utils import formatdate
    from email import encoders
    import google.auth.transport.requests
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    GMAIL_SCOPES = ['https://www.googleapis.com/auth/gmail.send']
    creds = None
    gmail_token = "token_gmail.json"

    if os.path.exists(gmail_token):
        creds = Credentials.from_authorized_user_file(gmail_token, GMAIL_SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(google.auth.transport.requests.Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file("credentials.json", GMAIL_SCOPES)
            creds = flow.run_local_server(port=0)
        with open(gmail_token, "w") as f:
            f.write(creds.to_json())

    service = build('gmail', 'v1', credentials=creds)

    if isinstance(recipients, str):
        recipients = [recipients]

    # Create a MIMEMultipart message for HTML
    msg = MIMEMultipart('alternative')
    msg['From'] = "me"
    msg['To'] = ", ".join(recipients)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    # Attach the HTML body
    part_html = MIMEText(html_body, 'html')
    msg.attach(part_html)

    # Optionally attach files
    if attachments:
        for file_path in attachments:
            if not os.path.isfile(file_path):
                continue
            filename = os.path.basename(file_path)
            with open(file_path, "rb") as fp:
                file_data = fp.read()
            part = MIMEBase("application", "octet-stream")
            part.set_payload(file_data)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="{filename}"')
            msg.attach(part)

    raw_message = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    send_req = {'raw': raw_message}

    try:
        sent = service.users().messages().send(userId='me', body=send_req).execute()
        print(f"HTML Email sent! ID: {sent['id']}")
    except Exception as e:
        print("[ERROR] Could not send HTML email:", e)

##############################################################################
# MAIN: ORCHESTRATE ALL STEPS
##############################################################################
def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Run the weekly Buzz automation job.")
    parser.add_argument(
        "--sales-source",
        choices=("api", "browser"),
        default="api",
        help="Where autoJob should pull sales data from. Default: api",
    )
    parser.add_argument(
        "--env-file",
        default=str(DEFAULT_API_ENV_FILE),
        help=f"Path to the Dutchie API .env file when using --sales-source api. Default: {DEFAULT_API_ENV_FILE}",
    )
    return parser


def main(argv: list[str] | None = None):
    args = build_parser().parse_args(argv)
    print("===== Starting autoJob.py =====")

    last_monday, last_sunday = get_last_monday_sunday()
    date_range_str = f"{last_monday} to {last_sunday}"
    print(f"Processing for last week range: {date_range_str}")

    # 1) Clean up files directory
    files_dir = Path("files")
    if files_dir.exists() and files_dir.is_dir():
        for file in files_dir.iterdir():
            try:
                if file.is_file():
                    file.unlink()
                    print(f"[CLEANUP] Deleted {file}")
            except Exception as e:
                print(f"[ERROR] Could not delete {file}: {e}")
    # 2) Sales
    if args.sales_source == "api":
        print(f"[AUTOJOB] Pulling weekly sales from the Dutchie API using {args.env_file}")
        run_sales_report_api(last_monday, last_sunday, env_file=args.env_file)
    else:
        print("[AUTOJOB] Pulling weekly sales from the browser export flow.")
        run_sales_report_browser(last_monday, last_sunday)

    # 3) Deals
    subprocess.run([sys.executable, str(BASE_DIR / "deals.py")], cwd=BASE_DIR)
    time.sleep(2)

    # 5) Drive Upload (both brand_reports + done/Hashish)
    run_drive_upload()


    # 6a) Parse links.txt to separate "Hashish" lines from "non-Hashish" lines,
    #     and build HTML bullet lists for each group.
   

    links_file = "links.txt"
    hashish_links = []
    non_hashish_links = []

    if os.path.exists(links_file):
        with open(links_file, "r", encoding="utf-8") as lf:
            lines = lf.readlines()
        for line in lines:
            line = line.strip()
            # Typical format: "filename.xlsx: https://drive.google.com/..."
            if "Hashish_" in line:
                hashish_links.append(line)
            else:
                non_hashish_links.append(line)

    subprocess.run([sys.executable, str(BASE_DIR / "brandDEALSEmailer.py")], cwd=BASE_DIR)


    print("\n===== autoJob.py completed successfully. =====")


if __name__ == "__main__":
    main()
