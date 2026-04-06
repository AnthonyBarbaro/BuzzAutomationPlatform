import os
import re
from collections import OrderedDict
from functools import lru_cache
import math

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ORDER_REPORT_WINDOWS = (7, 14, 30)
ORDER_REPORT_PATTERN = re.compile(
    r"^inventory_order_(?P<days>7d|14d|30d)_(?P<store>[A-Za-z0-9]+)\.(?P<ext>xlsx|xls|csv)$",
    re.IGNORECASE,
)

BRAND_COLUMN_CANDIDATES = (
    "brand",
    "product brand",
    "brand name",
    "vendor",
    "vendor name",
    "producer",
    "supplier",
)
PRODUCT_COLUMN_CANDIDATES = (
    "product",
    "product name",
    "inventory name",
    "item name",
    "name",
    "sku",
)
SORT_COLUMN_CANDIDATES = ("category", "product", "product name", "sku")
LOCATION_COLUMN_CANDIDATES = ("location", "store", "dispensary")
SKU_COLUMN_CANDIDATES = ("product sku", "sku", "product id", "inventory id")
MASTER_CATEGORY_COLUMN_CANDIDATES = ("master category",)
QOH_COLUMN_CANDIDATES = ("quantity on hand", "on hand", "qty on hand", "available")
PRICE_COLUMN_CANDIDATES = ("price", "unit price", "retail price")
QTY_SOLD_COLUMN_CANDIDATES = ("quantity sold", "qty sold", "units sold")
AVG_DAILY_SALES_COLUMN_CANDIDATES = ("avg daily sales",)
SOLD_PER_DAY_COLUMN_CANDIDATES = ("sold per day", "units per day")
DAYS_REMAINING_COLUMN_CANDIDATES = ("days remaining", "days on hand")
DAYS_SINCE_RECEIVED_COLUMN_CANDIDATES = ("days since last received", "days since received")
LAST_WHOLESALE_COST_COLUMN_CANDIDATES = ("last wholesale cost", "wholesale cost", "cost")
LAST_ORDERED_QTY_COLUMN_CANDIDATES = ("last ordered quantity", "last order qty", "ordered quantity")
LAST_AUDIT_COLUMN_CANDIDATES = ("last audit",)
GRAMS_COLUMN_CANDIDATES = ("grams concentration", "grams", "weight")
VENDOR_COLUMN_CANDIDATES = ("vendor", "vendor name", "supplier")
STRAIN_COLUMN_CANDIDATES = ("strain",)
FLOWER_TYPE_COLUMN_CANDIDATES = ("flower type",)
CONCENTRATE_TYPE_COLUMN_CANDIDATES = ("concentrate type",)
UPC_COLUMN_CANDIDATES = ("upc/gtin", "upc", "gtin")
PROVINCIAL_SKU_COLUMN_CANDIDATES = ("provincial sku",)

ORDER_SHEET_PREFIX = "Order_"
ORDER_SUMMARY_TITLE = "Ordering Summary"
ORDER_DETAIL_TITLE = "Product Detail"
ORDER_SUMMARY_HEADER_FILL = PatternFill(
    start_color="D3D3D3",
    end_color="D3D3D3",
    fill_type="solid",
)
ORDER_SUMMARY_TITLE_FILL = PatternFill(
    start_color="153B34",
    end_color="153B34",
    fill_type="solid",
)
ORDER_GROUP_FILL = PatternFill(
    start_color="E6E6FA",
    end_color="E6E6FA",
    fill_type="solid",
)
ORDER_PRIORITY_COLORS = {
    "Urgent": "FECACA",
    "Reorder Now": "FED7AA",
    "Reorder Soon": "FEF08A",
    "Check PO": "BFDBFE",
    "Watch": "FDE68A",
    "Healthy": "BBF7D0",
    "No Recent Sales": "E5E7EB",
}
LOW_COST_PROMO_THRESHOLD = 1.01
PROMO_SAMPLE_PATTERN = re.compile(
    r"\b(sample|samples|promo|promos|promotional|display|tester)\b",
    re.IGNORECASE,
)
STRAIN_CODE_TOKENS = {"S", "H", "I", "HH", "IN"}
KNOWN_PACK_LABELS = {
    1.0: "1g",
    3.5: "3.5g",
    7.0: "7g",
    14.0: "14g",
    28.0: "28g",
}

STORE_NAME_TO_ABBR = {
    "Buzz Cannabis - Mission Valley": "MV",
    "Buzz Cannabis-La Mesa": "LM",
    "Buzz Cannabis - SORRENTO VALLEY": "SV",
    "Buzz Cannabis - Lemon Grove": "LG",
    "Buzz Cannabis (National City)": "NC",
    "Buzz Cannabis Wildomar Palomar": "WP",
}
STORE_ABBR_TO_NAME = {abbr.upper(): name for name, abbr in STORE_NAME_TO_ABBR.items()}


def normalize_text(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def canonical_store_code(value):
    if value is None:
        return ""

    raw = str(value).strip()
    if not raw:
        return ""

    if raw in STORE_NAME_TO_ABBR:
        return STORE_NAME_TO_ABBR[raw].upper()

    upper = raw.upper()
    if upper in STORE_ABBR_TO_NAME:
        return upper

    match = re.search(r"_([A-Z]{2,3})$", upper)
    if match:
        return match.group(1)

    return upper


def order_report_filename(store_name_or_code, window_days, extension=".xlsx"):
    store_code = canonical_store_code(store_name_or_code)
    if not store_code:
        safe_store = re.sub(r"[^A-Za-z0-9]+", "_", str(store_name_or_code)).strip("_") or "UNK"
        store_code = safe_store.upper()

    ext = extension if extension.startswith(".") else f".{extension}"
    return f"inventory_order_{int(window_days)}d_{store_code}{ext.lower()}"


def extract_store_code_from_filename(name):
    base_name = os.path.splitext(os.path.basename(name))[0]
    parts = base_name.split("_")
    if len(parts) >= 2:
        return canonical_store_code(parts[-1])
    return ""


def find_matching_column(df, candidates):
    candidate_map = {normalize_text(col): col for col in df.columns}
    for candidate in candidates:
        match = candidate_map.get(normalize_text(candidate))
        if match:
            return match
    return None


def _score_header(columns):
    normalized = {normalize_text(col) for col in columns}
    score = 0
    if normalized & set(BRAND_COLUMN_CANDIDATES):
        score += 4
    if normalized & set(PRODUCT_COLUMN_CANDIDATES):
        score += 3
    if "category" in normalized:
        score += 1
    if len(normalized) >= 4:
        score += 1
    return score


@lru_cache(maxsize=64)
def _load_order_report_table_cached(path):
    ext = os.path.splitext(path)[1].lower()
    readers = []
    if ext in (".xlsx", ".xls"):
        readers = [("excel", header_idx) for header_idx in range(0, 7)]
    elif ext == ".csv":
        readers = [("csv", 0)]
    else:
        raise ValueError(f"Unsupported order report format: {path}")

    best_df = None
    best_score = -1

    for reader_type, header_idx in readers:
        try:
            if reader_type == "excel":
                df = pd.read_excel(path, header=header_idx)
            else:
                df = pd.read_csv(path)
        except Exception:
            continue

        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")
        if df.empty:
            continue

        score = _score_header(df.columns)
        if score > best_score:
            best_score = score
            best_df = df

    if best_df is None:
        if ext in (".xlsx", ".xls"):
            best_df = pd.read_excel(path)
        else:
            best_df = pd.read_csv(path)
        best_df = best_df.dropna(axis=1, how="all").dropna(axis=0, how="all")

    return best_df


def load_order_report_table(path):
    return _load_order_report_table_cached(os.path.abspath(path)).copy()


def discover_order_report_files(directory):
    report_files = OrderedDict((days, OrderedDict()) for days in ORDER_REPORT_WINDOWS)
    if not directory or not os.path.isdir(directory):
        return report_files

    for name in sorted(os.listdir(directory)):
        match = ORDER_REPORT_PATTERN.match(name)
        if not match:
            continue

        days = int(match.group("days")[:-1])
        store = canonical_store_code(match.group("store"))
        report_files.setdefault(days, OrderedDict())[store] = os.path.join(directory, name)

    return report_files


def summarize_order_report_files(directory):
    discovered = discover_order_report_files(directory)
    parts = []
    for days, files in discovered.items():
        if files:
            parts.append(f"{days}d ({len(files)} store{'s' if len(files) != 1 else ''})")
    return ", ".join(parts)


def _filter_brand_rows(df, brand_aliases):
    if df.empty or not brand_aliases:
        return pd.DataFrame(columns=df.columns)

    normalized_aliases = {normalize_text(alias) for alias in brand_aliases if normalize_text(alias)}
    if not normalized_aliases:
        return pd.DataFrame(columns=df.columns)

    brand_col = find_matching_column(df, BRAND_COLUMN_CANDIDATES)
    if brand_col:
        brand_series = df[brand_col].map(normalize_text)
        return df[brand_series.isin(normalized_aliases)].copy()

    product_col = find_matching_column(df, PRODUCT_COLUMN_CANDIDATES)
    if product_col:
        product_series = df[product_col].map(normalize_text)
        mask = product_series.apply(
            lambda value: any(alias and alias in value for alias in normalized_aliases)
        )
        return df[mask].copy()

    return pd.DataFrame(columns=df.columns)


def _sort_order_rows(df):
    sort_cols = []
    for col in df.columns:
        if normalize_text(col) in SORT_COLUMN_CANDIDATES and col not in sort_cols:
            sort_cols.append(col)

    if sort_cols:
        return df.sort_values(by=sort_cols, na_position="last")
    return df


def _to_numeric_series(df, candidates):
    col = find_matching_column(df, candidates)
    if not col:
        return pd.Series([pd.NA] * len(df), index=df.index, dtype="object"), None
    return pd.to_numeric(df[col], errors="coerce"), col


def _ceil_non_negative(value):
    if pd.isna(value):
        return 0
    return max(0, int(math.ceil(float(value))))


def _priority_from_metrics(suggested_qty, days_remaining, sold_per_day, qty_sold, days_since_received):
    sold_per_day = 0.0 if pd.isna(sold_per_day) else float(sold_per_day)
    qty_sold = 0.0 if pd.isna(qty_sold) else float(qty_sold)
    suggested_qty = 0.0 if pd.isna(suggested_qty) else float(suggested_qty)
    days_remaining = None if pd.isna(days_remaining) else float(days_remaining)
    days_since_received = None if pd.isna(days_since_received) else float(days_since_received)

    if sold_per_day <= 0 and qty_sold <= 0:
        return "No Recent Sales"
    if suggested_qty <= 0 and days_remaining is not None and days_remaining > 30:
        return "Healthy"
    if days_remaining is not None and days_remaining <= 3:
        return "Urgent"
    if days_remaining is not None and days_remaining <= 7:
        return "Reorder Now"
    if days_remaining is not None and days_remaining <= 14:
        return "Reorder Soon"
    if suggested_qty > 0 and days_since_received is not None and days_since_received >= 14:
        return "Check PO"
    if suggested_qty > 0:
        return "Watch"
    return "Healthy"


def _priority_rank(priority):
    ranks = {
        "Urgent": 0,
        "Reorder Now": 1,
        "Reorder Soon": 2,
        "Check PO": 3,
        "Watch": 4,
        "Healthy": 5,
        "No Recent Sales": 6,
    }
    return ranks.get(priority, 99)


def _build_reorder_note(suggested_qty, days_remaining, sold_per_day, qty_sold, days_since_received):
    notes = []
    if not pd.isna(days_remaining):
        if float(days_remaining) <= 7:
            notes.append("low cover")
        elif float(days_remaining) <= 14:
            notes.append("tight cover")
    if not pd.isna(days_since_received) and float(days_since_received) >= 14:
        notes.append("not received recently")
    if not pd.isna(sold_per_day) and float(sold_per_day) > 0:
        notes.append(f"{float(sold_per_day):.2f}/day")
    elif not pd.isna(qty_sold) and float(qty_sold) > 0:
        notes.append(f"{float(qty_sold):.0f} sold")
    if not pd.isna(suggested_qty) and float(suggested_qty) > 0:
        notes.append(f"order {int(float(suggested_qty))}")
    return ", ".join(notes)


def _reorder_columns(df, ordered_names):
    normalized_map = {normalize_text(col): col for col in df.columns}
    ordered_cols = []
    for name in ordered_names:
        actual = normalized_map.get(normalize_text(name))
        if actual and actual not in ordered_cols:
            ordered_cols.append(actual)
    remaining = [col for col in df.columns if col not in ordered_cols]
    return df[ordered_cols + remaining]


def _drop_columns_by_candidates(df, candidate_groups):
    drop_cols = []
    for candidates in candidate_groups:
        col = find_matching_column(df, candidates)
        if col and col not in drop_cols:
            drop_cols.append(col)
    if not drop_cols:
        return df
    return df.drop(columns=drop_cols)


def _safe_display_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _sort_display_text(value):
    return _safe_display_text(value).lower()


def _summary_sort_cost(value):
    if pd.isna(value):
        return (1, float("inf"))
    return (0, float(value))


def _find_exact_or_matching_column(df, names):
    normalized_map = {normalize_text(col): col for col in df.columns}
    for name in names:
        actual = normalized_map.get(normalize_text(name))
        if actual:
            return actual
    return None


def is_order_sheet_name(name):
    return str(name or "").startswith(ORDER_SHEET_PREFIX)


def _is_sample_or_promo(value):
    return bool(PROMO_SAMPLE_PATTERN.search(_safe_display_text(value)))


def _drop_sample_and_promo_rows(df):
    if df.empty:
        return df
    product_col = find_matching_column(df, PRODUCT_COLUMN_CANDIDATES)
    if not product_col:
        return df
    return df[~df[product_col].map(_is_sample_or_promo)].copy()


def _drop_low_cost_rows(df, threshold=LOW_COST_PROMO_THRESHOLD):
    if df.empty:
        return df
    cost_col = find_matching_column(df, LAST_WHOLESALE_COST_COLUMN_CANDIDATES)
    if not cost_col:
        return df

    numeric_cost = pd.to_numeric(df[cost_col], errors="coerce")
    keep_mask = numeric_cost.isna() | (numeric_cost > float(threshold))
    return df[keep_mask].copy()


def _extract_pack_grams(product_name):
    product_text = _safe_display_text(product_name).upper()
    matches = re.findall(r"(\d+(?:\.\d+)?)\s*G\b", product_text)
    if not matches:
        return None

    try:
        grams = float(matches[0])
    except ValueError:
        return None
    return grams if grams > 0 else None


def _format_grams(grams):
    if grams is None or pd.isna(grams):
        return ""
    grams = float(grams)
    if grams.is_integer():
        return f"{int(grams)}g"
    return f"{grams:.1f}".rstrip("0").rstrip(".") + "g"


def _pack_label_from_metrics(category, grams):
    category_text = _safe_display_text(category)
    normalized_category = normalize_text(category_text)

    if normalized_category == "ounces" or (grams is not None and abs(float(grams) - 28.0) < 0.01):
        return "28g"
    if normalized_category == "halves" or (grams is not None and abs(float(grams) - 14.0) < 0.01):
        return "14g"
    if normalized_category == "quarters" or (grams is not None and abs(float(grams) - 7.0) < 0.01):
        return "7g"
    if normalized_category == "eighths" or (grams is not None and abs(float(grams) - 3.5) < 0.01):
        return "3.5g"
    if normalized_category == "prerolls" or (grams is not None and abs(float(grams) - 1.0) < 0.01):
        return "1g"
    if grams is not None:
        rounded = round(float(grams), 1)
        return KNOWN_PACK_LABELS.get(rounded, _format_grams(rounded))
    return ""


def _unit_display_name(category, grams, quantity):
    normalized_category = normalize_text(category)
    quantity = int(quantity or 0)

    if normalized_category == "ounces" or (grams is not None and abs(float(grams) - 28.0) < 0.01):
        return "ounce" if quantity == 1 else "ounces"
    if normalized_category == "halves" or (grams is not None and abs(float(grams) - 14.0) < 0.01):
        return "half" if quantity == 1 else "halves"
    if normalized_category == "quarters" or (grams is not None and abs(float(grams) - 7.0) < 0.01):
        return "quarter" if quantity == 1 else "quarters"
    if normalized_category == "eighths" or (grams is not None and abs(float(grams) - 3.5) < 0.01):
        return "eighth" if quantity == 1 else "eighths"
    if normalized_category == "prerolls":
        return "preroll" if quantity == 1 else "prerolls"
    return "unit" if quantity == 1 else "units"


def _format_need_summary(category, grams, units_needed):
    units_needed = int(units_needed or 0)
    if units_needed <= 0:
        return ""

    unit_name = _unit_display_name(category, grams, units_needed)
    if grams is None or pd.isna(grams):
        return f"Need {units_needed} {unit_name}"

    total_grams = float(grams) * units_needed
    return f"Need {units_needed} {unit_name} total ({_format_grams(total_grams)})"


def _clean_summary_product_name(product_name):
    product_text = _safe_display_text(product_name)
    if not product_text:
        return ""

    segments = [segment.strip() for segment in product_text.split("|") if segment.strip()]
    candidate = ""
    for segment in reversed(segments):
        segment_upper = segment.upper()
        if segment_upper in STRAIN_CODE_TOKENS:
            continue
        if re.fullmatch(r"\d+(?:\.\d+)?\s*G(?:\s*\(.+\))?", segment_upper):
            continue
        candidate = segment
        break

    if not candidate:
        candidate = segments[-1] if segments else product_text

    candidate = re.sub(r"\([^)]*\b(?:sample|samples|promo|promos)\b[^)]*\)", "", candidate, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", candidate).strip(" -|")


def prepare_reorder_sheet(df, window_days):
    if df.empty:
        return df

    work = df.copy()

    qty_on_hand, qoh_col = _to_numeric_series(work, QOH_COLUMN_CANDIDATES)
    qty_sold, qty_sold_col = _to_numeric_series(work, QTY_SOLD_COLUMN_CANDIDATES)
    sold_per_day, sold_per_day_col = _to_numeric_series(work, SOLD_PER_DAY_COLUMN_CANDIDATES)
    avg_daily_sales, _avg_sales_col = _to_numeric_series(work, AVG_DAILY_SALES_COLUMN_CANDIDATES)
    days_remaining, days_remaining_col = _to_numeric_series(work, DAYS_REMAINING_COLUMN_CANDIDATES)
    days_since_received, days_since_received_col = _to_numeric_series(work, DAYS_SINCE_RECEIVED_COLUMN_CANDIDATES)
    last_ordered_qty, _last_ordered_col = _to_numeric_series(work, LAST_ORDERED_QTY_COLUMN_CANDIDATES)

    target_qty_col = f"Target Qty ({window_days}d)"
    suggested_qty_col = f"Suggested Order Qty ({window_days}d)"
    priority_col = "Reorder Priority"
    note_col = "Reorder Notes"

    target_qty = sold_per_day.fillna(0).apply(lambda value: _ceil_non_negative(value * window_days))
    suggested_qty = (target_qty - qty_on_hand.fillna(0)).apply(_ceil_non_negative)

    work[target_qty_col] = target_qty
    work[suggested_qty_col] = suggested_qty
    work[priority_col] = [
        _priority_from_metrics(sugg, remain, spd, sold, dsr)
        for sugg, remain, spd, sold, dsr in zip(
            suggested_qty, days_remaining, sold_per_day, qty_sold, days_since_received
        )
    ]
    work[note_col] = [
        _build_reorder_note(sugg, remain, spd, sold, dsr)
        for sugg, remain, spd, sold, dsr in zip(
            suggested_qty, days_remaining, sold_per_day, qty_sold, days_since_received
        )
    ]

    work["_priority_rank"] = work[priority_col].map(_priority_rank)
    category_col = find_matching_column(work, ("category",))
    product_col = find_matching_column(work, PRODUCT_COLUMN_CANDIDATES)
    vendor_col = find_matching_column(work, VENDOR_COLUMN_CANDIDATES)
    work["_sort_category"] = (
        work[category_col].fillna("").astype(str).str.lower() if category_col else ""
    )
    work["_sort_product"] = (
        work[product_col].fillna("").astype(str).str.lower() if product_col else ""
    )
    work["_sort_vendor"] = (
        work[vendor_col].fillna("").astype(str).str.lower() if vendor_col else ""
    )
    work["_sort_days_remaining"] = days_remaining.fillna(999999)
    work["_sort_suggested_qty"] = suggested_qty.fillna(0)
    work["_sort_sold_per_day"] = sold_per_day.fillna(0)
    work["_sort_qty_sold"] = qty_sold.fillna(0)
    work["_sort_days_since_received"] = days_since_received.fillna(-1)

    work = work.sort_values(
        by=[
            "_sort_category",
            "_priority_rank",
            "_sort_days_remaining",
            "_sort_suggested_qty",
            "_sort_sold_per_day",
            "_sort_qty_sold",
            "_sort_days_since_received",
            "_sort_vendor",
            "_sort_product",
        ],
        ascending=[True, True, True, False, False, False, False, True, True],
        na_position="last",
    )

    work = _drop_columns_by_candidates(
        work,
        [
            LOCATION_COLUMN_CANDIDATES,
            BRAND_COLUMN_CANDIDATES,
            SKU_COLUMN_CANDIDATES,
            MASTER_CATEGORY_COLUMN_CANDIDATES,
            STRAIN_COLUMN_CANDIDATES,
            FLOWER_TYPE_COLUMN_CANDIDATES,
            GRAMS_COLUMN_CANDIDATES,
        ],
    )

    work = work.drop(
        columns=[
            "_sort_category",
            "_priority_rank",
            "_sort_days_remaining",
            "_sort_suggested_qty",
            "_sort_sold_per_day",
            "_sort_qty_sold",
            "_sort_days_since_received",
            "_sort_vendor",
            "_sort_product",
        ]
    )

    preferred_order = [
        "Reorder Priority",
        note_col,
        target_qty_col,
        suggested_qty_col,
        "Category",
        "Product Name",
        "Quantity on Hand",
        "Quantity Sold",
        "Sold Per Day",
        "Avg Daily Sales",
        "Days Remaining",
        "Days Since Last Received",
        "Last Ordered Quantity",
        "Last Wholesale Cost",
        "Price",
        "Vendor",
        "Concentrate Type",
        "UPC/GTIN",
        "Provincial SKU",
        "Last Audit",
    ]
    work = _reorder_columns(work, preferred_order)

    # Keep source numeric columns in their original names/casing; these locals
    # only exist to force column discovery during sheet preparation.
    _ = (
        qoh_col,
        qty_sold_col,
        sold_per_day_col,
        days_remaining_col,
        days_since_received_col,
        last_ordered_qty,
        avg_daily_sales,
        last_ordered_qty,
    )

    return work


def build_grouped_order_summary(detail_df, window_days):
    suggested_qty_col = f"Suggested Order Qty ({window_days}d)"
    summary_columns = [
        "Category",
        "Unit Cost",
        "Units Needed",
        "Total Quantity Sold",
    ]

    if detail_df.empty:
        return pd.DataFrame(columns=summary_columns)

    work = detail_df.copy()
    category_col = find_matching_column(work, ("category",))
    product_col = find_matching_column(work, PRODUCT_COLUMN_CANDIDATES)
    qty_sold_col = find_matching_column(work, QTY_SOLD_COLUMN_CANDIDATES)
    cost_col = find_matching_column(work, LAST_WHOLESALE_COST_COLUMN_CANDIDATES)
    suggested_col = _find_exact_or_matching_column(work, (suggested_qty_col,))

    if not suggested_col:
        return pd.DataFrame(columns=summary_columns)

    work["_display_category"] = work[category_col].map(_safe_display_text) if category_col else ""
    work["_product_name"] = work[product_col].map(_safe_display_text) if product_col else ""
    work["_is_sample_or_promo"] = work["_product_name"].map(_is_sample_or_promo)
    work["_unit_cost"] = (
        pd.to_numeric(work[cost_col], errors="coerce")
        if cost_col
        else pd.Series([pd.NA] * len(work), index=work.index, dtype="object")
    )
    work["_qty_sold"] = (
        pd.to_numeric(work[qty_sold_col], errors="coerce").fillna(0)
        if qty_sold_col
        else 0
    )
    work["_suggested_qty"] = pd.to_numeric(work[suggested_col], errors="coerce").fillna(0)

    # Keep the top block focused on what actually needs to be ordered.
    work = work[
        (work["_suggested_qty"] > 0)
        & (~work["_is_sample_or_promo"])
        & (work["_unit_cost"].isna() | (work["_unit_cost"] > LOW_COST_PROMO_THRESHOLD))
    ].copy()
    if work.empty:
        return pd.DataFrame(columns=summary_columns)

    work["_group_category"] = work["_display_category"].map(normalize_text)
    work["_group_cost"] = work["_unit_cost"].map(
        lambda value: "__NA__" if pd.isna(value) else round(float(value), 2)
    )

    summary_rows = []
    group_cols = ["_group_category", "_group_cost"]
    for _, group in work.groupby(group_cols, dropna=False, sort=False):
        category_display = next((value for value in group["_display_category"] if value), "")
        unit_cost_value = group["_unit_cost"].iloc[0]
        if not pd.isna(unit_cost_value):
            unit_cost_value = round(float(unit_cost_value), 2)
        total_qty_sold = int(round(float(group["_qty_sold"].sum()))) if not group.empty else 0
        units_needed = int(round(float(group["_suggested_qty"].sum()))) if not group.empty else 0

        summary_rows.append(
            {
                "Category": category_display,
                "Unit Cost": unit_cost_value,
                "Units Needed": units_needed,
                "Total Quantity Sold": total_qty_sold,
            }
        )

    summary_df = pd.DataFrame(summary_rows, columns=summary_columns)
    if summary_df.empty:
        return summary_df

    summary_df = summary_df.iloc[
        sorted(
            range(len(summary_df)),
            key=lambda idx: (
                _sort_display_text(summary_df.iloc[idx]["Category"]),
                _summary_sort_cost(summary_df.iloc[idx]["Unit Cost"]),
                -int(summary_df.iloc[idx]["Units Needed"] or 0),
            ),
        )
    ].reset_index(drop=True)

    return summary_df


def write_order_section_sheet(writer, sheet_name, section_payload):
    summary_df = (section_payload or {}).get("summary")
    detail_df = (section_payload or {}).get("detail")
    if summary_df is None:
        summary_df = pd.DataFrame()
    if detail_df is None:
        detail_df = pd.DataFrame()

    summary_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=1)
    detail_title_row = len(summary_df.index) + 4
    detail_df.to_excel(writer, index=False, sheet_name=sheet_name, startrow=detail_title_row)

    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=ORDER_SUMMARY_TITLE)
    ws.cell(row=detail_title_row, column=1, value=ORDER_DETAIL_TITLE)


def write_order_sections(writer, order_sections):
    for sheet_name, section_payload in order_sections.items():
        write_order_section_sheet(writer, sheet_name, section_payload)


def _style_title_row(ws, row_idx):
    max_col = max(1, ws.max_column)
    if max_col > 1:
        try:
            ws.unmerge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
        except Exception:
            pass
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=max_col)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = ORDER_SUMMARY_TITLE_FILL
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_header_row(ws, row_idx):
    for cell in ws[row_idx]:
        cell.font = Font(bold=True)
        cell.fill = ORDER_SUMMARY_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_fit_sheet_columns(ws):
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 3


def _group_detail_rows_by_category(ws, detail_header_row):
    header_map = {
        normalize_text(cell.value): idx
        for idx, cell in enumerate(ws[detail_header_row], start=1)
        if cell.value
    }
    category_index = header_map.get("category")
    if not category_index or detail_header_row >= ws.max_row:
        return

    rows_data = list(
        ws.iter_rows(min_row=detail_header_row + 1, max_row=ws.max_row, values_only=True)
    )
    if not rows_data:
        return

    row_num = detail_header_row + 1
    cat_list = []
    current_cat = None
    for row_vals in rows_data:
        cat_val = row_vals[category_index - 1]
        if cat_val != current_cat:
            cat_list.append((row_num, cat_val))
            current_cat = cat_val
        row_num += 1

    for pos, cat_val in reversed(cat_list):
        ws.insert_rows(pos, 1)
        c = ws.cell(row=pos, column=1)
        c.value = str(cat_val)
        c.font = Font(bold=True, size=14)
        c.fill = ORDER_GROUP_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")


def format_order_sheet(ws):
    if not is_order_sheet_name(ws.title):
        return False

    detail_title_row = None
    for row_idx in range(1, ws.max_row + 1):
        if _safe_display_text(ws.cell(row=row_idx, column=1).value) == ORDER_DETAIL_TITLE:
            detail_title_row = row_idx
            break
    if detail_title_row is None:
        return False

    summary_header_row = 2
    detail_header_row = detail_title_row + 1

    _style_title_row(ws, 1)
    _style_title_row(ws, detail_title_row)
    _style_header_row(ws, summary_header_row)
    _style_header_row(ws, detail_header_row)
    _group_detail_rows_by_category(ws, detail_header_row)
    _auto_fit_sheet_columns(ws)

    ws.freeze_panes = None

    summary_header_map = {
        normalize_text(cell.value): idx
        for idx, cell in enumerate(ws[summary_header_row], start=1)
        if cell.value
    }
    detail_header_map = {
        normalize_text(cell.value): idx
        for idx, cell in enumerate(ws[detail_header_row], start=1)
        if cell.value
    }

    for summary_column_name, minimum_width in (
        ("need summary", 28),
        ("products to order", 48),
        ("vendors", 28),
    ):
        col_idx = summary_header_map.get(summary_column_name)
        if not col_idx:
            continue
        ws.column_dimensions[get_column_letter(col_idx)].width = max(
            minimum_width,
            ws.column_dimensions[get_column_letter(col_idx)].width or minimum_width,
        )
        for row_idx in range(summary_header_row + 1, detail_title_row):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    units_needed_idx = summary_header_map.get("units needed")
    if units_needed_idx:
        for row_idx in range(summary_header_row + 1, detail_title_row):
            ws.cell(row=row_idx, column=units_needed_idx).font = Font(bold=True)

    product_name_idx = detail_header_map.get("product name") or detail_header_map.get("product")
    if product_name_idx:
        ws.column_dimensions[get_column_letter(product_name_idx)].width = max(
            42,
            min(ws.column_dimensions[get_column_letter(product_name_idx)].width or 42, 52),
        )
        for row_idx in range(detail_header_row + 1, ws.max_row + 1):
            ws.cell(row=row_idx, column=product_name_idx).alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

    notes_idx = detail_header_map.get("reorder notes")
    if notes_idx:
        ws.column_dimensions[get_column_letter(notes_idx)].width = max(
            28,
            min(ws.column_dimensions[get_column_letter(notes_idx)].width or 28, 34),
        )

    estimated_cost_idx = next(
        (
            idx
            for name, idx in summary_header_map.items()
            if name.startswith("estimated order cost")
        ),
        None,
    )
    if estimated_cost_idx:
        all_blank = True
        for row_idx in range(summary_header_row + 1, detail_title_row):
            cell = ws.cell(row=row_idx, column=estimated_cost_idx)
            if cell.value not in (None, ""):
                cell.font = Font(bold=True)
                cell.number_format = "$#,##0.00"
                all_blank = False
        if all_blank:
            ws.column_dimensions[get_column_letter(estimated_cost_idx)].hidden = True

    unit_cost_idx = summary_header_map.get("unit cost")
    if unit_cost_idx:
        for row_idx in range(summary_header_row + 1, detail_title_row):
            cell = ws.cell(row=row_idx, column=unit_cost_idx)
            if cell.value not in (None, ""):
                cell.number_format = "$#,##0.00"

    detail_suggested_idx = next(
        (
            idx
            for name, idx in detail_header_map.items()
            if name.startswith("suggested order qty")
        ),
        None,
    )
    if detail_suggested_idx:
        for row_idx in range(detail_header_row + 1, ws.max_row + 1):
            ws.cell(row=row_idx, column=detail_suggested_idx).font = Font(bold=True)

    priority_idx = detail_header_map.get("reorder priority")
    if priority_idx:
        for row_idx in range(detail_header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=priority_idx)
            color = ORDER_PRIORITY_COLORS.get(_safe_display_text(cell.value))
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    return True


def build_brand_order_sections(order_reports_dir, brand_aliases, store_code=None):
    sections = OrderedDict()
    discovered = discover_order_report_files(order_reports_dir)
    wanted_store = canonical_store_code(store_code) if store_code else ""

    for days in ORDER_REPORT_WINDOWS:
        store_map = discovered.get(days, {})
        if wanted_store:
            store_items = [(wanted_store, store_map[wanted_store])] if wanted_store in store_map else []
        else:
            store_items = list(store_map.items())

        if not store_items:
            continue

        detail_frames = []
        include_store_column = len(store_items) > 1 and not wanted_store
        for store, path in store_items:
            try:
                df = load_order_report_table(path)
            except Exception as exc:
                print(f"[WARN] Could not read order report '{path}': {exc}")
                continue

            brand_rows = _filter_brand_rows(df, brand_aliases)
            if brand_rows.empty:
                continue

            brand_rows = _drop_sample_and_promo_rows(brand_rows)
            brand_rows = _drop_low_cost_rows(brand_rows)
            if brand_rows.empty:
                continue

            brand_rows = _sort_order_rows(brand_rows)
            brand_rows = prepare_reorder_sheet(brand_rows, days)
            if include_store_column:
                brand_rows.insert(0, "Store", store)
            detail_frames.append(brand_rows)

        if not detail_frames:
            continue

        detail_df = (
            pd.concat(detail_frames, ignore_index=True)
            if len(detail_frames) > 1
            else detail_frames[0]
        )
        sections[f"Order_{days}d"] = {
            "summary": build_grouped_order_summary(detail_df, days),
            "detail": detail_df,
        }

    return sections
